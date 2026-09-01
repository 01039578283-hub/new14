from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import random
import re
import shutil
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from urllib.parse import quote, unquote, urlsplit

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
COMMON = ROOT.parent / "참고자료" / "공통자료"
SOURCE_DIR = Path.home() / "Desktop" / "구글시트로 뽑은거"
CENTER_CSV = COMMON / "센터정보 정리.csv"
IMAGE_CSV = COMMON / "이미지링크.csv"
REP_SOURCE = COMMON / "대표이미지"
REP_TARGET = ROOT / "assets" / "representative"
MAP_DIR = ROOT / "assets" / "maps"
TARGET_ROOT = ROOT / "과목별학원"
ORIGIN = "https://xn--3e0bz50b1zcyxat54c.com"
SITE_NAME = "와와학습코칭센터"
TITLE_SUFFIX = "와와학습코칭센터 영어수학 전문학원"
PHONE = "010-6839-8283"
CONTENT_DATE = "2026-08-15"
EXPECTED_ROWS = 371
HIGH_ENGLISH_PARAGRAPH_DF_LIMIT = 30
REGION_ORDER = ("서울", "경기", "인천", "충청", "대전", "대구", "울산", "부산", "경상", "광주", "전라", "강원", "제주")


@dataclass(frozen=True)
class CategoryConfig:
    slug: str
    label: str
    workbook: str
    level: str
    grade_prefix: str
    school_name: str
    school_columns: tuple[str, ...]
    subjects: tuple[str, ...]
    english: str
    focus: str
    process: str
    rep_prefix: str
    source_mode: str = "single"

    @property
    def source_path(self) -> Path:
        return SOURCE_DIR / self.workbook

    @property
    def is_elementary(self) -> bool:
        return self.grade_prefix == "초"


CONFIGS: tuple[CategoryConfig, ...] = (
    CategoryConfig("영수학원", "영수학원", "영수학원 원고.xlsx", "초·중·고", "", "학교", ("타깃학교\n(초)", "타깃학교\n(중)", "타깃학교\n(고)"), ("영어", "수학"), "ENGLISH · MATH", "과목별 병목·주간 균형", "두 과목 실행 점검", "ys", "combined"),
    CategoryConfig("초등영어학원", "초등 영어학원", "초등 영어학원.xlsx", "초등", "초", "초등학교", ("타깃학교\n(초)",), ("영어",), "ELEMENTARY ENGLISH", "어휘·문장 읽기·기초 표현", "교과 이해와 습관", "es-eng"),
    CategoryConfig("초등수학학원", "초등 수학학원", "초등 수학학원.xlsx", "초등", "초", "초등학교", ("타깃학교\n(초)",), ("수학",), "ELEMENTARY MATH", "연산 원리·개념 설명·문장제", "풀이 과정 점검", "es-math"),
    CategoryConfig("중등영어학원", "중등 영어학원", "중등 영어학원.xlsx", "중등", "중", "중학교", ("타깃학교\n(중)",), ("영어",), "MIDDLE ENGLISH", "어휘·문법 적용·독해 근거", "내신 자료 연결", "ms-eng"),
    CategoryConfig("중등수학학원", "중등 수학학원", "중등 수학학원.xlsx", "중등", "중", "중학교", ("타깃학교\n(중)",), ("수학",), "MIDDLE MATH", "개념 연결·조건 해석·오답", "내신 풀이 점검", "ms-math"),
    CategoryConfig("고등영어학원", "고등 영어학원", "고등 영어학원.xlsx", "고등", "고", "고등학교", ("타깃학교\n(고)",), ("영어",), "HIGH SCHOOL ENGLISH", "어휘 누적·구문·독해 근거", "내신·모의고사 연결", "hs-eng"),
    CategoryConfig("고등수학학원", "고등 수학학원", "고등 수학학원.xlsx", "고등", "고", "고등학교", ("타깃학교\n(고)",), ("수학",), "HIGH SCHOOL MATH", "개념 통합·조건 해석·서술형", "내신·모의고사 오답", "hs-math"),
)

CONFIG_BY_SLUG = {config.slug: config for config in CONFIGS}
ALL_CATEGORY_SLUGS = (
    "고등학생국영수학원", "중학생국영수학원", "초등학생국영수학원",
    *(config.slug for config in CONFIGS),
)
ALL_CATEGORY_LABELS = {
    "고등학생국영수학원": "고등학생 국영수학원",
    "중학생국영수학원": "중학생 국영수학원",
    "초등학생국영수학원": "초등학생 국영수학원",
    **{config.slug: config.label for config in CONFIGS},
}


@dataclass(frozen=True)
class TopicSignal:
    code: str
    label: str
    keywords: tuple[str, ...]
    check: str
    evidence: str
    practice: str
    home_action: str


@dataclass(frozen=True)
class HighEnglishIntent:
    code: str
    label: str
    keywords: tuple[str, ...]
    concern: str
    evidence: str
    action: str
    checkpoint: str
    exam_use: str
    consult_question: str


@dataclass(frozen=True)
class HighEnglishProfile:
    focus: str
    source_title: str
    intents: tuple[HighEnglishIntent, ...]
    source_markers: tuple[str, ...]


@dataclass(frozen=True)
class MiddleEnglishProfile:
    focus: str
    source_title: str
    intents: tuple[HighEnglishIntent, ...]
    source_markers: tuple[str, ...]


@dataclass(frozen=True)
class HighMathIntent:
    code: str
    label: str
    keywords: tuple[str, ...]
    concern: str
    evidence: str
    action: str
    checkpoint: str
    exam_use: str
    consult_question: str


@dataclass(frozen=True)
class HighMathProfile:
    focus: str
    source_title: str
    intents: tuple[HighMathIntent, ...]
    source_markers: tuple[str, ...]


ENGLISH_SIGNALS: tuple[TopicSignal, ...] = (
    TopicSignal("vocabulary", "어휘 회상", ("어휘", "단어", "암기"), "뜻을 외운 단어를 문장 안에서 다시 알아보는지", "단어 시험과 독해 지문에서 같은 어휘를 놓친 기록", "뜻·품사·예문을 한 묶음으로 확인하기", "하루 뒤 예문 속 단어 뜻을 다시 말해 보기"),
    TopicSignal("sentence", "문장 구조", ("구문", "문장", "해석", "구조"), "주어와 동사, 수식 범위를 구분해 문장을 읽는지", "긴 문장에서 해석이 끊긴 위치와 표시한 문장 성분", "문장 뼈대를 먼저 적고 수식어를 붙여 읽기", "한 문장을 짧게 끊어 소리 내어 설명하기"),
    TopicSignal("grammar", "문법 적용", ("문법", "어법", "시제", "관계사"), "규칙을 아는 문제와 실제 문장에 적용하지 못한 문제를 나눴는지", "고른 답의 근거를 문법 규칙으로 설명한 흔적", "규칙 한 줄 뒤에 맞는 예문과 틀린 예문을 함께 적기", "틀린 문장을 고치고 이유를 한 문장으로 남기기"),
    TopicSignal("reading", "독해 근거", ("독해", "지문", "주제", "근거"), "답을 고른 문장과 선택지의 표현을 연결하는지", "정답 근거에 밑줄을 긋고 선택지를 지운 이유를 적은 기록", "문단별 핵심 문장과 연결어를 표시하며 읽기", "짧은 지문 한 편에서 근거 문장만 다시 찾기"),
    TopicSignal("listening", "듣기 확인", ("듣기", "음원", "발음"), "놓친 소리와 뜻을 모르는 표현을 구분했는지", "받아쓰기와 다시 들은 뒤 수정한 부분", "짧은 구간을 듣고 핵심 표현을 따라 말하기", "한 문장을 듣고 의미 단위로 끊어 적기"),
    TopicSignal("writing", "서술형 쓰기", ("서술형", "쓰기", "영작", "작문"), "조건에 맞는 문장을 스스로 구성하고 검토하는지", "서술형 답안에서 빠진 조건과 고친 문장", "핵심 표현을 넣어 짧은 답안을 완성하기", "답안을 다시 써 보고 빠진 조건을 표시하기"),
    TopicSignal("error", "영어 오답 유형", ("오답", "틀린", "실수", "재시험"), "어휘·문법·독해 중 어디에서 같은 실수가 이어지는지", "첫 풀이와 다시 푼 답 사이에 달라진 근거", "오답을 유형별로 나누고 다시 볼 날짜 정하기", "같은 유형 한 문제를 며칠 뒤 다시 풀기"),
    TopicSignal("pace", "시험 시간 배분", ("시간", "시험", "속도", "시간배분"), "어느 지문과 문항에서 시간이 길어지는지", "문항별 소요 시간과 끝까지 풀지 못한 구간", "읽기와 답 확인 시간을 나누어 연습하기", "짧은 세트를 정해진 시간 안에 풀고 기록하기"),
    TopicSignal("routine", "영어 학습 루틴", ("습관", "플래너", "기록", "과제"), "계획한 분량과 실제 완료한 분량이 구분되는지", "학습 시작 시각과 완료 여부를 적은 주간 기록", "매일 할 최소 분량과 다시 볼 항목을 나누기", "완료한 분량과 남은 질문을 짧게 적기"),
)


HIGH_ENGLISH_INTENTS: tuple[HighEnglishIntent, ...] = (
    HighEnglishIntent(
        "vocabulary", "문맥 어휘", ("어휘", "단어", "품사", "유의어", "반의어", "암기"),
        "외운 단어를 지문에서 만났을 때 품사와 문맥 의미를 바로 떠올리는지",
        "단어 시험 결과와 같은 단어가 나온 지문에서 멈춘 위치",
        "뜻·품사·문장 속 쓰임을 한 묶음으로 적고 다음 날 예문에서 다시 찾기",
        "처음 본 문장에서도 해당 어휘의 의미를 근거와 함께 설명하는지",
        "내신 본문 어휘와 모의고사 빈출 표현을 분리해 누적하는 기준",
        "암기량이 아니라 지문 안에서 다시 알아보는 비율을 어떻게 확인하는지",
    ),
    HighEnglishIntent(
        "sentence", "문장 구조", ("구문", "문장 구조", "문장", "해석", "수식", "주어", "동사"),
        "단어 뜻은 알지만 긴 문장의 주어·동사와 수식 범위를 놓치는지",
        "해석이 끊긴 문장에 표시한 주어·동사와 잘못 묶은 수식어",
        "문장 뼈대를 먼저 표시한 뒤 수식어를 붙여 의미 단위로 다시 읽기",
        "비슷한 구조의 새 문장을 도움 없이 끊어 읽고 설명하는지",
        "교과서 변형 문장과 모의고사 장문을 같은 구조 표시법으로 연결하는 기준",
        "해석 정답보다 문장 구조를 스스로 표시하는 과정을 어떻게 피드백하는지",
    ),
    HighEnglishIntent(
        "grammar", "문법·어법 적용", ("문법", "어법", "관계사", "시제", "수동태", "품사"),
        "문법 규칙은 말하지만 실제 문장과 선택지에서 적용 근거를 찾지 못하는지",
        "오답 선택지에 적은 규칙과 서술형 문장에서 스스로 고친 부분",
        "규칙 한 줄, 맞는 예문, 틀린 예문을 나란히 놓고 차이를 설명하기",
        "형태가 달라진 문장에서도 같은 규칙을 찾아 수정 이유를 말하는지",
        "내신 어법·서술형과 모의고사 어법 문항의 요구 차이를 나누는 기준",
        "암기 확인과 문장 적용 확인을 어떤 자료로 구분하는지",
    ),
    HighEnglishIntent(
        "reading", "근거 독해", ("독해", "지문", "주제", "요지", "근거", "빈칸", "순서", "삽입", "흐름"),
        "해석한 문장을 글의 주장과 연결하고 선택지 판단 근거를 찾는지",
        "문단별 핵심 문장, 연결어 표시와 선택지를 지운 이유",
        "문단 역할을 한 줄로 요약하고 답을 지지하는 문장을 선택지와 연결하기",
        "새 지문에서도 정답과 오답 선택지의 근거를 각각 짚는지",
        "내신 본문 암기와 모의고사 처음 보는 지문 독해를 구분하는 기준",
        "정답 수가 아니라 근거 문장을 찾는 과정을 어떻게 기록하는지",
    ),
    HighEnglishIntent(
        "pace", "장문 독해·시간 배분", ("시간", "속도", "장문", "집중", "끝까지", "긴 지문", "시험 시간"),
        "정확도 문제와 읽는 속도 문제를 구분하고 어느 문항에서 시간이 길어지는지",
        "문항별 소요 시간, 다시 읽은 문장과 끝까지 풀지 못한 구간",
        "짧은 세트에서 읽기·판단·검토 시간을 따로 기록하고 병목 구간만 재연습하기",
        "정확도를 유지한 채 같은 유형의 소요 시간이 줄었는지",
        "내신의 꼼꼼한 본문 확인과 모의고사의 제한 시간 판단을 따로 연습하는 기준",
        "빠르게 풀기 전에 정확도와 소요 시간을 어떤 순서로 확인하는지",
    ),
    HighEnglishIntent(
        "school_exam", "내신 본문·서술형", ("내신", "교과서", "본문", "학교 시험", "중간고사", "기말고사", "변형", "단원평가"),
        "학교 범위의 본문 이해, 어법 변형과 서술형 조건을 따로 준비하는지",
        "시험 범위표, 교과서 표시, 학교 자료와 최근 서술형 답안",
        "범위표를 기준으로 본문 이해·어법 변형·서술형 표현을 세 칸으로 나누기",
        "표현이 바뀐 문제에서도 본문 근거와 문법 이유를 설명하는지",
        "학교별 범위와 출제 자료를 먼저 확인하고 일반 독해 연습과 분리하는 기준",
        "학교 자료를 수업 계획에 반영하는 시점과 재시험 기준이 무엇인지",
    ),
    HighEnglishIntent(
        "mock_exam", "모의고사 오답", ("모의고사", "모고", "등급", "성적 기복", "전국연합", "오답 분석"),
        "모의고사 등급만 보지 않고 어휘·구문·유형·시간 중 원인을 나누는지",
        "최근 모의고사 시험지, 문항별 소요 시간과 처음 고른 오답 근거",
        "오답을 지식 부족과 판단 과정으로 나누고 같은 유형을 며칠 뒤 다시 풀기",
        "새 지문에서 같은 오답 유형이 줄고 근거 설명이 달라졌는지",
        "내신 범위 학습과 모의고사 누적 약점 보완의 주간 비중을 조정하는 기준",
        "점수 변화가 없을 때 문제 수보다 먼저 바꾸는 학습 절차가 무엇인지",
    ),
    HighEnglishIntent(
        "suneung", "수능·기출", ("수능", "기출", "입시", "고3", "입시로드맵", "입시 준비"),
        "수능 기출을 풀고도 출제 의도와 반복되는 유형별 약점을 남기는지",
        "최근 기출의 근거 표시, 유형별 정답률과 시간 기록",
        "기출 한 세트를 유형별로 분류하고 틀린 판단을 새 지문에 다시 적용하기",
        "회차가 달라도 같은 유형의 접근 순서와 근거가 안정되는지",
        "학년·목표 시점에 맞춰 내신 기간과 수능 누적 학습의 비중을 바꾸는 기준",
        "기출 진도보다 분석과 재풀이 완료 여부를 어떻게 확인하는지",
    ),
    HighEnglishIntent(
        "writing", "서술형·영작", ("서술형", "영작", "작문", "쓰기", "요약문", "첨삭", "문장력"),
        "답의 핵심 내용은 알지만 조건에 맞는 영어 문장을 완성하고 고치지 못하는지",
        "서술형 답안의 누락 조건, 어순·시제 오류와 첨삭 뒤 다시 쓴 문장",
        "필수 표현과 조건을 먼저 표시하고 초안·수정안의 차이를 스스로 설명하기",
        "유사한 질문에서도 조건을 빠뜨리지 않고 문장을 다시 구성하는지",
        "내신 서술형의 지정 조건과 모의고사 요약문에서 요구하는 핵심어를 구분하는 기준",
        "첨삭을 받은 뒤 학생이 스스로 다시 쓰고 이유를 설명하는 단계가 있는지",
    ),
    HighEnglishIntent(
        "performance", "수행평가 준비", ("수행평가", "과제", "발표 평가", "평가 기준", "루브릭"),
        "수행평가의 주제·분량·평가 조건을 읽고 준비 단계를 나누는지",
        "학교 안내문, 평가 기준표와 초안에서 빠진 조건",
        "마감일부터 역산해 자료 찾기·초안·수정·연습 날짜를 따로 적기",
        "제출 전 평가 조건을 학생이 직접 대조하고 빠진 항목을 고치는지",
        "시험 공부와 수행평가 준비 시간을 학교 일정에 맞춰 분리하는 기준",
        "결과물을 대신 만드는 것이 아니라 조건 확인과 수정 과정을 어떻게 돕는지",
    ),
    HighEnglishIntent(
        "listening", "듣기 평가", ("듣기", "음원", "받아쓰기", "발음", "들리는"),
        "소리를 놓친 문제와 표현의 뜻을 몰라 틀린 문제를 구분하는지",
        "오답 음원에서 놓친 구간, 받아쓰기와 다시 들은 뒤 수정한 부분",
        "짧은 구간을 의미 단위로 받아쓰고 핵심 표현을 따라 말한 뒤 다시 듣기",
        "새 음원에서도 같은 연결 발음과 핵심 표현을 알아듣는지",
        "학교 듣기 평가와 모의고사 듣기의 오답 원인을 같은 기록표로 비교하는 기준",
        "음원을 반복 재생하는 횟수보다 놓친 이유를 어떻게 분류하는지",
    ),
    HighEnglishIntent(
        "speaking", "발표·말하기", ("발표", "말하기", "질문하고 답", "표현력", "스피킹"),
        "알고 있는 표현을 발표 조건에 맞춰 말하고 질문에 이어 답할 수 있는지",
        "발표 대본, 제한 시간 기록과 말하다 멈춘 문장",
        "핵심 문장을 짧게 만들고 녹음 뒤 발음·속도·누락 내용을 스스로 표시하기",
        "대본을 그대로 읽지 않고 핵심 순서에 맞춰 다시 설명하는지",
        "학교 발표 평가의 조건과 일반 회화 연습의 목적을 구분하는 기준",
        "발표 연습 여부와 피드백 방식은 실제 개설 범위에서 어떻게 확인하는지",
    ),
    HighEnglishIntent(
        "transition", "고등 과정 전환", ("예비고1", "고등 전환", "새 학년", "선행", "중3", "겨울방학", "방학"),
        "중학교식 암기에서 고등 지문 분석으로 넘어갈 준비가 되어 있는지",
        "최근 시험지, 고등 수준 예시 지문과 일주일 학습 기록",
        "어휘·문장 구조의 현재선을 확인한 뒤 고등 지문 한 편에 적용해 보기",
        "무리한 선행보다 새 난도의 문장을 스스로 읽는 범위가 넓어졌는지",
        "입학 전 기초 보완과 학교 일정 이후 내신 준비를 나누는 기준",
        "선행 범위보다 고등 과정에서 혼자 수행할 수 있는 단계를 어떻게 확인하는지",
    ),
    HighEnglishIntent(
        "routine", "학습 루틴", ("습관", "루틴", "계획", "피드백", "오답 관리", "학습 태도", "점검표", "포트폴리오", "진도"),
        "계획한 영어 학습이 실제 완료 기록과 오답 재확인으로 이어지는지",
        "주간 계획표, 시작·완료 시각과 다음 날 남은 질문",
        "매일 할 최소 분량과 다시 볼 항목을 나누고 완료 근거를 한 줄로 남기기",
        "한 주 뒤 미완료 이유와 반복 오답을 바탕으로 분량을 조정하는지",
        "내신 기간의 집중 과제와 평소 누적 학습을 달력에서 분리하는 기준",
        "과제량보다 완료·재확인·계획 수정의 피드백 주기를 어떻게 확인하는지",
    ),
)
HIGH_ENGLISH_INTENT_BY_CODE = {intent.code: intent for intent in HIGH_ENGLISH_INTENTS}


MIDDLE_ENGLISH_INTENTS: tuple[HighEnglishIntent, ...] = (
    HighEnglishIntent(
        "vocabulary", "문맥 어휘", ("어휘", "단어", "품사", "유의어", "반의어", "암기"),
        "외운 단어를 교과서와 새 지문에서 만났을 때 품사와 문맥 의미를 바로 떠올리는지",
        "단어 시험 결과와 같은 단어가 나온 지문에서 멈춘 위치",
        "뜻·품사·문장 속 쓰임을 한 묶음으로 적고 다음 날 예문에서 다시 찾기",
        "처음 보는 문장에서도 해당 어휘의 의미를 근거와 함께 설명하는지",
        "학교 본문 어휘와 평소 독해에서 다시 만난 표현을 연결해 누적하는 기준",
        "암기량이 아니라 지문 안에서 다시 알아보는 비율을 어떻게 확인하는지",
    ),
    HighEnglishIntent(
        "sentence", "문장 구조", ("구문", "문장 구조", "문장", "해석", "수식", "주어", "동사"),
        "단어 뜻은 알지만 문장의 주어·동사와 수식 범위를 놓쳐 해석이 끊기는지",
        "해석이 끊긴 문장에 표시한 주어·동사와 잘못 묶은 수식어",
        "문장 뼈대를 먼저 표시한 뒤 수식어를 붙여 의미 단위로 다시 읽기",
        "비슷한 구조의 새 문장을 도움 없이 끊어 읽고 설명하는지",
        "교과서 변형 문장과 처음 보는 장문을 같은 구조 표시법으로 연결하는 기준",
        "해석 정답보다 문장 구조를 스스로 표시하는 과정을 어떻게 피드백하는지",
    ),
    HighEnglishIntent(
        "grammar", "문법·어법 적용", ("문법", "어법", "관계사", "시제", "수동태", "품사"),
        "문법 규칙은 말하지만 실제 문장과 선택지에서 적용 근거를 찾지 못하는지",
        "오답 선택지에 적은 규칙과 서술형 문장에서 스스로 고친 부분",
        "규칙 한 줄, 맞는 예문, 틀린 예문을 나란히 놓고 차이를 설명하기",
        "형태가 달라진 문장에서도 같은 규칙을 찾아 수정 이유를 말하는지",
        "학교 어법·서술형과 누적 문법 연습의 요구 차이를 나누는 기준",
        "암기 확인과 문장 적용 확인을 어떤 자료로 구분하는지",
    ),
    HighEnglishIntent(
        "reading", "근거 독해", ("독해", "지문", "주제", "요지", "근거", "빈칸", "순서", "삽입", "흐름"),
        "해석한 문장을 글의 주장과 연결하고 선택지 판단 근거를 찾는지",
        "문단별 핵심 문장, 연결어 표시와 선택지를 지운 이유",
        "문단 역할을 한 줄로 요약하고 답을 지지하는 문장을 선택지와 연결하기",
        "새 지문에서도 정답과 오답 선택지의 근거를 각각 짚는지",
        "학교 본문 이해와 처음 보는 지문의 근거 찾기를 구분해 연습하는 기준",
        "정답 수가 아니라 근거 문장을 찾는 과정을 어떻게 기록하는지",
    ),
    HighEnglishIntent(
        "school_exam", "학교 내신·서술형", ("내신", "교과서", "본문", "학교 시험", "중간고사", "기말고사", "변형", "단원평가"),
        "학교 범위의 본문 이해, 어법 변형과 서술형 조건을 따로 준비하는지",
        "시험 범위표, 교과서 표시, 학교 자료와 최근 서술형 답안",
        "범위표를 기준으로 본문 이해·어법 변형·서술형 표현을 세 칸으로 나누기",
        "표현이 바뀐 문제에서도 본문 근거와 문법 이유를 설명하는지",
        "학교별 범위와 출제 자료를 먼저 확인하고 평소 독해 연습과 분리하는 기준",
        "학교 자료를 학습 계획에 반영하는 시점과 재확인 기준이 무엇인지",
    ),
    HighEnglishIntent(
        "writing", "서술형·영작", ("서술형", "영작", "작문", "쓰기", "요약문", "첨삭", "문장력", "수행평가"),
        "답의 핵심 내용은 알지만 조건에 맞는 영어 문장을 완성하고 고치지 못하는지",
        "서술형 답안의 누락 조건, 어순·시제 오류와 수정 뒤 다시 쓴 문장",
        "필수 표현과 조건을 먼저 표시하고 초안·수정안의 차이를 스스로 설명하기",
        "유사한 질문에서도 조건을 빠뜨리지 않고 문장을 다시 구성하는지",
        "학교 서술형과 수행평가의 지정 조건을 답안에 빠짐없이 반영하는 기준",
        "수정을 받은 뒤 학생이 스스로 다시 쓰고 이유를 설명하는 단계가 있는지",
    ),
    HighEnglishIntent(
        "listening", "듣기·말하기", ("듣기", "음원", "받아쓰기", "발음", "말하기", "발표", "스피킹"),
        "소리를 놓친 문제와 표현의 뜻을 몰라 틀린 문제를 구분하고 직접 말할 수 있는지",
        "오답 음원에서 놓친 구간, 받아쓰기와 다시 들은 뒤 수정한 부분",
        "짧은 구간을 의미 단위로 받아쓰고 핵심 표현을 따라 말한 뒤 다시 듣기",
        "새 음원에서도 같은 연결 발음과 핵심 표현을 알아듣고 설명하는지",
        "학교 듣기·발표 평가의 조건과 평소 듣고 말하기 연습을 구분하는 기준",
        "반복 재생 횟수보다 놓친 이유와 말하다 멈춘 위치를 어떻게 분류하는지",
    ),
    HighEnglishIntent(
        "error", "오답 원인", ("오답", "틀린", "실수", "재시험", "재풀이", "성적 기복"),
        "같은 오답을 지식 부족·문장 해석·선택지 판단·시간 문제로 나누는지",
        "처음 고른 답의 근거와 해설 뒤 바뀐 판단, 며칠 뒤 재풀이 기록",
        "오답 원인을 한 줄로 적고 같은 유형의 새 문제를 며칠 뒤 다시 풀기",
        "답을 외운 것이 아니라 새 문제에서도 판단 순서와 근거가 달라졌는지",
        "학교 범위 오답과 평소 독해 오답의 재확인 날짜를 따로 관리하는 기준",
        "점수 변화보다 먼저 바꿔야 할 풀이 절차를 어떤 기록으로 찾는지",
    ),
    HighEnglishIntent(
        "routine", "복습·과제 루틴", ("습관", "루틴", "계획", "피드백", "오답 관리", "학습 태도", "점검표", "포트폴리오", "진도", "과제"),
        "계획한 영어 학습이 실제 완료 기록과 오답 재확인으로 이어지는지",
        "주간 계획표, 시작·완료 시각과 다음 날 남은 질문",
        "매일 할 최소 분량과 다시 볼 항목을 나누고 완료 근거를 한 줄로 남기기",
        "한 주 뒤 미완료 이유와 반복 오답을 바탕으로 분량을 조정하는지",
        "시험 기간의 집중 과제와 평소 누적 학습을 달력에서 분리하는 기준",
        "과제량보다 완료·재확인·계획 수정의 피드백 주기를 어떻게 확인하는지",
    ),
    HighEnglishIntent(
        "pace", "시험 시간 배분", ("시간", "속도", "장문", "집중", "끝까지", "긴 지문", "시험 시간", "시간배분"),
        "정확도 문제와 읽는 속도 문제를 구분하고 어느 문항에서 시간이 길어지는지",
        "문항별 소요 시간, 다시 읽은 문장과 끝까지 풀지 못한 구간",
        "짧은 세트에서 읽기·판단·검토 시간을 따로 기록하고 병목 구간만 재연습하기",
        "정확도를 유지한 채 같은 유형의 소요 시간이 줄었는지",
        "학교 시험의 꼼꼼한 범위 확인과 처음 보는 지문의 시간 판단을 따로 연습하는 기준",
        "빠르게 풀기 전에 정확도와 소요 시간을 어떤 순서로 확인하는지",
    ),
    HighEnglishIntent(
        "entry", "중1 첫 시험 적응", ("예비중", "예비 중", "중1", "중학교 첫", "입학", "첫 시험"),
        "초등 영어 학습 방식에서 중학교 시험 범위와 서술형 준비로 넘어갈 준비가 되어 있는지",
        "현재 교재, 중학교 예시 범위표와 문법·서술형 문제의 첫 풀이",
        "어휘·문장 구조의 현재선을 확인한 뒤 짧은 학교 시험형 문제에 적용해 보기",
        "무리한 선행보다 중학교 영어 문제를 스스로 읽고 근거를 설명하는 범위가 넓어졌는지",
        "입학 전 기초 보완과 첫 시험 범위 학습을 나누는 기준",
        "선행 범위보다 중1 학습에서 혼자 수행할 수 있는 단계를 어떻게 확인하는지",
    ),
    HighEnglishIntent(
        "transition", "고등 영어 전환 준비", ("예비고1", "예비 고1", "고등 전환", "새 학년", "선행", "중3", "겨울방학", "방학"),
        "중학교식 암기에서 고등 지문 분석으로 넘어갈 영어 기초가 연결되어 있는지",
        "최근 시험지, 고등 수준 예시 지문과 일주일 학습 기록",
        "어휘·문장 구조의 현재선을 확인한 뒤 고등 지문 한 편에 적용해 보기",
        "무리한 선행보다 새 난도의 문장을 스스로 읽는 범위가 넓어졌는지",
        "중3 학교 일정과 입학 전 기초 보완을 나누는 기준",
        "선행 범위보다 고등 영어에서 혼자 수행할 수 있는 단계를 어떻게 확인하는지",
    ),
    HighEnglishIntent(
        "diagnosis", "학습 우선순위", ("상담", "진단", "현재 실력", "우선순위", "방향", "로드맵", "기초", "시작", "점검"),
        "점수 한 줄이 아니라 어휘·문장·문법·독해 중 먼저 막히는 지점을 자료로 구분하는지",
        "최근 시험지와 교재에서 학생 설명이 멈춘 위치, 다시 풀며 바꾼 근거",
        "가장 먼저 막힌 영역 하나를 고르고 설명·연습·재확인의 순서를 정하기",
        "일주일 뒤 같은 영역의 새 문제에서 학생이 혼자 근거를 설명하는지",
        "학교 범위와 평소 누적 학습에서 같은 병목이 반복되는지 대조하는 기준",
        "교재나 분량보다 먼저 바꿀 학습 절차를 어떤 자료로 판단하는지",
    ),
)
MIDDLE_ENGLISH_INTENT_BY_CODE = {intent.code: intent for intent in MIDDLE_ENGLISH_INTENTS}

HIGH_MATH_INTENTS: tuple[HighMathIntent, ...] = (
    HighMathIntent(
        "concept", "개념 연결", ("개념", "정의", "공식", "원리", "이해", "기본기"),
        "공식의 이름은 알지만 적용 조건과 개념 사이의 관계를 자기 말로 설명하지 못하는지",
        "개념 설명 메모와 대표 문제에서 해당 공식을 선택한 이유",
        "정의·조건·대표식을 한 줄씩 연결한 뒤 조건이 바뀐 문제에 같은 개념을 적용하기",
        "풀이를 보지 않고 개념의 적용 조건과 사용 이유를 설명하는지",
        "내신 범위의 필수 개념과 모의고사에서 다시 등장한 선수 개념을 구분해 연결하는 기준",
        "공식 암기와 개념 이해를 어떤 질문과 풀이 기록으로 나누어 확인하는지",
    ),
    HighMathIntent(
        "calculation", "계산 정확도", ("계산", "연산", "부호", "괄호", "실수", "정확도", "검산"),
        "풀이 방향은 맞지만 부호·괄호·전개·약분 단계에서 같은 계산 오류가 반복되는지",
        "첫 풀이의 중간 계산과 답을 고친 뒤 다시 쓴 계산 과정",
        "계산 단계를 한 줄씩 분리하고 오류가 시작된 위치와 검산 방법을 함께 표시하기",
        "숫자와 식이 바뀐 새 문제에서도 같은 계산 오류 없이 끝까지 수행하는지",
        "내신의 제한 시간과 모의고사의 문항 선택 안에서 계산·검산 시간을 따로 확보하는 기준",
        "단순 실수로 넘기지 않고 반복되는 계산 오류의 시작 단계를 어떻게 기록하는지",
    ),
    HighMathIntent(
        "condition", "조건 해석·식 세우기", ("조건", "식", "문장제", "응용", "문제해결", "해석", "모델링"),
        "문제에서 주어진 조건과 구할 값을 구분해 식·그림·표로 바꾸지 못하는지",
        "밑줄 친 조건, 처음 세운 식과 빠뜨리거나 잘못 연결한 조건",
        "조건을 주어진 값·관계·구할 값으로 나눈 뒤 식이나 그림으로 바꾸고 이유를 설명하기",
        "표현이 달라진 새 문제에서도 필요한 조건만 골라 식을 세우는지",
        "내신 서술형의 조건 누락과 모의고사 복합 문항의 조건 연결을 따로 확인하는 기준",
        "정답 풀이를 보여 주기 전에 학생이 조건을 식으로 바꾸는 과정을 어떻게 점검하는지",
    ),
    HighMathIntent(
        "error", "오답 원인·재풀이", ("오답", "틀린", "재풀이", "복습", "오답노트", "틀린 이유"),
        "틀린 문제를 다시 풀어도 개념·조건·계산 중 원인을 구분하지 않아 같은 실수가 남는지",
        "첫 풀이, 오류 원인 표시와 해설을 덮고 다시 푼 풀이의 차이",
        "오답을 개념·조건·계산·시간으로 분류하고 같은 날 수정과 며칠 뒤 새 문제 확인을 나누기",
        "정답을 기억하지 않은 상태에서 비슷한 문제의 풀이 순서를 다시 설명하는지",
        "내신 범위 오답과 모의고사 누적 오답의 재확인 날짜를 다른 일정으로 관리하는 기준",
        "오답노트의 분량보다 원인 분류와 새 문제 재확인을 어떻게 연결하는지",
    ),
    HighMathIntent(
        "written", "서술형 풀이", ("서술형", "풀이 과정", "풀이과정", "과정", "논리", "답안", "설명"),
        "답은 구하지만 사용한 조건과 식의 이유를 채점 가능한 순서로 적지 못하는지",
        "서술형 초안에서 생략된 조건·등식·결론과 수정한 답안",
        "조건 표시·사용한 개념·계산·결론을 단계별로 적고 불필요한 문장을 줄이기",
        "유사 문제에서도 풀이 순서를 빠뜨리지 않고 각 식의 이유를 설명하는지",
        "학교별 서술형 채점 요소와 모의고사의 객관식 풀이 근거를 서로 다른 형식으로 남기는 기준",
        "정답뿐 아니라 서술형의 논리 순서와 누락 조건을 어떤 기준으로 첨삭하는지",
    ),
    HighMathIntent(
        "graph", "함수·그래프·도형 해석", ("함수", "그래프", "도형", "좌표", "표", "기하", "그림"),
        "식·그래프·표·도형이 같은 관계를 나타낸다는 점을 연결해 읽지 못하는지",
        "문제의 조건을 옮겨 그린 그림과 좌표·교점·범위를 잘못 표시한 위치",
        "주어진 관계를 식과 그림으로 각각 나타내고 두 표현에서 같은 조건을 가리키기",
        "표현 방식이 바뀐 문제에서도 그래프의 변화와 식의 조건을 연결하는지",
        "내신의 단원별 그래프 문제와 모의고사의 여러 개념이 섞인 시각 자료를 구분하는 기준",
        "그림을 대신 그려 주기보다 학생이 조건을 시각화하는 과정을 어떻게 확인하는지",
    ),
    HighMathIntent(
        "units", "선수 개념·단원 연결", ("단원", "누적", "선수", "연결", "취약 단원", "기초", "빈틈"),
        "현재 단원의 문제에서 이전 학년·앞 단원의 개념이 필요한 지점을 알아보지 못하는지",
        "최근 문제에서 다시 필요해진 선수 개념과 현재 단원 풀이가 끊긴 위치",
        "막힌 문제에 필요한 앞 개념만 짧게 복습하고 현재 문제로 돌아와 연결 이유를 적기",
        "단원명이 달라도 필요한 선수 개념을 스스로 찾아 현재 풀이에 적용하는지",
        "내신 범위의 진도와 수능형 누적 문제에 필요한 선수 개념 보완을 따로 배치하는 기준",
        "진도를 되돌릴 때 전체 복습이 아니라 필요한 선수 개념만 고르는 기준이 무엇인지",
    ),
    HighMathIntent(
        "pace", "시험 시간 배분", ("시간", "속도", "시간 배분", "시간배분", "시험 시간", "끝까지", "집중"),
        "쉬운 문제에서 계산 시간이 길어지거나 어려운 문제를 붙잡아 검토 시간을 잃는지",
        "문항별 시작·종료 시각, 건너뛴 문제와 다시 돌아온 순서",
        "문항을 바로 풀 문제·표시 후 돌아올 문제·마지막에 검산할 문제로 나누어 시간 제한 세트 풀기",
        "정확도를 유지하면서도 정한 시각에 문제를 넘기고 검산 시간을 확보하는지",
        "내신의 범위형 문항과 모의고사의 난도별 문항 선택 순서를 따로 연습하는 기준",
        "빠르게 푸는 연습보다 문항 선택과 검산 시간을 어떤 기록으로 조정하는지",
    ),
    HighMathIntent(
        "school_exam", "내신 범위·학교별 출제 유형", ("내신", "학교 시험", "중간고사", "기말고사", "시험 범위", "교과서", "학교별"),
        "학교 시험 범위 안에서 교과서·학교 자료·서술형 유형을 구분해 준비하는지",
        "시험 범위표, 교과서 예제·학교 프린트와 최근 서술형 답안",
        "범위별 개념·대표 유형·서술형·누적 오답을 네 칸으로 나누고 완료 기준 정하기",
        "숫자나 조건이 바뀐 학교 시험 문제에서도 개념과 풀이 이유를 설명하는지",
        "학교별 범위와 출제 자료를 우선 확인하고 일반 문제집 진도와 분리하는 기준",
        "학교 자료를 수업 계획에 반영하는 시점과 시험 전 재확인 기준이 무엇인지",
    ),
    HighMathIntent(
        "mock_exam", "모의고사 오답·문항 선택", ("모의고사", "모고", "전국연합", "등급", "성적 기복", "실전"),
        "모의고사 점수만 보고 개념·조건·계산·문항 선택 중 실제 원인을 나누지 못하는지",
        "최근 모의고사 시험지, 문항별 소요 시간과 처음 선택한 풀이 경로",
        "오답을 지식 부족과 풀이 판단으로 나누고 같은 유형의 새 문제를 제한 시간에 다시 풀기",
        "회차가 달라져도 문항 선택 순서와 풀이 근거가 안정되는지",
        "내신 범위 학습과 모의고사 누적 약점 보완의 주간 비중을 조정하는 기준",
        "등급 변화가 없을 때 문제 수보다 먼저 바꿀 풀이 절차와 시간 기록이 무엇인지",
    ),
    HighMathIntent(
        "suneung", "수능·기출 분석", ("수능", "기출", "입시", "고3", "입시 준비", "입시전략", "입시 전략"),
        "기출을 풀고도 출제 조건·개념 조합·문항 선택의 근거를 남기지 않는지",
        "기출 문제의 조건 표시, 사용 개념, 첫 풀이 경로와 소요 시간",
        "기출을 단원과 요구 행동으로 분류하고 틀린 판단을 새 문제에서 다시 검증하기",
        "회차와 표현이 달라도 필요한 개념 조합과 풀이 시작점을 스스로 찾는지",
        "학년·목표 시점에 맞춰 내신 기간과 수능 누적 학습의 비중을 바꾸는 기준",
        "기출 진도보다 분석·재풀이·문항 선택 기록을 어떻게 확인하는지",
    ),
    HighMathIntent(
        "advanced", "고난도 다단계 문제", ("고난도", "심화", "킬러", "준킬러", "상위권", "응용력", "다단계"),
        "한 단계씩은 풀 수 있지만 여러 조건과 개념이 결합되면 첫 풀이 방향을 정하지 못하는지",
        "고난도 문제에서 처음 세운 전략, 버린 접근과 다음 단계로 넘어간 근거",
        "문제를 조건 묶음과 중간 목표로 나눈 뒤 각 단계에 필요한 개념을 연결하기",
        "숫자와 조건 배열이 달라져도 중간 목표를 세워 풀이를 이어 가는지",
        "내신의 변형 심화 문항과 수능형 다단계 문항의 접근 시간을 구분하는 기준",
        "어려운 문제의 해설 암기보다 첫 접근과 중간 목표를 어떻게 피드백하는지",
    ),
    HighMathIntent(
        "performance", "수학 수행평가·탐구", ("수행평가", "탐구", "보고서", "발표", "평가 기준", "루브릭"),
        "수행평가의 주제·과정·표현 조건을 읽고 계산 결과와 설명을 함께 준비하는지",
        "학교 안내문, 평가 기준표와 탐구 초안에서 빠진 조건",
        "마감일부터 역산해 자료 정리·계산 확인·설명 작성·수정 날짜를 나누기",
        "제출 전 평가 조건을 학생이 직접 대조하고 계산과 설명의 누락을 고치는지",
        "지필 시험 공부와 수행평가 준비 시간을 학교 일정에 맞춰 분리하는 기준",
        "결과물을 대신 만드는 것이 아니라 조건 확인과 수학적 설명을 어떻게 돕는지",
    ),
    HighMathIntent(
        "transition", "고등 과정 전환", ("예비고1", "예비 고1", "고등 전환", "새 학년", "선행", "중3", "방학"),
        "중학교 문제 풀이에서 고등 수학의 개념 연결과 긴 계산 과정으로 넘어갈 준비가 되어 있는지",
        "최근 시험지, 고등 수준 예시 문제와 일주일 학습 기록",
        "선수 개념과 계산 현재선을 확인한 뒤 고등 문제 한 세트에 적용하기",
        "무리한 선행보다 새 난도의 조건을 읽고 풀이를 끝내는 범위가 넓어졌는지",
        "입학 전 기초 보완과 학교 일정 이후 내신 준비를 나누는 기준",
        "선행 범위보다 고등 과정에서 혼자 수행할 수 있는 단계를 어떻게 확인하는지",
    ),
    HighMathIntent(
        "routine", "수학 학습 루틴", ("습관", "루틴", "계획", "스케줄", "학습량", "과제", "피드백", "관리"),
        "계획한 문제 수가 실제 완료·질문·오답 재확인 기록으로 이어지는지",
        "주간 계획표, 시작·완료 시각과 다음 날 남은 질문",
        "새 문제·오답 재풀이·질문 정리를 다른 칸에 두고 완료 근거를 한 줄로 남기기",
        "한 주 뒤 미완료 이유와 반복 오답을 바탕으로 분량과 난도를 조정하는지",
        "내신 기간의 집중 과제와 평소 누적 문제를 달력에서 분리하는 기준",
        "과제량보다 완료·질문·재확인·계획 수정의 주기를 어떻게 확인하는지",
    ),
    HighMathIntent(
        "anxiety", "시험 긴장·수학 자신감", ("불안", "긴장", "자신감", "포기", "부담", "회복", "다시 시작"),
        "어려운 문제를 만나면 아는 단계까지 쓰지 못하고 풀이 전체를 멈추는지",
        "멈춘 문항, 다시 시작한 시각과 도움 없이 적은 첫 단계",
        "짧은 시간 제한 세트에서 아는 조건부터 적고 건너뛴 뒤 돌아오는 순서를 연습하기",
        "같은 난도의 새 문제에서 멈춘 뒤 풀이를 다시 시작하는 시간이 줄었는지",
        "내신 범위의 익숙한 문제와 모의고사의 처음 보는 문제를 다른 난도로 연습하는 기준",
        "정답 수보다 멈춘 뒤 회복하는 행동과 첫 풀이 단계를 어떻게 확인하는지",
    ),
)
HIGH_MATH_INTENT_BY_CODE = {intent.code: intent for intent in HIGH_MATH_INTENTS}


def middle_math_intent_text(value: str, code: str) -> str:
    """Convert high-school-only planning language into middle-school math language."""

    if code == "transition":
        return value
    replacements = (
        ("내신과 모의고사", "학교 시험과 누적 유형 문제"),
        ("내신·모의고사", "학교 시험·누적 유형 문제"),
        ("모의고사의", "누적 유형 문제의"),
        ("모의고사에서", "누적 유형 문제에서"),
        ("모의고사를", "누적 유형 문제를"),
        ("모의고사", "누적 유형 문제"),
        ("수능형", "여러 단원이 섞인"),
        ("수능", "누적 학습"),
        ("고등학교", "중학교"),
        ("고등 수학", "중등 수학"),
        ("고등 과정", "중등 과정"),
    )
    for before, after in replacements:
        value = value.replace(before, after)
    value = value.replace(
        "어떤 기준으로 첨삭하는지",
        "어떤 기준으로 풀이 과정을 확인하는지",
    )
    value = value.replace("첨삭", "풀이 확인")
    return value


_MIDDLE_MATH_EXCLUDED_KEYWORDS = {
    "모의고사", "모고", "전국연합", "등급", "수능", "기출", "입시",
    "고3", "입시 준비", "입시전략", "입시 전략", "킬러", "준킬러",
}
_MIDDLE_MATH_TRANSITION_KEYWORDS = {
    "예비고1", "예비 고1", "예비고등", "예비 고등", "고등 전환",
    "고등 과정", "고등 수학", "중3 고등", "중3에서 고등",
}
MIDDLE_MATH_INTENTS: tuple[HighMathIntent, ...] = tuple(
    HighMathIntent(
        intent.code,
        middle_math_intent_text(intent.label, intent.code),
        (
            tuple(sorted(_MIDDLE_MATH_TRANSITION_KEYWORDS))
            if intent.code == "transition"
            else tuple(
                keyword
                for keyword in intent.keywords
                if keyword not in _MIDDLE_MATH_EXCLUDED_KEYWORDS
            )
        ),
        middle_math_intent_text(intent.concern, intent.code),
        middle_math_intent_text(intent.evidence, intent.code),
        middle_math_intent_text(intent.action, intent.code),
        middle_math_intent_text(intent.checkpoint, intent.code),
        middle_math_intent_text(intent.exam_use, intent.code),
        middle_math_intent_text(intent.consult_question, intent.code),
    )
    for intent in HIGH_MATH_INTENTS
    if intent.code not in {"mock_exam", "suneung"}
) + (
    HighMathIntent(
        "entry", "중1 첫 시험 적응", ("예비중", "예비 중", "중1", "중학교 첫", "입학", "첫 시험"),
        "초등 수학의 연산 중심 학습에서 중학교의 개념·조건·서술형 문제로 넘어갈 준비가 되어 있는지",
        "현재 교재, 중학교 예시 범위표와 계산·문장제·서술형 문제의 첫 풀이",
        "연산과 개념 설명의 현재선을 확인한 뒤 짧은 학교 시험형 문제에 적용해 보기",
        "무리한 선행보다 중학교 문제의 조건을 읽고 풀이 이유를 설명하는 범위가 넓어졌는지",
        "입학 전 기초 보완과 첫 시험 범위 학습을 나누는 기준",
        "선행 범위보다 중1 수학에서 혼자 수행할 수 있는 단계를 어떻게 확인하는지",
    ),
)
MIDDLE_MATH_INTENT_BY_CODE = {intent.code: intent for intent in MIDDLE_MATH_INTENTS}

HIGH_ENGLISH_INTENT_OVERRIDES: dict[str, tuple[str, str, str, str]] = {
    "마곡동": ("school_exam", "transition", "routine", "vocabulary"),
    "중계동": ("routine", "reading", "sentence", "school_exam"),
    "신정동": ("transition", "school_exam", "routine", "vocabulary"),
    "정자동": ("sentence", "vocabulary", "grammar", "transition"),
    "금오동": ("transition", "vocabulary", "sentence", "routine"),
    "금촌동": ("transition", "vocabulary", "sentence", "routine"),
    "신현동": ("routine", "reading", "sentence", "grammar"),
    "신천동": ("school_exam", "grammar", "reading", "routine"),
    "석동": ("school_exam", "transition", "vocabulary", "routine"),
}

ELEMENTARY_ENGLISH_SIGNALS: tuple[TopicSignal, ...] = (
    TopicSignal("phonics", "소리와 철자 연결", ("파닉스", "소리", "철자", "발음"), "알파벳 소리와 낱말 철자를 연결해 읽는지", "처음 본 낱말을 소리 내어 읽고 고친 흔적", "소리 단위로 끊어 읽고 철자를 함께 적기", "짧은 낱말을 보고 소리와 뜻을 말해 보기"),
    TopicSignal("basic_words", "기초 어휘", ("어휘", "단어", "뜻"), "배운 낱말을 그림과 짧은 문장에서 알아보는지", "교재에서 반복해 헷갈린 낱말과 다시 읽은 기록", "낱말·그림·짧은 예문을 한 묶음으로 익히기", "낱말 세 개로 짧은 문장을 만들어 보기"),
    TopicSignal("sentence_reading", "문장 읽기", ("문장", "읽기", "해석"), "짧은 문장의 주어와 행동을 구분해 의미를 말하는지", "멈춰 읽은 문장과 뜻을 다시 설명한 기록", "짧은 문장을 의미 단위로 끊어 읽기", "한 문장을 소리 내어 읽고 우리말로 설명하기"),
    TopicSignal("basic_grammar", "기초 문장 규칙", ("문법", "시제", "동사", "복수"), "배운 문장 규칙을 예문에 맞게 적용하는지", "규칙은 알지만 문장에서 틀린 부분을 고친 기록", "맞는 문장과 틀린 문장을 비교하며 규칙 찾기", "짧은 문장 하나를 바르게 고쳐 쓰기"),
    TopicSignal("reading_flow", "짧은 글 이해", ("독해", "지문", "내용", "이야기"), "짧은 글의 인물·장소·행동을 순서대로 말하는지", "질문에 답한 근거가 있는 문장을 찾은 흔적", "문단마다 중요한 낱말과 문장을 표시하기", "읽은 내용을 두 문장으로 설명해 보기"),
    *ENGLISH_SIGNALS[4:],
)

MATH_SIGNALS: tuple[TopicSignal, ...] = (
    TopicSignal("concept", "개념 연결", ("개념", "정의", "원리", "이해"), "정의와 공식을 말로 설명한 뒤 문제에 적용하는지", "기본 문제와 조건이 달라진 문제의 풀이 근거", "개념 한 줄과 대표 문제를 한 묶음으로 정리하기", "해설 없이 개념을 설명하고 한 문제를 다시 풀기"),
    TopicSignal("calculation", "계산 정확도", ("계산", "연산", "부호", "실수"), "부호·괄호·연산 순서에서 같은 실수가 반복되는지", "중간 계산과 답을 고친 뒤의 재풀이 기록", "계산 단계를 한 줄씩 나누고 틀린 위치 표시하기", "짧은 계산 세트를 정확도 기준으로 다시 풀기"),
    TopicSignal("condition", "문제 조건 해석", ("조건", "문장제", "응용", "문제해결"), "문장에서 필요한 조건을 골라 식이나 그림으로 바꾸는지", "조건을 빠뜨린 문제와 식을 잘못 세운 문제의 구분", "조건에 밑줄을 긋고 구하려는 값을 먼저 적기", "비슷한 문제에서 식을 세우는 과정만 다시 연습하기"),
    TopicSignal("error", "수학 오답 재학습", ("오답", "틀린", "재풀이", "복습"), "틀린 이유를 개념·조건·계산으로 나누어 기록하는지", "첫 풀이와 다시 푼 풀이에서 달라진 단계", "오답 원인을 표시하고 다시 볼 날짜 정하기", "같은 유형 한 문제를 며칠 뒤 다시 풀기"),
    TopicSignal("written", "서술형 풀이", ("서술형", "풀이과정", "과정", "설명"), "답뿐 아니라 식을 세운 이유와 풀이 순서를 적는지", "서술형 답안에서 생략된 조건과 고친 표현", "풀이 단계마다 근거를 짧은 문장으로 남기기", "완성한 풀이를 소리 내어 설명해 보기"),
    TopicSignal("graph", "그래프와 자료 해석", ("그래프", "도형", "표", "좌표"), "그림·표·그래프의 정보를 식과 연결하는지", "자료를 잘못 읽은 부분과 다시 표시한 값", "주어진 정보를 그림에 옮기고 관계를 설명하기", "표나 그림을 보고 조건을 한 문장으로 적기"),
    TopicSignal("units", "단원 간 연결", ("단원", "누적", "선행", "연결"), "앞 단원의 개념이 현재 문제에서 어떻게 쓰이는지 설명하는지", "최근 문제에서 다시 필요해진 이전 단원 기록", "현재 단원과 연결된 앞 개념을 짧게 복습하기", "오늘 문제에 쓰인 이전 개념을 한 줄로 적기"),
    TopicSignal("pace", "수학 시간 배분", ("시간", "시험", "속도", "시간배분"), "어느 유형과 계산 단계에서 시간이 길어지는지", "문항별 소요 시간과 끝까지 풀지 못한 구간", "풀이와 검산 시간을 나누어 연습하기", "짧은 세트를 풀고 오래 걸린 문제를 표시하기"),
    TopicSignal("routine", "수학 학습 루틴", ("습관", "플래너", "기록", "과제"), "계획한 문제와 실제 완료한 문제를 구분하는지", "학습 시작 시각과 재풀이 완료 여부를 적은 기록", "매일 풀 분량과 다시 볼 문제를 나누기", "완료한 문제와 남은 질문을 짧게 적기"),
)

ELEMENTARY_MATH_SIGNALS: tuple[TopicSignal, ...] = (
    TopicSignal("number", "수 감각과 연산 원리", ("연산", "계산", "수", "받아올림", "받아내림"), "계산 절차를 외우기보다 수가 바뀌는 이유를 설명하는지", "계산 과정에서 자릿값과 부호를 표시한 흔적", "수 모형과 식을 연결한 뒤 같은 계산을 다시 풀기", "짧은 연산을 풀고 틀린 단계만 말로 설명하기"),
    TopicSignal("word_problem", "문장제 조건", ("문장제", "문제", "조건", "응용"), "문장에서 필요한 수와 질문을 구분해 표시하는지", "식을 세우기 전 그린 그림과 밑줄 친 조건", "질문을 한 문장으로 바꾸고 필요한 수를 고르기", "생활 장면 문제 하나를 그림과 식으로 나타내기"),
    TopicSignal("fraction", "분수·소수 이해", ("분수", "소수", "비율"), "전체와 부분의 관계를 그림과 수로 함께 설명하는지", "분수나 소수를 수직선과 그림에 표시한 기록", "그림·말·식 세 가지로 같은 값을 나타내기", "생활 속 양을 분수나 소수로 표현해 보기"),
    TopicSignal("geometry", "도형과 측정", ("도형", "각도", "넓이", "길이", "측정"), "도형의 성질과 측정 단위를 구분해 사용하는지", "도형에 표시한 길이·각도와 풀이 식", "그림에 조건을 옮기고 사용한 성질을 적기", "주변 물건의 길이와 모양을 직접 설명해 보기"),
    TopicSignal("explain", "수학 개념 설명", ("개념", "설명", "이해", "원리"), "답을 구한 이유를 자기 말로 설명할 수 있는지", "풀이 뒤에 이유를 한 문장으로 적은 흔적", "대표 문제를 풀고 사용한 개념을 말로 정리하기", "가족에게 풀이 순서를 짧게 설명해 보기"),
    *MATH_SIGNALS[3:],
)


def esc(value: object) -> str:
    return html.escape(str(value or ""), quote=True)


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", html.unescape(str(value or ""))).strip()


def final_consonant_index(value: str) -> int | None:
    """Return the Korean jongseong index for the final pronounced character."""
    digit_jongseong = {
        "0": 21,  # 영
        "1": 8,   # 일
        "2": 0,   # 이
        "3": 16,  # 삼
        "4": 0,   # 사
        "5": 0,   # 오
        "6": 1,   # 육
        "7": 8,   # 칠
        "8": 8,   # 팔
        "9": 0,   # 구
    }
    for character in reversed(clean(value)):
        codepoint = ord(character)
        if 0xAC00 <= codepoint <= 0xD7A3:
            return (codepoint - 0xAC00) % 28
        if character in digit_jongseong:
            return digit_jongseong[character]
        if character.isalnum():
            return None
    return None


def particle_for(value: str, consonant_form: str, vowel_form: str) -> str:
    jongseong = final_consonant_index(value)
    if consonant_form == "으로" and jongseong == 8:
        return vowel_form
    return consonant_form if jongseong else vowel_form


def normalize_particle_joins(text: str, tokens: Iterable[str]) -> str:
    """Correct particles attached to page-specific generated nouns/phrases."""
    pairs = (("으로", "로"), ("은", "는"), ("이", "가"), ("을", "를"), ("과", "와"))
    particle_pairs = {particle: pair for pair in pairs for particle in pair}
    values = sorted({clean(token) for token in tokens if clean(token)}, key=len, reverse=True)
    for value in values:
        escaped = re.escape(value)
        pattern = rf"{escaped}([’'\"”]?)(으로|로|은|는|이|가|을|를|과|와)(?![가-힣])"

        def replace_particle(match: re.Match[str]) -> str:
            consonant_form, vowel_form = particle_pairs[match.group(2)]
            return value + match.group(1) + particle_for(value, consonant_form, vowel_form)

        text = re.sub(pattern, replace_particle, text)
    return text


def page_particle_tokens(
    config: CategoryConfig,
    center: dict[str, object],
    signals: tuple[TopicSignal, ...],
    student: str = "",
) -> tuple[str, ...]:
    schools = relevant_schools(config, center)
    school_names = "·".join(schools[:3]) if schools else "실제 재학 학교"
    material = material_label(config, schools)
    grade_values = [
        f"{subject} {'·'.join(relevant_grades(config, center, subject)) or '상담 시 확인'}"
        for subject in config.subjects
    ]
    values: list[str] = [
        str(center["address"]),
        str(center["center_name"]),
        config.label,
        config.level,
        config.school_name,
        f"{center['locality']} {config.label}",
        "·".join(config.subjects),
        material,
        school_names,
        f"{school_names} 자료",
        ", ".join(grade_values),
        " / ".join(grade_values),
        *grade_values,
    ]
    if student:
        values.append(student)
    for signal in signals:
        values.extend((signal.label, signal.check, signal.evidence, signal.practice, signal.home_action))
    return tuple(dict.fromkeys(value for value in values if value))


def normalize_generated_value(value: object, tokens: tuple[str, ...]) -> object:
    if isinstance(value, str):
        return normalize_particle_joins(value, tokens)
    if isinstance(value, list):
        return [normalize_generated_value(item, tokens) for item in value]
    if isinstance(value, tuple):
        return tuple(normalize_generated_value(item, tokens) for item in value)
    if isinstance(value, dict):
        return {key: normalize_generated_value(item, tokens) for key, item in value.items()}
    return value


def stable_int(seed: str, label: str) -> int:
    return int(hashlib.sha256(f"{seed}|{label}".encode("utf-8")).hexdigest()[:16], 16)


def stable_pick(seed: str, label: str, choices: tuple[str, ...] | list[str]) -> str:
    return choices[stable_int(seed, label) % len(choices)]


def list_values(value: object) -> list[str]:
    return list(dict.fromkeys(clean(part) for part in re.split(r"[,/\n]+", str(value or "")) if clean(part)))


def normalize_school_values(value: object) -> list[str]:
    text = clean(value)
    if "지역내 모든 고등학교 가능" in text:
        return []
    text = re.sub(r"[.·]+", ",", text)
    known_joined = {"오현초호매실중": ("오현초", "호매실중")}
    values: list[str] = []
    for chunk in re.split(r"[,/\s]+", text):
        chunk = clean(chunk)
        if not chunk:
            continue
        if chunk in known_joined:
            values.extend(known_joined[chunk])
            continue
        values.append(chunk)
    return list(dict.fromkeys(value for value in values if len(value) >= 2))


def folder_slug(locality: str) -> str:
    return re.sub(r"\s+", "", locality)


def display_region_label(center: dict[str, object]) -> str:
    """Avoid repeating a city prefix already present in region/district."""
    region = str(center["region"])
    district = str(center["district"])
    locality = str(center["locality"])
    district_base = re.sub(r"(?:특별자치시|특별시|광역시|시|군|구)$", "", district)
    display_locality = locality
    for prefix in (region, district_base):
        if prefix and display_locality.startswith(prefix + " "):
            display_locality = display_locality[len(prefix):].strip()
            break
    return " ".join(part for part in (region, district, display_locality) if part)


def route_for(category_slug: str, locality_slug: str | None = None) -> str:
    path = f"/과목별학원/{category_slug}/"
    return path + (f"{locality_slug}/" if locality_slug else "")


def absolute_route(category_slug: str, locality_slug: str | None = None) -> str:
    return ORIGIN + quote(route_for(category_slug, locality_slug), safe="/%:@")


def center_entity_id(center: dict[str, object]) -> str:
    key = "|".join(str(center.get(name, "")) for name in ("center_name", "address", "registration"))
    return ORIGIN + "/#center-" + hashlib.sha256(key.encode("utf-8")).hexdigest()[:16]


def source_text(value: str) -> str:
    value = re.sub(r"(?is)<(?:script|style).*?>.*?</(?:script|style)>", " ", value)
    return clean(re.sub(r"<[^>]+>", " ", value))


def source_tag_texts(value: str, tag: str) -> list[str]:
    return [
        clean(re.sub(r"<[^>]+>", " ", match))
        for match in re.findall(rf"(?is)<{tag}[^>]*>(.*?)</{tag}>", value)
        if clean(re.sub(r"<[^>]+>", " ", match))
    ]


HIGH_ENGLISH_FOCUS_OVERRIDES = {
    "마곡동": "고등 영어 첫 시험의 준비 순서를 세우는 방법",
    "중계동": "현재 영어 학습의 병목을 찾는 상담 기준",
    "신정동": "새 학년 고등 영어의 우선순위를 정하는 방법",
    "정자동": "고등 영어 문장 구조의 기초를 점검하는 방법",
    "금오동": "고등 영어 기초를 다시 연결하는 학습 순서",
    "금촌동": "고등 영어로 넘어가기 전 준비할 학습 기준",
    "신현동": "고등 영어 오답 원인을 구분하는 방법",
    "신천동": "고등 영어 내신의 우선순위를 정하는 방법",
    "석동": "고등 영어 첫 시험을 준비하는 학습 기준",
    "길음동": "고등 영어 학습 상담에서 우선순위를 정하는 방법",
    "삼각산동": "내신과 모의고사를 학습 과정으로 나누는 방법",
    "하계동": "내신 범위와 모의고사 독해의 주간 균형",
    "위례": "내신 일정과 모의고사 누적 학습을 배치하는 방법",
    "금정동": "내신과 모의고사의 오답 근거를 구분하는 방법",
    "태평동": "내신과 모의고사의 오답 원인을 나누는 방법",
    "사파동": "내신 범위와 처음 보는 지문을 준비하는 순서",
    "신방화": "문장 구조를 내신 변형 문제로 연결하는 방법",
    "탄벌동": "문장 구조를 독해 속도와 내신에 연결하는 방법",
    "용강동": "구문 이해를 내신과 모의고사에 적용하는 방법",
    "심곡동": "풀이 과정으로 내신과 모의고사 약점을 찾는 방법",
    "후곡마을": "영어 기초의 빈틈부터 내신까지 연결하는 순서",
    "만촌동": "문장 구조와 문법 기초를 내신에 연결하는 순서",
    "칠성동": "내신 진도와 학습 점검을 함께 설계하는 방법",
    "수성동": "내신과 과제 점검을 함께 정리하는 방법",
    "역북동": "학생별 영어 학습 과정을 설계하는 방법",
    "용산동": "입시 영어 상담에서 학습 방향을 찾는 방법",
    "신도림동": "새 학기 고등 영어의 첫 점검 순서를 세우는 방법",
    "염리동": "상담 전 비용·시간표와 학습 조건을 구분하는 방법",
    "공덕동": "중3에서 고등 영어로 넘어가는 준비 순서",
    "상암동": "꾸준한 영어 학습을 이어 가는 복습 기준",
    "돈암동": "진단 결과로 영어 학습 순서를 설계하는 방법",
    "목동": "어려운 문법 용어를 문제 풀이와 연결하는 방법",
    "위례신도시": "개념과 기출을 연결해 내신을 준비하는 순서",
    "광명동": "교과서 학습을 새 지문 독해로 확장하는 방법",
    "산본동": "영어 학습 흐름을 다시 세우는 점검 기준",
    "부천 중동": "끊어진 영어 학습 과정을 다시 연결하는 방법",
    "성남 금곡동": "학생별 영어 취약점을 찾는 진단 순서",
    "수진동": "첫 학습에서 영어 공부 방향을 정하는 방법",
    "호매실": "서술형까지 연결하는 영어 학습 점검 순서",
    "목감": "진단 결과로 학생별 영어 학습을 설계하는 방법",
    "대야동": "학생별 진도에 맞춰 영어 학습 순서를 설계하는 방법",
    "초지동": "원서 읽기 전 확인할 고등 영어 기초",
    "석우동": "학습 준비도를 확인하고 첫 과제를 정하는 방법",
    "청계동": "영어 학습 흐름을 정리하는 방법",
    "인천 갈산동": "영어 학습의 빈틈을 찾는 점검 방법",
    "원내동": "학습 점검표로 학생별 우선순위를 정하는 방법",
    "불당동": "질문하기 어려운 학생의 영어 학습 점검 방법",
    "호암동": "시험 자료로 고등 영어 학습 기준을 세우는 방법",
    "충주 용산동": "입시 영어의 학습 우선순위를 정하는 방법",
    "양덕동": "흐릿한 영어 개념을 선명하게 정리하는 방법",
    "상남동": "학습 방법을 비교해 영어 공부 방향을 세우는 기준",
    "두호동": "영어 학습 방법과 복습 흐름을 점검하는 기준",
    "전주혁신도시": "고등 영어 학습 기준을 세우는 방법",
    "중화산동": "학생별 피드백 방법을 확인하는 상담 기준",
    "송천동": "학생별 영어 학습 흐름을 세우는 방법",
    "해운대 중동": "기초 확인부터 학생별 지도로 연결하는 방법",
    "병영동": "풀이 습관부터 영어 기초를 다지는 방법",
    "석사동": "시험 불안을 줄이는 평일·방학 학습 계획",
    "갈매동": "영어 문장 해석 실수를 줄이는 점검 방법",
    "죽전동": "교과서 개념과 독해 근거를 함께 점검하는 방법",
    "부발읍": "단원별 취약점부터 내신까지 연결하는 학습 점검",
    "구산동": "시험 시간 배분을 다시 세우는 학습 방법",
    "운양동": "최근 시험지에서 영어 학습의 우선순위를 정하는 방법",
    "죽백동": "학습 과정을 점검해 영어 이해를 깊게 만드는 방법",
    "당진 읍내동": "수행평가 글쓰기의 완성 기준을 점검하는 방법",
    "반송동": "문제를 끝까지 읽는 힘을 기르는 독해 점검",
    "서신동": "기초부터 시작하는 시험 긴장 완화 학습 점검",
    "성정동": "듣기 평가 오답과 입시 학습 우선순위를 함께 점검하는 방법",
}

# Source headings sometimes become grammatically valid but still read like
# stacked marketing nouns.  These page angles preserve the source intent in
# plain Korean and keep each locality's editorial focus distinct.
HIGH_ENGLISH_FOCUS_OVERRIDES.update({
    "경산사동": "점수보다 학습 과정을 먼저 점검하는 방법",
    "괴안동": "영어 공부 흐름을 세밀하게 점검하는 방법",
    "구월동": "영어 이해도를 기초부터 차근히 점검하는 방법",
    "구파발": "시험 결과를 다음 학습 계획으로 연결하는 방법",
    "금곡동": "학습 평가로 시작하는 학생별 영어 점검",
    "금광동": "학생에게 맞춘 영어 학습 방향 세우기",
    "금릉": "진도와 학습 태도를 함께 점검하는 방법",
    "다정동": "고등 영어 입시 상담 준비 안내",
    "대구장기동": "집중력과 영어 공부 흐름을 함께 점검하는 방법",
    "도안동": "내신 성취와 학습 흐름을 함께 점검하는 방법",
    "둔산동": "개별 학습 기록으로 영어의 빈틈을 채우는 방법",
    "마두동": "어휘부터 내신 독해까지 공부 흐름을 잡는 방법",
    "망월동": "학생별 영어 학습 흐름을 설계하는 방법",
    "범어동": "방학에도 영어 학습의 핵심을 유지하는 방법",
    "병점": "학습 태도부터 내신 준비까지 점검하는 방법",
    "병점동": "문장력과 영어 공부 흐름을 함께 잡는 방법",
    "본리동": "학습 문제의 원인을 찾아 방향을 세우는 방법",
    "부개동": "학생별 진도에 맞춰 공부 흐름을 잡는 방법",
    "비래동": "영어 학습 진척도를 확인하는 방법",
    "산남동": "입시 영어의 학습 우선순위를 세우는 방법",
    "송촌동": "학습 시간표부터 꼼꼼히 살피는 영어 점검",
    "수곡동": "입시 영어 전략을 세우는 학습 점검",
    "수완지구": "집중 학습 전 영어 공부 흐름 점검",
    "수택동": "고2 영어 학습량과 이해도를 함께 점검하는 방법",
    "신창지구": "영어 학습 흐름을 다시 세우는 방법",
    "안흥동": "공부 기록으로 찾는 학생별 영어 학습 방향",
    "야탑동": "성적보다 학습 습관을 먼저 점검하는 방법",
    "옥정동": "차분한 이해와 실천력을 중심에 둔 영어 학습",
    "용인중동": "영어 학습 로드맵으로 방향을 세우는 방법",
    "용인 중동": "영어 학습 로드맵으로 방향을 세우는 방법",
    "운정": "학습 목표를 점검하며 영어 실력을 쌓는 방법",
    "원동": "학생의 강점부터 확인하는 영어 학습",
    "월계동": "현재 공부 방법을 점검해 학습 방향을 세우는 방법",
    "율하동": "입시 준비와 학습 습관을 함께 점검하는 방법",
    "은행동": "영어 학습 자신감을 다시 세우는 방법",
    "인창동": "실시간 학습 기록으로 시작하는 영어 점검",
    "첨단지구": "학습 과정을 점검하는 영어 상담",
    "호평동": "느린 문제 풀이를 안정된 학습 흐름으로 바꾸는 방법",
    "흥덕마을": "학습 포트폴리오로 공부 흐름을 점검하는 방법",
    "대구 장기동": "집중력과 영어 공부 흐름을 함께 점검하는 방법",
})


def safe_high_english_phrase(value: str, locality: str = "") -> str:
    value = clean(value)
    if locality:
        value = re.sub(rf"{re.escape(locality)}\s*고등\s*영어학원", " ", value)
        value = re.sub(rf"(?<![가-힣]){re.escape(locality)}(?![가-힣])", " ", value)
    value = re.sub(r"고등\s*영어학원", " ", value)
    replacements = (
        (r"와와학습코칭센터(?:\s*\S*점)?", "학습 상담"),
        (r"하계점\s*설명회", "고등 영어 상담"),
        (r"(?:입시준비반|정규관리반|성적관리반)", "입시 준비 계획"),
        (r"(?:방학집중반|집중반)", "방학 집중 계획"),
        (r"(?:내신특강|방학특강|특강수업|특강)", "집중 학습"),
        (r"(?:실시간수업|녹화수업|오전수업|그룹수업|집중수업|참여형\s*수업)", "학습 방법"),
        (r"학원정규반", "정규 학습 계획"),
        (r"시험대비반", "시험 대비 계획"),
        (r"학원보충", "보충"),
        (r"학원비", "비용과 이용 조건"),
        (r"맞춤형?", "학생별"),
        (r"개별진도", "학생별 진도"),
        (r"학원진도", "현재 진도"),
        (r"(?:성적\s*향상|내신향상)", "성적 변화의 원인 점검"),
        (r"성적을\s*바꾸는", "성적 변화에 영향을 주는"),
        (r"학습성과", "학습 과정"),
        (r"입시지원전략", "입시 학습 우선순위"),
        (r"입시합격전략", "입시 학습 전략"),
        (r"입시맞춤상담", "입시 학습 상담"),
        (r"1:1", "학생별"),
        (r"소수정예", "학생별 피드백"),
        (r"학습코칭", "학습 점검"),
        (r"코칭수업", "피드백 방법"),
        (r"학습관리", "학습 기록 점검"),
        (r"영어관리", "영어 학습 점검"),
        (r"오답관리", "오답 점검"),
        (r"성적관리", "성적 변화 점검"),
        (r"시간관리", "시간 배분 점검"),
        (r"내신관리", "내신 준비 점검"),
        (r"시험관리", "시험 준비 점검"),
        (r"관리앱", "학습 기록"),
        (r"관리", "점검"),
        (r"코칭", "피드백"),
        (r"함께\s*준비하는\s*수업", "함께 준비하는 방법"),
        (r"수업", "학습"),
        (r"매일", "꾸준히"),
    )
    for pattern, replacement in replacements:
        value = re.sub(pattern, replacement, value)
    value = re.sub(r"(?:학습\s*){2,}점검", "학습 점검", value)
    value = re.sub(r"학생별학습", "학생별 학습", value)
    value = re.sub(r"학생별상담", "학생별 상담", value)
    value = re.sub(r"시험시간", "시험 시간", value)
    value = re.sub(r"학습성적", "학습 기록과 성적", value)
    value = re.sub(r"학습점검", "학습 점검", value)
    value = value.replace("점검로", "점검으로").replace("점검와", "점검과")
    value = value.replace("학습 방법 과정", "학습 과정").replace("학습 방식 과정", "학습 과정")
    value = value.replace("학습 방법 첫날부터", "첫 학습부터").replace("학습 방식 첫날부터", "첫 학습부터")
    value = value.replace("비용과 이용 조건와", "비용과 이용 조건과")
    value = re.sub(r"학습\s*방법\s*학습\s*방법", "학습 방법", value)
    spacing_replacements = (
        ("고등영어", "고등 영어"), ("먼저확인", "먼저 확인"), ("함께준비", "함께 준비"),
        ("학습상담", "학습 상담"), ("학습컨설팅", "학습 상담"), ("학습멘토링", "학습 상담"),
        ("학습평가", "학습 평가"), ("학습설계", "학습 설계"), ("학습계획", "학습 계획"),
        ("학습전략", "학습 전략"), ("학습루틴", "학습 루틴"), ("학습지속력", "학습 지속력"),
        ("학습실천력", "학습 실천력"), ("학습주도성", "학습 주도성"), ("학습회복력", "학습 회복력"),
        ("학습강점", "학습 강점"), ("학습약점", "학습 약점"), ("학습결과", "학습 결과"),
        ("학습일지", "학습 일지"), ("학습스케줄", "학습 일정"), ("학습로드맵", "학습 로드맵"),
        ("학습포트폴리오", "학습 포트폴리오"), ("학습통계", "학습 통계"), ("학습오답", "오답"),
        ("학습목표점검", "학습 목표 점검"), ("학습일정", "학습 일정"), ("학습습관", "학습 습관"),
        ("학습태도", "학습 태도"), ("학습집중점검", "학습 집중도 점검"),
        ("학습문제점검", "학습 문제 점검"), ("학습암기", "암기 학습"), ("학습이해", "학습 이해"),
        ("학습응용", "학습 응용"), ("학습심화", "심화 학습"), ("학습훈련", "학습 훈련"),
        ("학습시간표", "학습 시간표"), ("학습진척도", "학습 진척도"), ("학습점검표", "학습 점검표"),
        ("집중학습", "집중 학습"), ("정기학습", "정기 학습"), ("주간학습", "주간 학습"),
        ("입시멘토링", "입시 학습 상담"), ("입시설계", "입시 설계"), ("입시전략", "입시 전략"),
        ("입시계획", "입시 계획"), ("입시로드맵", "입시 로드맵"), ("입시점검", "입시 점검"),
        ("입시진단", "입시 진단"), ("입시평가", "입시 평가"), ("입시일정", "입시 일정"),
        ("입시자료분석", "입시 자료 분석"), ("내신전략", "내신 전략"), ("내신분석", "내신 분석"),
        ("내신평가", "내신 평가"), ("내신집중점검", "내신 집중 점검"), ("내신밀착점검", "내신 밀착 점검"),
        ("내신클리닉", "내신 점검"), ("내신컨설팅", "내신 상담"), ("시험계획", "시험 계획"),
        ("시험준비", "시험 준비"), ("시험성적", "시험 성적"), ("시험범위", "시험 범위"),
        ("시험집중점검", "시험 집중 점검"), ("성적상담", "성적 상담"), ("성적분석", "성적 분석"),
        ("집중점검", "집중 점검"), ("밀착점검", "밀착 점검"), ("반복학습", "반복 학습"),
        ("학생별피드백", "학생별 피드백"), ("장기점검", "장기 점검"), ("자기주도학습", "자기주도 학습"),
        ("차분히준비", "차분히 준비"), ("예비고1", "예비 고1"), ("단기방학", "단기 방학"),
        ("과정점검", "과정 점검"), ("학습우선순위", "학습 우선순위"),
        ("학습완성도", "학습 완성도"), ("학습몰입도", "학습 몰입도"),
        ("듣기평가", "듣기 평가"), ("입시분석", "입시 분석"),
    )
    for before, after in spacing_replacements:
        value = value.replace(before, after)
    value = value.replace("꾸준히의 공부", "꾸준한 공부")
    value = value.replace("우선순위을", "우선순위를")
    value = re.sub(r"\s+", " ", value).strip()
    value = re.sub(r"\s+", " ", value).strip(" ,·|:-")
    if value.endswith("확인하세요"):
        value = value.removesuffix("확인하세요").strip() + " 확인하는 방법"
    if value.endswith("준비하는 학습"):
        value = value.removesuffix("준비하는 학습").strip() + " 준비하는 방법"
    return value


def normalize_generated_tree(value: object, tokens: tuple[str, ...]) -> object:
    """Normalize a generated JSON-shaped tree in one pass per token."""

    serialized = json.dumps(value, ensure_ascii=False)
    normalized = normalize_particle_joins(serialized, tokens)
    return json.loads(normalized)


def high_english_profile(raw: str, locality: str) -> HighEnglishProfile:
    h1_values = source_tag_texts(raw, "h1")
    if not h1_values:
        raise ValueError(f"{locality}: high English source H1 missing")
    source_title = h1_values[0]
    if locality in HIGH_ENGLISH_FOCUS_OVERRIDES:
        focus = HIGH_ENGLISH_FOCUS_OVERRIDES[locality]
    else:
        focus = safe_high_english_phrase(source_title, locality)
    if (
        not focus
        or len(focus) > 64
        or re.search(r"(?:중1|중학생|중학교|초등|수학|차량|주차)", focus)
    ):
        focus = "고등 영어 학습의 빈틈을 찾는 방법"

    strong_values = source_tag_texts(raw, "strong")
    h3_values = [re.sub(r"^\d+\.\s*", "", value) for value in source_tag_texts(raw, "h3")]
    li_values = source_tag_texts(raw, "li")
    primary_parts: list[tuple[str, int]] = [(focus, 40)]
    primary_parts.extend((value, 10) for value in strong_values)
    primary_scores = {
        intent.code: sum(weight * sum(text.count(keyword) for keyword in intent.keywords) for text, weight in primary_parts)
        for intent in HIGH_ENGLISH_INTENTS
    }
    detail_parts: list[tuple[str, int]] = [(source_title, 7), (focus, 14)]
    detail_parts.extend((value, 4) for value in strong_values)
    detail_parts.extend((value, 3) for value in h3_values[-4:])
    detail_parts.extend((value, 1) for value in li_values)
    detail_scores = {
        intent.code: sum(weight * sum(text.count(keyword) for keyword in intent.keywords) for text, weight in detail_parts)
        for intent in HIGH_ENGLISH_INTENTS
    }
    full_text = source_text(raw)
    full_scores = {
        intent.code: sum(full_text.count(keyword) for keyword in intent.keywords)
        for intent in HIGH_ENGLISH_INTENTS
    }
    if locality in HIGH_ENGLISH_INTENT_OVERRIDES:
        ranked = [HIGH_ENGLISH_INTENT_BY_CODE[code] for code in HIGH_ENGLISH_INTENT_OVERRIDES[locality]]
    else:
        focus_ranked = sorted(
            HIGH_ENGLISH_INTENTS,
            key=lambda intent: (-primary_scores[intent.code], -detail_scores[intent.code], -full_scores[intent.code], intent.code),
        )
        detail_ranked = sorted(
            HIGH_ENGLISH_INTENTS,
            key=lambda intent: (-detail_scores[intent.code], -full_scores[intent.code], -primary_scores[intent.code], intent.code),
        )
        full_ranked = sorted(
            HIGH_ENGLISH_INTENTS,
            key=lambda intent: (-full_scores[intent.code], -detail_scores[intent.code], intent.code),
        )
        if primary_scores[focus_ranked[0].code] > 0:
            primary = focus_ranked[0]
        elif detail_scores[detail_ranked[0].code] > 0:
            primary = detail_ranked[0]
        elif full_scores[full_ranked[0].code] > 0:
            primary = full_ranked[0]
        else:
            primary = HIGH_ENGLISH_INTENT_BY_CODE["routine"]
        ranked = [primary]
        ranked.extend(
            intent
            for intent in detail_ranked
            if intent not in ranked and (detail_scores[intent.code] > 0 or full_scores[intent.code] > 0)
        )
        ranked.extend(
            HIGH_ENGLISH_INTENT_BY_CODE[code]
            for code in ("reading", "sentence", "routine", "school_exam", "vocabulary")
            if HIGH_ENGLISH_INTENT_BY_CODE[code] not in ranked
        )

    # Source headings can contain unsupported operating claims. They are used
    # only to classify the search intent; visible labels come from the audited
    # high-school English intent bank.
    markers = [intent.label for intent in ranked[:4]]
    return HighEnglishProfile(
        focus=focus,
        source_title=source_title,
        intents=tuple(ranked[:4]),
        source_markers=tuple(markers[:4]),
    )


MIDDLE_ENGLISH_FOCUS_OVERRIDES = {
    "불당동": "다음 학년 준비의 핵심을 잡는 학습",
    "탄벌동": "문장 구조의 어려움을 학습 흐름으로 점검하는 방법",
    "정평동": "본문 변형 문제까지 대비하는 집중 점검",
    "진월동": "과제 점검과 개별 피드백을 연결하는 방법",
    "안흥동": "학습 기록으로 찾는 학생별 영어 피드백",
    "수완동": "봄방학 영어 계획을 세우는 방법",
    "인창동": "학기 초 공부 습관을 세우는 방법",
    "화정동": "겨울방학 영어 기초를 다지는 방법",
    "쌍암동": "오후 학습 시간을 활용하는 영어 공부법",
    "삼산동": "학습과 연습을 연결하는 영어 점검",
    "삼송동": "학원 학습을 일주일 계획으로 연결하는 방법",
    "영천동": "영어 학습을 반복으로 정착시키는 방법",
    "천천동": "학습 점검으로 영어 빈틈을 찾는 방법",
    "중산동": "영어 학원 프로그램을 비교할 때 확인할 기준",
    "신월성": "다음 학년 준비를 위한 종합 점검",
    "대봉동": "시험 준비 과정을 끝까지 살피는 학습 점검",
    "탄방동": "학생별 학습 기록으로 공부 흐름을 점검하는 방법",
    "자은동": "영어 학습을 꾸준히 이어 가는 기준",
    "석동": "현재 영어 목표를 찾는 학습 점검",
    "심곡동": "단어 기록과 학습 루틴을 함께 점검하는 방법",
    "동천동": "내신 준비 상태를 밀착 점검하는 방법",
    "경산사동": "시험 대비와 학습 습관을 함께 점검하는 방법",
    "쌍용동": "다음 학년 준비와 현재 학습을 함께 점검하는 방법",
    "단계동": "틀린 문제를 성장의 기준으로 바꾸는 방법",
    "장곡동": "오래 공부해도 성적이 낮을 때 확인할 기준",
    "죽전동": "어법부터 내신까지 체계적으로 점검하는 방법",
    "일산동": "교과서 본문부터 탄탄하게 다지는 방법",
    "양덕동": "흐릿한 영어 개념을 선명하게 정리하는 방법",
    "창곡동": "독해 속도보다 개념을 먼저 선명하게 정리하는 방법",
    "덕이지구": "오답 점검부터 시작하는 학습 계획",
    "두호동": "실력 변화를 기록하는 학습 점검",
    "갈매동": "학교 학습과 기초를 함께 점검하는 방법",
    "금릉": "학습 흐름을 점검하는 방법",
    "금정동": "숙제와 내신을 함께 점검하는 방법",
    "금촌동": "학습 흐름을 잇는 점검 방법",
    "망포동": "학생별 학습 순서를 설계하는 방법",
    "미금": "공부의 틀을 세우는 방법",
    "세교": "학습 주도성을 키우는 방법",
    "수진동": "공부가 막히는 원인을 찾는 방법",
    "진관동": "학년별 학습 기준을 세우는 방법",
    "해운대중동": "개별 학습 상태를 확인하는 방법",
    "해운대 중동": "개별 학습 상태를 확인하는 방법",
    "사파동": "성적이 흔들리는 원인을 찾는 학습 점검",
    "송촌동": "꾸준한 학습 기록을 만드는 점검",
    "신불당": "성적 변동 원인을 확인하는 학습 점검",
    "구월동": "학습 이해도부터 태도까지 점검하는 방법",
    "노형동": "시험 대비 계획을 세우는 학습 안내",
    "신기동": "내신 준비 과정을 점검하는 방법",
    "조남동": "잊지 않는 복습 습관을 만드는 학습 점검",
    "하대원동": "꾸준한 단어 학습의 기준 세우기",
    "시지동": "내신 학습 점검 안내",
    "방화동": "암기보다 이해를 다지는 학습",
    "본리동": "학습이 막히는 원인부터 살피는 다음 학년 준비",
    "야탑": "문법 오답을 실력으로 바꾸는 복습",
    "주월동": "시험 대비 우선순위를 정하는 학습법",
    "호매실": "학습 기록으로 현재 실력을 점검하는 방법",
    "호매실동": "학습 방향을 바르게 잡는 점검",
    "연동": "다음 학년 준비까지 연결하는 학습 점검",
    "산내마을": "학원을 옮길 때 확인할 학습 기록",
    "새롬동": "기본기부터 다음 학년 준비와 피드백까지",
    "신도림동": "학교 시험과 처음 보는 지문을 함께 준비하는 방법",
    "신정동": "영어 성적 변동 원인을 자료로 구분하는 방법",
    "목동": "중3에서 고등 영어로 넘어가는 학습 흐름",
    "구파발": "학교 시험과 근거 독해를 함께 점검하는 방법",
    "후곡마을": "기초의 빈틈과 오답 흐름을 점검하는 방법",
    "구미동": "학습 우선순위를 정하고 꾸준함으로 연결하는 방법",
    "가정동": "막힌 이유부터 찾는 영어 출발점 점검",
    "영통동": "예비 고1을 위한 중등 영어 마무리 순서",
    "서정동": "예비 고1 입학 전 영어 기초 확인",
    "동춘동": "봄방학 동안 중등 영어 기초와 학습 목표를 정리하는 방법",
    "연수동": "독해를 근거 중심으로 점검하는 방법",
    "석우동": "현재 위치에서 다음 학습 단계로 가는 점검 기준",
    "성화동": "문법을 영작으로 연결하는 학습 순서",
    "신천동": "답보다 풀이 과정을 확인하는 학교 시험 준비",
    "화봉동": "예비 고1이 중등 영어 기초를 마무리하는 순서",
    "복산동": "중3 영어를 고등 학습으로 연결하는 준비 순서",
    "수완지구": "방학 영어 학습의 우선순위를 정하는 방법",
    "치평동": "학교 공부와 서술형을 연결하는 학습 흐름",
    "무실동": "기초에서 고등 독해로 넘어가는 순서",
    "석사동": "현재 학습을 고등 독해로 연결하는 점검 순서",
}


def safe_middle_english_phrase(value: str, locality: str = "") -> str:
    value = clean(value)
    if locality:
        value = re.sub(rf"{re.escape(locality)}\s*중등\s*영어학원", " ", value)
        value = re.sub(rf"(?<![가-힣]){re.escape(locality)}(?![가-힣])", " ", value)
    value = re.sub(r"중등\s*영어학원", " ", value)
    value = safe_high_english_phrase(value)
    replacements = (
        (r"내신과\s*모의고사", "학교 시험과 처음 보는 지문"),
        (r"내신·모의고사", "학교 시험·누적 독해"),
        (r"모의고사", "처음 보는 지문"),
        (r"입시(?:준비|전략|상담|로드맵)?", "다음 학년 준비"),
        (r"고3\s*영어", "중등 영어"),
        (r"고2\s*수능\s*대비", "중등 영어 독해 점검"),
        (r"수능\s*영어", "고등 전환 영어"),
        (r"수능\s*기출", "누적 독해 자료"),
        (r"수능", "고등 전환 학습"),
        (r"봄방학\s*수학\s*공부법", "방학 영어 학습 목표를 정리하는 방법"),
        (r"수학", "영어"),
    )
    for pattern, replacement in replacements:
        value = re.sub(pattern, replacement, value)
    compounds = (
        ("학습취약점", "학습 취약점"),
        ("자기주도", "자기 주도"),
        ("시험분석", "시험 분석"),
        ("시험오답", "시험 오답"),
        ("내신시험", "내신 시험"),
        ("일정점검", "일정 점검"),
        ("학습자립도", "학습 자립도"),
        ("내신보충학습", "내신 보충 학습"),
        ("학습진단", "학습 진단"),
        ("학습목표", "학습 목표"),
        ("학습기록", "학습 기록"),
        ("학습이력", "학습 이력"),
        ("대면학습", "대면 학습"),
        ("학습자율성", "학습 자율성"),
        ("학습플래너", "학습 계획표"),
        ("학습실전", "실전 학습"),
        ("화상학습", "화상 학습"),
        ("학습성취도", "학습 성취도"),
        ("내신학습", "내신 학습"),
        ("학교시험", "학교 시험"),
        ("시험복습", "시험 복습"),
        ("시험분석", "시험 분석"),
        ("시험오답", "시험 오답"),
        ("시험전략", "시험 전략"),
        ("구문독해", "구문 독해"),
        ("목표점검", "목표 점검"),
        ("일정점검", "일정 점검"),
        ("내신진도점검", "내신 진도 점검"),
        ("내신상담", "내신 상담"),
        ("문제풀이", "문제 풀이"),
        ("오답노트", "오답 노트"),
        ("자기주도", "자기 주도"),
        ("내신시험", "내신 시험"),
    )
    for before, after in compounds:
        value = value.replace(before, after)
    value = value.replace("고등 영어학원의", "고등 영어의")
    if value.count("학습") > 1:
        value = re.sub(r"학생별\s+학습$", "학생별 계획", value)
        value = re.sub(r"\s+학습\s+상담$", " 상담", value)
        value = re.sub(r"\s+학습\s+점검$", " 점검", value)
        value = re.sub(r"\s+학습$", " 방법", value)
        value = value.replace("학습 방법으로 학습 방향", "공부 방법으로 학습 방향")
        value = value.replace("학습으로 학습 흐름", "학습으로 영어 흐름")
    value = value.replace("과정를", "과정을").replace("점검를", "점검을")
    value = re.sub(r"(?<![가-힣])([가-힣]{2,12})\s+\1(?![가-힣])", r"\1", value)
    value = re.sub(r"\s+", " ", value).strip(" ,·|:-")
    if (
        not value
        or len(value) > 64
        or re.search(r"(?:초등학생|초등학교|고[123]|차량|주차|녹화수업|관리앱|학습리포트)", value)
    ):
        return "중등 영어 학습의 빈틈을 찾는 방법"
    return value


def middle_english_profile(raw: str, locality: str) -> MiddleEnglishProfile:
    h1_values = source_tag_texts(raw, "h1")
    if not h1_values:
        raise ValueError(f"{locality}: middle English source H1 missing")
    source_title = h1_values[0]
    focus = MIDDLE_ENGLISH_FOCUS_OVERRIDES.get(
        locality,
        safe_middle_english_phrase(source_title, locality),
    )

    strong_values = source_tag_texts(raw, "strong")
    h3_values = [re.sub(r"^\d+\.\s*", "", value) for value in source_tag_texts(raw, "h3")]
    li_values = source_tag_texts(raw, "li")
    primary_parts: list[tuple[str, int]] = [(focus, 40)]
    primary_parts.extend((value, 10) for value in strong_values)
    detail_parts: list[tuple[str, int]] = [(source_title, 7), (focus, 14)]
    detail_parts.extend((value, 4) for value in strong_values)
    detail_parts.extend((value, 3) for value in h3_values[-4:])
    detail_parts.extend((value, 1) for value in li_values)
    full_text = source_text(raw)

    def scores(parts: list[tuple[str, int]]) -> dict[str, int]:
        return {
            intent.code: sum(
                weight * sum(text.count(keyword) for keyword in intent.keywords)
                for text, weight in parts
            )
            for intent in MIDDLE_ENGLISH_INTENTS
        }

    primary_scores = scores(primary_parts)
    detail_scores = scores(detail_parts)
    full_scores = {
        intent.code: sum(full_text.count(keyword) for keyword in intent.keywords)
        for intent in MIDDLE_ENGLISH_INTENTS
    }
    focus_ranked = sorted(
        MIDDLE_ENGLISH_INTENTS,
        key=lambda intent: (
            -primary_scores[intent.code],
            -detail_scores[intent.code],
            -full_scores[intent.code],
            intent.code,
        ),
    )
    detail_ranked = sorted(
        MIDDLE_ENGLISH_INTENTS,
        key=lambda intent: (
            -detail_scores[intent.code],
            -full_scores[intent.code],
            -primary_scores[intent.code],
            intent.code,
        ),
    )
    focus_compact = re.sub(r"\s+", "", focus)
    forced_primary_code = None
    if any(marker in focus_compact for marker in ("예비고1", "고등영어전환", "고등과정전환")):
        forced_primary_code = "transition"
    elif any(marker in focus_compact for marker in ("예비중", "중1", "중학교첫", "첫시험")):
        forced_primary_code = "entry"
    primary = (
        MIDDLE_ENGLISH_INTENT_BY_CODE[forced_primary_code]
        if forced_primary_code
        else (
            focus_ranked[0]
            if primary_scores[focus_ranked[0].code] > 0
            else MIDDLE_ENGLISH_INTENT_BY_CODE["diagnosis"]
        )
    )
    # A page aimed at the transition into high school should not drift back to
    # first-year middle-school preparation, and vice versa. These are distinct
    # search intents even when the source title contains the broad word "입학".
    blocked_codes = {
        "entry" if primary.code == "transition" else
        "transition" if primary.code == "entry" else
        ""
    }
    ranked = [primary]
    ranked.extend(
        intent
        for intent in detail_ranked
        if (
            intent not in ranked
            and intent.code not in blocked_codes
            and (detail_scores[intent.code] > 0 or full_scores[intent.code] > 0)
        )
    )
    ranked.extend(
        MIDDLE_ENGLISH_INTENT_BY_CODE[code]
        for code in ("reading", "sentence", "school_exam", "routine", "vocabulary", "error")
        if code not in blocked_codes and MIDDLE_ENGLISH_INTENT_BY_CODE[code] not in ranked
    )
    selected = tuple(ranked[:4])
    return MiddleEnglishProfile(
        focus=focus,
        source_title=source_title,
        intents=selected,
        source_markers=tuple(intent.label for intent in selected),
    )


HIGH_MATH_FOCUS_OVERRIDES = {
    "방화동": "고등 수학 첫 내신을 준비하는 기본기 점검",
    "성복동": "중등 선수 개념을 고등 수학의 풀이 과정으로 연결하는 방법",
    "신중동": "단원별 취약점과 풀이 과정을 연결하는 방법",
    "야당동": "문제 이해와 수학 학습 습관을 함께 점검하는 방법",
    "자은동": "고등 수학 시험 범위와 오답 원인을 함께 점검하는 방법",
    "위례신도시": "학교별 내신 유형과 모의고사 문항 선택을 연결하는 방법",
}


def safe_high_math_phrase(value: str, locality: str = "") -> str:
    """Keep the source angle while removing unsupported operating claims."""

    value = clean(value)
    if locality:
        value = re.sub(rf"{re.escape(locality)}\s*고등\s*수학학원", " ", value)
        value = re.sub(rf"(?<![가-힣]){re.escape(locality)}(?![가-힣])", " ", value)
    value = re.sub(r"고등\s*수학학원", " ", value)
    replacements = (
        (r"와와학습코칭센터(?:\s*\S*점)?", "학습 상담"),
        (r"(?:입시준비반|정규관리반|성적관리반)", "입시 준비 계획"),
        (r"(?:방학집중반|집중반)", "방학 집중 계획"),
        (r"(?:입시특강|내신특강|방학특강|특강수업|특강)", "집중 학습 계획"),
        (r"(?:실시간수업|녹화수업|오전수업|그룹수업|집중수업|참여형\s*수업)", "학습 방법"),
        (r"학원정규반", "정규 학습 계획"),
        (r"시험대비반", "시험 대비 계획"),
        (r"학원보충", "보충 학습"),
        (r"학원비", "비용과 이용 조건"),
        (r"(?:일대일|1:1|소수정예)", "학생별 피드백"),
        (r"맞춤형?", "학생별"),
        (r"개별진도", "학생별 진도"),
        (r"학원진도", "현재 진도"),
        (r"(?:성적\s*향상|내신향상|성적향상)", "성적 변화의 원인 점검"),
        (r"성적을\s*바꾸는", "성적 변화에 영향을 주는"),
        (r"(?:합격\s*전략|합격)", "입시 학습 전략"),
        (r"학습성과", "학습 과정"),
        (r"입시지원전략", "입시 학습 우선순위"),
        (r"입시합격전략", "입시 학습 전략"),
        (r"입시맞춤상담", "입시 학습 상담"),
        (r"밀착", "세밀한"),
        (r"학습코칭", "학습 점검"),
        (r"코칭수업", "피드백 방법"),
        (r"학습관리", "학습 기록 점검"),
        (r"수학관리", "수학 학습 점검"),
        (r"오답관리", "오답 점검"),
        (r"성적관리", "성적 변화 점검"),
        (r"시간관리", "시간 배분 점검"),
        (r"내신관리", "내신 준비 점검"),
        (r"시험관리", "시험 준비 점검"),
        (r"관리앱", "학습 기록"),
        (r"관리", "점검"),
        (r"코칭", "피드백"),
        (r"수업", "학습"),
    )
    for pattern, replacement in replacements:
        value = re.sub(pattern, replacement, value)
    spacing_replacements = (
        ("고등수학", "고등 수학"), ("문제풀이", "문제 풀이"),
        ("풀이과정", "풀이 과정"), ("오답분석", "오답 분석"),
        ("오답복습", "오답 복습"), ("학습방향", "학습 방향"),
        ("학습흐름", "학습 흐름"), ("학습계획", "학습 계획"),
        ("학습훈련", "학습 훈련"),
        ("학습전략", "학습 전략"), ("학습루틴", "학습 루틴"),
        ("학습습관", "학습 습관"), ("학습태도", "학습 태도"),
        ("학습스케줄", "학습 일정"), ("학습로드맵", "학습 로드맵"),
        ("학습약점", "학습 약점"), ("학습목표", "학습 목표"),
        ("학습과정", "학습 과정"), ("학습완성도", "학습 완성도"),
        ("개념이해", "개념 이해"), ("개념연결", "개념 연결"),
        ("개념완성도", "개념 완성도"), ("계산실수", "계산 실수"),
        ("시간배분", "시간 배분"), ("내신전략", "내신 전략"),
        ("내신분석", "내신 분석"), ("시험계획", "시험 계획"),
        ("시험준비", "시험 준비"), ("시험범위", "시험 범위"),
        ("입시전략", "입시 전략"), ("입시계획", "입시 계획"),
        ("입시로드맵", "입시 로드맵"), ("입시점검", "입시 점검"),
        ("입시진단", "입시 진단"), ("입시자료분석", "입시 자료 분석"),
        ("학생별피드백", "학생별 피드백"), ("자기주도학습", "자기주도 학습"),
        ("반복학습", "반복 학습"), ("집중학습", "집중 학습"),
        ("예비고1", "예비 고1"), ("단원별취약점", "단원별 취약점"),
    )
    for before, after in spacing_replacements:
        value = value.replace(before, after)
    value = re.sub(r"(?:학습\s*){2,}", "학습 ", value)
    value = value.replace("학습 방법 학습", "학습 방법").replace("피드백 학습", "피드백 방법")
    value = re.sub(r"\s+", " ", value).strip(" ,·|:-")
    value = re.sub(r"학습 흐름을 세우는 학습 상담$", "학습 흐름을 세우는 상담 기준", value)
    value = re.sub(r"학습 상담$", "상담 기준", value)
    value = re.sub(r"점검하는 학습$", "점검하는 방법", value)
    value = re.sub(r"살피는 학습$", "살피는 방법", value)
    value = re.sub(r"이어지는 학습$", "이어지는 학습 방법", value)
    if value.endswith("안내"):
        value = value.removesuffix("안내").strip() + " 확인 기준"
    elif value.endswith("관리"):
        value = value.removesuffix("관리").strip() + " 점검 방법"
    elif value.endswith("수업"):
        value = value.removesuffix("수업").strip() + " 학습 방법"
    return value


def repair_high_math_surface(value: str) -> str:
    """Repair source-derived Korean joins at the final mathematics boundary."""

    repairs = (
        ("기’을", "기’를"),
        ("점검를", "점검을"),
        ("점검로", "점검으로"),
        ("점검가", "점검이"),
        ("과정를", "과정을"),
        ("풀이 풀이", "풀이"),
        ("문제 문제", "문제"),
        ("새 문제과", "새 문제와"),
        ("수학 학습 루틴의 원인", "수학 학습 루틴의 문제"),
        ("상담 상담", "상담"),
        ("입시입시", "입시"),
        ("전략점검", "전략 점검"),
    )
    for before, after in repairs:
        value = value.replace(before, after)
    return value


MIDDLE_MATH_FOCUS_OVERRIDES = {
    "명일동": "개념을 설명한 뒤 문제에 적용하는 학습 순서",
    "송도": "중3 수학의 기초를 고등 과정 준비로 연결하는 방법",
    "갈현동": "학교 시험과 누적 유형 문제를 위한 수학 학습 점검",
    "갈산동": "학습 데이터로 학생별 상태를 점검하는 방법",
    "금암동": "스스로 공부하는 습관을 기르는 방법",
    "내발산동": "현재 실력부터 차근히 높이는 학습",
    "동패동": "학습 기록으로 성적 변화 원인을 점검하는 방법",
    "북가좌동": "출발 상태부터 살피는 학습 설계",
    "성남 금곡동": "출발 상태와 오답을 함께 점검하는 방법",
    "시흥동": "학교 시험과 누적 유형 문제의 오답 원인을 나누는 방법",
    "목동": "학교 시험과 누적 유형 문제의 풀이 시간을 점검하는 방법",
    "목동동": "학원 변경 전 확인할 오답 점검 기준",
    "경산사동": "수준별 학습 상태부터 살피는 수학 점검",
    "소하동": "학생별 피드백 방법을 살펴 개념부터 확인하기",
    "수원 금곡동": "공부 습관을 기르는 학습 점검",
    "수월동": "학생별 피드백으로 공부 흐름을 잡는 방법",
    "신현동": "풀이 과정을 탄탄하게 만드는 방법",
    "옥정동": "이해와 학습 실행 습관을 함께 기르는 방법",
    "용곡동": "실수 원인부터 정리하는 진학 준비 학습 순서",
    "주월동": "시험 대비 계획과 학습 기록을 함께 점검하는 방법",
    "양평동": "시험 시간 배분부터 점검하는 방법",
    "오산동": "학생의 학습 속도에 맞춘 수학 계획",
    "죽전동": "학교 시험과 누적 유형 문제에서 조건 해석을 비교하는 방법",
    "옥산동": "교과서 유형 변형 문제의 조건을 해석하는 방법",
}


def repair_middle_math_focus(value: str) -> str:
    """Keep a source-led focus compact without nearby repeated content words."""

    value = clean(value)
    repairs = (
        ("학생별 학생별", "학생별"),
        ("진학 준비 준비", "진학 준비"),
        ("학습 전 진단평가", "사전 진단평가"),
        ("학습 전 진단", "사전 진단"),
        ("집중 학습 기록 점검으로 만드는 학습 흐름", "집중 기록 점검으로 만드는 학습 흐름"),
        ("집중 학습 계획과 4주 학습 기록 점검", "집중 계획과 4주 학습 기록 점검"),
        ("학습 방법으로 학습 방향", "학습 방법으로 공부 방향"),
        ("학습 방법으로 학습 흐름", "학습 방법으로 공부 흐름"),
        ("학습 리듬", "공부 리듬"),
        ("학생별 취약점을 찾는 학생별 학습", "학생별 취약점을 찾는 학습 방법"),
        ("성적 기복을 줄이는", "성적 변화의 원인을 점검하는"),
        ("첨삭", "풀이 확인"),
        ("오답노트", "오답 기록"),
        ("준비피드백", "준비 과정 피드백"),
        ("준비종합점검", "준비 전반 점검"),
        ("과제점검학습", "과제 점검 방법"),
        ("내신학습", "내신 학습"),
        ("학습클리닉", "학습 보완"),
        ("학습자립도", "스스로 학습하는 정도"),
        ("학습성취도", "학습 이해도"),
        ("학습기록", "학습 기록"),
        ("학습취약점", "학습 취약점"),
        ("오답 원인과 재풀이와", "오답 원인·재풀이와"),
        ("학습데이터", "학습 데이터"),
        ("학습점검", "학습 점검"),
        ("학생별 점검하는", "학생별로 점검하는"),
        ("조건 해석과 식 세우기", "조건 해석·식 세우기"),
        ("점검하는 학습 점검", "살피는 학습 점검"),
        ("원인을 점검하는 진학 준비 학습 점검", "원인을 살피는 진학 준비 학습 점검"),
        ("시험집중점검", "시험 집중 점검"),
        ("시험성적", "시험 성적"),
        ("학원점검", "학원 점검"),
        ("학습컨설팅", "학습 상담"),
        ("학습이해", "학습 이해"),
        ("학습과제점검", "학습 과제 점검"),
        ("학습성적", "학습 성적"),
        ("개별학습", "개별 학습"),
        ("학습상담", "학습 상담"),
        ("과목별공부", "과목별 공부"),
        ("학습집중점검", "학습 집중 점검"),
        ("내신성적", "내신 성적"),
        ("진학 준비정보", "진학 준비 정보"),
        ("진학 준비일정", "진학 준비 일정"),
        ("진학 준비정기상담", "진학 준비 정기 상담"),
        ("시험시간", "시험 시간"),
        ("과제점검", "과제 점검"),
        ("학습자신감", "학습 자신감"),
        ("내신클리닉", "내신 보완"),
        ("학습시간표", "학습 시간표"),
        ("학습실천력", "학습 실천력"),
        ("학습일정", "학습 일정"),
        ("학습동기", "학습 동기"),
        ("집중점검학습", "집중 점검 학습"),
        ("시험대비학습", "시험 대비 학습"),
        ("시험점검", "시험 점검"),
        ("학습복습", "복습"),
        ("학습예습", "예습"),
        ("피드백로", "피드백으로"),
        ("진학 준비를 준비하는", "진학 준비 과정을 설계하는"),
        ("현재 성적 기록 학습 전략", "현재 성적 기록을 활용한 학습 전략"),
        ("자기주도", "자기 주도"),
        ("시험분석", "시험 분석"),
        ("시험오답", "시험 오답"),
        ("내신시험", "내신 시험"),
        ("일정점검", "일정 점검"),
    )
    for before, after in repairs:
        value = value.replace(before, after)
    value = re.sub(r"[가-힣]+점$", "방법", value)
    if value.endswith(" 학습") and value.count("학습") > 1:
        value = value.removesuffix(" 학습") + " 방법"
    if re.search(r"(?:하는|돕는|키우는|찾는)$", value):
        value = f"{value} 방법"
    value = re.sub(
        r"(?:기초부터|학생별|시험 대비)?\s*학습 흐름을 (세우는|바로잡는) 방법$",
        lambda match: clean(
            f"{match.group(0).split('학습 흐름', 1)[0]}공부 흐름을 {match.group(1)} 방법"
        ),
        value,
    )
    value = value.replace("학생별 학습 변화 점검에 맞춘 방법", "학생별 변화 점검에 맞춘 학습 방법")
    value = re.sub(r"\s+", " ", value).strip(" ,·|:-")
    return value


def safe_middle_math_phrase(value: str, locality: str = "") -> str:
    """Keep the source angle while removing unsupported middle-math claims."""

    value = clean(value)
    if locality:
        value = re.sub(
            rf"{re.escape(locality)}\s*중등\s*수학학원",
            f"{locality} 고등 수학학원",
            value,
        )
    value = re.sub(r"중등\s*수학학원", "고등 수학학원", value)
    value = safe_high_math_phrase(value, locality)
    replacements = (
        ("고등 수능까지 보는 학습 전략", "중3 수학을 고등 과정 준비로 연결하는 방법"),
        ("고등 누적 학습까지 보는 학습 전략", "중3 수학을 고등 과정 준비로 연결하는 방법"),
        ("모의고사", "누적 유형 문제"),
        ("수능형", "여러 단원이 섞인"),
        ("수능", "누적 학습"),
        ("입시", "진학 준비"),
        ("본문 변형문제", "유형 변형 문제"),
        ("본문 변형 문제", "유형 변형 문제"),
        ("학습향상", "학습 변화 점검"),
        ("학습 향상", "학습 변화 점검"),
        ("학습료", "비용과 이용 조건"),
        ("시험성적 향상", "시험 결과의 원인 점검"),
        ("시험 성적 향상", "시험 결과의 원인 점검"),
        ("흔들리지 않는 실력 완성", "기초와 풀이 과정 점검"),
        ("2등급", "현재 성적 기록"),
        ("성과", "학습 변화"),
        ("정규반", "정규 학습 계획"),
        ("실시간 학습", "학습 방법"),
        ("주말 학습 점검", "주간 학습 기록 점검"),
        ("주말 학습 피드백", "주간 학습 피드백"),
        ("오후 학습", "학습 시간대 확인"),
        ("오후학습", "학습 시간대 확인"),
        ("진학 준비 정기 상담", "진학 준비 질문"),
        ("진학 준비 종합 점검", "진학 준비 학습 점검"),
        ("진학 준비 분석", "진학 준비 자료 점검"),
        ("진학 준비 정보", "진학 준비 확인 자료"),
        ("진학 준비 로드맵", "진학 준비 학습 순서"),
        ("진학 준비 진단", "진학 준비 기초 점검"),
        ("진학 준비 평가", "진학 준비 학습 점검"),
        ("진학 준비 일정", "진학 준비 학습 일정"),
    )
    for before, after in replacements:
        value = value.replace(before, after)
    value = value.replace("중등수학", "중등 수학")
    value = re.sub(r"(?:학습\s*){2,}", "학습 ", value)
    value = re.sub(r"\s+", " ", value).strip(" ,·|:-")
    return repair_middle_math_focus(value)


def middle_math_profile(raw: str, locality: str) -> HighMathProfile:
    h1_values = source_tag_texts(raw, "h1")
    if not h1_values:
        raise ValueError(f"{locality}: middle Math source H1 missing")
    source_title = h1_values[0]
    focus = MIDDLE_MATH_FOCUS_OVERRIDES.get(
        locality,
        safe_middle_math_phrase(source_title, locality),
    )
    focus = repair_middle_math_focus(focus)

    strong_values = source_tag_texts(raw, "strong")
    h3_values = [re.sub(r"^\d+\.\s*", "", value) for value in source_tag_texts(raw, "h3")]
    li_values = source_tag_texts(raw, "li")
    primary_parts: list[tuple[str, int]] = [(focus, 40)]
    primary_parts.extend((value, 10) for value in strong_values)
    detail_parts: list[tuple[str, int]] = [(source_title, 7), (focus, 14)]
    detail_parts.extend((value, 4) for value in strong_values)
    detail_parts.extend((value, 3) for value in h3_values[-4:])
    detail_parts.extend((value, 1) for value in li_values)
    full_text = source_text(raw)

    def scores(parts: list[tuple[str, int]]) -> dict[str, int]:
        return {
            intent.code: sum(
                weight * sum(text.count(keyword) for keyword in intent.keywords)
                for text, weight in parts
            )
            for intent in MIDDLE_MATH_INTENTS
        }

    primary_scores = scores(primary_parts)
    detail_scores = scores(detail_parts)
    full_scores = {
        intent.code: sum(full_text.count(keyword) for keyword in intent.keywords)
        for intent in MIDDLE_MATH_INTENTS
    }
    focus_ranked = sorted(
        MIDDLE_MATH_INTENTS,
        key=lambda intent: (
            -primary_scores[intent.code],
            -detail_scores[intent.code],
            -full_scores[intent.code],
            intent.code,
        ),
    )
    detail_ranked = sorted(
        MIDDLE_MATH_INTENTS,
        key=lambda intent: (
            -detail_scores[intent.code],
            -full_scores[intent.code],
            -primary_scores[intent.code],
            intent.code,
        ),
    )
    focus_compact = re.sub(r"\s+", "", focus)
    forced_primary_code = None
    if any(marker in focus_compact for marker in ("예비고1", "예비고등", "고등")):
        forced_primary_code = "transition"
    elif any(marker in focus_compact for marker in ("예비중", "중1", "중학교첫", "첫시험")):
        forced_primary_code = "entry"
    eligible_focus_ranked = [
        intent
        for intent in focus_ranked
        if forced_primary_code == "transition" or intent.code != "transition"
    ]
    primary = (
        MIDDLE_MATH_INTENT_BY_CODE[forced_primary_code]
        if forced_primary_code
        else (
            eligible_focus_ranked[0]
            if primary_scores[eligible_focus_ranked[0].code] > 0
            else MIDDLE_MATH_INTENT_BY_CODE["concept"]
        )
    )
    blocked_codes: set[str] = set()
    if primary.code == "transition":
        blocked_codes.add("entry")
    else:
        blocked_codes.add("transition")
    ranked = [primary]
    ranked.extend(
        intent
        for intent in detail_ranked
        if (
            intent not in ranked
            and intent.code not in blocked_codes
            and (detail_scores[intent.code] > 0 or full_scores[intent.code] > 0)
        )
    )
    ranked.extend(
        MIDDLE_MATH_INTENT_BY_CODE[code]
        for code in ("condition", "error", "school_exam", "units", "concept", "routine", "calculation", "written")
        if code not in blocked_codes and MIDDLE_MATH_INTENT_BY_CODE[code] not in ranked
    )
    selected = tuple(ranked[:4])
    if (
        not focus
        or len(focus) > 64
        or re.search(r"(?:초등학생|초등학교|영어|독해|구문|영작|듣기|말하기|고[23])", focus)
    ):
        focus = f"{selected[0].label}과 {selected[1].label}을 중등 수학 계획에 함께 반영하는 방법"
    focus = repair_high_math_surface(
        normalize_particle_joins(focus, tuple(intent.label for intent in selected))
    )
    focus = repair_middle_math_focus(focus)
    return HighMathProfile(
        focus=focus,
        source_title=source_title,
        intents=selected,
        source_markers=tuple(intent.label for intent in selected),
    )


def high_math_profile(raw: str, locality: str) -> HighMathProfile:
    h1_values = source_tag_texts(raw, "h1")
    if not h1_values:
        raise ValueError(f"{locality}: high Math source H1 missing")
    source_title = h1_values[0]
    focus = HIGH_MATH_FOCUS_OVERRIDES.get(locality, safe_high_math_phrase(source_title, locality))

    strong_values = source_tag_texts(raw, "strong")
    h3_values = [re.sub(r"^\d+\.\s*", "", value) for value in source_tag_texts(raw, "h3")]
    li_values = source_tag_texts(raw, "li")
    primary_parts: list[tuple[str, int]] = [(focus, 40)]
    primary_parts.extend((value, 10) for value in strong_values)
    primary_scores = {
        intent.code: sum(weight * sum(text.count(keyword) for keyword in intent.keywords) for text, weight in primary_parts)
        for intent in HIGH_MATH_INTENTS
    }
    detail_parts: list[tuple[str, int]] = [(source_title, 7), (focus, 14)]
    detail_parts.extend((value, 4) for value in strong_values)
    detail_parts.extend((value, 3) for value in h3_values[-4:])
    detail_parts.extend((value, 1) for value in li_values)
    detail_scores = {
        intent.code: sum(weight * sum(text.count(keyword) for keyword in intent.keywords) for text, weight in detail_parts)
        for intent in HIGH_MATH_INTENTS
    }
    full_text = source_text(raw)
    full_scores = {
        intent.code: sum(full_text.count(keyword) for keyword in intent.keywords)
        for intent in HIGH_MATH_INTENTS
    }
    focus_ranked = sorted(
        HIGH_MATH_INTENTS,
        key=lambda intent: (-primary_scores[intent.code], -detail_scores[intent.code], -full_scores[intent.code], intent.code),
    )
    detail_ranked = sorted(
        HIGH_MATH_INTENTS,
        key=lambda intent: (-detail_scores[intent.code], -full_scores[intent.code], -primary_scores[intent.code], intent.code),
    )
    full_ranked = sorted(
        HIGH_MATH_INTENTS,
        key=lambda intent: (-full_scores[intent.code], -detail_scores[intent.code], intent.code),
    )
    if primary_scores[focus_ranked[0].code] > 0:
        primary = focus_ranked[0]
    elif detail_scores[detail_ranked[0].code] > 0:
        primary = detail_ranked[0]
    elif full_scores[full_ranked[0].code] > 0:
        primary = full_ranked[0]
    else:
        primary = HIGH_MATH_INTENT_BY_CODE["concept"]
    ranked = [primary]
    ranked.extend(
        intent
        for intent in detail_ranked
        if intent not in ranked and (detail_scores[intent.code] > 0 or full_scores[intent.code] > 0)
    )
    ranked.extend(
        HIGH_MATH_INTENT_BY_CODE[code]
        for code in ("condition", "error", "school_exam", "units", "concept", "routine")
        if HIGH_MATH_INTENT_BY_CODE[code] not in ranked
    )
    if (
        not focus
        or len(focus) > 64
        or re.search(r"(?:초등|중1|중2|중3|중학생|중학교|영어|독해|구문|차량|주차)", focus)
    ):
        focus = f"{ranked[0].label}과 {ranked[1].label}을 고등 수학 계획에 함께 반영하는 방법"
    focus = repair_high_math_surface(
        normalize_particle_joins(focus, tuple(intent.label for intent in ranked[:4]))
    )
    return HighMathProfile(
        focus=focus,
        source_title=source_title,
        intents=tuple(ranked[:4]),
        source_markers=tuple(intent.label for intent in ranked[:4]),
    )


def high_math_particle_tokens(profile: HighMathProfile) -> tuple[str, ...]:
    values: list[str] = [profile.focus, *profile.source_markers]
    for intent in profile.intents:
        values.extend((
            intent.label,
            intent.concern,
            intent.evidence,
            intent.action,
            intent.checkpoint,
            intent.exam_use,
            intent.consult_question,
        ))
    return tuple(dict.fromkeys(value for value in values if value))


def mathify_text(value: str) -> str:
    replacements = (
        ("범위·본문·어법·서술형", "범위·개념·유형·서술형"),
        ("본문·어법·서술형", "개념·유형·서술형"),
        ("교과서 변형 문장", "조건이 바뀐 교과서 문제"),
        ("학교 범위와 새 지문", "학교 범위와 처음 보는 문제"),
        ("처음 보는 지문", "처음 보는 문제"),
        ("새 지문", "새 문제"),
        ("지문", "문제"),
        ("읽기·판단·검토", "풀이·판단·검산"),
        ("행동과 새 문제에서 확인할 행동", "연습 항목과 새 문제에서 확인할 기준"),
        ("학교·학년", "학교와 학년"),
        ("현재선을", "출발 상태를"),
        ("현재선", "출발 상태"),
        ("고등 영어", "고등 수학"),
        ("영어", "수학"),
    )
    for before, after in replacements:
        value = value.replace(before, after)
    return clean(value)


def mathify_tree(value: object) -> object:
    if isinstance(value, str):
        return mathify_text(value)
    if isinstance(value, list):
        return [mathify_tree(item) for item in value]
    if isinstance(value, tuple):
        return tuple(mathify_tree(item) for item in value)
    if isinstance(value, dict):
        return {key: mathify_tree(item) for key, item in value.items()}
    return value


def naturalize_high_math_text(value: str, profile: HighMathProfile) -> str:
    value = mathify_text(value)
    primary, secondary, support, extra = profile.intents
    for first, second in ((primary, secondary), (secondary, primary), (support, extra), (extra, support)):
        value = value.replace(f"{first.label}·{second.label}", f"{first.label}과 {second.label}")
    for intent in profile.intents:
        object_particle = particle_for(intent.label, "을", "를")
        for particle in ("을", "를"):
            value = value.replace(
                f"{intent.label}{particle} 먼저 설명",
                f"{intent.label}에서 막힌 이유를 먼저 설명",
            )
        value = value.replace(
            f"{intent.label}의 재확인 날짜",
            f"{intent.label}{object_particle} 다시 확인할 날짜",
        )
    # English-template particles and noun joins can become malformed after the
    # subject vocabulary is converted to mathematics. Repair those joins at
    # the final math-only boundary so the English renderer remains unchanged.
    value = repair_high_math_surface(value)
    return repair_high_math_surface(
        normalize_particle_joins(value, tuple(intent.label for intent in profile.intents))
    )


def naturalize_high_math_tree(value: object, profile: HighMathProfile) -> object:
    if isinstance(value, str):
        return naturalize_high_math_text(value, profile)
    if isinstance(value, list):
        return [naturalize_high_math_tree(item, profile) for item in value]
    if isinstance(value, tuple):
        return tuple(naturalize_high_math_tree(item, profile) for item in value)
    if isinstance(value, dict):
        return {key: naturalize_high_math_tree(item, profile) for key, item in value.items()}
    return value


def high_math_focus_guidance(profile: HighMathProfile) -> tuple[str, str, str]:
    focus = profile.focus
    compact = re.sub(r"\s+", "", focus)
    primary, secondary = profile.intents[:2]
    if any(token in compact for token in ("불안", "긴장", "자신감", "다시시작")):
        return (
            "어려운 문제에서 멈춘 위치와 다시 풀이를 시작한 시각, 도움 없이 적은 첫 단계를 표시하세요",
            "짧은 시간 제한 세트에서 아는 조건부터 쓰고 건너뛴 뒤 돌아오는 순서를 연습하세요",
            "일주일 뒤 같은 난도의 새 문제에서 풀이를 다시 시작하는 시간이 줄었는지 비교하세요",
        )
    if any(token in compact for token in ("서술형", "풀이과정", "답안", "논리")):
        return (
            "최근 서술형 답안에서 빠진 조건·등식·결론을 다른 표시로 나누세요",
            "조건, 사용 개념, 계산, 결론을 단계별로 다시 쓰고 각 식의 이유를 한 줄로 남기세요",
            "유사 문제에서도 같은 풀이 순서를 유지하고 생략한 조건 없이 설명하는지 확인하세요",
        )
    if any(token in compact for token in ("시간", "속도", "끝까지", "시험관리")):
        return (
            "문항별 시작·종료 시각과 건너뛴 문제, 다시 돌아온 순서를 시험지에 표시하세요",
            "바로 풀 문제·표시 후 돌아올 문제·마지막에 검산할 문제로 나누어 짧은 세트를 푸세요",
            "다음 세트에서 정확도를 유지하면서 검산 시간을 확보했는지 비교하세요",
        )
    if any(token in compact for token in ("그래프", "함수", "도형", "좌표", "기하")):
        return (
            "식·그래프·표·도형 중 조건을 잘못 옮긴 위치와 그때 세운 식을 함께 남기세요",
            "한 관계를 식과 그림으로 각각 나타내고 두 표현에서 같은 조건을 연결하세요",
            "표현 방식이 바뀐 새 문제에서도 교점·범위·변화 방향을 스스로 설명하는지 확인하세요",
        )
    if any(token in compact for token in ("계산", "실수", "정확")):
        return (
            "첫 풀이의 중간 계산을 지우지 말고 부호·괄호·전개 중 오류가 시작된 줄을 표시하세요",
            "오류가 난 한 단계만 고친 뒤 검산 방법을 적고 숫자가 바뀐 문제를 다시 푸세요",
            "다음 점검에서는 같은 계산 오류 없이 풀이와 검산을 끝냈는지 확인하세요",
        )
    if any(token in compact for token in ("오답", "복습", "재풀이", "틀린")):
        return (
            "틀린 문제를 개념·조건·계산·시간으로 나누고 첫 풀이와 수정 풀이를 함께 보존하세요",
            "같은 날의 풀이 수정과 며칠 뒤의 새 문제 재확인을 서로 다른 일정으로 적으세요",
            "정답을 기억하지 않은 상태에서도 비슷한 문제의 풀이 순서를 설명하는지 확인하세요",
        )
    if any(token in compact for token in ("내신", "중간고사", "기말고사", "학교시험")):
        return (
            "시험 범위표를 개념·대표 유형·서술형·누적 오답으로 나누고 미완료 항목을 표시하세요",
            "교과서 예제와 학교 자료에서 조건이 바뀐 문제를 골라 풀이 이유까지 다시 적으세요",
            "시험 전에는 새로운 문제 수보다 각 범위의 설명·재풀이 완료 여부를 확인하세요",
        )
    if any(token in compact for token in ("수능", "모의고사", "기출", "입시")):
        return (
            "최근 기출이나 모의고사에서 사용 개념·첫 풀이 경로·소요 시간을 문항별로 적으세요",
            "오답을 지식 부족과 문항 선택 문제로 나누고 같은 유형을 제한 시간 안에 다시 푸세요",
            "다음 회차에서 풀이 시작점과 문항 선택 순서가 안정되는지 비교하세요",
        )
    if any(token in compact for token in ("루틴", "습관", "학습량", "스케줄", "계획")):
        return (
            "계획한 문제와 실제 완료한 문제, 남은 질문과 오답 재확인 항목을 다른 칸에 적으세요",
            "새 문제·재풀이·질문 정리의 최소 분량을 정하고 완료 근거를 한 줄씩 남기세요",
            "일주일 뒤 미완료 이유와 반복 오답을 보고 다음 분량과 난도를 조정하세요",
        )
    return (
        f"최근 시험지에서 ‘{primary.evidence}’와 ‘{secondary.evidence}’가 드러난 위치를 한 곳씩 표시하세요",
        f"이번 주에는 ‘{primary.action}’을 먼저 하고 ‘{secondary.action}’은 별도 문제에서 연습하세요",
        f"다음 점검에서 ‘{primary.checkpoint}’와 ‘{secondary.checkpoint}’에 학생이 직접 답하는지 확인하세요",
    )


def high_math_meta_description(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighMathProfile,
) -> str:
    locality = str(center["locality"])
    title = f"{locality} {config.label}"
    primary, secondary = profile.intents[:2]
    candidates = (
        f"{title} 선택 전 {profile.focus}, {primary.label}과 {secondary.label} 점검, 내신·모의고사 계획과 센터 정보를 확인하세요.",
        f"{title}에서 {primary.label}과 {secondary.label}을 진단할 자료, 내신·모의고사 학습 순서와 센터·가능 학년 정보를 안내합니다.",
        f"{title}의 {profile.focus}, 최근 풀이 기록을 확인하는 방법과 내신·모의고사 계획, 센터 정보를 정리했습니다.",
        f"{title} 상담 전 {primary.label}과 {secondary.label}의 풀이 기록, 시험별 학습 순서와 확인된 센터 정보를 살펴보세요.",
    )
    for candidate in candidates:
        candidate = naturalize_high_math_text(candidate, profile)
        if 70 <= len(candidate) <= 100:
            return candidate
    raise ValueError(f"high Math meta description invalid: {title} / {profile.focus}")


def high_math_student_type(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighMathProfile,
    seed: str,
) -> str:
    return naturalize_high_math_text(high_english_student_type(config, center, profile, seed), profile)


def high_math_quick_answer(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighMathProfile,
    seed: str,
) -> str:
    locality = str(center["locality"])
    diagnosis, action, checkpoint = high_math_focus_guidance(profile)
    introduction = stable_pick(seed, "high-math-focus-introduction", (
        f"{locality}에서 ‘{profile.focus}’을 판단할 때는 최근 점수보다 풀이 흔적을 먼저 봐야 합니다.",
        f"{locality} 고등 수학의 이번 초점은 ‘{profile.focus}’이며, 실제 시험지에서 시작합니다.",
        f"‘{profile.focus}’이 필요한 {locality} 학생이라면 개념 설명과 첫 풀이 기록을 함께 확인하세요.",
        f"이 페이지는 {locality} 학생의 ‘{profile.focus}’을 자료·행동·재확인의 순서로 정리합니다.",
        f"{locality} 고등 수학에서 ‘{profile.focus}’을 살필 때는 문제 수보다 학생이 남긴 풀이 과정을 봅니다.",
        f"{locality} 학생에게 필요한 ‘{profile.focus}’은 최근 문제 한 장에서 구체화할 수 있습니다.",
    ))
    stage_labels = stable_pick(seed, "high-math-focus-stage-labels", (
        ("진단 자료", "이번 행동", "재확인"),
        ("먼저 볼 기록", "7일 실행", "다음 판단"),
        ("현재 상태", "연습 방법", "확인 기준"),
        ("출발 자료", "바꿀 행동", "일주일 뒤"),
    ))
    return " ".join((
        introduction,
        f"{stage_labels[0]}: {diagnosis}.",
        f"{stage_labels[1]}: {action}.",
        f"{stage_labels[2]}: {checkpoint}.",
    ))


def high_math_proxy_center(center: dict[str, object]) -> dict[str, object]:
    proxy = dict(center)
    grades = dict(center["grades"])  # type: ignore[arg-type]
    grades["영어"] = list(grades.get("수학", []))
    proxy["grades"] = grades
    return proxy


def high_math_sections(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighMathProfile,
    seed: str,
) -> list[dict[str, object]]:
    sections = naturalize_high_math_tree(
        high_english_sections(config, high_math_proxy_center(center), profile, seed),
        profile,
    )
    primary, secondary = profile.intents[:2]
    locality = str(center["locality"])
    diagnosis_heading = f"{locality} 고등 수학 진단: {profile.focus}"
    if len(diagnosis_heading) > 50:
        diagnosis_heading = f"{locality} 고등 수학: {primary.label} 진단과 {secondary.label} 확인"
    facts_heading = stable_pick(seed, "high-math-heading-facts", (
        f"{profile.focus}: 확인된 센터 사실과 이용 조건",
        f"{profile.focus} 상담과 센터 정보를 구분하는 법",
        f"{profile.focus}에 필요한 학교·학년 확인 기준",
        f"{profile.focus}: 학습 질문과 센터 사실 나누기",
        f"{profile.focus} 상담 전 확인할 주소·학년 정보",
        f"{profile.focus}와 센터 이용 조건을 따로 확인하는 법",
    ))
    if len(facts_heading) > 50:
        facts_heading = f"{primary.label}과 {secondary.label}: 센터 사실과 이용 조건"
    headings = (
        diagnosis_heading,
        stable_pick(seed, "high-math-heading-evidence", (
            f"{primary.label}과 {secondary.label}의 풀이 기록을 비교하는 방법",
            f"최근 시험지에서 {primary.label}과 {secondary.label}을 구분하는 기준",
            f"{primary.label}과 {secondary.label}을 보여 주는 두 가지 학습 자료",
            f"정답보다 먼저 확인할 {primary.label}과 {secondary.label}의 풀이 흔적",
            f"{primary.label} 기록과 {secondary.label} 문제를 함께 살피는 순서",
            f"{primary.label}과 {secondary.label}을 다시 확인할 자료 고르기",
        )),
        stable_pick(seed, "high-math-heading-exam", (
            f"내신과 모의고사에서 {primary.label}과 {secondary.label}을 나누는 방법",
            f"학교 시험과 모의고사에 {primary.label}과 {secondary.label}을 배치하는 기준",
            f"시험 전후 {primary.label}과 {secondary.label}의 비중을 조정하는 방법",
            f"내신 범위와 누적 학습에서 {primary.label}과 {secondary.label}을 확인하는 순서",
            f"{primary.label}과 {secondary.label}의 내신 일정과 재확인일을 나누는 기준",
            f"내신 자료와 모의고사 기록으로 {primary.label}과 {secondary.label}을 비교하는 방법",
        )),
        facts_heading,
        stable_pick(seed, "high-math-heading-plan", (
            f"7일 동안 {primary.label}을 연습하고 {secondary.label}을 확인하는 방법",
            f"{primary.label}에서 {secondary.label}으로 이어지는 7일 실행안",
            f"한 가지 풀이 병목부터 시작하는 {primary.label} 7일 계획",
            f"{primary.label}과 {secondary.label}의 분량을 정하는 7일 점검표",
            f"첫 풀이와 새 문제 확인을 연결하는 {primary.label} 7일 계획",
            f"{primary.label} 연습일과 {secondary.label} 확인일을 정하는 방법",
        )),
        stable_pick(seed, "high-math-heading-consult", (
            f"{primary.label}과 {secondary.label} 상담 전 준비할 자료와 질문",
            f"상담에서 {primary.label}과 {secondary.label}을 확인하는 순서",
            f"최근 시험지로 준비하는 {primary.label} 상담 질문",
            f"{secondary.label} 확인 질문과 센터 이용 조건을 나누는 방법",
            f"{primary.label}과 {secondary.label} 답변을 실행 계획으로 바꾸는 법",
            f"{primary.label}과 {secondary.label} 상담 전 가져갈 자료와 확인할 사실",
        )),
    )
    compact_fallbacks = (
        f"{locality} 고등 수학: {primary.label} 진단 기준",
        f"{primary.label}과 {secondary.label}의 풀이 기록 비교",
        f"내신·모의고사 {primary.label}과 {secondary.label} 점검",
        f"{primary.label}과 {secondary.label}: 센터 사실 확인",
        f"{primary.label}과 {secondary.label}의 7일 실행안",
        f"{primary.label}과 {secondary.label} 상담 질문",
    )
    heading_repairs = (
        ("함수·그래프·도형 해석", "함수·그래프 해석"),
        ("내신 범위·학교별 출제 유형", "학교별 내신 유형"),
        ("모의고사 오답·문항 선택", "모의고사 문항 선택"),
        ("선수 개념·단원 연결", "선수 개념 연결"),
        ("시험 긴장·수학 자신감", "시험 긴장과 자신감"),
        ("조건 해석·식 세우기", "조건 해석과 식 세우기"),
        ("오답 원인·재풀이", "오답 원인과 재풀이"),
    )
    for section, heading, fallback in zip(sections, headings, compact_fallbacks):
        normalized = naturalize_high_math_text(heading, profile)
        for before, after in heading_repairs:
            normalized = normalized.replace(before, after)
        compact_fallback = naturalize_high_math_text(fallback, profile)
        for before, after in heading_repairs:
            compact_fallback = compact_fallback.replace(before, after)
        section["heading"] = normalized if len(normalized) <= 50 and normalized.count("·") <= 3 else compact_fallback
    authored_text = clean(" ".join(
        str(value)
        for section in sections
        for value in (section["heading"], *section["paragraphs"])
    ))
    if profile.focus not in authored_text:
        sections[0]["paragraphs"][0] = clean(
            f"{locality} 고등 수학의 이번 초점은 ‘{profile.focus}’입니다. "
            f"{sections[0]['paragraphs'][0]}"
        )
    return sections  # type: ignore[return-value]


def high_math_faq(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighMathProfile,
    seed: str,
) -> list[tuple[str, str]]:
    return naturalize_high_math_tree(
        high_english_faq(config, high_math_proxy_center(center), profile, seed),
        profile,
    )  # type: ignore[return-value]


def high_math_scenarios(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighMathProfile,
    seed: str,
) -> list[str]:
    scenarios = naturalize_high_math_tree(
        high_english_scenarios(config, high_math_proxy_center(center), profile, seed),
        profile,
    )
    if not relevant_schools(config, center):
        locality = str(center["locality"])
        grades = relevant_grades(config, center, "수학")
        grade_condition = f"수학 가능 학년({'·'.join(grades)})" if grades else "고등 수학의 현재 개설 학년"
        replacement = (
            f"{locality} 페이지에 참고 고등학교 목록이 제공되지 않은 상황을 가정했습니다. "
            f"학부모는 실제 재학 학교의 시험 범위표와 {grade_condition}을 제시하고, "
            "현재 수업에 해당 자료를 반영할 수 있는지 별도로 확인합니다."
        )
        scenarios = [
            replacement if "참고 학교와 실제 재학 학교가 다른 경우" in value else value
            for value in scenarios
        ]
    return scenarios  # type: ignore[return-value]


def middle_math_particle_tokens(profile: HighMathProfile) -> tuple[str, ...]:
    return high_math_particle_tokens(profile)


def naturalize_middle_math_text(value: str, profile: HighMathProfile) -> str:
    """Convert shared high-math prose into evidence-led middle-math prose."""

    value = naturalize_high_math_text(value, profile)
    protected_values: list[str] = [profile.focus, *profile.source_markers]
    if profile.intents[0].code == "transition":
        protected_values.extend((
            "고등 수학에서 다시 쓰일 중등 선수 개념",
            "고등 수학에도 다시 쓰이는 중등 선수 개념",
            "고등 수준 예시 문제",
        ))
    for intent in profile.intents:
        protected_values.extend((
            intent.label,
            intent.concern,
            intent.evidence,
            intent.action,
            intent.checkpoint,
            intent.exam_use,
            intent.consult_question,
        ))
    protected: list[tuple[str, str]] = []
    for index, protected_value in enumerate(sorted(set(protected_values), key=len, reverse=True)):
        if protected_value and protected_value in value:
            marker = f"__MIDDLE_MATH_VALUE_{index}__"
            value = value.replace(protected_value, marker)
            protected.append((marker, protected_value))
    replacements = (
        ("내신과 모의고사", "학교 시험과 누적 유형 문제"),
        ("내신·모의고사", "학교 시험·누적 유형 문제"),
        ("모의고사 시험지", "누적 유형 문제 기록"),
        ("모의고사 새 문제", "새 누적 유형 문제"),
        ("모의고사 누적", "누적 유형 문제"),
        ("모의고사의", "누적 유형 문제의"),
        ("모의고사에서", "누적 유형 문제에서"),
        ("모의고사에는", "누적 유형 문제에는"),
        ("모의고사에", "누적 유형 문제에"),
        ("모의고사는", "누적 유형 문제는"),
        ("모의고사를", "누적 유형 문제를"),
        ("모의고사와", "누적 유형 문제와"),
        ("모의고사", "누적 유형 문제"),
        ("수능형", "여러 단원이 섞인"),
        ("수능", "누적 학습"),
        ("기출", "단원별 대표 문제"),
        ("입시", "진학 준비"),
        ("고등 수학", "중등 수학"),
        ("고등학생", "중학생"),
        ("고등학교", "중학교"),
    )
    for before, after in replacements:
        value = value.replace(before, after)
    value = value.replace("학교 시험과 누적 유형 문제 자료", "학교 시험 자료와 누적 유형 문제 기록")
    value = value.replace("학교 시험·누적 유형 문제 학습", "학교 시험과 누적 유형 문제 학습")
    value = value.replace("누적 유형 문제 누적 유형 문제", "누적 유형 문제")
    value = value.replace("누적 학습 누적 학습", "누적 학습")
    for marker, protected_value in protected:
        value = value.replace(marker, protected_value)
    primary, secondary = profile.intents[:2]

    def sentence_label(intent: HighMathIntent) -> str:
        label = intent.label
        if label.startswith("내신 범위·"):
            return label.removeprefix("내신 범위·")
        if label.startswith("조건 해석·"):
            return label.removeprefix("조건 해석·")
        return label

    primary_sentence_label = sentence_label(primary)
    secondary_sentence_label = sentence_label(secondary)
    for left in profile.intents:
        for right in profile.intents:
            if left is right:
                continue
            value = value.replace(
                f"{left.label}·{right.label}",
                f"{left.label}{particle_for(left.label, '과', '와')} {right.label}",
            )
    value = value.replace(
        "답은 성적표가 아니라 최근 시험지에서 확인합니다.",
        "답은 최근 시험지의 풀이 흔적에서 확인합니다.",
    )
    value = value.replace(
        "시험지의 표시와 재풀이 결과를 각각 이어서 봅니다.",
        "첫 풀이 표시와 재풀이 결과를 영역별로 비교합니다.",
    )
    value = re.sub(
        r"다음 점검에서도 같은 시험지를 사용하되 질문은 ‘([^’]+)’와 ‘([^’]+)’로 고정합니다\. "
        r"이 기준이면 처음 세운 학습 방향의 변화 여부를 확인하기 쉽습니다\.",
        r"다음 점검에서는 숫자와 조건이 바뀐 새 문제를 사용하고 ‘\1’와 ‘\2’에 같은 기준으로 답하는지 봅니다. "
        r"첫 기록과 비교하면 학습 방향의 변화를 확인하기 쉽습니다.",
        value,
    )
    value = value.replace(
        "다음 점검에서도 같은 시험지를 사용하되",
        "다음 점검에서는 숫자와 조건이 바뀐 새 문제를 사용하고",
    )

    def planning_evidence(evidence: str) -> bool:
        return any(marker in evidence for marker in (
            "계획표", "학습 기록", "완료 시각", "시작·완료", "달력",
            "진단표", "일주일 학습 기록", "마감일", "완료일",
        ))

    def replace_score_sheet(match: re.Match[str]) -> str:
        evidence = match.group(1)
        if planning_evidence(evidence):
            subject_particle = particle_for(evidence, "이", "가")
            return (
                f"점수표보다 최근 시험지와 ‘{evidence}’{subject_particle} 담긴 "
                "계획 기록을 별도 자료로 준비합니다."
            )
        object_particle = particle_for(evidence, "을", "를")
        return f"점수표보다 최근 시험지에서 ‘{evidence}’{object_particle} 먼저 확인합니다."

    def replace_exam_sheet_baseline(match: re.Match[str]) -> str:
        evidence = match.group(1)
        if planning_evidence(evidence):
            subject_particle = particle_for(evidence, "이", "가")
            return (
                f"최근 시험지와 ‘{evidence}’{subject_particle} 담긴 계획 기록을 "
                "별도 자료로 두고 기준선을 정합니다."
            )
        object_particle = particle_for(evidence, "을", "를")
        return f"최근 시험지에서 ‘{evidence}’{object_particle} 기준선으로 확인합니다."

    def replace_nested_exam_material(match: re.Match[str]) -> str:
        evidence = match.group(1)
        if planning_evidence(evidence):
            subject_particle = particle_for(evidence, "이", "가")
            return (
                f"최근 시험지와 ‘{evidence}’{subject_particle} 담긴 계획 기록을 "
                "따로 준비하고"
            )
        if "최근 시험지" in evidence:
            return f"‘{evidence}’ 안의 시험지와 보조 자료를 먼저 구분하고"
        return f"최근 시험지와 교재에서 ‘{evidence}’를 먼저 확인하고"

    value = re.sub(
        r"점수표보다 ‘([^’]+)’[이가] 남은 시험지를 먼저 봅니다\.",
        replace_score_sheet,
        value,
    )
    value = re.sub(
        r"최근 시험지 한 장에서 ‘([^’]+)’를 기준선으로 남깁니다\.",
        replace_exam_sheet_baseline,
        value,
    )
    value = re.sub(
        r"최근 시험지와 교재에서 ‘([^’]+)’를 먼저 찾고",
        replace_nested_exam_material,
        value,
    )
    value = re.sub(
        r"한 표 안에서도 ([^.!?]+)의 문제 번호와 다음 행동을 다른 열에 둡니다\.",
        r"한 표 안에서도 \1의 확인 근거와 다음 행동을 각각 다른 열에 적습니다.",
        value,
    )
    value = re.sub(
        r"([가-힣· ]+)의 최소 과제를 현재 시간표와 대조합니다\.",
        r"\1의 최소 행동을 현재 시간표에서 실행할 수 있는지 확인합니다.",
        value,
    )
    value = value.replace(
        "최소 과제를 현재 시간표와 대조",
        "최소 행동을 현재 시간표에서 실행할 수 있는지 확인",
    )
    value = value.replace("조건과 조건", "조건")
    value = re.sub(r"조건을 조건 해석(?:·|과 )식 세우기", "조건을 식 세우기", value)
    value = value.replace("내신은 내신 범위·학교별 출제 유형", "내신은 학교별 출제 유형")
    value = value.replace("수학 학습 루틴 학습 질문", "수학 학습 루틴 질문")
    value = value.replace("조건 해석과 식 세우기와 ", "조건 해석과 식 세우기, ")
    value = value.replace(
        "어떤 기준으로 첨삭하는지",
        "어떤 기준으로 풀이 과정을 확인하는지",
    ).replace("첨삭", "풀이 확인")
    surface_repairs = (
        ("오답노트", "오답 기록"),
        ("오답 원인과 재풀이", "오답 원인·재풀이"),
        ("조건 해석과 식 세우기", "조건 해석·식 세우기"),
        ("자기주도", "자기 주도"),
        ("시험분석", "시험 분석"),
        ("시험오답", "시험 오답"),
        ("내신시험", "내신 시험"),
        ("일정점검", "일정 점검"),
        ("준비피드백", "준비 과정 피드백"),
        ("준비종합점검", "준비 전반 점검"),
        ("과제점검학습", "과제 점검 방법"),
        ("내신학습", "내신 학습"),
        ("오후학습", "학습 시간대 확인"),
        ("학습클리닉", "학습 보완"),
        ("학습자립도", "스스로 학습하는 정도"),
        ("학습성취도", "학습 이해도"),
        ("학습기록", "학습 기록"),
        ("학습취약점", "학습 취약점"),
        ("성적 기복을 줄이는", "성적 변화의 원인을 점검하는"),
    )
    for before, after in surface_repairs:
        value = value.replace(before, after)
    value = value.replace(
        "내신 기간에는 학교 범위의 완료 흔적을, 평소에는 누적 유형 문제 기록을 남깁니다.",
        f"내신 기간에는 학교 범위 완료 흔적과 {primary_sentence_label} 기록을, "
        f"평소에는 누적 유형 문제의 {secondary_sentence_label} 기록을 남깁니다.",
    )
    value = value.replace(
        "내신은 정해진 범위의 이해와 변형 조건을, 누적 유형 문제는 새 문제의 근거와 시간을 확인합니다.",
        f"내신은 정해진 범위의 이해와 변형 조건을 {primary_sentence_label} 기준으로, "
        f"누적 유형 문제는 새 문제의 근거와 시간을 {secondary_sentence_label} 기준으로 확인합니다.",
    )
    value = value.replace(
        "내신은 범위 자료와 마감일을, 누적 유형 문제는 새 문제와 재확인일을 중심으로 기록합니다.",
        f"내신은 {primary_sentence_label}에 필요한 범위 자료와 마감일을, "
        f"누적 유형 문제는 {secondary_sentence_label}에 필요한 새 문제와 재확인일을 중심으로 기록합니다.",
    )
    value = value.replace(
        "중등 수학학원의 ",
        "중등 수학학원에서 확인할 ",
    )
    value = value.replace("과정를", "과정을").replace("점검를", "점검을")
    value = repair_high_math_surface(value)
    value = re.sub(r"\s+", " ", value).strip()
    return normalize_particle_joins(value, middle_math_particle_tokens(profile))


def naturalize_middle_math_tree(value: object, profile: HighMathProfile) -> object:
    if isinstance(value, str):
        return naturalize_middle_math_text(value, profile)
    if isinstance(value, list):
        return [naturalize_middle_math_tree(item, profile) for item in value]
    if isinstance(value, tuple):
        return tuple(naturalize_middle_math_tree(item, profile) for item in value)
    if isinstance(value, dict):
        return {key: naturalize_middle_math_tree(item, profile) for key, item in value.items()}
    return value


def middle_math_focus_guidance(profile: HighMathProfile) -> tuple[str, str, str]:
    compact = re.sub(r"\s+", "", profile.focus)
    primary = profile.intents[0]
    if primary.code == "entry":
        return (
            "현재 교재와 중학교 예시 범위표에서 계산·문장제·서술형의 첫 풀이를 각각 표시하세요",
            "중1 첫 시험형 문제를 짧게 풀고 조건 표시와 식을 세운 이유를 한 줄씩 남기세요",
            "일주일 뒤 숫자와 표현이 바뀐 새 문제에서도 같은 풀이 순서를 설명하는지 확인하세요",
        )
    if primary.code == "transition" or any(token in compact for token in ("예비고1", "고등과정", "고등수학")):
        return (
            "중3 시험지에서 고등 수학에도 다시 쓰이는 중등 선수 개념과 계산이 끊긴 줄을 함께 표시하세요",
            "필요한 중등 개념만 짧게 복습한 뒤 고등 수준 예시 문제의 조건을 식으로 바꾸어 보세요",
            "일주일 뒤 새 난도의 문제에서도 도움 없이 첫 식과 풀이 순서를 시작하는지 확인하세요",
        )
    if "방학" in compact:
        return (
            "방학 첫날의 시험지·교재 기록과 개학 전 다시 볼 문제를 같은 형식으로 남기세요",
            "개념·조건·계산 중 한 가지 병목만 정해 수정 풀이와 새 문제 확인을 다른 날짜에 실행하세요",
            "중간 점검과 개학 전 점검에서 같은 오류가 줄었는지 비교한 뒤 다음 분량을 정하세요",
        )
    return tuple(
        naturalize_middle_math_text(
            re.sub(r"^(?:이번 주에는|다음 점검에서는|다음 점검에서)\s*", "", sentence), profile
        )
        for sentence in high_math_focus_guidance(profile)
    )  # type: ignore[return-value]


def middle_math_meta_description(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighMathProfile,
) -> str:
    title = f"{center['locality']} {config.label}"
    primary, secondary = profile.intents[:2]
    connector = particle_for(primary.label, "과", "와")
    pair = f"{primary.label}{connector} {secondary.label}"
    candidates = (
        f"{title} 선택 전 {profile.focus}, {pair} 진단, 학교 시험·누적 유형 문제 순서와 센터·가능 학년 확인 기준을 살펴보세요.",
        f"{title}에서 {profile.focus}{particle_for(profile.focus, '을', '를')} 판단하는 자료와 {pair} 확인 순서, 학교 시험 준비, 센터·학년 정보를 정리했습니다.",
        f"{title} 상담 전 {pair} 풀이 기록과 학교 시험 학습 순서, 확인된 센터·가능 학년 정보를 살펴보세요.",
        f"{title}에서 {pair}{particle_for(pair, '을', '를')} 점검하는 법과 누적 유형 문제 재확인, 센터·학년 확인 기준을 안내합니다.",
    )
    for candidate in candidates:
        candidate = naturalize_middle_math_text(clean(candidate), profile)
        if 70 <= len(candidate) <= 100:
            return candidate
    raise ValueError(f"middle Math meta description invalid: {title} / {profile.focus}")


def middle_math_student_type(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighMathProfile,
    seed: str,
) -> str:
    locality = str(center["locality"])
    primary, secondary = profile.intents[:2]
    return naturalize_middle_math_text(stable_pick(seed, "middle-math-student", (
        f"{locality}에서 최근 시험지의 {primary.label} 문제와 {secondary.label} 문제를 나누어 봐야 하는 중학생",
        f"점수만으로 원인을 단정하지 않고 {primary.label} 풀이와 {secondary.label} 재확인 기록으로 다음 순서를 정해야 하는 {locality} 중학생",
        f"학교 범위와 누적 유형 문제 사이에서 {primary.label}{particle_for(primary.label, '과', '와')} {secondary.label}의 비중을 실제 자료로 판단해야 하는 {locality} 중학생",
        f"교재 진도보다 {primary.label}의 근거와 {secondary.label}의 다음 행동을 먼저 정리해야 하는 {locality} 중학생",
        f"최근 수학 시험에서 막힌 위치를 {primary.label}과 {secondary.label}로 구분하고 일주일 뒤 확인 기준까지 세워야 하는 {locality} 중학생",
        f"학교 일정 안에서 {profile.source_markers[0]}와 {profile.source_markers[1]}을 이어 갈 학습 순서가 필요한 {locality} 중학생",
    )), profile)


def middle_math_quick_answer(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighMathProfile,
    seed: str,
) -> str:
    locality = str(center["locality"])
    diagnosis, action, checkpoint = middle_math_focus_guidance(profile)
    introduction = stable_pick(seed, "middle-math-focus-introduction", (
        f"{locality}에서 ‘{profile.focus}’을 판단할 때는 점수보다 최근 수학 풀이를 먼저 봐야 합니다.",
        f"{locality} 중등 수학의 이번 초점은 ‘{profile.focus}’이며, 실제 시험지와 교재 기록에서 시작합니다.",
        f"‘{profile.focus}’이 필요한 {locality} 중학생이라면 개념 설명과 첫 풀이 흔적을 함께 확인하세요.",
        f"이 페이지는 {locality} 학생의 ‘{profile.focus}’을 자료·행동·재확인의 순서로 정리합니다.",
    ))
    labels = stable_pick(seed, "middle-math-focus-stage-labels", (
        ("진단 자료", "이번 행동", "재확인"),
        ("먼저 볼 기록", "7일 실행", "다음 판단"),
        ("현재 상태", "연습 방법", "확인 기준"),
        ("출발 자료", "바꿀 행동", "일주일 뒤"),
    ))
    return naturalize_middle_math_text(" ".join((
        introduction,
        f"{labels[0]}: {diagnosis}.",
        f"{labels[1]}: {action}.",
        f"{labels[2]}: {checkpoint}.",
    )), profile)


def middle_math_sections(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighMathProfile,
    seed: str,
) -> list[dict[str, object]]:
    sections = naturalize_middle_math_tree(
        high_math_sections(config, center, profile, seed),
        profile,
    )
    locality = str(center["locality"])
    primary, secondary = profile.intents[:2]
    _diagnosis, action, checkpoint = middle_math_focus_guidance(profile)
    facts_heading = stable_pick(seed, "middle-math-heading-facts", (
        f"{locality} 학교·학년 정보와 {primary.label} 질문을 구분하는 법",
        f"{locality} 참고 학교·가능 학년을 {primary.label} 질문과 나누는 순서",
        f"{locality} 센터 이용 조건과 {primary.label} 학습 질문을 따로 적는 방법",
        f"{primary.label}{particle_for(primary.label, '과', '와')} {secondary.label} 적용 전 확인할 {locality} 학교·학년 정보",
        f"{locality} 센터 정보 확인과 {primary.label} 상담의 구분",
        f"{locality} 이용 조건을 확인한 뒤 {primary.label} 질문을 비교하는 방법",
    ))
    if len(facts_heading) > 50:
        facts_heading = f"{locality} 학교·학년 정보와 {primary.label} 질문"
    for section in sections:
        key = str(section["key"])
        if key == "direct-answer":
            section["paragraphs"][-1] = clean(
                f"{section['paragraphs'][-1]} 일주일 뒤에는 ‘{profile.focus}’에 해당하는 새 문제도 "
                "혼자 시작하는지 다시 확인합니다."
            )
        elif key == "school-center-facts":
            section["heading"] = facts_heading
        elif key == "four-week-plan":
            section["paragraphs"][-1] = clean(
                f"{section['paragraphs'][-1]} ‘{profile.focus}’의 첫 행동은 {action}. 다음 판단에서는 {checkpoint}."
            )
        elif key == "consultation-checklist":
            checklist = list(section.get("checklist", []))
            if len(checklist) >= 2:
                checklist[0] = ("학생 자료", primary.evidence)
                checklist[1] = (
                    "시험 계획",
                    f"학교 범위표와 {secondary.label} 연습 완료일·재확인일",
                )
                section["checklist"] = checklist
        heading = naturalize_middle_math_text(str(section["heading"]), profile)
        heading = heading.replace("수학 학습 루틴 학습 질문", "수학 학습 루틴 질문")
        heading = heading.replace("조건 해석과 식 세우기와 ", "조건 해석과 식 세우기, ")
        heading = heading.replace("첨삭", "풀이 확인")
        if len(heading) > 50 or heading.count("·") > 3:
            heading = f"{primary.label}{particle_for(primary.label, '과', '와')} {secondary.label}: {locality} 확인 기준"
        section["heading"] = heading
    return naturalize_middle_math_tree(sections, profile)  # type: ignore[return-value]


def middle_math_faq(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighMathProfile,
    seed: str,
) -> list[tuple[str, str]]:
    return naturalize_middle_math_tree(
        high_math_faq(config, center, profile, seed),
        profile,
    )  # type: ignore[return-value]


def middle_math_scenarios(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighMathProfile,
    seed: str,
) -> list[str]:
    scenarios = naturalize_middle_math_tree(
        high_math_scenarios(config, center, profile, seed),
        profile,
    )
    if not relevant_schools(config, center):
        locality = re.escape(str(center["locality"]))
        scenarios = [
            re.sub(
                rf"(?:{locality}의|{locality} 페이지에|페이지의) 참고 중학교(?: 목록이 제공되지 않은 상황|와 실제 재학 학교가 다른 경우)를 가정했습니다\.",
                f"{center['locality']} 페이지에 참고 학교명이 없는 경우를 가정했습니다.",
                value,
            )
            for value in scenarios
        ]
    return scenarios  # type: ignore[return-value]


def physical_region(address: str, fallback: str) -> str:
    first = clean(address).split(" ", 1)[0]
    return first or fallback


SCHEMA_REGION_ALIASES = {
    "서울특별시": "서울", "서울": "서울",
    "경기도": "경기", "경기": "경기",
    "인천광역시": "인천", "인천": "인천",
    "부산광역시": "부산", "부산": "부산",
    "대구광역시": "대구", "대구": "대구",
    "대전광역시": "대전", "대전": "대전",
    "광주광역시": "광주", "광주": "광주",
    "울산광역시": "울산", "울산": "울산",
    "세종특별자치시": "세종", "세종": "세종",
    "강원특별자치도": "강원", "강원도": "강원", "강원": "강원",
    "충청북도": "충청", "충북": "충청", "충청남도": "충청", "충남": "충청", "충청": "충청",
    "전북특별자치도": "전라", "전라북도": "전라", "전북": "전라",
    "전라남도": "전라", "전남": "전라", "전라": "전라",
    "경상북도": "경상", "경북": "경상", "경상남도": "경상", "경남": "경상", "경상": "경상",
    "제주특별자치도": "제주", "제주": "제주",
}


def canonical_schema_region(value: str) -> str:
    value = clean(value)
    return SCHEMA_REGION_ALIASES.get(value, value)


def physical_schema_locality(address: str) -> str:
    parts = clean(address).split()
    if len(parts) < 2:
        return ""
    if parts[0].startswith("세종"):
        return "세종시"
    if parts[1].endswith("시"):
        return parts[1]
    return parts[1]


def safe_district(region: str, district: str, address: str) -> str:
    if region == "세종" or address.startswith("세종"):
        return "세종시"
    if district.endswith(("로", "길")):
        parts = address.split()
        return parts[1] if len(parts) > 1 and parts[1].endswith(("시", "군", "구")) else region
    return district


def load_source_rows(config: CategoryConfig) -> list[str]:
    if not config.source_path.is_file():
        raise FileNotFoundError(config.source_path)
    workbook = load_workbook(config.source_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    # Google Sheets exports can retain formatting far below the 371 content
    # rows. Read the contractual range directly so styled empty rows do not
    # turn a small workbook into a million-row scan.
    values = [
        str(row[0]).strip()
        for row in sheet.iter_rows(
            min_row=1,
            max_row=EXPECTED_ROWS,
            min_col=1,
            max_col=1,
            values_only=True,
        )
        if row and isinstance(row[0], str) and row[0].strip()
    ]
    workbook.close()
    if len(values) != EXPECTED_ROWS:
        raise ValueError(f"{config.workbook}: expected {EXPECTED_ROWS}, found {len(values)}")
    return values


def load_centers() -> list[dict[str, object]]:
    with IMAGE_CSV.open(encoding="utf-8-sig", newline="") as handle:
        image_rows = list(csv.DictReader(handle))
    with CENTER_CSV.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if len(rows) != EXPECTED_ROWS or len(image_rows) != EXPECTED_ROWS:
        raise ValueError(f"common rows center={len(rows)} image={len(image_rows)}")
    image_by_locality = {clean(row["제목"]): row for row in image_rows}
    centers: list[dict[str, object]] = []
    seen_localities: set[str] = set()
    seen_slugs: set[str] = set()
    for row in rows:
        locality = clean(row["근처 수업가능 동네"])
        slug = folder_slug(locality)
        if locality in seen_localities or slug in seen_slugs:
            raise ValueError(f"duplicate locality/slug: {locality}/{slug}")
        seen_localities.add(locality)
        seen_slugs.add(slug)
        english_slug = clean(row["동 영어"]).replace(" ", "-")
        maps = sorted(MAP_DIR.glob(english_slug + ".*"))
        if len(maps) != 1:
            raise FileNotFoundError(f"map {locality}: {english_slug} -> {len(maps)}")
        image = image_by_locality.get(locality)
        if not image:
            raise ValueError(f"image mapping missing: {locality}")
        address = clean(row["센터 주소"])
        region = clean(row["지역"])
        district = safe_district(region, clean(row["시or구"]), address)
        centers.append({
            "locality": locality,
            "slug": slug,
            "english_slug": english_slug,
            "region": region,
            "district": district,
            "address_region": physical_region(address, region),
            "center_name": clean(row["센터명"]),
            "tuition_url": clean(row["센터 교습비"]),
            "office_name": clean(row["교육지원청명칭"]),
            "registration": clean(row["교육지원청 등록번호"]),
            "address": address,
            "schools": {
                "타깃학교\n(초)": normalize_school_values(row["타깃학교\n(초)"]),
                "타깃학교\n(중)": normalize_school_values(row["타깃학교\n(중)"]),
                "타깃학교\n(고)": normalize_school_values(row["타깃학교\n(고)"]),
            },
            "grades": {
                "국어": list_values(row["가능학년\n(국어)"]),
                "영어": list_values(row["가능학년\n(영어)"]),
                "수학": list_values(row["가능학년\n(수학)"]),
            },
            "map_name": maps[0].name,
            "body_image": "seoul6839.webp" if clean(image["본문"]).lower() == "seoul.jpg" else "local6839.webp",
        })
    return centers


def relevant_grades(config: CategoryConfig, center: dict[str, object], subject: str) -> list[str]:
    values = list(center["grades"][subject])  # type: ignore[index]
    if not config.grade_prefix:
        return values
    return [value for value in values if value.startswith(config.grade_prefix)]


def relevant_schools(config: CategoryConfig, center: dict[str, object]) -> list[str]:
    result: list[str] = []
    for column in config.school_columns:
        result.extend(center["schools"][column])  # type: ignore[index]
    return list(dict.fromkeys(result))


def signal_bank(config: CategoryConfig, subject: str) -> tuple[TopicSignal, ...]:
    if subject == "영어":
        return ELEMENTARY_ENGLISH_SIGNALS if config.is_elementary else ENGLISH_SIGNALS
    return ELEMENTARY_MATH_SIGNALS if config.is_elementary else MATH_SIGNALS


def rank_signals(config: CategoryConfig, raw: str, seed: str) -> tuple[TopicSignal, ...]:
    text = source_text(raw)
    ranked_by_subject: dict[str, list[TopicSignal]] = {}
    for subject in config.subjects:
        bank = signal_bank(config, subject)
        scored = sorted(
            bank,
            key=lambda signal: (
                -sum(text.count(keyword) for keyword in signal.keywords),
                stable_int(seed, signal.code),
            ),
        )
        ranked_by_subject[subject] = scored
    if len(config.subjects) == 2:
        english = ranked_by_subject["영어"]
        math = ranked_by_subject["수학"]
        # English and math banks intentionally share a few generic actions
        # (for example, retrying the same question or timing a short set).
        # Avoid selecting an identical action/evidence pair for a combined
        # page because it would ask families to compare the same thing twice.
        english_primary = english[0]
        critical_fields = ("evidence", "practice", "home_action")
        math_primary = next(
            (
                candidate
                for candidate in math
                if all(
                    getattr(candidate, field) != getattr(english_primary, field)
                    for field in critical_fields
                )
            ),
            math[0],
        )
        math_secondary = next(candidate for candidate in math if candidate != math_primary)
        return (english_primary, math_primary, english[1], math_secondary)
    return tuple(ranked_by_subject[config.subjects[0]][:3])


def representative_pool() -> list[Path]:
    unique: list[Path] = []
    seen: set[str] = set()
    for path in sorted(REP_SOURCE.iterdir(), key=lambda item: item.name):
        if not path.is_file() or path.suffix.lower() not in {".gif", ".jpg", ".jpeg", ".png", ".webp"}:
            continue
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        if digest in seen:
            continue
        seen.add(digest)
        unique.append(path)
    if len(unique) < EXPECTED_ROWS:
        raise ValueError(f"representative pool unique={len(unique)}")
    return unique


def existing_row_digests() -> list[set[str]]:
    rows = [set() for _ in range(EXPECTED_ROWS)]
    for prefix in ("hs-kem", "ms-kem", "es-kem"):
        for index in range(EXPECTED_ROWS):
            matches = sorted(REP_TARGET.glob(f"{prefix}-{index + 1:03d}.*"))
            if len(matches) == 1:
                rows[index].add(hashlib.sha256(matches[0].read_bytes()).hexdigest())
    return rows


def assign_representatives(configs: Iterable[CategoryConfig]) -> dict[str, list[str]]:
    REP_TARGET.mkdir(parents=True, exist_ok=True)
    pool = representative_pool()
    pool_digests = [hashlib.sha256(path.read_bytes()).hexdigest() for path in pool]
    row_digests = existing_row_digests()
    assignments: dict[str, list[str]] = {}
    for config in configs:
        existing: list[Path] = []
        for index in range(EXPECTED_ROWS):
            matches = sorted(REP_TARGET.glob(f"{config.rep_prefix}-{index + 1:03d}.*"))
            if len(matches) != 1:
                existing = []
                break
            existing.append(matches[0])
        if existing:
            names = [path.name for path in existing]
            digests = [hashlib.sha256(path.read_bytes()).hexdigest() for path in existing]
            if len(set(digests)) != EXPECTED_ROWS:
                raise ValueError(f"{config.slug}: existing representative duplicates")
            for index, digest in enumerate(digests):
                if digest in row_digests[index]:
                    raise ValueError(f"{config.slug}: same-locality representative collision at {index + 1}")
                row_digests[index].add(digest)
            assignments[config.slug] = names
            continue
        used: set[str] = set()
        names: list[str] = []
        for index in range(EXPECTED_ROWS):
            start = stable_int(config.slug, f"rep-{index}") % len(pool)
            selected = None
            for step in range(len(pool)):
                position = (start + step) % len(pool)
                digest = pool_digests[position]
                if digest not in used and digest not in row_digests[index]:
                    selected = pool[position]
                    used.add(digest)
                    row_digests[index].add(digest)
                    break
            if selected is None:
                raise ValueError(f"{config.slug}: representative allocation failed at {index + 1}")
            name = f"{config.rep_prefix}-{index + 1:03d}{selected.suffix.lower()}"
            shutil.copy2(selected, REP_TARGET / name)
            names.append(name)
        assignments[config.slug] = names
    return assignments


def material_label(config: CategoryConfig, schools: list[str]) -> str:
    if config.is_elementary:
        return "현재 교재·학교 진도·단원평가 자료" if schools else "현재 교재·학습지·단원 기록"
    if config.grade_prefix in {"중", "고"}:
        return "최근 시험지·시험 범위표·현재 교재" if schools else "최근 시험지·현재 교재·오답 기록"
    return "최근 영어·수학 교재·학교 자료·오답 기록"


def grade_summary(config: CategoryConfig, center: dict[str, object]) -> list[tuple[str, str]]:
    return [
        (subject, "·".join(relevant_grades(config, center, subject)) or "상담 시 확인")
        for subject in config.subjects
    ]


def student_level_label(config: CategoryConfig) -> str:
    """Use standard compound labels while keeping the multi-level label readable."""
    return f"{config.level}학생" if config.grade_prefix else f"{config.level} 학생"


def student_type(config: CategoryConfig, signals: tuple[TopicSignal, ...], seed: str, locality: str) -> str:
    primary, secondary = signals[:2]
    subject_text = "·".join(config.subjects)
    student_level = student_level_label(config)
    situation = stable_pick(seed, "student-type", [
        f"{primary.label}에서 겪는 어려움과 {secondary.label}에서 막히는 원인을 구분해 {subject_text} 학습 순서를 정해야 하는 {student_level}",
        f"현재 교재의 진도는 나가지만 {primary.label} 상태를 스스로 점검하거나 {secondary.label} 관련 기록을 남기기 어려운 {student_level}",
        f"과제량보다 {primary.label}과 {secondary.label}을 실제 자료로 확인한 뒤 주간 계획을 세워야 하는 {student_level}",
        f"학습 결과만 보지 않고 {primary.label}이 드러난 자료와 {secondary.label}을 다시 살필 시점을 함께 점검해야 하는 {student_level}",
        f"혼자 공부할 때 {primary.label}에서 멈추고 {secondary.label} 항목을 미루는 경향을 함께 점검할 필요가 있는 {student_level}",
        f"학교 진도와 현재 수준 사이에서 {primary.label} 및 {secondary.label}의 우선순위를 정하기 어려운 {student_level}",
    ])
    return f"{locality}에서 {config.level} {subject_text} 상담을 준비하며 {situation}"


def meta_description(config: CategoryConfig, center: dict[str, object], signals: tuple[TopicSignal, ...]) -> str:
    locality = str(center["locality"])
    title = f"{locality} {config.label}"
    candidates = (
        clean(
            f"{title} 선택 전 {signals[0].label}·{signals[1].label} 진단, {config.school_name} 자료, "
            "가능 학년과 센터 위치를 확인하세요. 실제 교재로 상담할 질문도 정리했습니다."
        ),
        clean(
            f"{title} 선택 전 {signals[0].label}·{signals[1].label} 진단과 {config.school_name} 자료, "
            "가능 학년, 센터 위치 및 상담 질문을 확인하세요."
        ),
        clean(
            f"{title}의 {signals[0].label}·{signals[1].label} 진단, 가능 학년, 센터 위치와 "
            "상담 전 질문을 실제 학생 자료 기준으로 정리했습니다."
        ),
    )
    for value in candidates:
        if 70 <= len(value) <= 100:
            return value
    eligible = [value for value in candidates if len(value) <= 100]
    if not eligible:
        raise ValueError(f"meta description too long: {title}")
    value = max(eligible, key=len)
    addition = " 실제 학생 자료를 기준으로 확인합니다."
    if len(value) + len(addition) <= 100:
        value += addition
    if len(value) < 70:
        raise ValueError(f"meta description too short: {title} ({len(value)})")
    return value


def quick_answer(config: CategoryConfig, center: dict[str, object], signals: tuple[TopicSignal, ...], seed: str) -> str:
    locality = str(center["locality"])
    schools = relevant_schools(config, center)
    material = material_label(config, schools)
    primary, secondary = signals[:2]
    first = stable_pick(seed, "quick-first", [
        f"{locality} {config.label}을 비교할 때는 진도보다 {primary.label}과 {secondary.label}이 실제 학습에서 어떻게 드러나는지 먼저 확인해야 합니다.",
        f"{locality}에서 {config.label}을 찾는다면 현재 점수만 보지 말고 {primary.label}이 드러난 자료와 {secondary.label} 재확인 과정을 나누어 살펴보세요.",
        f"{locality} {config.level} 학습의 출발점은 분량을 늘리는 일이 아니라 {primary.label}과 {secondary.label}의 상태를 자료로 구분하는 것입니다.",
        f"좋은 {locality} {config.label} 상담은 {primary.label}에서 막히는 원인과 {secondary.label}에 필요한 연습을 같은 계획 안에서 구분하는 데서 시작합니다.",
        f"{locality} {config.label} 선택 전에는 학생이 {primary.label}에서 멈추는 장면과 {secondary.label}을 다시 확인하는 방식을 먼저 기록해야 합니다.",
        f"{locality}의 {config.label} 안내를 볼 때는 광고 문구보다 {primary.label}·{secondary.label}을 어떤 자료로 판단하는지 확인하는 편이 정확합니다.",
    ])
    school_note = f"{schools[0]} 등 공개된 참고 학교명은" if schools else f"{config.school_name} 정보가 비어 있는 이 지역은"
    second = stable_pick(seed, "quick-second", [
        f"{material}를 준비하면 {center['center_name']} 상담에서 현재 수준과 다음 점검일을 구체적으로 비교할 수 있습니다.",
        f"{school_note} 수업 가능 여부를 보장하지 않으므로 실제 학생 자료와 센터의 현재 개설 범위를 함께 확인해야 합니다.",
        f"제공 주소인 {center['address']}와 과목별 가능 학년, 통학 시간은 등록 전에 다시 확인하고 학습 계획과 분리해 판단하세요.",
        f"상담에서는 {material}에 표시된 어려운 부분을 바탕으로 첫 학습 행동과 재확인 기준을 정하는 것이 좋습니다.",
        f"학생이 사용 중인 자료를 가져가면 {primary.evidence}와 {secondary.evidence}, 두 자료를 바탕으로 우선순위를 질문할 수 있습니다.",
        f"센터 정보와 학습 자료를 함께 보면 확인된 사실, 상담에서 물을 조건과 학생의 학습 과제를 구분할 수 있습니다.",
    ])
    return first + " " + second


def high_english_meta_description(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighEnglishProfile,
) -> str:
    locality = str(center["locality"])
    title = f"{locality} {config.label}"
    primary, secondary = profile.intents[:2]
    candidates = (
        f"{title} 선택 전 {profile.focus}, {primary.label}·{secondary.label} 진단, 내신·모의고사 학습 구분과 센터 확인 정보를 살펴보세요.",
        f"{title}의 {profile.focus}, {primary.label}·{secondary.label} 진단과 내신·모의고사 준비, 센터·학년 정보를 정리했습니다.",
        f"{title} 상담 전 {primary.label}·{secondary.label}의 진단 자료, 내신·모의고사 학습 순서와 확인된 센터 정보를 살펴보세요.",
        f"{title}에서 {primary.label}·{secondary.label}을 점검하는 법과 내신·모의고사 준비, 센터·가능 학년 확인 기준을 안내합니다.",
    )
    for value in candidates:
        value = clean(value)
        if 70 <= len(value) <= 100:
            return value
    raise ValueError(f"high English meta description invalid: {title} / {profile.focus}")


def high_english_student_type(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighEnglishProfile,
    seed: str,
) -> str:
    locality = str(center["locality"])
    primary, secondary = profile.intents[:2]
    return stable_pick(seed, "high-student", [
        f"{locality}에서 최근 시험지의 {primary.label} 문제와 {secondary.label} 문제를 나누어 봐야 하는 고등학생",
        f"점수만으로 원인을 단정하지 않고 {primary.label} 기록과 {secondary.label} 수행 과정을 비교해 다음 공부 순서를 정해야 하는 {locality} 고등학생",
        f"내신 범위 학습과 모의고사 누적 학습 사이에서 {primary.label}·{secondary.label}의 우선순위를 실제 자료로 판단해야 하는 {locality} 고등학생",
        f"현재 교재의 진도보다 {primary.label}의 증거와 {secondary.label}의 재확인 행동을 먼저 정리해야 하는 {locality} 고등학생",
        f"최근 영어 시험에서 막힌 위치를 {primary.label}과 {secondary.label}로 구분하고 일주일 뒤의 확인 기준까지 세워야 하는 {locality} 고등학생",
        f"학교 일정 안에서 {profile.source_markers[0]}와 {profile.source_markers[1]}을 무리 없이 이어 갈 학습 순서가 필요한 {locality} 고등학생",
    ])


def high_english_particle_tokens(profile: HighEnglishProfile) -> tuple[str, ...]:
    """Return every variable phrase that can receive a Korean particle."""

    values: list[str] = [profile.focus, *profile.source_markers]
    for intent in profile.intents:
        values.extend((
            intent.label,
            intent.concern,
            intent.evidence,
            intent.action,
            intent.checkpoint,
            intent.exam_use,
            intent.consult_question,
        ))
    return tuple(dict.fromkeys(value for value in values if value))


def naturalize_high_english_text(value: str, profile: HighEnglishProfile) -> str:
    """Apply Korean particle repairs without crossing into math vocabulary."""

    value = normalize_particle_joins(value, high_english_particle_tokens(profile))
    value = value.replace("과정를", "과정을").replace("점검를", "점검을")
    return re.sub(r"\s+", " ", value).strip()


def naturalize_high_english_tree(value: object, profile: HighEnglishProfile) -> object:
    if isinstance(value, str):
        return naturalize_high_english_text(value, profile)
    if isinstance(value, list):
        return [naturalize_high_english_tree(item, profile) for item in value]
    if isinstance(value, tuple):
        return tuple(naturalize_high_english_tree(item, profile) for item in value)
    if isinstance(value, dict):
        return {key: naturalize_high_english_tree(item, profile) for key, item in value.items()}
    return value


def high_english_focus_guidance(profile: HighEnglishProfile) -> tuple[str, str, str]:
    """Turn the page's unique source focus into a concrete reader workflow."""

    focus = profile.focus
    compact = re.sub(r"\s+", "", focus)
    primary, secondary = profile.intents[:2]

    if (
        any(token in compact for token in ("시험불안", "불안을줄", "긴장을줄"))
        and any(token in compact for token in ("평일", "방학", "겨울방학", "여름방학"))
    ):
        return (
            "평일의 실제 가능 시간과 방학의 연장 가능 시간을 나누고, 두 기간에 불안 때문에 멈춘 순간을 함께 표시하세요",
            "평일에는 짧은 시간 제한 세트와 회복 순서를, 방학에는 약점 유형별 연습과 재확인일을 따로 계획하세요",
            "일주일 뒤에는 같은 난도의 새 세트에서 다시 시작하는 시간이 줄었는지 보고 평일·방학 분량을 조정하세요",
        )

    if any(token in compact for token in ("질문하기어려", "질문이어려", "질문하는힘", "질문습관")):
        return (
            "학생이 멈춘 문장 옆에 무엇을 모르겠는지 한 줄로 적을 수 있는지 확인하세요",
            "교재에 물음표를 붙이고 ‘알고 있는 것·모르는 것·먼저 물을 것’을 각각 한 칸씩 채우게 하세요",
            "다음 점검에서는 학생이 메모를 보지 않고 질문 하나를 자기 말로 설명하는지 살펴보세요",
        )
    if any(token in compact for token in ("시험불안", "불안을줄", "긴장을줄", "자신감")):
        return (
            "어려운 문항에서 바로 답을 바꿨는지, 건너뛴 뒤 돌아왔는지와 같은 행동 기록부터 확인하세요",
            "짧은 세트를 정해진 시간에 풀고 막힌 위치·다시 시작한 시각·근거를 찾은 순간을 표시하세요",
            "일주일 뒤 같은 난도의 새 세트에서 멈춘 뒤 다시 풀이를 이어 가는 시간이 달라졌는지 비교하세요",
        )
    if "고2" in compact and any(token in compact for token in ("수능", "입시", "모의고사")):
        return (
            "고2 학교 일정과 최근 모의고사의 누적 약점을 같은 달력에 놓고 어느 쪽이 밀리는지 확인하세요",
            "내신 기간의 최소 누적 학습과 시험이 없는 주의 기출 분석 분량을 다른 줄에 적어 실행하세요",
            "다음 모의고사 전에는 푼 회차보다 반복 유형의 근거 설명과 소요 시간이 달라졌는지 비교하세요",
        )
    if any(token in compact for token in ("평일", "방학", "겨울방학", "여름방학")):
        return (
            "평일의 실제 가능 시간과 방학에 늘릴 수 있는 시간을 구분하고 각각 지킨 기록을 확인하세요",
            "평일에는 끊기지 않을 최소 행동을, 방학에는 약점 보완 묶음을 정해 서로 다른 계획표에 적으세요",
            "방학 계획을 확정하기 전 평일 기록에서 미완료 원인을 찾아 늘릴 분량과 유지할 분량을 나누세요",
        )
    if any(token in compact for token in ("복습", "재확인", "오답", "다시푸", "틀린문제")):
        return (
            "처음 틀린 이유와 해설 뒤 바뀐 이유가 함께 남은 오답을 골라 복습의 시작점을 확인하세요",
            "같은 날의 재풀이와 며칠 뒤의 새 문제 확인을 분리하고 두 기록에 각각 근거를 적으세요",
            "다음 점검에서는 정답을 기억하는지가 아니라 비슷한 문제에서 판단 순서를 다시 설명하는지 보세요",
        )
    if any(token in compact for token in ("글쓰기", "서술", "영작", "작문", "초안")):
        return (
            "수행평가 조건표·초안·수정본을 놓고 빠진 요구 조건과 근거가 약한 문단을 먼저 확인하세요",
            "주장·근거·예시를 세 칸으로 나눈 뒤 조건에 필요한 표현을 넣고 문장 연결을 다시 고치세요",
            "다음 글에서는 조건표를 보지 않고도 필요한 요소를 모두 넣고 수정 이유를 설명하는지 살펴보세요",
        )
    if any(token in compact for token in ("발표", "말하기", "수행평가", "스피킹")):
        return (
            "시험지보다 발표 대본·녹음·평가 조건표에서 멈춘 문장과 빠진 조건을 먼저 확인하세요",
            "핵심 순서를 세 문장으로 줄여 녹음하고 대본을 보지 못한 구간과 수정할 표현을 표시하세요",
            "다음 녹음에서는 제한 시간 안에 핵심 순서를 유지하고 이어지는 질문에도 답하는지 살펴보세요",
        )
    if any(token in compact for token in ("듣기", "음원", "받아쓰기")):
        return (
            "오답 음원에서 소리를 놓친 구간과 표현의 뜻을 몰랐던 구간을 다른 표시로 나누세요",
            "짧은 구간을 의미 단위로 받아쓴 뒤 듣기 대본과 비교하고 놓친 연결 발음을 따라 말해 보세요",
            "새 음원에서는 반복 재생 횟수보다 같은 표현을 처음 듣고 알아차렸는지 확인하세요",
        )
    if any(token in compact for token in ("문장", "구문", "해석", "어법", "문법")):
        return (
            "해석이 끊긴 문장에서 주어·동사·수식 범위를 학생이 직접 표시하는지 확인하세요",
            "문장 뼈대를 먼저 적고 수식어를 붙인 뒤 같은 구조의 짧은 문장을 한 번 더 읽게 하세요",
            "다음 점검에서는 처음 보는 문장에서도 같은 표시 순서를 도움 없이 재현하는지 살펴보세요",
        )
    if any(token in compact for token in ("독해", "지문", "글의흐름", "읽는힘", "빈칸", "순서")):
        return (
            "최근 지문에서 핵심 문장과 연결어, 선택지를 지운 근거가 모두 남아 있는지 확인하세요",
            "문단 역할을 한 줄로 요약하고 답을 지지하는 문장을 선택지 표현과 연결해 보세요",
            "다음 점검에서는 새 지문에서도 정답과 오답의 근거를 각각 짚는지 살펴보세요",
        )
    if any(token in compact for token in ("시간배분", "풀이시간", "읽는시간", "검토시간", "속도")):
        return (
            "최근 시험지에 문항별 읽기·판단·검토 시간을 적고 어느 구간에서 시간이 길어지는지 확인하세요",
            "짧은 세트에서 세 시간을 따로 기록한 뒤 정확도를 해치지 않는 범위에서 병목 구간만 다시 연습하세요",
            "다음 세트에서는 정답 수와 함께 지문별 소요 시간과 마지막 검토 시간을 같은 표에서 비교하세요",
        )
    if any(token in compact for token in ("내신", "시험", "서술형", "학교", "성적")):
        return (
            "학교 범위표와 최근 답안에서 본문 이해·어법 변형·서술형 조건 중 막힌 곳을 구분하세요",
            "시험일까지 남은 날에 맞춰 세 영역의 마감일을 나누고 완료한 자료의 위치를 함께 기록하세요",
            "시험 뒤에는 총점만 보지 말고 바뀐 근거와 다시 설명하지 못한 조건을 다음 계획에 반영하세요",
        )
    if any(token in compact for token in ("루틴", "습관", "자기주도", "계획", "시간", "꾸준", "집중")):
        return (
            "계획표의 분량과 실제 완료 기록을 대조해 시작하지 못한 날과 끝내지 못한 날을 구분하세요",
            "매일 지킬 최소 행동 하나와 미완료 때 줄일 분량을 정하고 완료 근거를 한 줄로 남기세요",
            "일주일 뒤 반복 오답과 미완료 이유를 보고 시간·분량·점검 순서 중 한 가지만 조정하세요",
        )
    return (
        f"‘{primary.evidence}’와 ‘{secondary.evidence}’를 나란히 놓고 현재의 병목을 자료로 확인하세요",
        f"이번 주에는 ‘{primary.action}’을 먼저 실행하고 ‘{secondary.action}’은 별도 날짜에 배치하세요",
        f"다음 점검에서 ‘{primary.checkpoint}’와 ‘{secondary.checkpoint}’에 학생이 직접 답하는지 비교하세요",
    )


def high_english_quick_answer(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighEnglishProfile,
    seed: str,
) -> str:
    locality = str(center["locality"])
    primary, secondary = profile.intents[:2]
    diagnosis, action, checkpoint = high_english_focus_guidance(profile)
    introductions = (
        f"{locality}에서 ‘{profile.focus}’을 판단할 때는 제목의 표현을 실제 학생 기록으로 바꾸어 봐야 합니다.",
        f"{locality} 고등 영어의 이번 초점은 ‘{profile.focus}’이며, 점수보다 확인 가능한 행동에서 시작합니다.",
        f"‘{profile.focus}’이 필요한 {locality} 학생이라면 최근 자료 한 장으로 현재선과 다음 행동을 구분하세요.",
        f"{locality} 학생의 ‘{profile.focus}’은 막연한 목표가 아니라 자료·실행·재확인의 세 단계로 확인합니다.",
        f"{locality} 고등 영어에서 ‘{profile.focus}’을 살필 때는 광고 문구보다 학생이 남긴 흔적을 먼저 봅니다.",
        f"이 페이지는 {locality} 학생의 ‘{profile.focus}’을 실제로 점검할 수 있는 순서로 정리합니다.",
    )
    endings = (
        f"이 과정을 마친 뒤 {primary.label} 기록과 {secondary.label} 기록을 비교해 다음 순서를 정하세요.",
        f"마지막에는 {primary.label}의 변화와 {secondary.label}의 남은 질문을 다른 칸에 적으세요.",
        f"확인 결과는 {primary.label}의 첫 행동과 {secondary.label}의 재확인 날짜로 나누어 남기면 됩니다.",
        f"센터 이용 조건은 별도 표에 두고, 학습 기록에서는 {primary.label}과 {secondary.label}만 비교하세요.",
        f"학생이 직접 설명한 내용은 {primary.label}·{secondary.label}의 다음 주 분량을 정하는 근거가 됩니다.",
        f"두 기록이 달라진 이유를 적으면 {primary.label}과 {secondary.label}을 함께 할지 순서를 나눌지 판단할 수 있습니다.",
    )
    introduction = stable_pick(seed, "high-focus-introduction", introductions)
    ending = stable_pick(seed, "high-focus-ending", endings)
    stage_labels = stable_pick(seed, "high-focus-stage-labels", (
        ("진단 자료", "이번 주 행동", "재확인 기준"),
        ("먼저 볼 기록", "바로 할 일", "다시 볼 시점"),
        ("현재선 확인", "실행 항목", "다음 판단"),
        ("출발 자료", "첫 실천", "일주일 뒤 확인"),
        ("문제 구분", "학습 행동", "조정 기준"),
        ("확인할 흔적", "적용할 계획", "변화 판단"),
    ))
    workflow = (
        f"{stage_labels[0]}: {diagnosis}. "
        f"{stage_labels[1]}: {action}. "
        f"{stage_labels[2]}: {checkpoint}."
    )
    return f"{introduction} {workflow} {ending}"


def _legacy_high_english_sections(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighEnglishProfile,
    seed: str,
) -> list[dict[str, object]]:
    locality = str(center["locality"])
    title = f"{locality} {config.label}"
    primary, secondary, support, extra = profile.intents
    marker_a, marker_b, marker_c, marker_d = profile.source_markers
    schools = relevant_schools(config, center)
    school_text = "·".join(schools[:4])
    school_reference = (
        f"공개 자료에는 {school_text} 등이 참고 학교로 적혀 있습니다. 이는 수업 가능을 보장하는 목록이 아니므로 실제 재학 학교의 시험 범위와 센터의 현재 개설 범위를 함께 확인해야 합니다."
        if schools else
        "이 지역은 공개된 참고 고등학교 목록이 없습니다. 실제 재학 학교와 시험 범위를 상담에서 알리고 현재 개설 범위와 자료 반영 방식을 확인해야 합니다."
    )
    grades = relevant_grades(config, center, "영어")
    grade_fact = (
        f"제공 자료의 고등 영어 가능 학년은 {'·'.join(grades)}입니다."
        if grades else
        "제공 자료에는 고등 영어 가능 학년이 기재되지 않아 상담 시 확인해야 합니다."
    )
    grade_object = (
        f"제공 자료의 고등 영어 가능 학년 표기({'·'.join(grades)})"
        if grades else
        "고등 영어 가능 학년이 기재되지 않은 상태"
    )
    materials = "최근 영어 시험지·학교 시험 범위표·현재 교재·오답 기록"

    direct = {
        "key": "direct-answer",
        "heading": stable_pick(seed, "high-h2-direct", [
            f"{profile.focus}, 무엇부터 확인해야 할까",
            f"{title}에서 {profile.focus}을 판단하는 첫 질문",
            f"{profile.focus}에 앞서 구분할 {primary.label}의 문제",
            f"{locality} 고등학생의 {profile.focus}: 점수보다 먼저 볼 근거",
            f"{primary.label} 기록으로 시작하는 {profile.focus}",
            f"{profile.focus}이 필요한 학생의 공통 신호",
        ]),
        "paragraphs": [
            stable_pick(seed, "high-direct-p1", [
                f"{profile.focus}이 필요한지는 총점만으로 알기 어렵습니다. 먼저 {primary.concern}를 확인하고, 같은 시험지에서 {secondary.concern}도 별도로 표시해야 두 문제의 우선순위를 정할 수 있습니다.",
                f"{title}을 비교할 때 첫 판단 자료는 광고 문구가 아니라 학생이 직접 푼 시험지입니다. {primary.evidence}를 찾은 뒤 {secondary.evidence}와 대조하면 {profile.focus}의 출발점이 구체화됩니다.",
                f"학생이 영어를 어렵다고 말해도 원인은 서로 다릅니다. {primary.label}에서는 {primary.concern}를, {secondary.label}에서는 {secondary.concern}를 묻고 답과 실제 표시가 일치하는지 살펴보세요.",
                f"최근 점수가 같아도 {profile.focus}에 필요한 학습은 달라질 수 있습니다. {primary.evidence}가 남아 있는지부터 보고 {secondary.label}은 새 문제에서 다시 수행할 수 있는지 확인합니다.",
                f"진단 자료는 많을 필요가 없습니다. {materials} 중 {primary.label}이 드러난 한 곳과 {secondary.label}이 드러난 한 곳을 골라 비교하면 첫 학습 행동을 정하기 쉽습니다.",
                f"{locality} 고등 영어의 시작점을 정할 때는 잘한 단원보다 멈춘 장면이 유용합니다. {primary.concern}를 확인한 뒤 {secondary.checkpoint}까지 살펴야 설명과 연습의 비중을 나눌 수 있습니다.",
                f"{profile.focus}을 상담 주제로 삼는다면 ‘왜 틀렸는가’를 한 단어로 적지 마세요. {primary.label}과 {secondary.label}에 해당하는 근거를 각각 남겨야 다음 점검에서 변화를 비교할 수 있습니다.",
                f"문제 수를 늘리기 전에 {primary.evidence}를 먼저 확보하세요. 그 기록과 {secondary.evidence} 사이의 차이가 {profile.focus}에 필요한 순서를 정하는 실제 근거가 됩니다.",
            ]),
            stable_pick(seed, "high-direct-p2", [
                f"첫 행동은 {primary.action}입니다. 한 번 설명을 들은 뒤에는 {primary.checkpoint}를 새 문장에서 확인하고, 통과하지 못하면 진도보다 같은 절차의 짧은 재연습을 먼저 배치합니다.",
                f"학생에게 정답을 다시 외우게 하기보다 {primary.action}를 직접 해 보게 하세요. 이어서 {secondary.action}를 수행한 기록이 남으면 두 영역을 함께 다룰지 순서를 나눌지 결정할 수 있습니다.",
                f"판단 기준은 맞힌 개수 하나가 아닙니다. {primary.checkpoint}와 {secondary.checkpoint}를 각각 확인한 뒤, 달라진 근거가 있는 영역만 다음 범위로 넓히는 편이 안전합니다.",
                f"{marker_a}와 {marker_b}를 같은 날 모두 끝내려 하지 않아도 됩니다. 첫 점검에서 {primary.action}를 실행하고 다음 날 {secondary.checkpoint}를 확인하는 식으로 학습 부담을 나눌 수 있습니다.",
                f"학부모는 학생 대신 원인을 단정하기보다 ‘{primary.concern}’를 질문으로 적어 두세요. 학생의 설명과 시험지 표시가 다르면 그 차이 자체가 첫 피드백 주제가 됩니다.",
                f"교재 교체는 마지막 판단입니다. 현재 자료에서 {primary.action}를 해 본 뒤에도 {primary.checkpoint}가 확인되지 않을 때 설명 난도·연습량·교재 난도를 순서대로 비교하세요.",
                f"{profile.focus}은 한 번의 진단으로 끝나지 않습니다. 이번 주의 {primary.evidence}와 다음 주의 {support.checkpoint}를 같은 형식으로 남겨야 계획을 유지할지 바꿀지 판단할 수 있습니다.",
                f"상담에서는 {primary.consult_question}를 먼저 묻고, 답변이 학생의 실제 시험지에 어떻게 적용되는지 확인하세요. 추상적인 ‘꼼꼼한 관리’보다 기록의 형태와 점검 시점이 중요합니다.",
            ]),
        ],
    }

    diagnosis = {
        "key": "diagnostic-evidence",
        "heading": stable_pick(seed, "high-h2-evidence", [
            f"{marker_a}와 {marker_b}, 최근 자료에서 어떻게 구분할까",
            f"시험지에서 찾는 {primary.label}·{secondary.label} 진단 증거",
            f"{marker_a}에서 멈춘 이유를 기록하는 방법",
            f"정답 수보다 유용한 {marker_b} 확인 기록",
            f"{locality} 고등 영어 상담에 가져갈 학습 흔적",
            f"{primary.label}과 {secondary.label}의 병목을 나누는 자료",
        ]),
        "paragraphs": [
            stable_pick(seed, "high-evidence-p1", [
                f"{primary.label}은 {primary.evidence}에서 확인하고, {secondary.label}은 {secondary.evidence}에서 확인합니다. 두 기록에 날짜와 문제 번호를 적으면 같은 실수가 반복된 것인지 일시적인 실수인지 구분하기 쉬워집니다.",
                f"{materials} 전체를 모두 복사할 필요는 없습니다. {primary.evidence}와 {secondary.evidence}가 보이는 부분을 각각 한 곳씩 골라 ‘알고 있던 것·막힌 것·다시 해 볼 것’ 세 칸으로 나누세요.",
                f"첫 풀이와 해설을 본 뒤의 답을 한 화면에 놓으면 {primary.label}의 변화가 보입니다. 여기에 {secondary.evidence}를 더하면 지식 부족과 선택 과정의 문제를 섞지 않고 기록할 수 있습니다.",
                f"{marker_a}을 점검할 때는 학생이 표시한 흔적을 지우지 않는 것이 좋습니다. 그 위에 {primary.evidence}와 {secondary.evidence}를 다른 표시로 남겨야 설명 전후의 차이를 비교할 수 있습니다.",
                f"시험지에는 틀린 이유보다 먼저 멈춘 위치를 표시하세요. {primary.label}의 증거와 {secondary.label}의 증거가 같은 문항에 함께 있어도 각각 기록하면 학습 순서를 세밀하게 조정할 수 있습니다.",
                f"{locality} 고등학생의 상담 자료에는 점수표보다 풀이 흔적이 필요합니다. {primary.evidence}를 통해 시작점을 잡고 {secondary.evidence}로 재확인 방식을 정하세요.",
                f"학생의 말과 실제 자료가 다를 때가 진단에 가장 유용합니다. ‘{primary.concern}’에 대한 답을 적은 뒤 {primary.evidence}에서 같은 내용이 확인되는지 대조해 보세요.",
                f"{profile.focus}과 연결되는 자료는 {marker_a} 하나만 골라도 됩니다. 다만 {primary.evidence}와 {secondary.evidence}를 구분해 놓아야 단순 반복과 필요한 설명을 나눌 수 있습니다.",
            ]),
            stable_pick(seed, "high-evidence-p2", [
                f"기록의 목적은 학생을 평가하는 것이 아니라 다음 행동을 고르는 데 있습니다. {primary.checkpoint}가 확인되면 {secondary.action}으로 넘어가고, 그렇지 않으면 같은 범위의 설명과 짧은 재시도를 먼저 배치합니다.",
                f"진단표에는 ‘완료’ 대신 학생이 보여 준 근거를 적으세요. {primary.checkpoint}와 {support.checkpoint} 중 확인된 항목만 표시하면 다음 주 분량을 과하게 잡는 일을 줄일 수 있습니다.",
                f"한 번 맞힌 문제는 끝난 항목이 아닙니다. 며칠 뒤 {primary.action}를 새 자료에서 다시 수행하고 {secondary.checkpoint}까지 설명할 수 있을 때 학습 범위를 넓히세요.",
                f"{marker_b} 관련 피드백은 정답을 알려 주는 문장보다 ‘어디서 판단이 달라졌는가’를 남겨야 합니다. 첫 풀이와 재풀이의 근거가 달라졌는지 비교하면 복습의 효과를 확인할 수 있습니다.",
                f"자료가 많아도 점검 시점이 없으면 개선 여부를 알기 어렵습니다. {primary.action}를 한 날짜와 {primary.checkpoint}를 다시 볼 날짜를 함께 적어 다음 상담에서 같은 기준으로 검토하세요.",
                f"{secondary.label}이 약하다고 단정하기 전 {secondary.action}를 짧게 수행하게 해 보세요. 독립 수행이 가능한 단계와 설명이 필요한 단계를 나누면 과제량보다 정확한 계획을 세울 수 있습니다.",
                f"학부모가 확인할 것은 문제 수가 아니라 기록의 연결입니다. {primary.evidence}가 다음 과제의 {support.action}으로 이어졌는지 보면 계획이 실제로 작동하는지 알 수 있습니다.",
                f"{primary.label}과 {secondary.label}의 우선순위가 애매하면 시험 일정이 가까운 영역을 먼저 두되, 다른 영역의 최소 복습 날짜를 남겨 학습 공백을 막으세요.",
            ]),
        ],
    }

    exam = {
        "key": "exam-strategy",
        "heading": stable_pick(seed, "high-h2-exam", [
            f"내신과 모의고사에서 {marker_c}을 다르게 적용하는 법",
            f"{primary.label}을 학교 시험과 모의고사에 연결하는 기준",
            f"범위가 정해진 내신, 처음 보는 지문의 준비 순서",
            f"{profile.focus}을 내신·모의고사 계획으로 옮기는 방법",
            f"학교 자료와 누적 독해를 한 주에 배치하는 법",
            f"{marker_c}·{marker_d}의 시험별 점검 기준",
        ]),
        "paragraphs": [
            stable_pick(seed, "high-exam-p1", [
                f"내신은 시험 범위표와 교과서·학교 자료를 기준으로 본문 이해, 어법 변형, 서술형 조건을 확인합니다. 모의고사는 처음 보는 지문에서 {primary.label}과 {secondary.label}을 시간 안에 적용하는지 살펴야 하므로 같은 문제 수로 계획하지 않습니다.",
                f"학교 시험은 정해진 범위를 빠짐없이 확인하는 일이 우선이고, 모의고사는 누적된 {primary.label}의 약점을 새 지문에서 확인하는 과정입니다. {primary.exam_use}에 따라 두 기록표를 나누세요.",
                f"내신 기간에는 범위표의 본문·어법·서술형을 세 칸으로 나누고, 모의고사 기록에는 유형·근거·소요 시간을 적습니다. {secondary.exam_use}을 적용하면 한쪽 준비가 다른 쪽을 밀어내는지 확인할 수 있습니다.",
                f"같은 영어 시험이라도 내신과 모의고사는 자료가 다릅니다. 학교 자료에서는 {marker_c}의 변형 가능성을, 모의고사에서는 {primary.checkpoint}가 낯선 지문에서도 유지되는지를 확인하세요.",
                f"{profile.focus}을 시험 계획으로 바꿀 때는 학교 범위와 누적 학습을 섞지 마세요. 내신 마감일을 먼저 표시하고 남은 날에 {secondary.action}을 짧게 반복하는 방식이 현실적입니다.",
                f"내신 점수가 높아도 처음 보는 지문에서 {primary.label}이 흔들릴 수 있고, 모의고사 성적이 좋아도 학교 서술형 조건을 놓칠 수 있습니다. 두 결과를 별도 원인표로 나눠야 합니다.",
                f"학교 시험지에는 범위 안에서 놓친 표현을, 모의고사 시험지에는 유형과 시간 기록을 남기세요. 두 자료를 비교하면 {marker_d}에 배정할 주간 시간을 정하기 쉬워집니다.",
                f"내신 대비는 학교별 범위를 확인한 뒤 시작하고, 모의고사 대비는 최근 회차의 반복 오답부터 시작합니다. {primary.exam_use}이 두 계획을 연결하는 공통 기준이 됩니다.",
            ]),
            stable_pick(seed, "high-exam-p2", [
                f"시험 직전에는 새로운 교재를 늘리기보다 틀린 근거를 다시 설명할 수 있는지 확인하세요. {support.action}을 한 뒤 {support.checkpoint}가 보이면 다음 유형으로 넘어가고, 아니면 같은 자료의 재풀이를 남깁니다.",
                f"주간 비중은 고정하지 않습니다. 학교 시험이 가까우면 범위 학습을 늘리되 {primary.action}의 최소 횟수는 유지하고, 시험 뒤에는 모의고사 오답과 누적 어휘로 중심을 옮기세요.",
                f"{secondary.label}이 두 시험에서 모두 약하다면 공통 기초를 먼저 보완합니다. 반대로 한 시험에서만 흔들리면 자료 해석이나 시간 배분처럼 시험별 절차를 따로 연습해야 합니다.",
                f"‘내신 몇 회, 모의고사 몇 회’보다 중요한 것은 완료 기준입니다. {primary.checkpoint}와 {extra.checkpoint}가 확인된 문제만 학습 완료로 표시하세요.",
                f"시험 뒤 성적표만 보관하지 말고 {primary.evidence}와 {secondary.evidence}를 남기세요. 다음 범위가 시작될 때 같은 약점이 이어지는지 확인할 기준선이 됩니다.",
                f"{marker_c}을 내신에서 익혔다면 모의고사 새 지문에 한 번 적용해 보세요. 반대로 모의고사에서 찾은 독해 약점은 학교 본문 변형 문제에서 다시 확인할 수 있습니다.",
                f"학년과 시험 시기에 따라 필요한 비중은 달라집니다. 따라서 센터 상담에서는 {primary.consult_question}와 학교 시험 기간의 계획 변경 방식을 함께 물어보세요.",
                f"시험 준비의 목표를 성과 보장 문구로 정하지 마세요. 이번 주에는 {support.action}, 다음 점검에는 {support.checkpoint}처럼 관찰 가능한 행동과 근거로 정하는 편이 정확합니다.",
            ]),
        ],
    }

    facts = {
        "key": "school-center-facts",
        "heading": stable_pick(seed, "high-h2-facts", [
            f"{locality} 학교 자료와 센터 정보를 구분해 읽는 법",
            f"{title}의 가능 학년·학교·주소 확인",
            f"학교명은 참고, 개설 여부는 상담에서 확인할 조건",
            f"{center['center_name']} 공개 정보와 학생 자료의 역할",
            f"{locality} 고등 영어 등록 전에 확인할 사실",
            f"학습 계획과 이용 조건을 한 문장으로 묶지 않는 이유",
        ]),
        "paragraphs": [
            stable_pick(seed, "high-facts-p1", [
                f"{school_reference} 학교 자료는 내신 범위를 이해하는 근거이고, 센터 운영 사실은 주소·등록 정보·가능 학년·현재 시간표로 따로 판단합니다.",
                f"{school_reference} 참고 학교명이 학생의 반 편성이나 자료 제공을 뜻하지는 않습니다. 실제 시험 범위표를 가져가 반영 방식부터 물어보세요.",
                f"{school_reference} {grade_fact} 공개 정보와 현재 운영 시간표가 다를 수 있으므로 등록 전 다시 대조해야 합니다.",
                f"학교명은 검색 편의를 위한 참고 정보로 읽어야 합니다. {school_reference} 학생별 계획은 {materials} 전체를 본 뒤 별도로 정합니다.",
                f"{school_reference} {primary.label} 진단은 센터 목록이 아니라 학생의 실제 시험지에서 진행해야 두 사실이 섞이지 않습니다.",
                f"{grade_fact} {school_reference} 학년과 학교 정보를 확인한 다음 {marker_a}에 필요한 자료를 상담에서 제시하세요.",
            ]),
            stable_pick(seed, "high-facts-p2", [
                f"제공 주소는 {center['address']}이고 센터 기준 명칭은 {center['center_name']}입니다. 방문 전 실제 운영 여부와 통학 동선, 영어 시간표를 확인하고 교습비는 연결된 공식 조회 경로에서 대조하세요.",
                f"이 페이지가 안내하는 센터 기준은 {center['center_name']}, 제공 주소는 {center['address']}입니다. {grade_fact} 수업 시간·반 구성·보완 방식은 현재 운영 범위를 직접 확인해야 합니다.",
                f"{center['center_name']}의 제공 주소는 {center['address']}입니다. 주소와 등록 정보는 사실 영역이고, {profile.focus}에 맞는 학습 계획은 시험지를 본 뒤 결정할 상담 영역입니다.",
                f"등록 전에는 {center['address']}의 방문 동선과 {grade_object}를 먼저 대조하세요. 차량·주차·보강처럼 공개 자료에 없는 조건은 추정하지 말고 센터에 직접 묻습니다.",
                f"센터명은 {center['center_name']}, 제공 주소는 {center['address']}입니다. 공개된 가능 학년과 실제 시간표를 구분하고, {primary.consult_question}도 별도 질문으로 준비하세요.",
                f"{center['center_name']} 상담에서는 주소·가능 학년·교습비 확인을 먼저 끝내고 학습 질문으로 넘어가세요. {profile.focus}을 위한 진단과 이용 조건을 섞지 않아야 비교가 쉬워집니다.",
            ]),
        ],
    }

    plan = {
        "key": "four-week-plan",
        "heading": stable_pick(seed, "high-h2-plan", [
            f"{marker_a}에서 {marker_d}로 이어지는 4주 점검 예시",
            f"{primary.label} 기록을 바꾸는 조건부 4주 계획",
            f"설명·연습·재확인을 한 달에 배치하는 방법",
            f"{profile.focus}을 주간 행동으로 바꾸는 순서",
            f"첫 진단부터 재풀이까지, 고등 영어 4주 흐름",
            f"{secondary.label}의 변화를 기록으로 확인하는 한 달",
        ]),
        "paragraphs": [
            stable_pick(seed, "high-plan-p1", [
                f"1주차에는 {primary.evidence}를 모아 기준선을 만들고 {primary.action}을 짧게 실행합니다. 2주차에는 {secondary.action}을 더해 {marker_a}와 {marker_b}가 같은 문제에서 어떻게 연결되는지 기록하세요.",
                f"첫 주는 진단 주간입니다. {materials}에서 {primary.label} 사례를 고르고, 둘째 주에는 {primary.action}을 비슷한 새 문제에 적용해 독립 수행 범위를 확인합니다.",
                f"1주차에는 {primary.concern}를 답할 자료를 준비합니다. 2주차에는 설명 뒤 {primary.checkpoint}가 확인되는지 보고, 통과하지 못한 부분만 분량을 줄여 다시 연습하세요.",
                f"처음 7일은 {marker_a} 기록을 남기는 데 집중하고 다음 7일은 {secondary.action}으로 연결합니다. 계획표에는 문제 수보다 시작 시각·완료 근거·남은 질문을 적는 편이 유용합니다.",
                f"1주차에 {primary.evidence}와 {secondary.evidence}를 분류한 뒤 우선순위 하나를 고릅니다. 2주차에는 선택한 영역의 짧은 연습과 재확인을 같은 요일에 배치하세요.",
                f"첫째 주에는 현재 자료를 바꾸지 말고 {primary.action}을 해 봅니다. 둘째 주에는 새 지문에서 {primary.checkpoint}를 확인해 설명이 필요한지 반복 연습이 필요한지 나눕니다.",
                f"4주 계획의 시작은 분량 약속이 아닙니다. 1주차에 {primary.label}의 기준선을 남기고 2주차에 {support.action}을 한 번 적용해 비교 가능한 기록을 만드세요.",
                f"{profile.focus}을 첫 두 주에 적용하려면 진단과 연습을 섞지 않는 편이 좋습니다. 1주차는 근거 수집, 2주차는 {secondary.action}과 재확인으로 역할을 나눕니다.",
            ]),
            stable_pick(seed, "high-plan-p2", [
                f"3주차에는 내신 범위와 모의고사 누적 약점을 나눠 {support.action}을 적용합니다. 4주차에는 {primary.checkpoint}와 {secondary.checkpoint}를 처음 기록과 비교해 유지·축소·확대 중 다음 선택을 정합니다.",
                f"셋째 주에는 {marker_c}을 학교 자료와 새 지문에 각각 적용하고, 넷째 주에는 처음 틀린 문제를 해설 없이 다시 풉니다. 근거가 달라지지 않았다면 진도를 늘리지 말고 설명 방식을 재검토하세요.",
                f"3주차는 시험 유형에 적용하는 기간입니다. {primary.exam_use}을 계획에 반영하고, 4주차에는 {extra.checkpoint}까지 확인한 뒤 다음 한 달의 우선순위를 다시 세웁니다.",
                f"세 번째 주에는 학생이 혼자 수행할 구간을 늘리고 네 번째 주에는 첫 시험지와 재풀이를 나란히 봅니다. 정답 수뿐 아니라 근거·소요 시간·수정한 문장이 달라졌는지 확인하세요.",
                f"3주차에는 {support.label}의 최소 루틴을 유지하면서 학교 일정에 맞춰 비중을 조정합니다. 4주차에는 미완료 이유까지 기록해 다음 계획의 분량과 점검 주기를 현실적으로 바꾸세요.",
                f"셋째 주에는 {secondary.checkpoint}를 다른 지문에서 확인하고, 넷째 주에는 학생이 변화 이유를 말로 설명하게 합니다. 설명할 수 없는 항목은 완료 처리하지 않는 편이 좋습니다.",
                f"후반 2주는 성과를 보장하는 단계가 아니라 가설을 검증하는 기간입니다. {support.action} 뒤의 기록과 {extra.evidence}를 비교해 무엇을 계속하고 무엇을 멈출지 결정하세요.",
                f"3주차에 내신·모의고사 자료를 분리해 연습하고 4주차에 같은 진단표로 돌아옵니다. {profile.focus}에 필요한 근거가 실제로 달라졌을 때만 다음 범위를 넓히세요.",
            ]),
        ],
    }

    consultation = {
        "key": "consultation-checklist",
        "heading": stable_pick(seed, "high-h2-consult", [
            f"{title} 상담 전에 적어 갈 네 가지 질문",
            f"{profile.focus}을 수업 설명과 비교하는 체크리스트",
            f"첫 상담에서 학습 진단과 이용 조건을 나누는 법",
            f"{primary.label} 피드백을 구체적으로 묻는 방법",
            f"등록 전, 학생 자료로 확인할 마지막 기준",
            f"{locality} 고등 영어 상담을 기록으로 끝내는 방법",
        ]),
        "paragraphs": [
            stable_pick(seed, "high-consult-p1", [
                f"상담에는 {materials} 중 실제로 표시가 남은 자료를 가져가세요. ‘{primary.consult_question}’와 ‘{secondary.consult_question}’를 차례로 물으면 설명·독립 연습·재확인의 범위를 비교할 수 있습니다.",
                f"첫 상담의 목적은 많은 약속을 듣는 것이 아니라 {profile.focus}에 필요한 근거를 확인하는 일입니다. {primary.evidence}를 보여 주고 어떤 행동을 언제 다시 확인하는지 질문하세요.",
                f"학생이 사용 중인 자료를 펼쳐 {marker_a}에서 멈춘 위치와 {marker_b}을 다시 해 본 흔적을 보여 주세요. 상담 답변이 실제 자료에 적용되는지 확인해야 추상적인 설명을 줄일 수 있습니다.",
                f"학습 질문은 ‘잘 가르치나요’보다 구체적이어야 합니다. {primary.consult_question}를 묻고, 답변 뒤에는 {primary.checkpoint}를 언제 어떤 형식으로 확인하는지 이어 물으세요.",
                f"{locality} 고등 영어 상담에서는 학교 일정, 가능한 가정 학습 시간과 {primary.evidence}를 함께 준비하세요. 학습 목표와 실제로 지킬 수 있는 분량이 맞는지 비교합니다.",
                f"{profile.focus}에 대한 답을 얻으려면 시험지 한 장이면 충분할 수 있습니다. 학생이 막힌 위치를 보여 주고 {secondary.action}을 수업 전후 어디에 배치하는지 확인하세요.",
            ]),
            stable_pick(seed, "high-consult-p2", [
                f"센터의 주소·가능 학년·시간표·교습비는 사실 확인 항목입니다. 피드백 방식과 주간 계획은 학생 자료를 바탕으로 정할 항목이므로 두 목록을 나눠 적고, 확인되지 않은 운영 조건은 추정하지 마세요.",
                f"상담 뒤에는 답변을 학생과 다시 읽어 보세요. {primary.action}을 실제로 할 수 있는지, 다음 점검일이 분명한지, 미완료 때 계획을 어떻게 바꾸는지가 남아 있어야 실행 가능한 안내입니다.",
                f"등록 판단에서는 통학 시간과 실제 시간표를 먼저 대조합니다. 이후 {primary.label}·{secondary.label}의 피드백 근거가 보이는지 살펴야 이용 조건과 학습 적합성을 균형 있게 비교할 수 있습니다.",
                f"성과나 등급 변화를 단정하는 표현보다 기록을 확인하세요. 첫 풀이, 수정한 근거, 며칠 뒤 재풀이를 누가 언제 보는지 설명할 수 있어야 {profile.focus}과 연결됩니다.",
                f"마지막으로 학생에게 계획을 자기 말로 설명하게 해 보세요. {marker_c}과 {marker_d}의 순서를 이해하지 못했다면 과제량을 정하기 전에 상담 내용을 다시 단순화할 필요가 있습니다.",
                f"상담 메모에는 확인된 센터 사실, 학생의 현재 어려움, 첫 행동, 재확인 날짜를 네 줄로 남기세요. 다음 상담에서도 같은 형식을 쓰면 계획이 달라진 이유를 추적하기 쉽습니다.",
            ]),
        ],
        "checklist": [
            ("학생 자료", f"{locality} 학생의 {primary.evidence}와 {secondary.evidence}를 보여 주고 우선순위를 묻습니다."),
            ("첫 행동", f"{locality} 계획에서 {primary.action}을 실제 수업과 복습 중 어디에 배치하는지 확인합니다."),
            ("재확인", f"{locality} 학생이 {primary.checkpoint}를 언제 어떤 기록으로 다시 보는지 질문합니다."),
            ("이용 조건", f"{locality} 페이지의 {grade_fact} 현재 시간표·통학·교습비는 등록 전에 별도로 대조합니다."),
        ],
    }
    result = [direct, diagnosis, exam, facts, plan, consultation]
    for section in result:
        heading = str(section["heading"])
        if locality not in heading:
            section["heading"] = f"{locality} 고등 영어: {heading}"
    return result


def high_english_sections(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighEnglishProfile,
    seed: str,
) -> list[dict[str, object]]:
    """Build six purpose-led sections without repeating the page angle."""

    locality = str(center["locality"])
    primary, secondary, support, extra = profile.intents
    schools = relevant_schools(config, center)
    grades = relevant_grades(config, center, "영어")
    grade_value = "·".join(grades) if grades else "기재 없음"

    direct_paragraphs = [
        stable_pick(seed, "high-direct-p1", [
            f"{locality} 고등 영어의 핵심 질문은 ‘{profile.focus}’입니다. 최근 점수보다 먼저 ‘{primary.concern}’를 시험지의 실제 표시와 대조하세요.",
            f"{locality}에서 살필 고등 영어 주제는 ‘{profile.focus}’입니다. 학생에게 ‘{primary.concern}’를 묻고 답을 최근 풀이 흔적과 비교하세요.",
            f"{locality} 고등 영어를 비교할 때는 ‘{profile.focus}’부터 구체화합니다. 첫 질문은 ‘{primary.concern}’이며, 답은 성적표가 아니라 최근 시험지에서 확인합니다.",
            f"{locality} 학생의 출발점을 정하는 질문은 ‘{profile.focus}’입니다. 먼저 ‘{primary.concern}’에 학생이 직접 답하게 하고 풀이 기록과 맞는지 봅니다.",
            f"{locality} 고등 영어 상담의 첫 주제는 ‘{profile.focus}’입니다. ‘{primary.concern}’를 묻고 정답 수보다 답을 고른 과정을 확인하세요.",
            f"{locality}에서 고등 영어 계획을 세울 때 ‘{profile.focus}’을 첫 질문으로 둡니다. ‘{primary.concern}’가 실제 자료에서도 드러나는지부터 확인하세요.",
        ]),
        stable_pick(seed, "high-direct-p2", [
            f"다음으로 ‘{secondary.concern}’를 따로 확인하고 두 답이 모두 흔들리면 {primary.label}과 {secondary.label}을 한 과제로 묶지 말고 먼저 설명할 영역을 고르세요.",
            f"이어 ‘{secondary.concern}’도 별도 질문으로 남기고 답변과 풀이 흔적이 어긋난 지점을 찾아 {primary.label}과 {secondary.label}의 순서를 현실적으로 정하세요.",
            f"두 번째 질문은 ‘{secondary.concern}’이며 한 영역에서만 막혔다면 그 부분부터 보완하고, 둘 다 어렵다면 {primary.label}을 먼저 설명한 뒤 {secondary.label}을 재확인하세요.",
            f"이후 ‘{secondary.concern}’를 물어 원인을 나누고 {primary.label}과 {secondary.label}의 어려움을 구분해 교재나 문제 수를 늘리기 전 첫 행동을 정하세요.",
            f"‘{secondary.concern}’까지 확인한 뒤 두 답을 나란히 적고 학생 설명과 기록이 일치하는 영역부터 {primary.label}·{secondary.label}의 학습 순서를 잡습니다.",
            f"끝으로 ‘{secondary.concern}’를 확인해 단순 실수와 반복되는 병목을 구분하고 {primary.label}에 설명이 필요한지, {secondary.label}을 혼자 연습할지 판단하세요.",
        ]),
    ]

    evidence_paragraphs = [
        stable_pick(seed, "high-evidence-p1", [
            f"진단 자료에서는 ‘{primary.evidence}’와 ‘{secondary.evidence}’를 한 곳씩 표시하고 자료 위치·날짜·처음 판단한 이유가 달라진 지점을 남기세요.",
            f"최근 시험지와 교재에서 ‘{primary.evidence}’를 먼저 찾고 ‘{secondary.evidence}’를 다른 색으로 표시해 정답보다 판단이 바뀐 위치를 기록하세요.",
            f"많은 자료보다 역할이 다른 두 기록이 낫기 때문에 ‘{primary.evidence}’는 현재선, ‘{secondary.evidence}’는 풀이 과정으로 두고 같은 날짜 기준으로 비교하세요.",
            f"첫 풀이를 지우지 말고 ‘{primary.evidence}’와 ‘{secondary.evidence}’를 나란히 두어 도움을 받기 전후로 달라진 근거를 한 줄씩 적으세요.",
            f"진단할 때는 ‘{primary.evidence}’와 ‘{secondary.evidence}’가 보이는 자료만 고르고 처음 답의 이유와 수정한 이유를 나눠 지식 부족과 절차 문제를 구분하세요.",
            f"시험지나 학습표에서 ‘{primary.evidence}’를 표시한 뒤 ‘{secondary.evidence}’와 비교하고 자료 위치·날짜·수정 이유를 남겨 우연한 실수인지 반복되는 문제인지 확인하세요.",
        ]),
        stable_pick(seed, "high-evidence-p2", [
            f"다시 볼 날짜에는 ‘{primary.checkpoint}’와 ‘{secondary.checkpoint}’를 차례로 묻습니다. 두 답의 변화가 이 학습 초점에 필요한 설명과 독립 연습을 나누는 기준입니다.",
            f"재확인표에는 ‘{primary.checkpoint}’를 첫 질문으로, ‘{secondary.checkpoint}’를 두 번째 질문으로 적습니다. 답을 학생이 혼자 설명해야 앞서 정한 초점의 다음 행동을 고를 수 있습니다.",
            f"첫 기록과 재풀이 기록을 비교할 때는 ‘{primary.checkpoint}’에 답한 근거와 ‘{secondary.checkpoint}’의 변화를 함께 봅니다. 그 차이로 계속 보완할 부분을 판단하세요.",
            f"다음 점검에서도 같은 시험지를 사용하되 질문은 ‘{primary.checkpoint}’와 ‘{secondary.checkpoint}’로 고정합니다. 이 기준이면 처음 세운 학습 방향의 변화 여부를 확인하기 쉽습니다.",
            f"각 표시 옆에는 ‘{primary.checkpoint}’ 또는 ‘{secondary.checkpoint}’ 중 해당 질문을 붙이고 재확인 날짜를 적습니다. 이 기록으로 {primary.label}과 {secondary.label}에 필요한 피드백을 구체화합니다.",
            f"도움 없이 다시 수행한 날에는 ‘{primary.checkpoint}’와 ‘{secondary.checkpoint}’에 학생이 직접 답하게 하세요. 첫 자료와 비교하면 설명·연습·재확인의 순서가 분명해집니다.",
        ]),
    ]

    exam_paragraphs = [
        stable_pick(seed, "high-exam-p1", [
            f"내신 기록에는 범위·본문·어법·서술형을, 모의고사 기록에는 유형·근거·소요 시간을 적습니다. 두 기록 모두 ‘{primary.exam_use}’와 ‘{secondary.exam_use}’ 중 실제 자료에 맞는 기준을 고르세요.",
            f"학교 시험과 모의고사는 같은 문제 수로 계획하지 않습니다. 각각의 자료에서 ‘{primary.exam_use}’와 ‘{secondary.exam_use}’가 필요한 지점을 찾아 완료 기준을 나누세요.",
            f"내신은 정해진 범위의 이해와 변형 조건을, 모의고사는 새 지문의 근거와 시간을 확인합니다. {primary.label}과 {secondary.label}은 한 시험에 고정하지 말고 두 자료에서 모두 대조하세요.",
            f"시험 일정표에서 내신 범위와 모의고사 누적 약점을 다른 줄에 둡니다. 이어 ‘{primary.exam_use}’와 ‘{secondary.exam_use}’를 각 자료에 적용할지 판단하세요.",
            f"학교 자료의 마감일과 다음 점검 날짜를 분리한 뒤 {primary.label}·{secondary.label}의 최소 행동을 각 일정에 맞게 배치하세요.",
            f"내신 기간에도 누적 학습을 완전히 멈추지 마세요. ‘{primary.exam_use}’와 ‘{secondary.exam_use}’를 살펴 학교 범위와 새 지문에 필요한 시간을 따로 남깁니다.",
        ]),
        stable_pick(seed, "high-exam-p2", [
            f"이번 주에는 학교 범위에 적용할 {primary.label}·{secondary.label} 행동과 새 지문에서 확인할 행동을 각각 한 줄씩 정합니다. 시험 뒤 네 기록을 비교해 공통 기초와 시험별 문제를 구분하세요.",
            f"내신 마감일과 모의고사 재확인일을 달력에 따로 표시하고 두 일정마다 {primary.label}과 {secondary.label}의 최소 분량을 적으세요. 한쪽이 밀리면 다음 주 비중을 조정합니다.",
            f"정해진 학교 범위와 처음 보는 지문에서 {primary.label}·{secondary.label}이 각각 어떻게 드러나는지 확인하세요. 자료별 완료 기준으로 시험 전후의 우선순위를 다시 세웁니다.",
            f"주간 계획에는 {primary.label}과 {secondary.label}의 최소 행동을 각각 한 줄로 남깁니다. 시험 뒤에는 점수보다 근거·시간·독립 수행 기록의 변화를 비교하세요.",
            f"{primary.label} 기록과 {secondary.label} 기록은 구분하되 내신·모의고사 자료를 모두 연결하세요. 남은 날짜와 실제 완료량으로 다음 주의 설명·연습 비중을 정합니다.",
            f"내신 종료 뒤 학교 범위 기록과 모의고사 누적 기록을 함께 봅니다. 두 자료에서 {primary.label}·{secondary.label}의 겹치는 약점과 다른 절차를 구분해 다음 계획에 반영하세요.",
        ]),
    ]
    exam_timing = stable_pick(seed, "high-exam-focus-timing", [
        f"‘{profile.focus}’을 시험 주간에 적용할 때는 {primary.label} 점검일과 {secondary.label} 재확인일을 다른 줄에 적으세요.",
        f"시험 달력에는 ‘{profile.focus}’에 필요한 {primary.label} 확인일과 {secondary.label} 다음 점검일을 따로 표시하세요.",
        f"‘{profile.focus}’을 계획으로 옮길 때 {primary.label}은 먼저 확인하고 {secondary.label}은 별도 재확인일을 정하세요.",
        f"시험 일정 옆에는 ‘{profile.focus}’의 {primary.label} 점검 날짜와 {secondary.label} 재확인 날짜를 구분해 남기세요.",
        f"‘{profile.focus}’에 맞춘 주간표에서는 {primary.label} 확인과 {secondary.label} 재확인을 같은 날에 몰아넣지 마세요.",
        f"내신 마감일을 적은 뒤 ‘{profile.focus}’과 관련된 {primary.label} 점검일, {secondary.label} 재확인일을 차례로 배치하세요.",
    ])
    exam_follow_up = stable_pick(seed, "high-exam-focus-follow-up", [
        f"시험이 끝난 뒤에는 {primary.label}·{secondary.label} 기록 가운데 학생이 근거를 설명하지 못한 영역만 다음 주 계획에 남깁니다.",
        f"시험 뒤 두 기록을 비교해 {primary.label}과 {secondary.label} 중 학생이 판단 근거를 다시 말하지 못한 영역부터 보완하세요.",
        f"다음 주 계획에는 {primary.label}·{secondary.label} 가운데 첫 풀이와 재확인 답이 달랐던 영역의 행동만 남깁니다.",
        f"점수 확인이 끝나면 {primary.label}과 {secondary.label}의 기록을 대조하고 독립 설명이 끊긴 영역의 분량만 조정하세요.",
        f"이후 학생이 {primary.label}·{secondary.label}의 판단 과정을 각각 설명하게 하고 근거가 불분명한 쪽만 다시 계획합니다.",
        f"마지막으로 {primary.label}과 {secondary.label}의 완료 기록을 나눠 보고 재확인 기준을 통과하지 못한 영역만 이어 가세요.",
    ])
    exam_paragraphs.append(f"{exam_timing} {exam_follow_up}")

    if schools:
        school_names = "·".join(schools[:4])
        school_open = stable_pick(seed, "high-school-open", [
            f"참고 학교로 {school_names} 등이 기재돼 있지만",
            f"제공된 학교 목록에는 {school_names} 등이 포함돼 있으나",
            f"페이지에서 확인되는 참고 학교는 {school_names} 등이며",
            f"상담 준비용 학교명은 {school_names} 등으로 정리돼 있으나",
            f"내신 자료를 대조할 참고 학교에는 {school_names} 등이 적혀 있지만",
            f"공개 자료에 {school_names} 등이 학교명으로 보이지만",
            f"학교 범위를 확인할 때 참고할 명칭에는 {school_names} 등이 포함돼 있으나",
            f"제공 목록에서는 {school_names} 등을 확인할 수 있지만",
        ])
        school_close = stable_pick(seed, "high-school-close", [
            "이 목록만으로 해당 학교 학생의 수업 가능 여부를 판단할 수는 없습니다",
            "실제 재학 학교의 범위와 현재 자료 반영 가능 여부는 따로 확인해야 합니다",
            "현재 개설 범위와 학교별 시험 자료 반영 여부는 상담에서 확인해야 합니다",
            "학생의 실제 학교와 시험 범위를 알려 자료 반영 가능 여부를 대조해야 합니다",
            "학교명보다 실제 시험 범위표와 현재 수업 가능 여부를 먼저 확인해야 합니다",
            "이 명칭은 참고용이므로 재학 학교의 범위와 개설 여부를 별도로 물어야 합니다",
            "학교별 진도와 자료 반영 방식은 목록이 아니라 상담 시점의 답변으로 판단해야 합니다",
            "해당 학교의 모든 과정이 열린다는 뜻은 아니므로 현재 범위를 다시 확인해야 합니다",
        ])
        school_note = f"{school_open}, {school_close}."
    else:
        school_note = stable_pick(seed, "high-school-blank", [
            "공개된 참고 고등학교 목록은 없습니다. 실제 재학 학교와 시험 범위표를 알려 현재 자료 반영 가능 여부를 확인하세요.",
            "제공 자료에 참고 고등학교가 기재돼 있지 않습니다. 학생의 학교와 시험 범위를 상담에서 직접 전달해야 합니다.",
            "이 페이지에는 참고 학교 목록이 없습니다. 학교명만 추정하지 말고 실제 범위표와 현재 개설 범위를 대조하세요.",
            "상담 준비용 고등학교 명칭이 공개돼 있지 않습니다. 재학 학교·시험 범위·현재 진도를 직접 알려 주세요.",
            "확인 가능한 참고 학교 정보가 없습니다. 학생이 다니는 학교와 최근 시험 자료를 기준으로 반영 가능 여부를 물어야 합니다.",
            "학교 목록은 제공되지 않았습니다. 실제 재학 학교의 시험 일정과 자료를 상담 시점에 확인하세요.",
            "공개 자료만으로 참고 고등학교를 특정할 수 없습니다. 학교 범위표를 가져가 현재 수업에 반영할 수 있는지 물어보세요.",
            "기재된 참고 학교가 없으므로 학생의 재학 학교와 범위 자료를 별도 확인 항목으로 준비해야 합니다.",
        ])

    if grades:
        grade_open = stable_pick(seed, "high-grade-open", [
            f"제공 자료의 영어 가능 학년 표기는 {grade_value}이며",
            f"영어 가능 학년 항목에서 {grade_value} 표기를 확인할 수 있어",
            f"현재 페이지에서 확인되는 영어 학년 범위는 {grade_value}이며",
            f"제공된 영어 학년 정보에는 {grade_value} 표기가 있어",
        ])
        grade_check = f"영어 가능 학년({grade_value})"
    else:
        grade_open = stable_pick(seed, "high-grade-open-blank", [
            "제공 자료에는 고등 영어 가능 학년이 기재돼 있지 않아",
            "현재 페이지에서 고등 영어 가능 학년 표기를 확인할 수 없으므로",
            "공개 자료만으로는 고등 영어 가능 학년을 판단할 수 없어",
            "제공된 센터 자료에 고등 영어 개설 학년이 표시돼 있지 않기 때문에",
        ])
        grade_check = "고등 영어 가능 학년 기재 여부"
    grade_close = stable_pick(seed, "high-grade-close", [
        f"{support.label}과 {extra.label}의 실제 개설 여부와 시간표는 상담 시점에 다시 확인하세요",
        f"{support.label}·{extra.label} 관련 반 구성과 현재 시간표는 별도 확인 항목입니다",
        f"이 표기만으로 {support.label}·{extra.label} 과정의 운영 여부를 단정할 수는 없습니다",
        f"{support.label}과 {extra.label}을 어느 학년에 적용하는지는 현재 개설 범위와 함께 물어보세요",
    ])
    grade_center_note = stable_pick(seed, "high-center-fact", [
        f"{grade_open} {grade_close}. 이 페이지에 연결된 센터는 {center['center_name']}이며 제공 주소는 {center['address']}입니다.",
        f"{grade_open} {grade_close}. 제공 센터명은 {center['center_name']}, 주소는 {center['address']}입니다.",
        f"{grade_open} {grade_close}. 확인할 센터는 {center['center_name']}이고 제공된 위치는 {center['address']}입니다.",
        f"{grade_open} {grade_close}. 센터 정보는 {center['center_name']}과 {center['address']}를 기준으로 확인합니다.",
        f"{grade_open} {grade_close}. 페이지가 안내하는 센터명은 {center['center_name']}, 제공 주소는 {center['address']}입니다.",
        f"{grade_open} {grade_close}. 방문 전 {center['center_name']}과 제공 주소 {center['address']}를 함께 대조하세요.",
    ])
    tuition_open = stable_pick(seed, "high-tuition-open", [
        f"{center['center_name']} 등록 전에는",
        f"{center['center_name']} 이용 조건을 비교할 때는",
        f"{center['center_name']} 상담을 마친 뒤에는",
        f"{center['center_name']} 방문을 결정하기 전에는",
        f"{center['center_name']}의 운영 정보를 확인할 때는",
        f"{center['center_name']} 학습 계획과 별도로",
        f"{center['center_name']} 등록 조건을 확인할 때는",
        f"{center['center_name']} 상담 메모를 정리할 때는",
    ])
    if center["tuition_url"]:
        tuition_close = stable_pick(seed, "high-tuition-close", [
            "연결된 교습비 자료와 실제 시간표를 비교하고 제공 주소를 기준으로 통학 동선을 확인합니다",
            "교습비 자료·현재 시간표·통학 동선을 서로 다른 항목으로 적습니다",
            "교습비 자료와 실제 시간표를 대조하고 방문 동선을 따로 살핍니다",
            "교습비·시간표·통학 조건을 확인된 사실로만 기록합니다",
            "교습비와 시간표를 상담 시점에 다시 확인하고 제공 주소에서의 이동 시간을 계산합니다",
            "페이지의 교습비 자료와 실제 시간표, 통학에 걸리는 시간을 차례로 확인합니다",
            "교습비 자료의 기준과 현재 시간표가 같은 시점의 정보인지 확인합니다",
            "학습 조언과 분리해 교습비·시간표·통학 가능 여부를 대조합니다",
        ])
    else:
        tuition_close = stable_pick(seed, "high-tuition-close-blank", [
            "공개된 교습비 자료가 없다는 점과 실제 시간표 확인 필요를 함께 적습니다",
            "교습비와 실제 시간표를 추정하지 않고 각각 직접 확인합니다",
            "현재 교습비·시간표·통학 동선을 별도 사실 항목으로 물어봅니다",
            "교습비와 운영 시간, 제공 주소까지의 이동 시간을 차례로 확인합니다",
            "교습비·시간표·통학 가능 여부를 학습 조언과 분리해 질문합니다",
            "실제 시간표와 교습비를 등록 전에 직접 대조합니다",
            "공개 자료에 없는 교습비와 운영 시간을 상담 질문으로 남깁니다",
            "제공 주소의 통학 동선과 현재 이용 조건을 함께 확인합니다",
        ])
    tuition_sentence = f"{tuition_open} {tuition_close}."

    plan_paragraphs = [
        stable_pick(seed, "high-plan-p1", [
            f"첫 7일에는 ‘{primary.evidence}’를 기준선으로 남깁니다. 이번 주 실행 항목은 ‘{primary.action}’이며 하루 분량보다 완료 흔적을 기록하는 일이 우선입니다.",
            f"7일 계획의 출발점은 ‘{primary.evidence}’입니다. 첫 행동을 ‘{primary.action}’으로 정하고 시작일·완료일·남은 질문을 한 줄씩 적으세요.",
            f"첫 주에는 ‘{primary.evidence}’에서 한 가지 병목만 고릅니다. 실행 항목은 ‘{primary.action}’이며 새 교재나 추가 문제는 재확인 뒤에 결정하세요.",
            f"기준선 자료로 ‘{primary.evidence}’를 보존하세요. 7일 동안 ‘{primary.action}’을 해 보고 도움을 받은 지점과 혼자 한 지점을 구분합니다.",
            f"처음 7일은 진도를 늘리는 기간이 아닙니다. ‘{primary.evidence}’를 남기고 실행 항목 ‘{primary.action}’의 완료 여부부터 확인하세요.",
            f"주간 계획표 첫 줄에 ‘{primary.evidence}’를 적습니다. 이후 ‘{primary.action}’을 실행하고 실제로 끝낸 날짜와 수정한 이유를 기록하세요.",
        ]),
        stable_pick(seed, "high-plan-p2", [
            f"7일 뒤에는 ‘{primary.checkpoint}’를 다시 묻습니다. 학생이 혼자 답하면 ‘{secondary.action}’으로 넘어가고, 어렵다면 첫 행동을 더 작은 분량으로 반복하세요.",
            f"다음 점검 질문은 ‘{primary.checkpoint}’입니다. 답이 불분명하면 ‘{primary.action}’의 설명 단계를 보완하고, 분명하면 ‘{secondary.action}’을 다음 주 행동으로 정합니다.",
            f"일주일 뒤 ‘{primary.checkpoint}’를 같은 자료로 확인합니다. 첫 기록과 달라진 근거가 있을 때만 ‘{secondary.action}’을 새 과제로 추가하세요.",
            f"재확인 기준은 ‘{primary.checkpoint}’입니다. 완료하지 못한 이유가 분량인지 이해인지 나눈 뒤 ‘{primary.action}’을 유지할지 ‘{secondary.action}’으로 옮길지 결정합니다.",
            f"마지막 날에는 ‘{primary.checkpoint}’에 학생이 직접 답합니다. 답을 근거로 첫 행동을 다시 설명할지, ‘{secondary.action}’을 시작할지 선택하세요.",
            f"7일째에는 ‘{primary.checkpoint}’를 첫 기록과 대조합니다. 변화가 없으면 문제 수를 늘리지 말고 ‘{primary.action}’의 순서를 단순화한 뒤 다시 확인합니다.",
        ]),
    ]

    focus_compact = re.sub(r"\s+", "", profile.focus)
    if "방학" in focus_compact and "불안" in focus_compact:
        plan_paragraphs.append(
            "평일에는 20~30분짜리 제한 세트를 풀고 멈춘 위치와 다시 시작한 시각을 함께 기록하세요. "
            "방학에는 약점 유형을 묶어 연습하되 재확인 날짜를 미리 정합니다. "
            "불안으로 풀이가 끊긴 날은 정답 수보다 회복까지 걸린 시간과 다시 시작한 문항을 남겨 다음 분량을 조정하세요."
        )

    consultation_paragraphs = [
        stable_pick(seed, "high-consult-p1", [
            f"상담에는 표시가 남아 있는 시험지와 학교 범위표를 가져갑니다. ‘{primary.consult_question}’와 ‘{secondary.consult_question}’를 차례로 물으세요.",
            f"최근 시험지 한 장과 현재 교재를 준비하세요. 첫 질문은 ‘{primary.consult_question}’, 두 번째는 ‘{secondary.consult_question}’입니다.",
            f"학교 범위표·최근 시험지·오답 기록 중 실제 표시가 있는 자료를 고릅니다. ‘{primary.consult_question}’와 ‘{secondary.consult_question}’를 메모해 가세요.",
            f"좋은 상담 질문은 자료에서 시작합니다. 학생이 멈춘 시험지를 보여 주며 ‘{primary.consult_question}’, 이어 ‘{secondary.consult_question}’를 묻습니다.",
            f"상담 전 ‘{profile.focus}’을 설명할 최근 영어 자료와 가능한 가정 학습 시간을 적어 두세요. ‘{primary.consult_question}’와 ‘{secondary.consult_question}’를 분리해 질문합니다.",
            f"시험지에서 막힌 위치를 표시해 상담에 가져갑니다. ‘{primary.consult_question}’를 먼저 묻고 ‘{secondary.consult_question}’를 다음 질문으로 둡니다.",
        ]),
        stable_pick(seed, "high-consult-p2", [
            f"답변은 {primary.label}의 첫 행동과 {secondary.label}의 재확인 날짜로 바꿔 적으세요. 주소·학년·시간표·교습비는 학습 질문과 다른 줄에 남깁니다.",
            f"상담 메모에는 {primary.label}의 실행 항목과 {secondary.label}의 확인일이 남아야 합니다. 센터 이용 조건은 별도 목록으로 구분하세요.",
            f"설명을 들은 뒤 {primary.label}의 시작 행동과 {secondary.label}의 완료 기준을 한 줄씩 적습니다. 학습 적합성과 등록 조건을 섞지 않아야 비교가 쉽습니다.",
            f"답변이 {primary.label}의 실제 행동과 {secondary.label}의 점검일로 이어지는지 확인하세요. 운영 정보는 확인된 사실만 따로 기록합니다.",
            f"마지막에는 {primary.label}을 누가 언제 확인하는지, {secondary.label}을 어떤 자료로 다시 볼지 적습니다. 시간표·교습비·통학은 별도 조건입니다.",
            f"상담 뒤 학생이 {primary.label}의 첫 행동을 설명하게 하세요. {secondary.label}의 재확인 날짜와 센터 이용 조건은 서로 다른 칸에 적습니다.",
        ]),
    ]

    headings = [
        f"{locality} 고등 영어 첫 진단: {profile.focus}",
        stable_pick(seed, "high-heading-evidence", [
            f"{secondary.label} 자료로 {primary.label}을 다시 확인하는 법",
            f"진단에 필요한 {primary.label}·{secondary.label} 기록",
            f"시험지에서 {primary.label}·{secondary.label} 자료를 남기는 방법",
            f"{primary.label}·{secondary.label} 기록이 보여 주는 영어 학습의 현재선",
            f"정답보다 먼저 볼 {primary.label}·{secondary.label} 풀이 흔적",
            f"{primary.label}과 {secondary.label} 재확인 자료를 고르는 기준",
        ]),
        stable_pick(seed, "high-heading-exam", [
            f"내신과 모의고사에서 {primary.label}·{secondary.label} 기록을 나누는 방법",
            f"{primary.label}·{secondary.label}을 내신·모의고사에 다르게 적용하는 기준",
            f"학교 시험과 모의고사의 {primary.label}·{secondary.label} 계획",
            f"내신 범위와 누적 학습에 {primary.label}·{secondary.label}을 배치하는 법",
            f"{primary.label}·{secondary.label}의 내신 일정과 재확인일을 나누는 기준",
            f"시험 전후 {primary.label}·{secondary.label}의 비중을 조정하는 방법",
        ]),
        stable_pick(seed, "high-heading-facts", [
            f"{locality} {primary.label}·{secondary.label} 상담 전 학교·학년 확인 순서",
            f"{locality} 센터 정보와 {primary.label}·{secondary.label} 질문을 구분하는 법",
            f"{locality} {primary.label} 상담 전 확인할 학교·학년·주소 정보",
            f"{secondary.label} 진단과 분리해 볼 {locality} 센터 이용 조건",
            f"{locality} 참고 학교와 {primary.label} 개설 범위를 대조하는 방법",
            f"{locality} {primary.label}의 확인된 사실과 상담에서 물을 항목",
        ]),
        stable_pick(seed, "high-heading-plan", [
            f"7일 동안 {primary.label}을 실행하고 {secondary.label}을 재확인하는 방법",
            f"{primary.label}에서 {secondary.label}으로 이어지는 7일 영어 계획",
            f"한 가지 병목부터 시작하는 {primary.label}·{secondary.label} 7일 점검",
            f"{primary.label}·{secondary.label}의 분량을 정하는 7일 실행안",
            f"첫 기록과 재풀이를 연결하는 {primary.label} 7일 계획",
            f"{primary.label}의 시작일과 {secondary.label}의 확인일을 정하는 법",
        ]),
        stable_pick(seed, "high-heading-consult", [
            f"{primary.label}·{secondary.label} 상담 전 준비할 자료와 등록 확인 항목",
            f"상담에서 {primary.label}·{secondary.label}을 묻고 이용 조건을 확인하는 순서",
            f"시험지와 범위표로 준비하는 {primary.label} 상담 질문",
            f"{secondary.label} 재확인 질문과 센터 이용 조건을 나누는 방법",
            f"{primary.label}·{secondary.label} 답변을 실행 계획으로 바꾸는 상담 메모",
            f"{primary.label}·{secondary.label} 상담 전 가져갈 자료와 확인할 사실",
        ]),
    ]
    headings = [
        (
            "·".join(heading.split("·")[:2]) + ", " + "·".join(heading.split("·")[2:])
            if heading.count("·") > 3
            else heading
        )
        for heading in headings
    ]
    return [
        {"key": "direct-answer", "heading": headings[0], "paragraphs": direct_paragraphs},
        {"key": "diagnostic-evidence", "heading": headings[1], "paragraphs": evidence_paragraphs},
        {"key": "exam-strategy", "heading": headings[2], "paragraphs": exam_paragraphs},
        {"key": "school-center-facts", "heading": headings[3], "paragraphs": [school_note, f"{grade_center_note} {tuition_sentence}"]},
        {"key": "four-week-plan", "heading": headings[4], "paragraphs": plan_paragraphs},
        {
            "key": "consultation-checklist",
            "heading": headings[5],
            "paragraphs": consultation_paragraphs,
            "checklist": [
                ("학생 자료", f"최근 시험지와 {primary.label} 표시 위치"),
                ("시험 계획", f"학교 범위표와 {secondary.label} 마감일"),
                ("재확인", f"{support.label} 확인 날짜와 학생 설명 기록"),
                ("이용 조건", f"{grade_check}·시간표·교습비·통학 동선"),
            ],
        },
    ]


def _legacy_high_english_faq(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighEnglishProfile,
    seed: str,
) -> list[tuple[str, str]]:
    locality = str(center["locality"])
    primary, secondary, support, extra = profile.intents
    schools = relevant_schools(config, center)
    grades = relevant_grades(config, center, "영어")
    grade_answer = (
        f"{locality} 페이지의 제공 자료에는 영어 {'·'.join(grades)}이 가능 학년으로 기재되어 있습니다. 실제 개설 시간표와 반 구성은 상담 시점에 다시 확인하세요."
        if grades else
        f"{locality} 페이지의 제공 자료에는 고등 영어 가능 학년이 기재되지 않았습니다. 현재 개설 학년과 시간표를 센터에 직접 확인해야 합니다."
    )
    school_answer = (
        f"{locality} 페이지의 { '·'.join(schools[:4]) } 등은 상담 준비를 위한 참고 학교명입니다. 수업 가능을 보장하지 않으므로 실제 재학 학교의 범위표와 현재 개설 범위를 함께 확인하세요."
        if schools else
        f"{locality} 페이지에는 공개된 참고 고등학교 목록이 없습니다. 실제 재학 학교와 시험 범위표를 가져가 자료 반영 방식과 현재 개설 범위를 확인하세요."
    )
    intents = [
        (
            f"{locality} 고등 영어에서 {profile.focus}은 어떻게 진단하나요?",
            f"{primary.evidence}를 먼저 보고 {secondary.evidence}를 별도로 확인합니다. 이후 {primary.action}을 짧게 실행해 {primary.checkpoint}가 나타나는지 비교하면 첫 학습 순서를 정할 수 있습니다.",
        ),
        (
            f"{locality} 고등 영어에서 {primary.label}과 {secondary.label} 중 무엇을 먼저 공부해야 하나요?",
            f"두 영역을 같은 분량으로 시작하지 않습니다. 시험 일정과 최근 기록을 함께 보고, {primary.concern}가 확인되면 {primary.action}부터 실행한 뒤 {secondary.checkpoint}를 다음 점검 기준으로 둡니다.",
        ),
        (
            f"{locality} 학생의 내신과 모의고사 영어는 같은 방식으로 준비해도 되나요?",
            f"내신은 범위표·교과서·학교 자료를 중심으로, 모의고사는 처음 보는 지문의 유형·근거·시간 기록을 중심으로 봅니다. {primary.exam_use}을 적용해 두 계획의 주간 비중을 조정하세요.",
        ),
        (
            f"{locality} 고등 영어 상담에 어떤 자료를 가져가야 하나요?",
            f"최근 시험지, 학교 시험 범위표, 현재 교재와 오답 기록 중 실제 표시가 남은 자료를 준비하세요. ‘{primary.consult_question}’를 묻고 첫 행동과 재확인 날짜를 함께 기록하면 됩니다.",
        ),
        (
            f"{locality}에서 {center['center_name']}의 고등 영어 가능 학년은 어떻게 확인하나요?",
            grade_answer,
        ),
        (
            f"{locality} 페이지에 적힌 고등학교는 모두 수업 가능한 학교인가요?",
            school_answer,
        ),
        (
            f"{locality} 고등 영어 4주 학습 계획 뒤에는 무엇을 비교해야 하나요?",
            f"성과를 보장하는 기간으로 보지 말고 첫 풀이와 재풀이의 근거를 비교하세요. {support.checkpoint}와 {extra.checkpoint}가 달라졌는지 확인한 뒤 분량·순서·피드백 주기를 조정합니다.",
        ),
        (
            f"{locality} 고등학생의 {secondary.label} 과제는 얼마나 해야 하나요?",
            f"정해진 문제 수보다 독립 수행과 재확인이 가능한 분량이 우선입니다. {secondary.action}을 마친 기록을 보고 {secondary.checkpoint}가 확인되는 범위 안에서 다음 분량을 정하세요.",
        ),
    ]
    chosen = [intents[index] for index in (0, 1, 2, 3)]
    chosen.append(intents[4 + stable_int(seed, "high-faq-fact") % 4])
    return chosen


def high_english_faq(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighEnglishProfile,
    seed: str,
) -> list[tuple[str, str]]:
    locality = str(center["locality"])
    primary, secondary, support, extra = profile.intents
    schools = relevant_schools(config, center)
    grades = relevant_grades(config, center, "영어")
    if grades:
        grade_text = "·".join(grades)
        grade_answer = stable_pick(seed, "high-faq-grade-answer", [
            f"{locality} 페이지 제공 자료에는 영어 {grade_text}이 가능 학년으로 기재돼 있습니다. 현재 고등 영어 시간표와 반 구성은 등록 전에 다시 확인하세요.",
            f"{locality}의 영어 가능 학년 표기는 {grade_text}입니다. 실제 수업 시간과 현재 개설 반은 상담 시점의 시간표로 대조하세요.",
            f"공개된 {locality} 페이지에서 확인되는 영어 학년 정보는 {grade_text}입니다. 등록 전에는 실제 개설 여부와 시간을 다시 확인해야 합니다.",
            f"{locality} 페이지의 제공 자료 기준 영어 가능 학년은 {grade_text}입니다. 이 표기와 현재 시간표가 일치하는지 상담에서 확인하세요.",
            f"제공 자료에 적힌 {locality} 영어 학년 범위는 {grade_text}입니다. 다만 현재 열려 있는 반과 수업 시간은 별도 확인 항목입니다.",
            f"{locality} 페이지에는 영어 {grade_text}이 가능 학년으로 표시돼 있습니다. 실제 운영 학년과 시간은 최신 시간표를 기준으로 물어보세요.",
            f"{locality}의 가능 학년 자료에는 영어 {grade_text}이 적혀 있습니다. 등록 전에 현재 수업 시간과 개설 범위를 다시 대조하세요.",
            f"제공된 {locality} 페이지의 영어 학년은 {grade_text}입니다. 실제 수업 가능 여부는 상담 시점의 시간표로 확인하세요.",
        ])
    else:
        grade_answer = stable_pick(seed, "high-faq-grade-answer-blank", [
            f"{locality} 페이지 제공 자료에는 고등 영어 가능 학년이 기재돼 있지 않습니다. 현재 개설 학년과 시간표를 직접 확인해야 합니다.",
            f"공개된 {locality} 자료만으로 고등 영어 가능 학년을 판단할 수 없습니다. 등록 전에 개설 학년과 시간을 따로 물어보세요.",
            f"{locality} 페이지에는 고등 영어 학년 표기가 없습니다. 현재 운영 학년과 실제 시간표를 센터 답변으로 확인하세요.",
            f"제공된 {locality} 학년 정보가 비어 있으므로 가능한 고등 영어 학년과 실제 시간을 등록 전에 대조해야 합니다.",
        ])
    if schools:
        school_text = "·".join(schools[:4])
        school_answer = stable_pick(seed, "high-faq-school-answer", [
            f"참고 학교에는 {school_text} 등이 적혀 있습니다. 이는 수업 가능 학교를 보장하는 목록이 아니므로 실제 재학 학교의 범위표와 현재 자료 반영 여부를 함께 확인하세요.",
            f"제공된 참고 학교명은 {school_text} 등입니다. 학교명만으로 개설 여부를 판단하지 말고 학생의 시험 범위와 현재 반영 가능한 자료를 따로 물어보세요.",
            f"{school_text} 등은 상담 준비에 쓰는 참고 학교입니다. 실제 재학 학교의 범위표를 제시하고 해당 범위를 현재 수업에 반영할 수 있는지 대조하세요.",
            f"페이지에서 확인되는 학교는 {school_text} 등이며 모든 과정의 개설을 뜻하지 않습니다. 학생 학교의 시험 일정과 자료 반영 가능 여부를 직접 확인하세요.",
            f"참고용 학교 목록에는 {school_text} 등이 포함돼 있습니다. 등록 판단은 학교명이 아니라 실제 범위표와 센터의 현재 개설 답변을 기준으로 해야 합니다.",
            f"{school_text} 등은 주변 학교를 확인하기 위한 명칭입니다. 재학 학교와 시험 범위를 알린 뒤 현재 자료로 학습 계획을 세울 수 있는지 물어보세요.",
            f"공개 자료에는 {school_text} 등이 참고 학교로 표시돼 있습니다. 수업 가능 여부는 학생의 학교·범위·현재 개설 상황을 함께 확인한 뒤 판단하세요.",
            f"확인 가능한 참고 학교는 {school_text} 등입니다. 이 명칭과 실제 운영 범위는 다를 수 있으므로 시험 범위표를 가져가 반영 방식을 확인하세요.",
            f"학교 정보에는 {school_text} 등이 보이지만 이는 가능 학교 확정 목록이 아닙니다. 학생이 다니는 학교의 범위와 현재 수업 자료를 함께 대조하세요.",
            f"{school_text} 등은 상담 질문을 준비하기 위한 학교명입니다. 실제 재학 학교의 시험 자료를 보여 주고 현재 개설 범위에 맞는지 직접 확인하세요.",
            f"참고 학교로 {school_text} 등이 기재돼 있습니다. 학교명 자체보다 학생의 범위표와 최근 시험지를 기준으로 자료 반영 가능 여부를 물어보세요.",
            f"제공 목록에서 {school_text} 등을 확인할 수 있습니다. 다만 모든 학교 과정이 열려 있다는 뜻은 아니므로 재학 학교의 실제 범위를 따로 대조하세요.",
        ])
    else:
        school_answer = stable_pick(seed, "high-faq-school-answer-blank", [
            "참고할 고등학교 명칭이 별도로 기재돼 있지 않습니다. 재학 학교와 시험 범위를 직접 전달하고 현재 자료 반영 방식을 물어보세요.",
            "제공 자료에는 참고 학교명이 없습니다. 학생이 다니는 학교의 범위표를 가져가 현재 수업에 반영할 수 있는지 확인하세요.",
            "확인 가능한 고등학교 목록이 없으므로 학교명을 추정하지 말고 실제 재학 학교와 시험 일정을 상담에서 알려 주세요.",
            "이 페이지에는 참고 학교 정보가 기재돼 있지 않습니다. 학생의 학교·시험 범위·현재 진도를 직접 제시해 자료 반영 여부를 확인하세요.",
            "공개된 학교 명칭이 없기 때문에 실제 범위표와 최근 시험지를 기준으로 현재 개설 범위를 물어봐야 합니다.",
            "참고 학교 목록은 제공되지 않았습니다. 재학 학교의 시험 일정과 범위 자료를 가져가 반영 가능 여부를 따로 대조하세요.",
            "학교 정보가 비어 있으므로 주변 학교를 임의로 판단하지 말고 학생 학교의 범위와 현재 수업 자료를 직접 확인하세요.",
            "제공 자료만으로 참고 학교를 특정할 수 없습니다. 재학 학교와 시험 범위를 알린 뒤 현재 수업에서 다룰 수 있는지 물어보세요.",
        ])
    fact_answer = f"{grade_answer} {school_answer}"
    first_answer = stable_pick(seed, "high-faq-answer-first", [
        f"최근 자료에서 ‘{primary.evidence}’를 표시하고 ‘{primary.checkpoint}’를 다음 점검 질문으로 씁니다. 그 답을 확인한 뒤 ‘{primary.action}’을 첫 실행 항목으로 정하세요.",
        f"첫 확인 자료는 ‘{primary.evidence}’입니다. 학생이 ‘{primary.checkpoint}’에 직접 답하게 한 뒤 이번 주에는 ‘{primary.action}’만 실행해 보세요.",
        f"점수표보다 ‘{primary.evidence}’가 남은 시험지를 먼저 봅니다. ‘{primary.action}’을 해 본 뒤 ‘{primary.checkpoint}’에 학생이 혼자 답할 수 있는지를 같은 자료로 다시 확인하세요.",
        f"학생에게 ‘{primary.evidence}’를 직접 설명하게 하세요. 설명이 끊긴 지점을 표시한 뒤 ‘{primary.action}’을 실행하고 ‘{primary.checkpoint}’로 재확인합니다.",
        f"최근 시험지 한 장에서 ‘{primary.evidence}’를 기준선으로 남깁니다. 첫 행동은 ‘{primary.action}’, 일주일 뒤 확인 질문은 ‘{primary.checkpoint}’입니다.",
        f"먼저 ‘{primary.evidence}’가 보이는 자료 위치와 날짜를 적습니다. 이후 ‘{primary.action}’을 수행하고 ‘{primary.checkpoint}’에 학생이 혼자 답하는지 확인하세요.",
    ])
    distinction_answer = stable_pick(seed, "high-faq-answer-distinction", [
        f"‘{primary.evidence}’와 ‘{secondary.evidence}’를 다른 칸에 적고 같은 날짜에 비교하세요. 각 기록에는 다시 설명한 날짜와 도움 없이 수행했는지를 남깁니다.",
        f"{primary.label}에는 ‘{primary.evidence}’, {secondary.label}에는 ‘{secondary.evidence}’를 기록합니다. 완료일과 재확인 결과도 영역별로 나누세요.",
        f"한 표 안에서도 {primary.label}과 {secondary.label}의 문제 번호와 다음 행동을 다른 열에 둡니다. 같은 날짜 기록끼리 비교하면 원인이 선명해집니다.",
        f"첫 풀이 기록과 다시 설명한 기록을 {primary.label}·{secondary.label} 영역별로 분리하세요. 각 기록에 날짜와 도움받은 지점을 함께 적습니다.",
        f"{primary.label}과 {secondary.label}은 시험지의 표시와 재풀이 결과를 각각 이어서 봅니다. 정답 수만 한 칸에 모으지 마세요.",
        f"진단표에는 ‘{primary.evidence}’와 ‘{secondary.evidence}’를, 실행표에는 두 영역의 시작일과 확인일을 적어 역할을 나누세요.",
    ])
    exam_answer = stable_pick(seed, "high-faq-answer-exam", [
        f"내신과 모의고사 모두에서 {primary.label}·{secondary.label}을 확인하되 자료와 마감일은 나눕니다. 판단할 때는 ‘{primary.exam_use}’와 ‘{secondary.exam_use}’를 각각 대조하세요.",
        f"학교 범위 학습과 누적 학습의 달력을 분리한 뒤 두 일정에서 {primary.label}과 {secondary.label}의 최소 행동을 따로 정하세요. 기준은 ‘{primary.exam_use}’와 ‘{secondary.exam_use}’입니다.",
        f"한 영역을 내신 전용, 다른 영역을 모의고사 전용으로 고정하지 않습니다. 시험 자료에 ‘{primary.exam_use}’와 ‘{secondary.exam_use}’가 어떻게 적용되는지 보고 주간 비중을 정하세요.",
        f"내신은 범위 자료와 마감일을, 모의고사는 새 지문과 재확인일을 중심으로 기록합니다. 두 기록 모두 ‘{primary.exam_use}’와 ‘{secondary.exam_use}’를 확인 기준으로 사용하세요.",
        f"시험별 계획표는 따로 두되 {primary.label}·{secondary.label}의 현재선을 같은 주에 비교하세요. ‘{primary.exam_use}’와 ‘{secondary.exam_use}’ 중 필요한 항목만 각 자료에 적용합니다.",
        f"내신 기간에는 학교 범위의 완료 흔적을, 평소에는 모의고사 누적 기록을 남깁니다. 어느 쪽이든 ‘{primary.exam_use}’와 ‘{secondary.exam_use}’를 자료에 맞게 선택하세요.",
    ])
    material_answer = stable_pick(seed, "high-faq-answer-material", [
        f"{support.label}·{extra.label}을 확인할 수 있는 최근 자료와 학교 범위표, 현재 교재를 준비합니다. ‘{extra.consult_question}’를 묻고 두 영역의 재확인 날짜를 적으세요.",
        f"{support.label}은 ‘{support.evidence}’, {extra.label}은 ‘{extra.evidence}’가 남은 최근 자료로 확인하세요. 이어 ‘{extra.consult_question}’를 묻고 확인일을 정합니다.",
        f"상담 자료는 {support.label}·{extra.label}의 수행 흔적이 보이는 교재와 주간 계획표면 충분합니다. ‘{support.consult_question}’도 메모해 가세요.",
        f"학교 범위표와 표시가 남은 학습 자료를 한 묶음으로 준비하세요. {support.label}에서 막힌 이유와 {extra.label}의 확인 방식을 묻고 다음 점검 날짜를 남깁니다.",
        f"{support.label}의 ‘{support.evidence}’와 {extra.label}의 ‘{extra.evidence}’가 보이는 자료를 각각 한 곳씩 고르세요. 두 자료를 언제 다시 확인할지도 상담 메모에 적습니다.",
        f"{support.label}은 ‘{support.evidence}’로, {extra.label}은 ‘{extra.evidence}’로 확인합니다. 자료를 많이 가져가기보다 표시가 남은 부분과 질문을 함께 준비하세요.",
    ])
    return [
        (
            f"{locality} 고등 영어에서 ‘{profile.focus}’은 어떤 자료부터 확인하나요?",
            first_answer,
        ),
        (
            f"{primary.label}·{secondary.label} 진단 기록은 어떻게 나누나요?",
            distinction_answer,
        ),
        (
            f"내신과 모의고사에서 {primary.label}·{secondary.label} 비중은 어떻게 나누나요?",
            exam_answer,
        ),
        (
            f"상담 전 {support.label}·{extra.label}을 확인하려면 어떤 자료를 가져가나요?",
            material_answer,
        ),
        (
            f"{locality} 고등 영어 상담에서 가능 학년과 참고 학교는 어떻게 확인하나요?",
            fact_answer,
        ),
    ]


def _legacy_high_english_scenarios(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighEnglishProfile,
    seed: str,
) -> list[str]:
    locality = str(center["locality"])
    primary, secondary, support = profile.intents[:3]
    first = stable_pick(seed, "high-scenario-first", [
        f"{locality} 학부모가 최근 시험지에서 {primary.evidence}를 표시해 상담에 가져오는 가상 상황입니다. 학생은 {primary.action}을 한 번 수행하고, 학부모는 {secondary.checkpoint}를 다음 점검에서 어떤 자료로 확인하는지 질문합니다.",
        f"{profile.focus}이 필요한지 고민하는 가상의 학부모 상담입니다. 학부모는 점수표보다 {primary.evidence}와 {secondary.evidence}를 보여 주며 설명이 필요한 부분과 혼자 연습할 부분을 나눠 달라고 요청합니다.",
        f"내신 준비와 모의고사 복습이 겹친 {locality} 고등학생을 가정했습니다. 학부모는 {primary.exam_use}에 따라 이번 주 우선순위를 묻고, 다른 영역의 최소 복습 날짜도 함께 정리합니다.",
        f"과제를 끝내도 같은 오답이 반복되는 가상 상황입니다. 학부모는 {primary.concern}를 질문으로 바꾸어 적고 {support.action} 뒤 무엇이 달라져야 학습 완료로 보는지 확인합니다.",
        f"현재 교재를 바꿔야 할지 고민하는 {locality} 학부모의 가상 상담입니다. 먼저 {primary.action}을 기존 자료에서 해 본 뒤, 설명 방식·연습량·교재 난도 중 무엇을 조정할지 차례로 묻습니다.",
        f"학생이 영어에서 막힌 이유를 말하기 어려운 가상 상황입니다. 학부모는 {primary.evidence}를 보여 주고 학생이 직접 {secondary.checkpoint}를 설명할 수 있을 때까지 어떤 피드백을 받는지 질문합니다.",
    ])
    second = stable_pick(seed, "high-scenario-second", [
        f"{center['center_name']}의 제공 주소와 가능 학년을 확인한 뒤 시간표·통학·교습비를 따로 비교하는 가상 장면입니다. 학습 질문에는 {primary.consult_question}를 적고, 이용 조건과 학생별 계획을 서로 다른 목록으로 남깁니다.",
        f"상담 뒤 학생과 학부모가 4주 계획을 다시 읽는 가상 상황입니다. {secondary.action}을 실제 일정 안에서 할 수 있는지 확인하고 어렵다면 성과를 단정하지 않고 분량과 점검일을 조정합니다.",
        f"페이지의 참고 학교와 실제 재학 학교가 다른 경우를 가정했습니다. 학부모는 시험 범위표를 제시해 자료 반영 방식을 묻고, 학교명만으로 수업 가능 여부를 판단하지 않습니다.",
        f"첫 주 계획을 지키지 못한 {locality} 고등학생의 가상 상담입니다. 미완료 이유를 기록한 뒤 {support.action}만 남길지, 학습 시간을 바꿀지 학생과 함께 선택합니다.",
        f"첫 진단 뒤 정답 수는 늘었지만 근거 설명은 달라지지 않은 가상 상황입니다. 학부모는 진도를 늘리기보다 {primary.checkpoint}를 새 지문에서 다시 확인해 달라고 질문합니다.",
        f"센터 설명과 학생 자료가 맞는지 대조하는 가상 장면입니다. 학부모는 확인되지 않은 차량·주차·보강을 추정하지 않고 주소·학년·시간표를 확인한 뒤 {secondary.label}의 피드백 주기를 따로 묻습니다.",
    ])
    if locality not in first:
        first = f"{locality} 고등 영어 상담을 가정한 장면입니다. {first}"
    if locality not in second:
        second = f"{locality} 고등 영어 상담을 가정한 장면입니다. {second}"
    return [first, second]


def high_english_scenarios(
    config: CategoryConfig,
    center: dict[str, object],
    profile: HighEnglishProfile,
    seed: str,
) -> list[str]:
    locality = str(center["locality"])
    primary, secondary, support, extra = profile.intents
    grades = relevant_grades(config, center, "영어")
    grade_condition = f"영어 가능 학년({'·'.join(grades)})" if grades else "고등 영어의 현재 개설 학년"
    first = stable_pick(seed, "high-compact-scenario-learning", [
        f"{locality} 학부모가 ‘{profile.focus}’을 상담 주제로 삼은 가상 상황입니다. 학생은 {primary.label}의 첫 행동을 해 보고, 학부모는 {secondary.label}의 재확인 날짜를 묻습니다.",
        f"{locality} 학생의 내신과 모의고사 일정이 겹친 가상 상황입니다. 학부모는 ‘{profile.focus}’을 기준으로 {primary.label}을 먼저 배치하고 {secondary.label}의 최소 복습일을 따로 정합니다.",
        f"같은 오답이 반복되는 {locality} 학생의 가상 상담입니다. ‘{profile.focus}’을 상담 주제로 정하고 영어가 어렵다는 말 대신 {primary.label} 질문과 {secondary.label} 기록을 남깁니다.",
        f"교재 변경을 고민하는 {locality} 학부모의 가상 상황입니다. ‘{profile.focus}’에 맞춰 {primary.label}의 현재선을 확인하고 {secondary.label}의 연습량을 조정합니다.",
        f"시험 뒤 공부 순서를 정하지 못한 {locality} 학생을 가정했습니다. 학부모는 ‘{profile.focus}’을 중심으로 {primary.label}에서 막힌 이유와 {secondary.label}의 다음 확인일을 나누어 적습니다.",
        f"{locality} 학생이 새 교재를 시작하기 전 상담하는 가상 장면입니다. ‘{profile.focus}’에 필요한 {primary.label} 자료를 보여 주고 {secondary.label}을 언제 다시 볼지 질문합니다.",
        f"내신 범위와 누적 학습이 겹친 {locality} 가상 사례입니다. ‘{profile.focus}’을 기준으로 {primary.label}의 마감일과 {secondary.label}의 최소 행동을 다른 줄에 둡니다.",
        f"{locality} 학부모가 최근 시험지 한 장으로 상담을 준비하는 가상 상황입니다. ‘{profile.focus}’을 구체화하기 위해 {primary.label} 기록과 {secondary.label} 설명을 비교합니다.",
    ])
    second = stable_pick(seed, "high-compact-scenario-decision", [
        f"{locality}의 {center['center_name']} 상담을 준비하는 가상 장면입니다. 학부모는 {grade_condition}과 실제 시간표를 확인하고 ‘{extra.consult_question}’를 별도 질문으로 남깁니다.",
        f"상담 뒤 {locality} 학생과 학부모가 계획을 다시 읽는 가상 상황입니다. {grade_condition}·제공 주소·통학 시간을 확인한 다음 {support.label}의 첫 행동을 실제 일정과 대조합니다.",
        f"{locality}의 참고 학교와 실제 재학 학교가 다른 경우를 가정했습니다. 학부모는 시험 범위표와 {grade_condition}을 대조한 뒤 {secondary.label}의 확인 방식과 {support.label}의 점검 주기를 질문합니다.",
        f"첫 주 계획을 지키지 못한 {locality} 학생의 가상 상담입니다. 학부모는 이용 조건과 미완료 원인을 나눠 적고, 학생은 {support.label} 기록을 본 뒤 {extra.label}의 분량을 바꿀지 결정합니다.",
        f"{locality} 상담 뒤 등록 여부를 비교하는 가상 장면입니다. 학부모는 {grade_condition}·교습비·시간표를 사실 항목으로 적고 {support.label}의 재확인 날짜는 학습 계획에 따로 남깁니다.",
        f"{locality} 학생의 학교 일정이 바뀐 경우를 가정했습니다. 학부모는 시험 범위와 {grade_condition}을 다시 확인하고 {extra.label}의 최소 과제를 현재 시간표와 대조합니다.",
        f"제공 주소까지의 이동 시간을 확인하는 {locality} 가상 상담입니다. 통학·시간표와 {grade_condition}을 먼저 적은 뒤 {secondary.label}의 확인 질문은 별도 칸에 둡니다.",
        f"{locality} 학부모가 센터 답변을 학생 계획으로 옮기는 가상 상황입니다. {grade_condition}과 실제 시간표를 확인하고 {support.label}의 첫 행동과 {extra.label}의 확인일을 정합니다.",
    ])
    return [first, second]


def middle_english_particle_tokens(profile: MiddleEnglishProfile) -> tuple[str, ...]:
    values: list[str] = [profile.focus, *profile.source_markers]
    for intent in profile.intents:
        values.extend((
            intent.label,
            intent.concern,
            intent.evidence,
            intent.action,
            intent.checkpoint,
            intent.exam_use,
            intent.consult_question,
        ))
    return tuple(dict.fromkeys(value for value in values if value))


def naturalize_middle_english_text(value: str, profile: MiddleEnglishProfile) -> str:
    protected_values: list[str] = [profile.focus, *profile.source_markers]
    for intent in profile.intents:
        protected_values.extend((
            intent.label,
            intent.concern,
            intent.evidence,
            intent.action,
            intent.checkpoint,
            intent.exam_use,
            intent.consult_question,
        ))
    protected: list[tuple[str, str]] = []
    for index, protected_value in enumerate(sorted(set(protected_values), key=len, reverse=True)):
        if protected_value and protected_value in value:
            marker = f"__MIDDLE_ENGLISH_VALUE_{index}__"
            value = value.replace(protected_value, marker)
            protected.append((marker, protected_value))
    replacements = (
        ("내신과 모의고사 영어", "학교 시험과 평소 독해"),
        ("내신과 모의고사", "학교 시험과 누적 독해"),
        ("내신·모의고사", "학교 시험·누적 독해"),
        ("모의고사 시험지", "누적 독해 자료"),
        ("모의고사 오답", "새 지문 오답"),
        ("모의고사 재풀이", "새 지문 재풀이"),
        ("모의고사 재확인", "누적 독해 재확인"),
        ("모의고사 누적", "누적 독해"),
        ("모의고사 장문", "처음 보는 장문"),
        ("모의고사 어법 문항", "처음 보는 문장의 어법 문제"),
        ("모의고사 요약문", "새 지문의 요약문"),
        ("모의고사 듣기", "새 음원 듣기"),
        ("모의고사의", "처음 보는 지문의"),
        ("모의고사에서", "처음 보는 지문에서"),
        ("모의고사에는", "누적 독해에는"),
        ("모의고사에", "누적 독해에"),
        ("모의고사는", "누적 독해는"),
        ("모의고사를", "누적 독해를"),
        ("모의고사와", "누적 독해와"),
        ("모의고사", "누적 독해"),
        ("고등 영어", "중등 영어"),
        ("고등학생", "중학생"),
        ("고등학교", "중학교"),
    )
    for before, after in replacements:
        value = value.replace(before, after)
    value = value.replace("학교 시험과 누적 독해 자료를", "학교 시험 자료와 누적 독해 기록을")
    value = value.replace("학교 시험과 누적 독해 자료에서", "학교 시험 자료와 누적 독해 기록에서")
    value = value.replace("누적 독해 새 지문", "처음 보는 지문")
    value = value.replace("평소 독해 누적 학습", "평소 누적 독해")
    value = value.replace("학교 시험·누적 독해 학습", "학교 시험과 누적 독해 학습")
    for marker, protected_value in protected:
        value = value.replace(marker, protected_value)
    for left in profile.intents:
        for right in profile.intents:
            if left is right:
                continue
            joined = f"{left.label}·{right.label}"
            natural_join = f"{left.label}{particle_for(left.label, '과', '와')} {right.label}"
            value = value.replace(joined, natural_join)
    value = value.replace("평소에는 평소 독해 누적 기록", "평소에는 누적 독해 기록")
    value = value.replace(
        "을 실제로 점검할 수 있는 순서로 정리합니다.",
        "을 구체적인 자료로 확인할 수 있는 순서로 정리합니다.",
    )
    value = value.replace("누적 학습에 학습 우선순위", "누적 독해에 학습 우선순위")
    value = value.replace("근거와 근거 독해", "확인 자료와 근거 독해")
    value = value.replace(
        "답은 성적표가 아니라 최근 시험지에서 확인합니다.",
        "답은 성적표가 아니라 최근 영어 학습 자료에서 확인합니다.",
    )
    value = value.replace(
        "다음 점검에서도 같은 시험지를 사용하되",
        "다음 점검에서도 같은 유형의 학습 자료를 사용하되",
    )
    value = re.sub(
        r"([은는]) 시험지의 표시와 재풀이 결과를 각각 이어서 봅니다\.",
        r"\1 영역별 학습 기록과 재확인 결과를 따로 이어서 봅니다.",
        value,
    )
    for left in profile.intents:
        for right in profile.intents:
            if left is right:
                continue
            connector = particle_for(left.label, "과", "와")
            value = value.replace(
                f"센터 정보와 {left.label}{connector} {right.label} 질문",
                f"센터 정보, {left.label}{connector} {right.label} 질문",
            )
            value = value.replace(
                f"{left.label}{connector} {right.label} 과정의 운영 여부를 단정할 수는 없습니다.",
                "두 학습 영역이 실제 수업에 어떻게 반영되는지는 알 수 없습니다.",
            )
            value = value.replace(
                f"{left.label}에 설명이 필요한지, {right.label}{particle_for(right.label, '을', '를')} 혼자 연습할지",
                f"{left.label}{particle_for(left.label, '은', '는')} 설명이 필요한 영역인지, "
                f"{right.label}{particle_for(right.label, '은', '는')} 혼자 연습할 영역인지",
            )
            value = value.replace(
                f"{left.label}{particle_for(left.label, '과', '와')} {right.label}의 우선순위를",
                f"{left.label}{particle_for(left.label, '과', '와')} {right.label}의 실행 순서를",
            )
        value = value.replace(
            f"{left.label}의 실행 기록",
            f"{left.label}{particle_for(left.label, '을', '를')} 확인한 기록",
        )
        value = value.replace(
            f"{left.label}의 실제 개설 여부",
            f"{left.label}의 실제 학습 적용 방식",
        )
        value = value.replace(
            f"{left.label}의 실제 학습 적용 방식와 시간표",
            f"{left.label}{particle_for(left.label, '을', '를')} 실제 학습에 적용하는 방식과 시간표",
        )
        value = value.replace(
            f"{left.label}의 최소 행동",
            f"{left.label}{particle_for(left.label, '을', '를')} 확인하는 최소 행동",
        )
    primary, secondary = profile.intents[:2]
    value = value.replace(
        f"참고 학교와 {primary.label} 개설 범위를 대조하는 방법",
        f"{primary.label}{particle_for(primary.label, '과', '와')} {secondary.label} 적용 전 확인할 학교·학년 정보",
    )
    value = re.sub(
        r"([가-힣· ]+)의 최소 과제를 현재 시간표와 대조합니다\.",
        r"\1의 최소 행동을 현재 시간표에서 실행할 수 있는지 확인합니다.",
        value,
    )
    if profile.intents[0].code == "routine":
        value = re.sub(
            r"(‘[^’]+’)에 학생이 직접 답하는지를 새 문제 기록과 비교하세요\.",
            r"\1를 주간 계획표와 완료 기록에서 다시 확인하세요.",
            value,
        )
    value = re.sub(
        r"(‘[^’]+’)에 학생이 직접 답하는지를 새 문제 기록과 비교하세요\.",
        r"\1에 대한 학생의 답변을 새 문제 풀이 기록과 대조하세요.",
        value,
    )
    value = re.sub(
        r"점수표보다 ‘([^’]+)’[이가] 남은 시험지를 먼저 봅니다\.",
        r"점수표보다 최근 학습 자료 묶음을 먼저 보고 ‘\1’에 해당하는 근거를 확인합니다.",
        value,
    )
    value = re.sub(
        r"최근 시험지 한 장에서 ‘([^’]+)’를 기준선으로 남깁니다\.",
        r"최근 학습 자료에서 기준선으로 삼을 근거는 ‘\1’입니다.",
        value,
    )
    value = re.sub(
        r"최근 시험지와 교재에서 ‘([^’]+)’를 먼저 찾고",
        r"‘\1’가 남아 있는 자료를 먼저 보고",
        value,
    )
    value = re.sub(
        r"한 표 안에서도 [^.!?]+의 문제 번호와 다음 행동을 다른 열에 둡니다\.",
        "한 표 안에서도 두 영역에 해당하는 자료 위치나 문제 번호와 다음 행동을 각각 다른 열에 둡니다.",
        value,
    )
    value = value.replace("과정를", "과정을").replace("점검를", "점검을")
    value = re.sub(r"(?<![가-힣])([가-힣]{2,12})\s+\1(?![가-힣])", r"\1", value)
    value = re.sub(r"\s+", " ", value).strip()
    return normalize_particle_joins(value, middle_english_particle_tokens(profile))


def naturalize_middle_english_tree(value: object, profile: MiddleEnglishProfile) -> object:
    if isinstance(value, str):
        return naturalize_middle_english_text(value, profile)
    if isinstance(value, list):
        return [naturalize_middle_english_tree(item, profile) for item in value]
    if isinstance(value, tuple):
        return tuple(naturalize_middle_english_tree(item, profile) for item in value)
    if isinstance(value, dict):
        return {key: naturalize_middle_english_tree(item, profile) for key, item in value.items()}
    return value


def middle_english_meta_description(
    config: CategoryConfig,
    center: dict[str, object],
    profile: MiddleEnglishProfile,
) -> str:
    title = f"{center['locality']} {config.label}"
    primary, secondary = profile.intents[:2]
    intent_pair = f"{primary.label}{particle_for(primary.label, '과', '와')} {secondary.label}"
    if profile.focus.endswith("학습"):
        intent_pair = intent_pair.replace("학습 우선순위", "우선순위")
    if str(center["locality"]) in {"동춘동", "화봉동"}:
        focused = (
            f"{title}: {profile.focus}. {intent_pair} 확인 순서와 "
            "센터·가능 학년 기준을 안내합니다."
        )
        if 70 <= len(focused) <= 160:
            return focused
        raise ValueError(f"middle English focused meta description invalid: {title} / {profile.focus}")
    candidates = (
        f"{title} 선택 전 {profile.focus}, {intent_pair} 진단, 학교 시험과 누적 독해 순서, 센터·가능 학년 확인 기준을 살펴보세요.",
        f"{title}의 {profile.focus}, {intent_pair}{particle_for(intent_pair, '을', '를')} 확인하는 기준과 학교 시험 준비 순서, 센터·학년 정보를 정리했습니다.",
        f"{title} 상담 전 {intent_pair} 자료와 학교 시험 학습 순서, 확인된 센터·가능 학년 정보를 살펴보세요.",
        f"{title}에서 {intent_pair}{particle_for(intent_pair, '을', '를')} 점검하는 법과 학교 시험 준비, 센터·가능 학년 확인 기준을 안내합니다.",
    )
    for value in candidates:
        value = clean(value)
        if 70 <= len(value) <= 100:
            return value
    raise ValueError(f"middle English meta description invalid: {title} / {profile.focus}")


def middle_english_student_type(
    config: CategoryConfig,
    center: dict[str, object],
    profile: MiddleEnglishProfile,
    seed: str,
) -> str:
    locality = str(center["locality"])
    primary, secondary = profile.intents[:2]
    return stable_pick(seed, "middle-student", [
        f"{locality}에서 최근 시험지의 {primary.label} 문제와 {secondary.label} 문제를 나누어 봐야 하는 중학생",
        f"점수만으로 원인을 단정하지 않고 {primary.label} 기록과 {secondary.label} 수행 과정을 비교해 다음 공부 순서를 정해야 하는 {locality} 중학생",
        f"학교 범위 학습과 평소 누적 독해 사이에서 {primary.label}{particle_for(primary.label, '과', '와')} {secondary.label} 중 먼저 점검할 영역과 순서를 실제 자료로 판단해야 하는 {locality} 중학생",
        f"현재 교재의 진도보다 {primary.label}의 근거와 {secondary.label}의 재확인 행동을 먼저 정리해야 하는 {locality} 중학생",
        f"최근 영어 시험에서 막힌 위치를 {primary.label}과 {secondary.label}로 구분하고 일주일 뒤의 확인 기준까지 세워야 하는 {locality} 중학생",
        f"학교 일정 안에서 {profile.source_markers[0]}와 {profile.source_markers[1]}을 무리 없이 이어 갈 학습 순서가 필요한 {locality} 중학생",
    ])


def middle_english_quick_answer(
    config: CategoryConfig,
    center: dict[str, object],
    profile: MiddleEnglishProfile,
    seed: str,
) -> str:
    return naturalize_middle_english_text(
        high_english_quick_answer(config, center, profile, seed),  # type: ignore[arg-type]
        profile,
    )


def middle_english_focus_guidance(
    profile: MiddleEnglishProfile,
    locality: str,
) -> tuple[str, str]:
    """Turn a source-derived page focus into two concrete, reader-facing checks."""

    focus = profile.focus
    compact = re.sub(r"\s+", "", focus)
    primary = profile.intents[0]
    if "봄방학" in compact:
        return (
            f"{locality}의 ‘{focus}’은 개학 전 남은 날짜를 먼저 세고, 첫날에는 어휘·문장 구조의 기준선을 기록하는 데서 시작합니다.",
            f"‘{focus}’의 중간 점검에는 같은 난도의 새 문장을 쓰고, 개학 전날에는 미완료 분량보다 혼자 설명하지 못한 항목만 다음 계획으로 넘기세요.",
        )
    if "여름방학" in compact:
        return (
            f"{locality}의 ‘{focus}’은 방학 시작 자료와 개학 직전 자료를 비교할 수 있도록 같은 형식의 어휘·문장·독해 기록을 남기는 것이 핵심입니다.",
            f"‘{focus}’은 시작·중간·마무리로 나누고, 매 구간 끝에는 새 지문 한 편에서 틀린 이유와 다시 읽은 근거가 달라졌는지 확인하세요.",
        )
    if "겨울방학" in compact:
        return (
            f"{locality}의 ‘{focus}’은 현재 학년의 빈틈과 다음 학년 준비를 한 계획표에 섞지 않고 두 칸으로 나누는 데서 출발합니다.",
            f"‘{focus}’의 전반에는 현재 교재의 반복 오류를 정리하고, 후반에는 다음 난도의 짧은 지문에 적용해 혼자 읽을 수 있는 범위를 확인하세요.",
        )
    if "방학" in compact:
        return (
            f"{locality}의 ‘{focus}’은 방학 첫날의 기준선, 중간 점검일, 개학 전 확인일을 먼저 정한 뒤 각 날짜에 같은 형식의 자료를 남겨야 판단할 수 있습니다.",
            f"‘{focus}’의 다음 점검 전에는 어휘·문장 구조·독해 중 한 병목만 다루고, 새 문제에서 같은 실수가 줄었을 때 다음 영역으로 넘어가세요.",
        )
    if any(word in compact for word in ("예비고", "고등", "전환")):
        return (
            f"{locality}의 ‘{focus}’은 중등 지문과 고등 수준의 짧은 예시 지문을 나란히 놓고 문장 길이·수식 범위·문단 연결에서 먼저 멈추는 지점을 찾는 과정입니다.",
            f"‘{focus}’의 재확인에서는 풀이 속도보다 낯선 지문에 표시한 주어·동사와 핵심 문장을 자기 말로 설명하는 범위가 넓어졌는지 보세요.",
        )
    if "기출" in compact:
        return (
            f"{locality}의 ‘{focus}’은 개념 이름을 외우는 데서 끝내지 않고, 기출 문항의 어느 문장·선택지에 그 개념이 쓰였는지 연결표에 적는 방식으로 확인합니다.",
            f"‘{focus}’의 다음 점검에서는 같은 개념이 표현만 바뀐 새 문항에도 적용되는지 다시 풀고, 적용 근거와 잘못 연결한 지점을 기록하세요.",
        )
    if any(word in compact for word in ("불안", "자신감", "부담")):
        return (
            f"{locality}의 ‘{focus}’은 틀린 개수만 세지 않고 멈춘 문항·멈춘 시각·다시 시작한 행동을 함께 기록해 지식 부족과 긴장 반응을 나누어 보는 과정입니다.",
            f"‘{focus}’의 재확인에서는 짧은 제한 시간 세트의 정답률과 회복 시간을 함께 비교하고, 설명 가능한 문제부터 분량을 조정하세요.",
        )
    if any(word in compact for word in ("내신", "시험", "중간고사", "기말고사", "서술형", "수행")):
        return (
            f"{locality}의 ‘{focus}’은 학교 범위표를 기준으로 교과서 본문·어법 변형·서술형 조건을 나누고 각 칸의 완료 흔적을 확인하는 방식으로 구체화합니다.",
            f"‘{focus}’의 시험 뒤 기록은 범위 미완료, 문장 해석, 조건 누락으로 분류하고 같은 조건의 새 문항에서 수정 이유를 다시 설명하게 하세요.",
        )
    if any(word in compact for word in ("문법", "어법", "영작")):
        return (
            f"{locality}의 ‘{focus}’은 규칙 한 줄, 맞는 예문, 틀린 예문을 나란히 두고 형태가 달라질 때 무엇이 바뀌는지 설명하는 기록으로 점검합니다.",
            f"‘{focus}’의 영작·서술형 재확인에서는 처음 문장과 수정 문장의 차이를 표시해 다음 문장에서도 같은 규칙을 적용하는지 보세요.",
        )
    if any(word in compact for word in ("독해", "지문", "읽기", "근거")):
        return (
            f"{locality}의 ‘{focus}’은 문단마다 핵심 문장과 연결어를 표시하고, 선택지를 고르거나 지운 근거를 지문 문장에 연결하는 기록으로 확인합니다.",
            f"‘{focus}’을 다시 확인할 때는 처음 보는 지문에서도 문단 역할과 정답·오답 근거를 각각 짚는지를 첫 기록과 비교하세요.",
        )
    if any(word in compact for word in ("어휘", "단어")):
        return (
            f"{locality}의 ‘{focus}’은 단어 뜻만 적지 않고 품사·문장 속 의미·함께 쓰인 표현을 묶어 기록한 뒤 다른 지문에서 다시 찾는 방식으로 확인합니다.",
            f"‘{focus}’의 다음 날 예문과 일주일 뒤 새 지문 기록을 비교해 같은 단어의 문맥 근거를 말할 수 있는지 확인하세요.",
        )
    if any(word in compact for word in ("오답", "틀린", "실수", "성적기복")):
        return (
            f"{locality}의 ‘{focus}’은 오답을 지식 부족·문장 해석·선택지 판단·시간 문제로 나누고 처음 선택한 근거까지 남겨야 원인을 구분할 수 있습니다.",
            f"‘{focus}’은 며칠 뒤 같은 유형의 새 문제를 다시 풀어 판단 순서의 변화를 기록하고, 반복된 원인 한 가지를 다음 계획에 반영하세요.",
        )
    if any(word in compact for word in ("시간", "속도", "집중")):
        return (
            f"{locality}의 ‘{focus}’은 문항별 읽기·판단·검토 시간을 나누어 적고 어느 문장에서 다시 읽기가 반복됐는지 찾는 방식으로 점검합니다.",
            f"‘{focus}’의 다음 점검에서는 짧은 세트를 같은 정확도로 다시 풀어 병목 구간의 시간이 줄었을 때만 전체 분량을 조정하세요.",
        )
    if any(word in compact for word in ("듣기", "말하기", "발음")):
        return (
            f"{locality}의 ‘{focus}’은 놓친 음원 구간을 받아쓰고, 소리를 못 들은 경우와 표현 뜻을 몰랐던 경우를 다른 표시로 남겨 점검합니다.",
            f"‘{focus}’의 재확인에서는 짧은 구간을 따라 말한 뒤 새 음원에서 같은 연결 발음과 핵심 표현을 알아듣는지 기록하세요.",
        )
    if any(word in compact for word in ("습관", "루틴", "꾸준", "자기주도", "계획")):
        return (
            f"{locality}의 ‘{focus}’은 계획 분량보다 시작·완료 시각, 남은 질문, 다시 볼 날짜를 한 줄씩 남겨 실제 실행 여부를 확인하는 데서 시작합니다.",
            f"‘{focus}’의 한 주 뒤 기록에서는 미완료 이유와 반복 오답을 함께 보고, 혼자 끝낼 최소 분량과 도움을 요청할 시점을 다시 정하세요.",
        )
    return (
        f"{locality}의 ‘{focus}’은 ‘{primary.evidence}’에서 확인할 위치를 표시하고 학습 전 기준선을 남기는 방식으로 시작합니다.",
        f"‘{focus}’의 다음 점검에서는 ‘{primary.action}’ 뒤 ‘{primary.checkpoint}’에 학생이 직접 답하는지를 새 문제 기록과 비교하세요.",
    )


def middle_english_sections(
    config: CategoryConfig,
    center: dict[str, object],
    profile: MiddleEnglishProfile,
    seed: str,
) -> list[dict[str, object]]:
    sections = naturalize_middle_english_tree(
        high_english_sections(config, center, profile, seed),  # type: ignore[arg-type]
        profile,
    )
    locality = str(center["locality"])
    primary, secondary, support, _extra = profile.intents
    focus_start, focus_check = middle_english_focus_guidance(profile, locality)
    fact_heading_options = [
        f"{locality} 학교·학년 정보와 {primary.label} 상담 질문을 나누는 법",
        f"{locality} 참고 학교·가능 학년을 {primary.label} 질문과 구분해 확인하는 순서",
        f"{locality} 센터 이용 조건과 {primary.label} 질문을 따로 적는 방법",
        f"{primary.label} 질문과 분리해 확인할 {locality} 학교·학년 정보",
        f"{locality} 상담 전에 확인할 학교·학년 정보와 {primary.label} 질문",
        f"{locality} 참고 학교와 가능 학년, {primary.label} 질문을 구분하는 기준",
        f"{locality} 학년·학교 정보와 {primary.label} 자료를 따로 확인하는 이유",
        f"{locality} 학교 정보 확인 뒤 {primary.label} 질문을 정리하는 순서",
        f"{primary.label}{particle_for(primary.label, '과', '와')} {secondary.label} 질문 전 확인할 {locality} 학교·학년 정보",
        f"{locality} 센터 정보 확인과 {primary.label} 상담의 구분",
        f"{locality} 학교·학년·교습비 정보와 {primary.label} 질문을 나누는 방법",
        f"{locality} 이용 조건을 확인한 뒤 {primary.label} 질문을 비교하는 방법",
    ]
    if primary.code == "transition":
        fact_heading_options.extend((
            f"{locality} 예비 고1 상담에서 학교·학년 정보와 영어 기초 질문 구분",
            f"{locality} {primary.label} 질문과 분리해 확인할 학교·학년 정보",
        ))
    elif primary.code == "entry":
        fact_heading_options.extend((
            f"{locality} 중1 첫 시험 상담에서 학교·학년 정보와 기초 질문 구분",
            f"{locality} {primary.label} 질문과 분리해 확인할 학교·학년 정보",
        ))
    fact_heading = stable_pick(seed, "middle-heading-facts-v2", fact_heading_options)
    procedure_notes = {
        "direct-answer": (
            f"{focus_start} 첫 기록에서 기준선을 표시하고 다음 점검일을 정하세요."
        ),
        "four-week-plan": (
            focus_check
        ),
    }
    for section in sections:
        if section["key"] == "school-center-facts":
            section["heading"] = fact_heading
        note = procedure_notes.get(str(section["key"]))
        if note:
            section["paragraphs"][-1] = clean(f"{section['paragraphs'][-1]} {note}")
        if section["key"] == "consultation-checklist":
            checklist = list(section["checklist"])
            checklist[0] = ("학생 자료", primary.evidence)
            schedule_label = {
                "diagnosis": "학습 우선순위 결정일·재확인일",
                "routine": "복습·과제 실행 완료일·재확인일",
                "error": "오답 원인 확인일·재확인일",
                "transition": "고등 영어 전환 준비 완료일·재확인일",
                "entry": "중1 첫 시험 적응 준비 완료일·재확인일",
            }.get(
                secondary.code,
                f"{secondary.label} 연습 완료일·재확인일",
            )
            checklist[1] = (
                "시험 계획",
                f"학교 범위표와 {schedule_label}",
            )
            section["checklist"] = checklist
    # The strict corpus review found one otherwise-valid pair with three fully
    # identical instructional paragraphs. Keep the same fact boundaries while
    # giving the Siheung-dong page its own evidence and scheduling procedure.
    if locality == "시흥동":
        for section in sections:
            if section["key"] == "diagnostic-evidence":
                section["paragraphs"][1] = (
                    f"시흥동 상담을 준비할 때는 ‘{primary.evidence}’와 ‘{secondary.evidence}’를 다른 줄에 적으세요. "
                    f"일주일 뒤에는 ‘{primary.checkpoint}’와 ‘{secondary.checkpoint}’를 각각 확인해 어떤 행동이 달라졌는지 비교합니다."
                )
            elif section["key"] == "exam-strategy":
                section["paragraphs"][0] = (
                    f"학교 시험 범위에는 ‘{primary.action}’을 먼저 배치하고, 누적 독해 시간에는 ‘{secondary.action}’을 짧게 이어 가세요. "
                    "두 기록을 분리하면 시험 준비 때문에 평소 영어 학습이 멈췄는지 확인할 수 있습니다."
                )
                section["paragraphs"][1] = (
                    f"시흥동 학생의 주간표에는 학교 범위 마감일과 누적 독해 재확인일을 따로 적고, "
                    f"각 날짜에 {primary.label}과 {secondary.label}의 최소 행동을 하나씩 배치하세요. 완료하지 못한 항목은 이유를 기록한 뒤 다음 주 분량을 조정합니다."
                )
    return naturalize_middle_english_tree(sections, profile)  # type: ignore[return-value]


def middle_english_faq(
    config: CategoryConfig,
    center: dict[str, object],
    profile: MiddleEnglishProfile,
    seed: str,
) -> list[tuple[str, str]]:
    return naturalize_middle_english_tree(
        high_english_faq(config, center, profile, seed),  # type: ignore[arg-type]
        profile,
    )  # type: ignore[return-value]


def middle_english_scenarios(
    config: CategoryConfig,
    center: dict[str, object],
    profile: MiddleEnglishProfile,
    seed: str,
) -> list[str]:
    scenarios = naturalize_middle_english_tree(
        high_english_scenarios(config, center, profile, seed),  # type: ignore[arg-type]
        profile,
    )
    if not relevant_schools(config, center):
        locality = re.escape(str(center["locality"]))
        scenarios = [
            re.sub(
                rf"(?:{locality}의|페이지의) 참고 학교와 실제 재학 학교가 다른 경우를 가정했습니다\.",
                f"{center['locality']} 페이지에 참고 학교명이 없는 경우를 가정했습니다.",
                value,
            )
            for value in scenarios
        ]
    return scenarios  # type: ignore[return-value]


def answer_cards(config: CategoryConfig, center: dict[str, object], signals: tuple[TopicSignal, ...], student: str) -> list[tuple[str, str]]:
    locality = str(center["locality"])
    service_region = display_region_label(center)
    address_region = physical_region(str(center["address"]), str(center["region"]))
    region_answer = (
        f"{service_region} 지역 자료와 {address_region} 소재 제공 센터 주소를 각각 확인합니다."
        if not service_region.startswith(address_region)
        else f"{service_region}의 제공 센터·주소 자료를 확인합니다."
    )
    grade_items = grade_summary(config, center)
    listed_items = [(subject, value) for subject, value in grade_items if value != "상담 시 확인"]
    missing_subjects = [subject for subject, value in grade_items if value == "상담 시 확인"]
    if missing_subjects:
        if listed_items:
            listed_text = " / ".join(f"{subject} {value}" for subject, value in listed_items)
            grade_answer = (
                f"공개 자료에는 {listed_text} 범위가 기재되어 있고, {'·'.join(missing_subjects)} 가능 학년은 상담에서 확인합니다. "
                "실제 시간표도 함께 확인합니다."
            )
        else:
            grade_answer = (
                f"공개 자료에는 {'·'.join(missing_subjects)} 가능 학년이 기재되지 않아 상담에서 확인해야 합니다. "
                "실제 시간표도 함께 확인합니다."
            )
    else:
        grade_text = " / ".join(f"{subject} {value}" for subject, value in grade_items)
        grade_answer = f"공개된 가능 학년은 {grade_text}이며 실제 시간표는 상담에서 확인합니다."
    return [
        ("01 / 지역", region_answer),
        ("02 / 학년", grade_answer),
        ("03 / 추천 학생", student),
    ]


def build_sections(config: CategoryConfig, center: dict[str, object], signals: tuple[TopicSignal, ...], seed: str) -> list[dict[str, object]]:
    locality = str(center["locality"])
    title = f"{locality} {config.label}"
    primary, secondary, support = signals[:3]
    extra = signals[3] if len(signals) > 3 else primary
    schools = relevant_schools(config, center)
    material = material_label(config, schools)
    school_text = "·".join(schools[:3]) if schools else "실제 재학 학교"
    school_materials = f"{school_text} 자료"
    school_reference = (
        f"{title} 페이지에 표시된 {school_text}은 상담 준비를 위한 참고 정보입니다. 학교명만으로 수업 가능 여부를 판단하지 말고 학생의 {material}를 센터의 현재 개설 범위와 함께 확인해야 합니다."
        if schools else
        f"{title} 페이지에는 참고 학교 목록이 제공되지 않았습니다. 실제 재학 학교와 {material}를 상담에서 알려 주고 센터의 현재 개설 범위와 함께 확인해야 합니다."
    )
    subject_grade_pairs = [
        (subject, relevant_grades(config, center, subject))
        for subject in config.subjects
    ]
    grades = ", ".join(
        f"{subject} {'·'.join(subject_grades) or '상담 시 확인'}"
        for subject, subject_grades in subject_grade_pairs
    )
    has_listed_grades = any(subject_grades for _, subject_grades in subject_grade_pairs)
    all_grades_listed = all(subject_grades for _, subject_grades in subject_grade_pairs)
    subject_text = "·".join(config.subjects)
    student_level = student_level_label(config)
    student = student_type(config, signals, seed, locality)

    diagnosis = {
        "key": "diagnosis",
        "heading": stable_pick(seed, "h2-diagnosis", [
            f"{title} 진단은 어디서 시작할까",
            f"{locality} {config.level} {subject_text} 학습의 출발점을 찾는 방법",
            f"{title} 상담 전에 확인할 첫 근거",
            f"{locality}에서 {config.label}을 비교하는 진단 기준",
        ]),
        "paragraphs": [
            stable_pick(seed, "diagnosis-p1", [
                f"{title}의 첫 진단에서는 {primary.label}과 {secondary.label}을 한 가지 문제로 묶지 않습니다. {primary.check}와 {secondary.check}를 각각 확인해야 보완 순서가 선명해집니다.",
                f"{locality} {student_level}의 현재 상태는 점수 한 줄보다 풀이와 읽기 과정에서 더 잘 드러납니다. {primary.evidence}와 {secondary.evidence}를 나란히 보면 반복되는 어려움을 구분할 수 있습니다.",
                f"{title} 상담은 학생이 막힌 장면을 구체적으로 설명하는 데서 시작합니다. {primary.label}은 현재 자료에서 확인하고 {secondary.label}은 다시 수행한 기록으로 비교하는 편이 좋습니다.",
                f"{locality}에서 {config.label}을 알아볼 때 진도표만 보면 시작점을 놓칠 수 있습니다. 먼저 {primary.label}이 드러난 자료와 {secondary.label} 실행 여부를 별도의 항목으로 표시하세요.",
                f"{title}의 진단 자료는 많을 필요가 없습니다. 최근 기록에서 {primary.label}과 {secondary.label}이 드러난 부분을 골라 오면 상담 질문을 구체화할 수 있습니다.",
                f"{locality} {config.level} 과정의 우선순위는 잘한 단원보다 멈춘 과정에서 찾습니다. {primary.check}를 살핀 뒤 {secondary.check}를 이어 확인하는 순서가 알맞습니다.",
            ]),
            stable_pick(seed, "diagnosis-p2", [
                f"첫 자료에서는 ‘{primary.evidence}’부터 살펴보고, 별도 기록에서는 ‘{support.evidence}’도 확인하세요. 두 자료의 차이가 {title} 상담에서 첫 학습 행동을 정하는 근거가 됩니다.",
                f"학생에게 답을 다시 외우게 하기보다 ‘{primary.practice}’와 ‘{secondary.practice}’ 두 활동을 짧게 수행하게 해 보세요. 결과가 아니라 과정의 변화를 기록해야 다음 점검 기준을 정할 수 있습니다.",
                f"{material}에서 어려운 부분을 한두 곳만 골라도 충분합니다. 상담에서는 {primary.label} 관련 안내와 {support.label}을 점검할 연습량을 구분해 질문하세요.",
                f"같은 오답이라도 {primary.label}의 문제인지 {secondary.label}의 문제인지에 따라 다음 과제가 달라집니다. {locality} 페이지의 센터 정보와 학생 기록을 함께 놓고 판단하세요.",
                f"진단 뒤에는 ‘{support.home_action}’를 첫 점검 행동으로 정할 수 있습니다. 이 행동이 현재 통학과 과제 일정 안에서 가능한지도 {center['center_name']}에 확인하세요.",
                f"학부모는 학생 대신 원인을 단정하지 말고 {primary.check}를 질문으로 바꾸어 적어 두는 편이 좋습니다. 학생의 설명과 실제 자료가 일치하는지를 상담에서 비교할 수 있습니다.",
            ]),
        ],
    }

    method = {
        "key": "method",
        "heading": stable_pick(seed, "h2-method", [
            f"{primary.label}과 {secondary.label}을 나누는 {locality} {config.level} {subject_text} 기준",
            f"{title}에서 과목별 병목을 구분하는 방법",
            f"{locality} {student_level}의 {subject_text} 학습 과정을 확인하는 순서",
            f"{title} 학습 기록을 다음 수업으로 연결하는 법",
        ]),
        "paragraphs": [
            stable_pick(seed, "method-p1", [
                f"{title} 학습에서는 {primary.label}과 {secondary.label}을 같은 분량으로 다루기보다 막힌 원인에 맞춰 시간을 나누어야 합니다. 먼저 ‘{primary.practice}’를 실행하고 기록이 남으면 ‘{secondary.practice}’로 이어 갑니다.",
                f"{locality} {student_level}에게 필요한 연습은 문제 수보다 확인 가능한 과정입니다. {primary.evidence}를 살핀 뒤 {secondary.evidence}를 비교하면 설명과 반복 연습의 비중을 정할 수 있습니다.",
                f"{title}의 수업 흐름은 진단, 짧은 연습, 재확인의 순서가 분명해야 합니다. {primary.label}을 점검할 때는 ‘{primary.practice}’를, {secondary.label}을 점검할 때는 ‘{secondary.practice}’를 활용합니다.",
                f"{locality}에서 학습 계획을 세울 때는 학생이 이미 아는 내용과 혼자 적용하지 못하는 내용을 나눕니다. {primary.check}를 확인한 다음 {secondary.check}를 별도로 살펴보세요.",
                f"{title} 상담에서 한 번에 많은 과제를 약속할 필요는 없습니다. {primary.label}과 {secondary.label} 중 먼저 바꿀 행동 하나를 고르고 다음 점검일에 실제 기록을 확인하면 됩니다.",
                f"과목 학습을 함께 관리하더라도 {locality} 학생의 병목은 서로 다를 수 있습니다. {primary.evidence}와 {secondary.evidence}를 바탕으로 설명 시간과 독립 연습 시간을 구분하세요.",
            ]),
            stable_pick(seed, "method-p2", [
                f"가정에서는 {primary.home_action}와 {secondary.home_action} 가운데 하나만 정해 실행해도 좋습니다. {title} 상담에서 완료 여부와 어려웠던 부분을 공유하면 다음 과제량을 조정하기 쉽습니다.",
                f"피드백은 정답 수보다 학생이 이유를 설명하고 다시 수행할 수 있는지에 초점을 맞춥니다. {support.label} 기록이 남아야 수업 뒤 복습이 이어졌는지 확인할 수 있습니다.",
                f"첫 주에는 {primary.label}, 다음 점검에서는 {secondary.label}을 살피는 식으로 순서를 정할 수 있습니다. 다만 이 흐름은 학생 자료와 실제 가능 일정에 따라 달라지는 조건부 예시입니다.",
                f"교재를 바꾸기 전에는 현재 자료에서 {support.check}를 먼저 확인하세요. 자료가 어려운 것인지 학습 절차가 빠진 것인지 구분한 뒤 교재와 분량을 결정해야 합니다.",
                f"{center['center_name']}에 문의할 때는 설명 방식, 독립 연습과 피드백 주기를 함께 물어보세요. 실제 수업 시간과 반 편성은 센터의 현재 운영 범위에 따라 달라질 수 있습니다.",
                f"학생이 혼자 다시 해 본 흔적을 남기면 같은 설명을 반복하는 시간을 줄일 수 있습니다. ‘{support.practice}’ 활동을 다음 수업 전에 해낼 수 있는지 현실적으로 확인하세요.",
            ]),
        ],
    }

    school = {
        "key": "school",
        "heading": stable_pick(seed, "h2-school", [
            f"{school_materials}로 살펴보는 {title} 학습 범위",
            f"{locality} {config.school_name} {subject_text} 자료와 현재 교재를 연결하는 법",
            f"{title}에서 학교 자료를 상담 근거로 쓰는 방법",
            f"{locality} 학생의 실제 진도와 {config.label} 계획 비교",
        ]),
        "paragraphs": [
            stable_pick(seed, "school-p1", [
                school_reference,
                f"{locality} 학생의 학교 진도는 같은 학년 안에서도 다를 수 있습니다. {material}에서 {primary.label}과 {secondary.label}이 드러난 부분을 표시해 실제 수업 계획과 대조하세요.",
                f"{school_materials}를 준비할 때는 이름을 나열하기보다 현재 단원과 평가 일정을 확인하는 편이 좋습니다. {title} 상담에서는 그 범위가 주간 과제에 어떻게 반영되는지를 물어보세요.",
                f"학교 자료가 없는 경우에도 {locality} 학생의 현재 교재와 학습 기록으로 시작할 수 있습니다. 실제 학교명과 진도는 상담에서 확인하고 임의로 추정하지 않습니다.",
                f"{title}의 학교 정보는 광고 문구가 아니라 상담 자료를 고르는 기준으로 사용합니다. {school_text}의 학습 자료와 학생이 푼 교재를 함께 보면 보완할 범위를 구체화할 수 있습니다.",
                f"{locality} 학부모는 {material} 가운데 학생이 설명하지 못한 부분을 표시해 가져가면 좋습니다. 학교 자료와 센터 운영 범위가 일치하는지는 등록 전에 별도로 확인해야 합니다.",
            ]),
            stable_pick(seed, "school-p2", [
                (
                    f"공개 자료에 적힌 가능 학년은 {grades}입니다. 실제 시간표가 필요한 경우에는 {center['center_name']}에서 과목·학년·요일을 다시 확인하세요."
                    if all_grades_listed else
                    (
                        f"공개 자료에는 일부 과목의 가능 학년이 기재되지 않았습니다. {center['center_name']}에서 희망 과목·학년·요일과 실제 개설 범위를 확인하세요."
                        if len(config.subjects) > 1 and has_listed_grades else
                        f"제공된 센터 자료만으로는 {subject_text} 가능 학년을 확인할 수 없습니다. {center['center_name']}에서 희망 학년·요일과 실제 개설 범위를 확인하세요."
                    )
                ),
                f"제공된 센터 기준 주소는 {center['address']}입니다. {locality}에서의 실제 통학 시간과 수업 시작 시각은 학생 일정에 맞춰 직접 확인해야 합니다.",
                f"{center['center_name']}의 제공 정보와 학생 학교 자료는 서로 다른 사실입니다. 센터 위치·교습비 경로와 학습 범위·과제 기준을 구분해서 상담 메모에 적어 두세요.",
                (
                    f"학년 표기가 있더라도 반 편성이나 시간표까지 확정된 뜻은 아닙니다. {title} 등록 전에는 표시된 학년 범위와 실제 개설 여부, 수업 방식을 함께 확인하세요."
                    if has_listed_grades else
                    f"공개 자료의 가능 학년 정보가 비어 있으므로 {title} 등록 전에는 희망 학년과 실제 개설 여부, 수업 방식을 함께 확인하세요."
                ),
                (
                    f"표시된 {school_text} 목록은 상담 준비용 참고 자료입니다. 학생의 실제 재학 학교와 현재 진도를 알려 주고 자료 반영 방식을 질문하세요."
                    if schools else
                    "학교명 목록이 비어 있으면 실제 재학 학교를 상담에서 알려 주고 자료 반영 방식을 질문하세요. 공개되지 않은 학교 정보나 시험 범위를 페이지가 임의로 보충하지 않습니다."
                ),
                f"교습비 링크가 제공된 경우에도 금액과 과정은 변경될 수 있습니다. {locality} 상담에서 희망 과목, 학년과 주당 일정을 먼저 말한 뒤 최신 자료를 확인하세요.",
            ]),
        ],
    }

    recommended = {
        "key": "recommended",
        "heading": stable_pick(seed, "h2-recommended", [
            f"{title}이 필요한 학생과 확인할 수업 조건",
            f"{locality}에서 {config.label} 상담이 특히 유용한 경우",
            f"{locality} {student_level}에게 맞는 {'·'.join(config.subjects)} 점검 기준",
            f"{title} 추천 학생을 판단하는 실제 기준",
        ]),
        "paragraphs": [
            stable_pick(seed, "recommended-p1", [
                f"{title}을 우선 살펴볼 학생은 {student}입니다. 이 표현은 등록을 권하는 기준이 아니라 현재 자료와 상담 질문을 정리하기 위한 학습 상황입니다.",
                f"{locality}에서 {config.label} 상담이 유용한 경우는 {student}일 때입니다. 학생이 사용 중인 교재와 실행 기록을 확인해 실제 원인과 일치하는지 먼저 살펴야 합니다.",
                f"{student}이라면 {title} 상담에서 도움을 받을 부분과 혼자 연습할 부분을 구분해 질문해 보세요. 필요한 수업 조건은 학생마다 달라질 수 있습니다.",
                f"{title}의 추천 학생 기준은 성적 구간이 아니라 {primary.label}과 {secondary.label}의 실행 상태입니다. 최근 자료에서 두 항목이 어떻게 나타나는지 확인한 뒤 판단하세요.",
                f"{locality} {student_level}이 과제를 마쳐도 같은 어려움을 반복한다면 원인 분류가 필요합니다. {primary.label}과 {support.label}을 나눠 보면 필요한 피드백을 구체적으로 질문할 수 있습니다.",
                f"{title}을 알아보는 학부모는 학생에게 많은 문제를 먼저 제시하지 말고 {primary.check}를 확인해 보세요. 실제 자료와 학생 설명이 다를 때 상담에서 그 차이를 다루면 됩니다.",
            ]),
            stable_pick(seed, "recommended-p2", [
                f"학부모는 정답을 대신 알려 주기보다 학생이 ‘{support.home_action}’ 활동을 했는지 기록할 수 있습니다. {locality} 상담에서 이 기록을 공유하면 가정과 수업의 역할을 구분하기 쉽습니다.",
                f"학생이 도움을 요청할 시점과 혼자 다시 해 볼 범위를 정해 두는 편이 좋습니다. {center['center_name']}의 피드백 방식이 이 기준과 맞는지 상담에서 확인하세요.",
                f"추천 여부는 한 번의 점수나 설명만으로 결정하지 않습니다. {primary.evidence}와 {secondary.evidence}를 일정 간격으로 비교해 실제 학습 흐름을 살펴보세요.",
                f"과제량이 많은 수업보다 학생이 ‘{support.practice}’ 활동을 해 보고 확인받을 수 있는지가 중요합니다. 실제 운영 주기와 보완 방식은 센터에 직접 질문해야 합니다.",
                f"{locality} 학생의 귀가 시간과 학교 과제를 함께 놓고 실행 가능한 분량을 정하세요. 계획을 지키지 못한 경우에는 의지로 단정하지 말고 시간·난도·방법을 나누어 확인합니다.",
                f"첫 상담에서는 학생이 바꾸고 싶은 행동도 직접 말하게 해 보세요. 학부모 질문, 학생 목표와 실제 자료가 한 방향인지 확인해야 계획이 현실적입니다.",
            ]),
        ],
    }

    plan = {
        "key": "plan",
        "heading": stable_pick(seed, "h2-plan", [
            f"{title} 상담에서 정하는 조건부 4주 계획",
            f"{locality} {config.level} {subject_text} 학습의 첫 점검 주기 설계",
            f"{title} 학습을 진단·연습·재확인으로 나누기",
            f"{locality} {student_level}에게 맞는 {subject_text} 첫 달 실행 순서",
        ]),
        "paragraphs": [
            stable_pick(seed, "plan-p1", [
                f"{title}의 4주 계획은 성과를 약속하는 프로그램이 아니라 상담 내용을 정리하는 조건부 예시입니다. 첫 주에는 {primary.label}, 둘째 주에는 {secondary.label}, 이후에는 {support.label} 기록을 살피고 마지막 점검에서 다음 범위를 조정할 수 있습니다.",
                f"{locality} 학생의 첫 학습 주기는 진단, 짧은 연습과 재확인으로 나눌 수 있습니다. 먼저 ‘{primary.practice}’를 실행하고 이어 ‘{secondary.practice}’도 수행할 수 있는지 확인해 분량을 조정하세요.",
                f"{title} 상담 뒤에는 한꺼번에 여러 행동을 바꾸지 않습니다. 먼저 ‘{primary.home_action}’ 활동을 하고 다음 점검에서 기록을 확인한 뒤 {support.label} 과제를 더할 수 있습니다.",
                f"{locality} {config.level} 학습의 첫 달은 현재 자료를 읽는 기간으로 사용할 수 있습니다. {primary.evidence}, {secondary.evidence}와 실제 완료 분량을 비교해 다음 계획을 정합니다.",
                f"{title}에서 제안받은 계획은 학생 일정에 맞게 줄이거나 순서를 바꿀 수 있어야 합니다. {primary.label}과 {secondary.label} 중 먼저 확인할 항목을 정하고 점검일을 남기세요.",
                f"첫 계획에서는 ‘{support.practice}’ 활동을 수행할 시간을 확보하는 것이 중요합니다. {locality} 학생의 학교 일정과 통학 시간을 반영해 무리하지 않는 주간 분량을 정해야 합니다.",
            ]),
            stable_pick(seed, "plan-p2", [
                f"계획을 확인할 때는 문제 수보다 학생이 이유를 설명하고 다시 수행한 흔적을 봅니다. {title} 상담에서 첫 점검일과 계획을 바꿀 조건을 미리 질문하세요.",
                f"넷째 주에는 점수만 비교하지 않고 {primary.label}과 {support.label}에서 달라진 점이 있는지 살펴봅니다. 학습 결과는 출발점과 실천 정도에 따라 달라질 수 있습니다.",
                f"실행 기록이 남지 않았다면 새 진도를 늘리기 전에 분량과 방법을 다시 조정합니다. {center['center_name']}의 실제 피드백 주기가 학생 일정과 맞는지도 확인하세요.",
                f"한 주의 계획은 학교 일정이 바뀌면 함께 조정되어야 합니다. {material}를 기준으로 집중할 범위와 이어 갈 복습 내용을 나누어 두세요.",
                f"다음 단계는 첫 계획을 모두 마쳤을 때만 정하지 않습니다. 학생이 어려움을 설명한 시점과 다시 수행한 기록을 보고 보완 순서를 바꿀 수 있습니다.",
                f"이 예시는 특정 학생의 성과나 실제 수강 결과가 아닙니다. {locality} 학생의 자료, 가능한 일정과 센터의 현재 운영 조건을 확인한 뒤 개별 계획을 정해야 합니다.",
            ]),
        ],
    }

    checklist_items = [
        ("현재 자료", f"{title} 확인용으로 {material}에서 {primary.label}과 {secondary.label}이 드러난 부분을 표시합니다."),
        ("오답 근거", f"{locality} {config.level} {subject_text} 상담에는 {primary.evidence}와 {support.evidence}를 서로 나누어 가져갑니다."),
        ("가능 일정", f"{locality} {config.level} {subject_text} 학습에 필요한 통학 시간, 학교 일정과 가정 복습 시간을 적습니다."),
        ("센터 확인", f"{title} 등록 전에는 {center['center_name']}의 과목·학년·시간표·교습비·보완 방식을 확인합니다."),
    ]
    checklist = {
        "key": "checklist",
        "heading": stable_pick(seed, "h2-checklist", [
            f"{title} 상담 전 체크리스트",
            f"{locality} {config.label} 방문 전에 준비할 네 가지",
            f"{title} 등록 판단 전에 확인할 질문",
            f"{locality} 학부모가 {config.level} {subject_text} 상담 메모에 남길 항목",
        ]),
        "paragraphs": [
            stable_pick(seed, "checklist-p1", [
                f"{title} 상담 전에는 학생의 실제 자료, 어려웠던 과정과 가능한 일정을 한 장에 정리하세요. 아래 네 항목을 준비하면 학습 문제와 센터 운영 조건을 섞지 않고 질문할 수 있습니다.",
                f"{locality} 학부모는 많은 자료보다 최근 상태를 보여 주는 자료를 고르는 편이 좋습니다. ‘{primary.label}·{secondary.label}’ 관련 질문과 실제 학습 시간 확인 항목으로 나누어 보세요.",
                f"{title} 등록 판단은 수업 설명만 듣고 끝내지 않습니다. 현재 교재와 학생 기록, 가능 학년과 통학 조건을 차례로 확인해야 합니다.",
                f"{locality} 상담 메모에는 학습 질문과 운영 질문을 구분해 적으세요. 학생 자료는 수업 방향을, 주소·시간표·교습비는 실제 이용 조건을 확인하는 데 사용합니다.",
                f"{title}에 문의할 때는 최근 자료와 함께 학생이 혼자 공부할 수 있는 시간도 알려 주세요. 계획의 분량과 피드백 주기를 현실적으로 비교할 수 있습니다.",
                f"상담 준비는 학생의 약점을 많이 나열하는 일이 아닙니다. {locality} 학생이 지금 바꿀 행동, 확인할 자료와 센터에 물을 조건을 네 묶음으로 정리하면 충분합니다.",
            ]),
            stable_pick(seed, "checklist-p2", [
                (
                    f"표시된 가능 학년은 {grades}이며 실제 개설 여부는 상담 시점에 달라질 수 있습니다. 확인되지 않은 시간표, 차량·주차와 보강 운영은 페이지에서 단정하지 않습니다."
                    if all_grades_listed else
                    f"과목별 가능 학년은 제공 자료에서 모두 확인되지 않습니다. 희망 과목·학년과 실제 개설 여부는 {center['center_name']}에서 확인하세요. 확인되지 않은 시간표, 차량·주차와 보강 운영은 페이지에서 단정하지 않습니다."
                ),
                f"제공 주소 {center['address']}를 기준으로 실제 이동 시간을 계산해 보세요. 온라인 지도와 현장 동선이 다를 수 있으므로 방문 전에 위치도 다시 확인합니다.",
                f"교습비 링크가 있으면 최신 과정과 금액을 확인하고, 링크가 없으면 센터에 직접 문의하세요. 상담 질문과 실제 계약 조건은 구분해서 기록하는 편이 좋습니다.",
                f"체크리스트의 목적은 한 번에 등록을 결정하는 것이 아니라 비교 기준을 남기는 데 있습니다. {title} 상담 뒤에는 확인된 사실과 추가 질문을 나누어 적으세요.",
                f"학교명과 가능 학년 표기는 참고 자료입니다. 학생의 실제 학교·학년과 희망 과목이 현재 운영 범위에 포함되는지 {center['center_name']}에서 재확인하세요.",
                f"상담이 끝나면 첫 점검일, 가정에서 볼 기록과 계획을 조정할 조건을 한 줄씩 남기세요. 설명이 실제 학습 과정으로 이어지는지 확인하는 기준이 됩니다.",
            ]),
        ],
        "checklist": checklist_items,
    }

    middle = [method, school, recommended, plan]
    rotation = stable_int(seed, "section-order") % len(middle)
    middle = middle[rotation:] + middle[:rotation]
    if stable_int(seed, "section-reverse") % 2:
        middle[1:3] = reversed(middle[1:3])
    result = [diagnosis, *middle, checklist]

    return result


def build_faq(config: CategoryConfig, center: dict[str, object], signals: tuple[TopicSignal, ...], seed: str) -> list[tuple[str, str]]:
    locality = str(center["locality"])
    title = f"{locality} {config.label}"
    primary, secondary, support = signals[:3]
    schools = relevant_schools(config, center)
    material = material_label(config, schools)
    grade_items = grade_summary(config, center)
    listed_items = [(subject, value) for subject, value in grade_items if value != "상담 시 확인"]
    missing_subjects = [subject for subject, value in grade_items if value == "상담 시 확인"]
    if missing_subjects:
        if listed_items:
            listed_text = ", ".join(f"{subject} {value}" for subject, value in listed_items)
            grade_answer = (
                f"공개 자료에는 {listed_text} 범위가 기재되어 있습니다. {'·'.join(missing_subjects)} 가능 학년은 "
                f"{center['center_name']}에 희망 학년을 알려 확인하세요. 실제 반 편성·요일도 함께 확인해야 합니다."
            )
        else:
            grade_answer = (
                f"공개 자료에는 {'·'.join(missing_subjects)} 가능 학년이 기재되지 않았습니다. "
                f"{center['center_name']}에 희망 학년을 알려 현재 개설 범위와 실제 반 편성·요일을 확인하세요."
            )
    else:
        grades = ", ".join(f"{subject} {value}" for subject, value in grade_items)
        grade_answer = (
            f"공개된 표기는 {grades}입니다. 실제 반 편성·요일이 필요한 경우에는 "
            f"{center['center_name']}에 희망 과목과 학년을 알려 현재 개설 범위를 확인하세요."
        )
    subject_text = "·".join(config.subjects)
    faq_context = f"{locality} {config.level} {subject_text}"
    faq_topic = f"{config.level} {subject_text}"
    school_text = "·".join(schools[:3]) if schools else "실제 재학 학교"

    intents: dict[str, tuple[str, str]] = {
        "diagnosis": (
            f"{faq_context} 상담에서는 무엇을 먼저 진단하나요?",
            f"최근 점수만 보지 말고 {primary.check}와 {secondary.check}를 먼저 살핍니다. {locality} 학생의 {material}를 가져오면 두 항목이 실제 학습에서 어떻게 나타나는지 구분할 수 있습니다.",
        ),
        "materials": (
            f"{faq_context} 상담 때 학생은 어떤 자료를 준비하면 좋나요?",
            f"{material}에서 어려웠던 부분을 표시해 가져오면 됩니다. 자료가 많지 않아도 {primary.evidence}와 {support.evidence}를 확인할 수 있으면 학습 순서를 질문하기에 충분합니다.",
        ),
        "grades": (
            f"{locality}의 {config.level} {subject_text} 가능 학년과 실제 수업 일정은 어떻게 확인하나요?",
            grade_answer,
        ),
        "school": (
            f"{school_text} 자료는 {faq_context} 계획에 어떻게 반영하나요?",
            f"학교명 자체보다 학생이 가져온 범위와 현재 교재의 진행 위치를 확인합니다. {primary.label}과 {secondary.label}이 드러난 부분을 표시해 주간 과제와 복습 순서에 반영되는지 질문하세요.",
        ),
        "tuition": (
            f"{faq_context} 상담의 교습비와 센터 주소는 어디에서 확인하나요?",
            (f"제공 주소는 {center['address']}이며 교습비 자료는 페이지의 확인 링크로 연결됩니다. " if center["tuition_url"] else f"제공 주소는 {center['address']}이며 교습비는 희망 센터에 직접 문의해야 합니다. ")
            + "과정과 금액, 실제 방문 동선은 변경될 수 있으므로 등록 전에 최신 정보를 다시 확인하세요.",
        ),
        "schedule": (
            f"{faq_context} 수업 시간과 결석 시 보완 방식은 어떻게 확인하나요?",
            f"시간표·반 편성·보완 운영은 센터별로 다를 수 있어 페이지에서 확정하지 않습니다. {locality} 학생의 학교 일정과 통학 시간을 알려 주고 가능한 요일, 피드백 주기와 결석 시 절차를 함께 질문하세요.",
        ),
        "homework": (
            f"{locality}에서 {config.level} {subject_text} 학습을 하는 학생의 과제량은 어떤 기준으로 정하나요?",
            f"많은 분량보다 ‘{primary.practice}’와 ‘{secondary.practice}’ 두 활동을 실제로 마칠 수 있는지를 먼저 봅니다. 완료 기록을 다음 점검에서 확인하고 학교 일정과 가정 학습 시간에 맞춰 분량을 조정하세요.",
        ),
        "feedback": (
            f"{faq_context} 학습에서는 피드백 주기를 어떻게 비교해야 하나요?",
            f"정답을 알려 주는 횟수보다 학생이 이유를 설명하고 다시 수행한 기록을 확인하는지가 중요합니다. ‘{support.home_action}’ 활동을 해 본 뒤의 결과를 언제 확인하는지와 계획을 바꾸는 기준을 상담에서 물어보세요.",
        ),
        "balance": (
            f"{faq_context} 학습 순서는 어떻게 정하나요?",
            f"모든 단원이나 과목을 같은 비중으로 다루지 않고 {primary.label}과 {secondary.label}의 어려움을 구분합니다. 학교 일정과 현재 자료를 바탕으로 집중할 범위와 이어 갈 복습 내용을 나누는 편이 좋습니다.",
        ),
        "four_weeks": (
            f"{faq_context} 4주 계획은 어떤 흐름으로 확인하나요?",
            f"4주 계획은 성과를 보장하는 과정이 아니라 진단·연습·재확인을 정리한 조건부 예시입니다. {primary.label}의 첫 기록과 {support.label}의 재확인 결과를 비교해 다음 범위를 조정합니다.",
        ),
        "consultation": (
            f"{faq_context} 첫 상담 전에 학부모가 정리할 질문은 무엇인가요?",
            f"학생이 막힌 과정, 실제 가능한 학습 시간과 센터 운영 조건을 구분해 적으세요. 이 상담에서는 첫 점검일, 가정에서 확인할 기록과 시간표·교습비 확인 방법을 차례로 질문하면 됩니다.",
        ),
        "results": (
            f"{faq_context} 상담 예시를 실제 후기나 성과로 보아도 되나요?",
            f"페이지의 상담 상황은 실제 이용 후기나 특정 학생의 성과가 아닌 준비용 가상 예시입니다. 학습 결과는 학생의 출발점과 실천 정도에 따라 달라질 수 있으므로 실제 자료와 센터 조건을 확인해 판단하세요.",
        ),
    }
    answer_contexts = {
        "diagnosis": f"{faq_topic} 상담에서 사용할 첫 진단 기준입니다.",
        "materials": f"{faq_topic} 상담 자료는 현재 학습 상태를 보여 주는 범위에서 고릅니다.",
        "grades": f"{faq_topic} 학년·시간표 정보는 공개 자료와 상담 확인 내용을 구분합니다.",
        "school": f"{faq_topic} 학교 자료 반영 방식은 실제 학생 자료를 기준으로 확인합니다.",
        "tuition": f"{faq_topic} 주소·교습비는 제공된 센터 자료를 기준으로 안내합니다.",
        "schedule": f"{faq_topic} 시간표·보완 방식은 현재 센터 운영 범위를 직접 확인해야 합니다.",
        "homework": f"{faq_topic} 과제량은 학생이 수행하고 다시 확인할 수 있는지를 기준으로 비교합니다.",
        "feedback": f"{faq_topic} 피드백은 기록과 다음 점검일이 연결되는지를 살펴봅니다.",
        "balance": f"{faq_topic} 학습 순서는 현재 병목과 학교 일정을 함께 놓고 정합니다.",
        "four_weeks": f"{faq_topic} 4주 흐름은 첫 기록과 마지막 재확인 결과를 이어 봅니다.",
        "consultation": f"{faq_topic} 첫 상담 질문은 학습 자료와 이용 조건을 나누어 준비합니다.",
        "results": f"{faq_topic} 상담 장면을 읽을 때는 출처와 성과 여부를 구분해야 합니다.",
    }
    required = ["diagnosis", "grades"]
    rotating = ["materials", "school", "tuition", "schedule", "homework", "feedback", "balance", "four_weeks", "consultation", "results"]
    start = stable_int(seed, "faq-intents") % len(rotating)
    # Every stride must be coprime with the ten rotating intents; otherwise a
    # stride of five would visit only two entries and never fill five slots.
    stride = (1, 3, 7, 9)[stable_int(seed, "faq-stride") % 4]
    chosen = list(required)
    cursor = start
    while len(chosen) < 5:
        intent = rotating[cursor % len(rotating)]
        if intent not in chosen:
            chosen.append(intent)
        cursor += stride
    if stable_int(seed, "faq-order") % 2:
        chosen[2:] = reversed(chosen[2:])
    return [
        (intents[intent][0], f"{answer_contexts[intent]} {intents[intent][1]}")
        for intent in chosen
    ]


def build_scenarios(config: CategoryConfig, center: dict[str, object], signals: tuple[TopicSignal, ...], seed: str) -> list[str]:
    locality = str(center["locality"])
    title = f"{locality} {config.label}"
    subject_text = "·".join(config.subjects)
    scenario_context = f"{locality} {config.level} {subject_text}"
    primary, secondary, support = signals[:3]
    material = material_label(config, relevant_schools(config, center))
    first = stable_pick(seed, "scenario-first", [
        f"{scenario_context} 상담을 준비하는 학부모가 {material}를 가져오는 장면을 가정했습니다. 학부모는 {primary.label}과 {secondary.label}이 드러난 부분을 보여 주고, 학생이 다음 점검까지 수행할 한 가지 행동이 무엇인지 질문합니다.",
        f"{locality}에서 {config.level} {subject_text} 학습을 하는 학생이 과제를 마쳐도 같은 어려움을 반복하는 상황을 가정했습니다. 학부모는 {primary.evidence}와 {support.evidence}, 두 자료를 나누어 제시하고 설명이 필요한 부분과 혼자 연습할 부분을 확인합니다.",
        f"학교 진도와 현재 실력 사이에서 우선순위를 정하기 어려운 {locality}의 {config.level} {subject_text} 학습 상담 장면입니다. 학부모는 {primary.label}을 먼저 보완할지 {secondary.label} 기록을 더 확인할지 실제 자료를 바탕으로 질문합니다.",
        f"가정 학습 시간이 일정하지 않은 {locality}의 {config.level} {subject_text} 학습 상황을 가정했습니다. 학부모는 많은 과제를 요청하기보다 ‘{support.home_action}’ 활동을 실천할 수 있는 분량과 다음 점검일을 상담에서 확인합니다.",
        f"{locality}에서 {config.level} {subject_text} 수업 설명을 비교하는 학부모 상황입니다. 학부모는 점수 향상 표현보다 {primary.check}와 {secondary.check}를 어떤 자료로 판단하는지 질문합니다.",
        f"{locality}에서 {config.level} {subject_text} 교재를 바꿔야 할지 고민하는 학부모의 상담 장면입니다. 학부모는 {primary.evidence}를 보여 주고 교재 난도와 학습 절차 중 무엇을 먼저 조정할지 확인합니다.",
    ])
    second = stable_pick(seed, "scenario-second", [
        f"{scenario_context} 상담에서 {center['center_name']}의 제공 주소와 가능 학년을 확인한 뒤 실제 시간표, 통학 시간과 교습비를 따로 비교하는 상황입니다. 학습 계획과 이용 조건을 한 문장으로 묶지 않고 확인된 사실과 추가 질문을 나누어 기록합니다.",
        f"{locality}에서 {config.level} {subject_text} 상담을 마친 학부모가 학생의 학교 일정과 가정 학습 시간을 대조하는 장면입니다. 첫 계획을 그대로 따르기보다 실행하지 못했을 때 분량과 순서를 어떻게 바꿀지 질문합니다.",
        f"{scenario_context} 학습 자료가 공개 학교 목록과 다를 수 있는 경우를 가정했습니다. 학부모는 실제 재학 학교와 현재 진도를 알려 주고 수업 가능 여부와 자료 반영 방식을 센터에서 다시 확인합니다.",
        f"{locality}에서 {config.level} {subject_text} 학습 시간이 겹치는 가정의 상담 상황입니다. 학부모는 {primary.label}에 집중할 시간과 {secondary.label}을 복습할 시간을 구분하고 일주일 뒤 어떤 기록으로 계획을 재검토할지 묻습니다.",
        f"{locality}에서 {config.level} {subject_text} 첫 상담을 마친 뒤 학생과 학부모가 내용을 다시 정리하는 상황입니다. ‘{support.practice}’를 실제로 해낼 수 있는지 살펴보고, 어려우면 다음 상담에서 분량과 피드백 주기를 조정합니다.",
        f"{scenario_context} 상담에서 센터 설명과 학생 자료가 일치하는지 비교하는 상황입니다. 확인되지 않은 차량·주차·보강 운영을 추정하지 않고 현재 개설 과목과 학년, 방문 조건을 직접 확인합니다.",
    ])
    return [first, second]


def offer_nodes(config: CategoryConfig, center: dict[str, object], service_id: str) -> list[dict[str, object]]:
    offers: list[dict[str, object]] = []
    for subject in config.subjects:
        grades = relevant_grades(config, center, subject)
        name = f"{config.level} {subject} 학습 상담" if config.grade_prefix else f"{subject} 학습 상담"
        offer: dict[str, object] = {
            "@type": "Offer",
            "name": name,
            "itemOffered": {"@id": service_id, "@type": "Service", "name": name},
        }
        if grades:
            offer["eligibleCustomerType"] = "·".join(grades)
        offers.append(offer)
    if center["tuition_url"]:
        offers.append({
            "@type": "Offer",
            "name": f"{center['center_name']} 교습과정·교습비 확인",
            "url": center["tuition_url"],
            "itemOffered": {"@type": "Service", "name": "센터별 교습과정 정보 확인"},
        })
    return offers


def related_links(config: CategoryConfig, center: dict[str, object], previous_slug: str, next_slug: str) -> list[tuple[str, str]]:
    locality = str(center["locality"])
    local_slug = str(center["slug"])
    related_slugs: list[str]
    if config.slug == "영수학원":
        related_slugs = ["초등학생국영수학원", "중학생국영수학원", "고등학생국영수학원", "중등영어학원", "중등수학학원"]
    elif config.grade_prefix == "초":
        related_slugs = ["초등영어학원", "초등수학학원", "초등학생국영수학원", "영수학원"]
    elif config.grade_prefix == "중":
        related_slugs = ["중등영어학원", "중등수학학원", "중학생국영수학원", "영수학원"]
    else:
        related_slugs = ["고등영어학원", "고등수학학원", "고등학생국영수학원", "영수학원"]
    links: list[tuple[str, str]] = [(f"{config.label} 전체 지역", route_for(config.slug))]
    for slug in related_slugs:
        if slug == config.slug:
            continue
        links.append((f"{locality} {ALL_CATEGORY_LABELS[slug]}", route_for(slug, local_slug)))
    links.extend([
        (f"이전 지역 · {previous_slug}", route_for(config.slug, previous_slug)),
        (f"다음 지역 · {next_slug}", route_for(config.slug, next_slug)),
        ("학습가이드", "/학습가이드/"),
        ("상담문의", "/상담문의/"),
    ])
    return links


def build_graph(
    config: CategoryConfig,
    center: dict[str, object],
    title: str,
    meta: str,
    rep_name: str,
    sections: list[dict[str, object]],
    faq: list[tuple[str, str]],
    links: list[tuple[str, str]],
    high_english: HighEnglishProfile | MiddleEnglishProfile | HighMathProfile | None = None,
) -> dict[str, object]:
    locality = str(center["locality"])
    slug = str(center["slug"])
    url = absolute_route(config.slug, slug)
    hub_url = absolute_route(config.slug)
    org_id = center_entity_id(center)
    service_id = url + "#service"
    rep_url = ORIGIN + f"/assets/representative/{quote(rep_name)}"
    body_url = ORIGIN + f"/assets/centers/common/{center['body_image']}"
    map_url = ORIGIN + "/assets/maps/" + quote(str(center["map_name"]))
    schools = relevant_schools(config, center)
    grades = list(dict.fromkeys(grade for subject in config.subjects for grade in relevant_grades(config, center, subject)))
    offers = offer_nodes(config, center, service_id)
    about = [
        {"@type": "Thing", "name": config.label},
        *({"@type": "Thing", "name": f"{config.level} {subject} 학습"} for subject in config.subjects),
        {"@type": "Thing", "name": "학습 진단"},
        {"@type": "Thing", "name": "오답 재학습"},
    ]
    if high_english is not None:
        about.extend({"@type": "Thing", "name": intent.label} for intent in high_english.intents)
        subject = config.subjects[0]
        cumulative_topic = "모의고사" if config.level == "고등" else "누적 학습"
        about.extend((
            {"@type": "Thing", "name": f"{config.level} {subject} 내신"},
            {"@type": "Thing", "name": f"{config.level} {subject} {cumulative_topic}"},
        ))
    mentions = [
        {"@type": "Place", "name": locality},
        *({"@type": "EducationalOrganization", "name": school} for school in schools),
        *({"@type": "Thing", "name": subject + " 학습"} for subject in config.subjects),
    ]
    section_parts = [
        {"@type": "WebPageElement", "@id": f"{url}#section-{index}", "name": str(section["heading"])}
        for index, section in enumerate(sections, 1)
    ]
    organization: dict[str, object] = {
        "@type": ["EducationalOrganization", "LocalBusiness"],
        "@id": org_id,
        "name": center["center_name"],
        "url": ORIGIN + "/",
        "telephone": PHONE,
        "address": {
            "@type": "PostalAddress",
            "streetAddress": center["address"],
            "addressRegion": center["address_region"],
            "addressCountry": "KR",
        },
        "areaServed": [{"@type": "Place", "name": locality}],
        "description": f"{center['center_name']}의 제공 주소·등록정보와 {locality} 학습 상담 확인 경로를 안내합니다.",
        "image": rep_url,
        "teaches": [f"{config.level} {subject}" for subject in config.subjects] + ["학습코칭"],
        "makesOffer": offers,
    }
    if center["registration"]:
        organization["identifier"] = {
            "@type": "PropertyValue",
            "name": center["office_name"] or "등록 학원명",
            "value": center["registration"],
        }
    if grades:
        organization["educationalLevel"] = grades
    breadcrumb = {
        "@type": "BreadcrumbList",
        "@id": url + "#breadcrumb",
        "itemListElement": [
            {"@type": "ListItem", "position": 1, "name": "홈", "item": ORIGIN + "/"},
            {"@type": "ListItem", "position": 2, "name": "과목별학원", "item": ORIGIN + quote("/과목별학원/", safe="/")},
            {"@type": "ListItem", "position": 3, "name": config.label, "item": hub_url},
            {"@type": "ListItem", "position": 4, "name": title, "item": url},
        ],
    }
    image_object = {
        "@type": "ImageObject",
        "@id": url + "#primaryimage",
        "url": rep_url,
        "contentUrl": rep_url,
        "caption": f"{title} 대표 이미지",
    }
    article = {
        "@type": "Article",
        "@id": url + "#article",
        "url": url,
        "headline": title,
        "description": meta,
        "abstract": meta,
        "inLanguage": "ko-KR",
        "mainEntityOfPage": {"@id": url + "#webpage"},
        "author": {"@id": org_id},
        "publisher": {"@id": org_id},
        "datePublished": CONTENT_DATE,
        "dateModified": "2026-09-01" if high_english is not None else CONTENT_DATE,
        "articleSection": [config.label, str(center["region"]), str(center["district"]), locality],
        "about": about,
        "mentions": mentions,
        "hasPart": section_parts,
        "image": [rep_url, body_url, map_url],
    }
    if grades:
        article["educationalLevel"] = grades
    if high_english is not None:
        subject = config.subjects[0]
        cumulative_topic = "모의고사" if config.level == "고등" else "누적 학습"
        article["keywords"] = [
            title,
            high_english.focus,
            *(intent.label for intent in high_english.intents),
            f"{config.level} {subject} 내신",
            f"{config.level} {subject} {cumulative_topic}",
        ]
    service = {
        "@type": "Service",
        "@id": service_id,
        "url": url,
        "name": f"{title} 학습 상담 안내",
        "serviceType": config.label,
        "provider": {"@id": org_id},
        "areaServed": {"@type": "Place", "name": locality},
        "audience": {"@type": "EducationalAudience", "educationalRole": "student", "audienceType": config.level},
        "description": meta,
        "offers": offers,
    }
    webpage = {
        "@type": "WebPage",
        "@id": url + "#webpage",
        "url": url,
        "name": title,
        "description": meta,
        "inLanguage": "ko-KR",
        "isPartOf": {"@id": hub_url + "#collection"},
        "breadcrumb": {"@id": breadcrumb["@id"]},
        "primaryImageOfPage": {"@id": image_object["@id"]},
        "about": about,
        "mentions": mentions,
        "hasPart": section_parts,
        "mainEntity": {"@id": article["@id"]},
    }
    faq_node = {
        "@type": "FAQPage",
        "@id": url + "#faq",
        "mainEntity": [
            {"@type": "Question", "name": question, "acceptedAnswer": {"@type": "Answer", "text": answer}}
            for question, answer in faq
        ],
    }
    item_list = {
        "@type": "ItemList",
        "@id": url + "#related",
        "name": f"{title} 관련 페이지",
        "numberOfItems": len(links),
        "itemListElement": [
            {"@type": "ListItem", "position": index, "name": name, "url": ORIGIN + quote(href, safe="/%:@")}
            for index, (name, href) in enumerate(links, 1)
        ],
    }
    return {
        "@context": "https://schema.org",
        "@graph": [organization, image_object, webpage, breadcrumb, article, service, faq_node, item_list],
    }


def header(prefix: str) -> str:
    items = [
        ("home", "홈", "index.html"),
        ("about", "학원소개", "학원소개/index.html"),
        ("guide", "학습가이드", "학습가이드/index.html"),
        ("contact", "상담문의", "상담문의/index.html"),
        ("subjects", "과목별학원", "과목별학원/index.html"),
    ]
    links = "".join(
        f'<a href="{prefix}{href}" data-nav="{key}"{" aria-current=\"page\"" if key == "subjects" else ""}>{label}</a>'
        for key, label, href in items
    )
    return f'''<a class="skip-link" href="#main">본문으로 건너뛰기</a>
  <header class="site-header"><div class="site-shell header-inner">
    <a class="brand" href="{prefix}index.html"><span class="brand-mark" aria-hidden="true">W</span><span class="brand-copy"><strong>{SITE_NAME}</strong><small>STUDY RECORD COACHING</small></span></a>
    <nav class="primary-nav" aria-label="주요 메뉴">{links}</nav>
  </div></header>'''


def footer(prefix: str) -> str:
    return f'''<footer class="site-footer"><div class="site-shell footer-grid">
    <div class="footer-brand"><h2>{SITE_NAME}</h2><p>학생별 진도와 교재, 실행 기록과 오답 재학습을 연결해 다음 공부 순서를 정리합니다. 개설 과목과 학년, 수업 방식은 센터별로 확인합니다.</p></div>
    <div><nav class="footer-links" aria-label="하단 메뉴"><a href="{prefix}학원소개/index.html">학원소개</a><a href="{prefix}학습가이드/index.html">학습가이드</a><a href="{prefix}과목별학원/index.html">과목별학원</a><a href="{prefix}상담문의/index.html">상담문의</a></nav><p class="footer-meta">대표 상담 {PHONE}<br>© {SITE_NAME}</p></div>
  </div></footer>
  <nav class="contact-dock" aria-label="빠른 상담"><a href="tel:010-6839-8283">전화문의</a><a href="https://blogsms.net/01068398283" target="_blank" rel="noopener">문자문의</a><a href="https://docs.google.com/forms/d/e/1FAIpQLSdb2oE5Qk5YS0TfYDxyV1w-IOTkhkjOCmmpAKTI9FmqpVj6Yg/viewform" target="_blank" rel="noopener">상담신청</a></nav>'''


def render_info(config: CategoryConfig, center: dict[str, object]) -> str:
    rows = [
        ("지역", display_region_label(center)),
        ("센터 기준", center["center_name"]),
        ("제공 주소", center["address"]),
    ]
    if center["registration"]:
        rows.append(("등록 정보", center["registration"]))
    html_rows = "".join(f"<div><dt>{esc(label)}</dt><dd>{esc(value)}</dd></div>" for label, value in rows if value)
    schools = relevant_schools(config, center)
    school_html = (
        f'<div><dt>{esc(config.school_name)} 참고</dt><dd><div class="local-tags">{"".join(f"<span>{esc(school)}</span>" for school in schools[:8])}</div></dd></div>'
        if schools else f'<div><dt>{esc(config.school_name)} 참고</dt><dd>제공 목록 없음 · 상담 시 실제 학교 자료 확인</dd></div>'
    )
    grades = "".join(
        f'<li><strong>{esc(subject)}</strong><span>{esc(value)}</span></li>'
        for subject, value in grade_summary(config, center)
    )
    grade_note = ", ".join(f"{subject} {value}" for subject, value in grade_summary(config, center))
    tuition = (
        f'<a class="button compact" href="{esc(center["tuition_url"])}" target="_blank" rel="noopener">센터별 교습비 확인 <span aria-hidden="true">↗</span></a>'
        if center["tuition_url"] else '<p class="info-note">교습비 자료는 희망 센터에서 확인합니다.</p>'
    )
    verification_text = (
        "센터명·주소·가능 학년·학교명은 제공 센터 자료를 기준으로 정리했습니다. 시간표와 교습비는 등록 전에 최신 안내를 다시 확인하며, 학습 방법은 학생 자료를 살펴보기 위한 일반 안내입니다."
        if config.slug == "고등영어학원" else
        "표기된 학년은 제공 자료 기준이며, 시간표와 실제 개설 여부는 상담 시 확인합니다."
    )
    return (
        f'<dl class="local-facts">{html_rows}{school_html}</dl><ul class="grade-list">{grades}</ul>'
        f'<p class="center-verified-note"><strong>제공 자료 확인 기준</strong><span>{esc(verification_text)}</span></p>{tuition}'
    )


def render_map_contact_overlays(map_name: str) -> str:
    from postprocess_center_entities import phone_banner_bands

    return "".join(
        (
            f'<span class="map-contact-correction" style="top:{top:.4f}%;height:{height:.4f}%" '
            f'aria-label="대표 상담 전화 {PHONE}">{PHONE}</span>'
        )
        for top, height in phone_banner_bands(MAP_DIR / map_name)
    )


def preserve_high_english_center_schema(rendered: str, existing: str) -> str:
    """Keep the established physical-center entity while refreshing page prose schema."""

    pattern = re.compile(
        r'(<script\s+type="application/ld\+json">)(.*?)(</script>)',
        re.DOTALL | re.IGNORECASE,
    )
    rendered_match = pattern.search(rendered)
    existing_match = pattern.search(existing)
    if not rendered_match or not existing_match:
        raise ValueError("high English JSON-LD script missing")
    rendered_graph = json.loads(rendered_match.group(2))
    existing_graph = json.loads(existing_match.group(2))
    new_nodes = {
        str(node.get("@id", "")): node
        for node in rendered_graph.get("@graph", [])
        if isinstance(node, dict) and node.get("@id")
    }
    merged_nodes: list[dict[str, object]] = []
    mutable_fields = {
        "Article": ("description", "abstract", "about", "mentions", "hasPart", "dateModified", "keywords"),
        "WebPage": ("description", "about", "mentions", "hasPart"),
        "Service": ("description",),
    }
    for existing_node in existing_graph.get("@graph", []):
        if not isinstance(existing_node, dict):
            continue
        node_id = str(existing_node.get("@id", ""))
        new_node = new_nodes.get(node_id)
        types = existing_node.get("@type", [])
        type_values = {types} if isinstance(types, str) else set(types)
        if new_node is None:
            merged_nodes.append(existing_node)
            continue
        if "FAQPage" in type_values:
            merged_nodes.append(new_node)
            continue
        merged = dict(existing_node)
        for schema_type, fields in mutable_fields.items():
            if schema_type not in type_values:
                continue
            for field in fields:
                if field in new_node:
                    merged[field] = new_node[field]
                else:
                    merged.pop(field, None)
        if {"EducationalOrganization", "LocalBusiness"} <= type_values:
            existing_address = existing_node.get("address")
            fresh_address = new_node.get("address")
            if isinstance(existing_address, dict) and isinstance(fresh_address, dict):
                street_address = str(fresh_address.get("streetAddress", ""))
                physical = physical_region(street_address, str(fresh_address.get("addressRegion", "")))
                recorded = str(existing_address.get("addressRegion", ""))
                if canonical_schema_region(physical) != canonical_schema_region(recorded):
                    corrected_address = dict(existing_address)
                    corrected_address["addressRegion"] = physical
                    locality = physical_schema_locality(street_address)
                    if locality:
                        corrected_address["addressLocality"] = locality
                    merged["address"] = corrected_address
        merged_nodes.append(merged)
    merged_graph = {
        "@context": existing_graph.get("@context", "https://schema.org"),
        "@graph": merged_nodes,
    }
    payload = json.dumps(merged_graph, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    return rendered[:rendered_match.start(2)] + payload + rendered[rendered_match.end(2):]


def render_page(
    config: CategoryConfig,
    record: dict[str, object],
    previous_record: dict[str, object],
    next_record: dict[str, object],
) -> str:
    center = record["center"]
    locality = str(center["locality"])
    title = str(record["title"])
    meta = str(record["meta"])
    rep_name = str(record["rep_name"])
    signals = record["signals"]
    sections = record["sections"]
    faq = record["faq"]
    scenarios = record["scenarios"]
    student = str(record["student"])
    quick = str(record["quick"])
    high_english = record.get("high_english_profile")
    middle_english = record.get("middle_english_profile")
    middle_math = record.get("middle_math_profile")
    high_math = record.get("high_math_profile")
    high_profile = (
        high_english if isinstance(high_english, HighEnglishProfile)
        else middle_english if isinstance(middle_english, MiddleEnglishProfile)
        else middle_math if isinstance(middle_math, HighMathProfile)
        else high_math if isinstance(high_math, HighMathProfile)
        else None
    )
    student_level = student_level_label(config)
    if high_profile is not None:
        primary, secondary, support, extra = high_profile.intents
        quick_student_note = stable_pick(title, "high-quick-note", [
            f"{locality} 상담 메모에는 {primary.label}의 첫 행동과 {secondary.label}의 재확인 날짜를 적고 이용 조건은 별도 줄로 나눕니다.",
            f"{locality} 학생의 {primary.label} 설명과 {secondary.label} 풀이 흔적이 이어지는지 확인하고 센터 정보는 따로 기록합니다.",
            f"{locality} 학습 진단에는 {primary.label}·{secondary.label}의 확인 기록을, 등록 조건에는 주소·학년·시간표를 적습니다.",
            f"{locality} 학생의 한 가지 병목과 {primary.label}의 첫 행동, {secondary.label}의 재확인 날짜가 분명해야 계획을 비교할 수 있습니다.",
            f"{locality} 시험지에서는 {primary.label}에서 막힌 이유를, 다음 계획에서는 {secondary.label}의 확인일을 찾아 이용 조건과 구분합니다.",
            f"{locality} 상담 답변을 {primary.label}의 실행 항목과 {secondary.label}의 완료 기준으로 바꾸고 주소·시간표는 별도 메모에 둡니다.",
            f"{locality} 학생 기록에는 {primary.label}의 문제 번호와 {secondary.label}의 재확인 날짜가 함께 있어야 합니다.",
            f"{locality} 학습 계획은 {primary.label} 자료와 {secondary.label} 질문에서 시작하고 센터 조건은 마지막에 따로 확인합니다.",
            f"{locality} 학생의 {primary.label} 시작일과 {secondary.label} 확인일을 나란히 적되 학년·시간표는 이용 조건으로 구분합니다.",
            f"{locality} 학습 메모에서는 {primary.label}의 확인 기록과 {secondary.label}의 다음 행동을 먼저 확인합니다.",
            f"{locality} 상담 전에는 {primary.label} 자료와 {secondary.label} 재확인 질문을 준비하고 등록 조건은 별도 표에 둡니다.",
            f"{locality} 학생 계획에 {primary.label} 실행 여부와 {secondary.label} 점검 결과가 모두 있어야 다음 순서를 정할 수 있습니다.",
        ])
        quick_note_tokens: list[str] = []
        for intent in high_profile.intents:
            quick_note_tokens.extend((intent.label, intent.evidence))
        quick_student_note = normalize_particle_joins(quick_student_note, quick_note_tokens)
        if isinstance(middle_english, MiddleEnglishProfile):
            quick_student_note = naturalize_middle_english_text(quick_student_note, middle_english)
        elif isinstance(middle_math, HighMathProfile):
            quick_student_note = naturalize_middle_math_text(quick_student_note, middle_math)
        elif isinstance(high_math, HighMathProfile):
            quick_student_note = naturalize_high_math_text(quick_student_note, high_math)
        subject_label = config.subjects[0]
        quick_heading = f"{locality} {config.level} {subject_label}: {high_profile.focus}"
        manuscript_intro = (
            f"핵심 주제는 ‘{high_profile.focus}’입니다. "
            f"{primary.label}·{secondary.label}의 현재선을 확인하고 {support.label}·{extra.label}의 실행 기록과 학교·센터 사실을 분리해 살펴봅니다."
        )
        if isinstance(middle_math, HighMathProfile):
            manuscript_intro = naturalize_middle_math_text(manuscript_intro, middle_math)
        elif isinstance(high_math, HighMathProfile):
            manuscript_intro = naturalize_high_math_text(manuscript_intro, high_math)
        elif isinstance(middle_english, MiddleEnglishProfile):
            manuscript_intro = naturalize_middle_english_text(manuscript_intro, middle_english)
    else:
        quick_student_note = (
            f"{locality} {student_level} 상담에서는 {signals[0].label}·{signals[1].label}의 현재 상태를 "
            "실제 자료에서 나누어 확인합니다."
        )
        quick_heading = f"{locality} {config.level} {'·'.join(config.subjects)} 학습에서 먼저 확인할 핵심 답변"
        manuscript_intro = f"{locality} {config.level} {'·'.join(config.subjects)} 학습 자료를 기준으로 진단, 학습 순서, 학교·센터 사실과 상담 질문을 차례로 살펴봅니다."
    links = related_links(config, center, str(previous_record["slug"]), str(next_record["slug"]))
    graph = build_graph(
        config,
        center,
        title,
        meta,
        rep_name,
        sections,
        faq,
        links,
        high_profile,
    )
    if isinstance(middle_math, HighMathProfile):
        graph = naturalize_middle_math_tree(graph, middle_math)  # type: ignore[assignment]
    elif isinstance(high_math, HighMathProfile):
        graph = naturalize_high_math_tree(graph, high_math)  # type: ignore[assignment]
    elif isinstance(middle_english, MiddleEnglishProfile):
        graph = naturalize_middle_english_tree(graph, middle_english)  # type: ignore[assignment]
    graph_json = json.dumps(graph, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    url = absolute_route(config.slug, str(center["slug"]))
    rep_url = ORIGIN + f"/assets/representative/{quote(rep_name)}"
    cards_html = "".join(
        f'<article><span>{esc(label)}</span><strong>{esc(value)}</strong></article>'
        for label, value in answer_cards(config, center, signals, student)
    )
    section_html_parts: list[str] = []
    for index, section in enumerate(sections, 1):
        paragraphs = "".join(f"<p>{esc(paragraph)}</p>" for paragraph in section["paragraphs"])
        checklist = section.get("checklist")
        if checklist:
            paragraphs += '<ol class="checklist">' + "".join(
                f'<li><div><strong>{esc(label)}</strong><span>{esc(value)}</span></div></li>'
                for label, value in checklist
            ) + "</ol>"
        section_html_parts.append(
            f'<section class="manuscript-section" id="section-{index}"><span class="section-kicker">{index:02d}</span><h2>{esc(section["heading"])}</h2>{paragraphs}</section>'
        )
    body_html = "".join(section_html_parts)
    faq_html = "".join(
        f'<details{" open" if index == 0 else ""}><summary>{esc(question)}</summary><p>{esc(answer)}</p></details>'
        for index, (question, answer) in enumerate(faq)
    )
    scenario_html = "".join(
        f'<article class="scenario-card"><span>학부모 관점 상담 예시 {index:02d}</span><p>{esc(value)}</p></article>'
        for index, value in enumerate(scenarios, 1)
    )
    related_html = "".join(
        f'<a class="local-nav-card{" is-parent" if index == 0 else ""}" href="{esc(href)}"><small>{"카테고리" if index == 0 else "관련 페이지"}</small><strong>{esc(name)}</strong><span>→</span></a>'
        for index, (name, href) in enumerate(links)
    )
    region_label = display_region_label(center)
    return f'''<!doctype html>
<html lang="ko"><head>
  <meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{esc(title)} | {TITLE_SUFFIX}</title><meta name="description" content="{esc(meta)}"><meta name="robots" content="index,follow,max-image-preview:large"><meta name="theme-color" content="#f7f3ec">
  <link rel="canonical" href="{url}"><meta property="og:type" content="article"><meta property="og:locale" content="ko_KR"><meta property="og:title" content="{esc(title)} | {TITLE_SUFFIX}"><meta property="og:description" content="{esc(meta)}"><meta property="og:url" content="{url}"><meta property="og:image" content="{rep_url}">
  <link rel="icon" href="../../../assets/favicon.png"><link rel="stylesheet" href="../../../assets/site14.css"><script type="application/ld+json">{graph_json}</script>
</head><body data-page="subjects">{header("../../../")}
  <main id="main">
    <section class="local-hero"><div class="site-shell">
      <nav class="breadcrumbs" aria-label="현재 위치"><a href="../../../index.html">홈</a><a href="../../index.html">과목별학원</a><a href="../index.html">{esc(config.label)}</a><span>{esc(title)}</span></nav>
      <p class="eyebrow">{esc(config.english)}</p><h1>{esc(title)}</h1><p class="local-lead">{esc(meta)}</p><div class="local-answer-grid">{cards_html}</div>
    </div></section>
    <section class="section local-overview"><div class="site-shell local-overview-grid">
      <div class="local-summary"><p class="chapter-label"><span>01</span> Quick answer</p><h2>{esc(quick_heading)}</h2><p>{esc(quick)}</p><div class="answer-note"><strong>확인 기준</strong><p>{esc(quick_student_note)}</p></div></div>
      <aside class="local-info-card"><p class="eyebrow">Center information</p><h2>지역·학년·센터 확인 정보</h2>{render_info(config, center)}</aside>
    </div></section>
    <section class="local-media-section"><div class="site-shell local-media-stack">
      <img src="../../../assets/representative/{esc(rep_name)}" alt="{esc(title)} {SITE_NAME} 대표" style="display:none;">
      <figure class="local-body-image"><img src="../../../assets/centers/common/{esc(center['body_image'])}" width="918" height="16116" alt="{esc(title)} 본문 학습 안내" loading="lazy" decoding="async"><figcaption>{esc(region_label)} {esc(config.label)} 학습 점검 안내</figcaption></figure>
      <figure class="local-map-image" id="center-map"><div class="map-art"><img src="../../../assets/maps/{esc(center['map_name'])}" alt="{esc(title)} 지도 {esc(center['center_name'])}" loading="lazy" decoding="async">{render_map_contact_overlays(str(center['map_name']))}</div><figcaption>센터 위치는 제공 주소를 기준으로 표시하며 방문 전 실제 운영 여부와 동선을 확인합니다.</figcaption></figure>
    </div></section>
    <section class="section manuscript-wrap"><article class="site-shell manuscript-article"><div class="manuscript-intro"><span>학습 답변 요약</span><p>{esc(manuscript_intro)}</p></div>{body_html}</article></section>
    <section class="section blue-wash"><div class="site-shell"><div class="section-heading"><div class="chapter-label"><span>02</span> Parent perspective</div><div><h2>{esc(locality)} {esc(config.level)} {esc('·'.join(config.subjects))} 학부모 상담 상황</h2><p>아래 {esc(locality)} {esc(config.level)} {esc('·'.join(config.subjects))} 상담 장면은 실제 이용 후기나 특정 학생의 성과가 아닌 준비용 가상 예시입니다. 학습 결과는 학생의 출발점과 실천 정도에 따라 달라질 수 있습니다.</p></div></div><div class="scenario-grid">{scenario_html}</div></div></section>
    <section class="section"><div class="site-shell"><div class="section-heading"><div class="chapter-label"><span>03</span> FAQ</div><div><h2>{esc(locality)} {esc(config.level)} {esc('·'.join(config.subjects))} 자주 묻는 질문</h2><p>학생 자료, 가능 학년과 실제 이용 조건을 나누어 답했습니다.</p></div></div><div class="faq-list">{faq_html}</div></div></section>
    <section class="section local-links-section"><div class="site-shell"><div class="section-heading compact-heading"><div class="chapter-label"><span>04</span> Related</div><div><h2>{esc(locality)} 관련 학원·상담 페이지</h2><p>같은 지역의 과목 조합과 학년 카테고리, 앞뒤 지역 안내를 함께 확인하세요.</p></div></div><div class="local-navigation">{related_html}</div></div></section>
  </main>{footer("../../../")}<script src="../../../assets/site14.js" defer></script>
</body></html>'''


def grouped_records(records: list[dict[str, object]]) -> dict[str, dict[str, list[dict[str, object]]]]:
    grouped: dict[str, dict[str, list[dict[str, object]]]] = defaultdict(lambda: defaultdict(list))
    for record in records:
        center = record["center"]
        grouped[str(center["region"])][str(center["district"])].append(record)
    return grouped


def render_category_hub(config: CategoryConfig, records: list[dict[str, object]]) -> str:
    url = absolute_route(config.slug)
    items = [
        {
            "@type": "ListItem",
            "position": index,
            "name": record["title"],
            "url": absolute_route(config.slug, str(record["slug"])),
        }
        for index, record in enumerate(records, 1)
    ]
    graph = {
        "@context": "https://schema.org",
        "@graph": [
            {"@type": "EducationalOrganization", "@id": ORIGIN + "/#organization", "name": SITE_NAME, "url": ORIGIN + "/", "telephone": PHONE, "teaches": [f"{config.level} {subject}" for subject in config.subjects]},
            {"@type": "BreadcrumbList", "@id": url + "#breadcrumb", "itemListElement": [
                {"@type": "ListItem", "position": 1, "name": "홈", "item": ORIGIN + "/"},
                {"@type": "ListItem", "position": 2, "name": "과목별학원", "item": ORIGIN + quote("/과목별학원/", safe="/")},
                {"@type": "ListItem", "position": 3, "name": config.label, "item": url},
            ]},
            {"@type": "CollectionPage", "@id": url + "#collection", "url": url, "name": f"{config.label} 지역 안내", "description": f"371개 동네별 {config.label} 학습 답변과 센터·학교·가능 학년 확인 정보를 제공합니다.", "inLanguage": "ko-KR", "breadcrumb": {"@id": url + "#breadcrumb"}, "hasPart": [{"@type": "WebPage", "name": record["title"], "url": absolute_route(config.slug, str(record["slug"]))} for record in records]},
            {"@type": "ItemList", "@id": url + "#directory", "name": f"{config.label} 371개 지역", "numberOfItems": len(items), "itemListElement": items},
        ],
    }
    graph_json = json.dumps(graph, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    grouped = grouped_records(records)
    region_filters = "".join(f'<button type="button" data-region-filter="{esc(region)}">{esc(region)}</button>' for region in REGION_ORDER if region in grouped)
    region_html: list[str] = []
    for region in REGION_ORDER:
        if region not in grouped:
            continue
        districts = grouped[region]
        district_html: list[str] = []
        for district, values in districts.items():
            cards = "".join(
                f'<a class="directory-card" href="{esc(record["slug"])}/index.html" data-locality="{esc(record["center"]["locality"])} {esc(record["title"])} {esc(record["center"]["center_name"])}"><strong>{esc(record["center"]["locality"])}</strong><span>{esc(config.label)} · {esc(record["signals"][0].label)}</span><i aria-hidden="true">→</i></a>'
                for record in values
            )
            district_html.append(f'<section class="directory-district"><div class="directory-district-head"><h2>{esc(district)}</h2><span>{len(values)}개 지역</span></div><div class="directory-grid">{cards}</div></section>')
        region_html.append(f'<details class="directory-region" data-region="{esc(region)}" open><summary><span><b>{esc(region)}</b><small>{sum(len(value) for value in districts.values())}개 지역</small></span><i aria-hidden="true">＋</i></summary><div class="directory-region-body">{"".join(district_html)}</div></details>')
    description = f"371개 동네별 {config.label} 페이지에서 {config.focus}, {config.school_name} 자료, 가능 학년과 센터 상담 체크리스트를 확인하세요."
    return f'''<!doctype html><html lang="ko"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>{esc(config.label)} 지역 안내 | {TITLE_SUFFIX}</title><meta name="description" content="{esc(description)}"><meta name="robots" content="index,follow,max-image-preview:large"><meta name="theme-color" content="#f7f3ec"><link rel="canonical" href="{url}"><meta property="og:type" content="website"><meta property="og:locale" content="ko_KR"><meta property="og:title" content="{esc(config.label)} 지역 안내 | {TITLE_SUFFIX}"><meta property="og:description" content="{esc(description)}"><meta property="og:url" content="{url}"><link rel="icon" href="../../assets/favicon.png"><link rel="stylesheet" href="../../assets/site14.css"><script type="application/ld+json">{graph_json}</script></head>
<body data-page="subjects">{header("../../")}<main id="main"><section class="directory-hero"><div class="site-shell"><nav class="breadcrumbs" aria-label="현재 위치"><a href="../../index.html">홈</a><a href="../index.html">과목별학원</a><span>{esc(config.label)}</span></nav><p class="eyebrow">{esc(config.english)} DIRECTORY</p><h1>동네별 {esc(config.label)}</h1><p>{esc(config.focus)}을 중심으로 학생 자료를 확인하고, 지역별 센터·학교·가능 학년과 상담 전 질문을 한 흐름에서 비교하세요.</p><div class="hub-metrics"><div><strong>371</strong><span>지역 페이지</span></div><div><strong>{esc(config.focus)}</strong><span>핵심 진단</span></div><div><strong>{esc(config.process)}</strong><span>상담 흐름</span></div></div></div></section>
<section class="section directory-section"><div class="site-shell"><div class="directory-toolbar"><label for="local-search-{esc(config.slug)}">동네·센터·학교 검색</label><div class="directory-search"><input id="local-search-{esc(config.slug)}" type="search" placeholder="예: 명일동, 강동구, 명일점" data-local-search><span data-directory-count>371개 지역</span></div><div class="region-filters"><button type="button" class="is-active" data-region-filter="all">전체</button>{region_filters}</div><div class="directory-actions"><button type="button" data-expand-all>모두 펼치기</button></div></div><p class="directory-empty" data-directory-empty hidden>검색 조건에 맞는 지역이 없습니다.</p><div class="directory-list">{"".join(region_html)}</div></div></section>
<section class="section ink"><div class="site-shell consult-cta"><div><h2>{esc(config.label)} 상담은 실제 자료에서 시작하세요</h2><p>현재 교재와 오답 기록, 학교 일정과 가능한 학습 시간을 준비하면 설명·연습·재확인 기준을 구체적으로 비교할 수 있습니다.</p></div><a class="button orange" href="../../상담문의/index.html">상담 방법 확인 <span aria-hidden="true">→</span></a></div></section></main>{footer("../../")}<script src="../../assets/site14.js" defer></script></body></html>'''


def build_records(config: CategoryConfig, centers: list[dict[str, object]], source_rows: list[str], reps: list[str]) -> list[dict[str, object]]:
    if not (len(centers) == len(source_rows) == len(reps) == EXPECTED_ROWS):
        raise ValueError(f"{config.slug}: row mismatch")
    records: list[dict[str, object]] = []
    for index, (center, raw, rep_name) in enumerate(zip(centers, source_rows, reps), 1):
        seed = f"{config.slug}|{center['locality']}|{index}"
        signals = rank_signals(config, raw, seed)
        title = f"{center['locality']} {config.label}"
        english_profile: HighEnglishProfile | None = None
        middle_profile: MiddleEnglishProfile | None = None
        middle_math_record_profile: HighMathProfile | None = None
        math_profile: HighMathProfile | None = None
        if config.slug == "중등영어학원":
            middle_profile = middle_english_profile(raw, str(center["locality"]))
            middle_grades = relevant_grades(config, center, "영어")
            if not middle_grades and any(
                intent.code in {"entry", "transition"} for intent in middle_profile.intents
            ):
                neutral_intents: list[HighEnglishIntent] = [
                    intent
                    for intent in middle_profile.intents
                    if intent.code not in {"entry", "transition"}
                ]
                neutral_intents.extend(
                    MIDDLE_ENGLISH_INTENT_BY_CODE[code]
                    for code in ("diagnosis", "reading", "sentence", "grammar", "routine", "error", "vocabulary")
                    if MIDDLE_ENGLISH_INTENT_BY_CODE[code] not in neutral_intents
                )
                middle_profile = MiddleEnglishProfile(
                    focus=middle_profile.focus,
                    source_title=middle_profile.source_title,
                    intents=tuple(neutral_intents[:4]),
                    source_markers=tuple(intent.label for intent in neutral_intents[:4]),
                )
            unsupported_focus_grades = {
                grade
                for grade in re.findall(r"중[123]", middle_profile.focus)
                if grade not in middle_grades
            }
            if unsupported_focus_grades:
                safe_focus = middle_profile.focus
                for grade in sorted(unsupported_focus_grades):
                    safe_focus = safe_focus.replace(grade, "중등 과정")
                middle_profile = MiddleEnglishProfile(
                    focus=clean(safe_focus.replace("중등 과정 중등 과정", "중등 과정")),
                    source_title=middle_profile.source_title,
                    intents=middle_profile.intents,
                    source_markers=middle_profile.source_markers,
                )
            used_focuses = {
                str(existing["middle_english_profile"].focus)
                for existing in records
                if isinstance(existing.get("middle_english_profile"), MiddleEnglishProfile)
            }
            if middle_profile.focus in used_focuses:
                primary, secondary, support, extra = middle_profile.intents
                focus_candidates = (
                    f"{primary.label}에서 {secondary.label}으로 이어지는 영어 학습 순서",
                    f"{primary.label}과 {support.label}을 {secondary.label}로 확인하는 방법",
                    f"{secondary.label} 기록으로 {primary.label}을 다시 점검하는 방법",
                    f"{primary.label}·{secondary.label} 뒤 {extra.label}을 확인하는 학습 기준",
                )
                focus_tokens = tuple(intent.label for intent in middle_profile.intents)
                normalized_candidates = tuple(
                    normalize_particle_joins(value, focus_tokens) for value in focus_candidates
                )
                replacement = next(
                    (value for value in normalized_candidates if value not in used_focuses),
                    None,
                )
                if replacement is None:
                    raise ValueError(f"unable to create unique middle English focus: {title}")
                middle_profile = MiddleEnglishProfile(
                    focus=replacement,
                    source_title=middle_profile.source_title,
                    intents=middle_profile.intents,
                    source_markers=middle_profile.source_markers,
                )
            student = middle_english_student_type(config, center, middle_profile, seed)
            profile_tokens = list(middle_english_particle_tokens(middle_profile))
            if middle_grades:
                profile_tokens.append("·".join(middle_grades))
            tokens = (*page_particle_tokens(config, center, signals, student), *profile_tokens)
            student = normalize_particle_joins(student, tokens)
            sections = normalize_generated_tree(
                middle_english_sections(config, center, middle_profile, seed), tokens
            )
            faq = normalize_generated_tree(
                middle_english_faq(config, center, middle_profile, seed), tokens
            )
            scenarios = normalize_generated_tree(
                middle_english_scenarios(config, center, middle_profile, seed), tokens
            )
            meta = normalize_particle_joins(
                middle_english_meta_description(config, center, middle_profile), tokens
            )
            quick = normalize_particle_joins(
                middle_english_quick_answer(config, center, middle_profile, seed), tokens
            )
        elif config.slug == "고등영어학원":
            english_profile = high_english_profile(raw, str(center["locality"]))
            student = high_english_student_type(config, center, english_profile, seed)
            profile_tokens = list(high_english_particle_tokens(english_profile))
            high_english_grades = relevant_grades(config, center, "영어")
            if high_english_grades:
                profile_tokens.append("·".join(high_english_grades))
            tokens = (*page_particle_tokens(config, center, signals, student), *profile_tokens)
            student = normalize_particle_joins(student, tokens)
            sections = normalize_generated_tree(high_english_sections(config, center, english_profile, seed), tokens)
            faq = normalize_generated_tree(high_english_faq(config, center, english_profile, seed), tokens)
            scenarios = normalize_generated_tree(high_english_scenarios(config, center, english_profile, seed), tokens)
            meta = normalize_particle_joins(high_english_meta_description(config, center, english_profile), tokens)
            quick = normalize_particle_joins(high_english_quick_answer(config, center, english_profile, seed), tokens)
        elif config.slug == "중등수학학원":
            middle_math_record_profile = middle_math_profile(raw, str(center["locality"]))
            middle_math_grades = relevant_grades(config, center, "수학")
            if not middle_math_grades and any(
                intent.code in {"entry", "transition"}
                for intent in middle_math_record_profile.intents
            ):
                neutral_intents: list[HighMathIntent] = [
                    intent
                    for intent in middle_math_record_profile.intents
                    if intent.code not in {"entry", "transition"}
                ]
                neutral_intents.extend(
                    MIDDLE_MATH_INTENT_BY_CODE[code]
                    for code in (
                        "condition", "error", "school_exam", "units",
                        "concept", "routine", "calculation", "written", "graph",
                    )
                    if MIDDLE_MATH_INTENT_BY_CODE[code] not in neutral_intents
                )
                middle_math_record_profile = HighMathProfile(
                    focus=middle_math_record_profile.focus,
                    source_title=middle_math_record_profile.source_title,
                    intents=tuple(neutral_intents[:4]),
                    source_markers=tuple(intent.label for intent in neutral_intents[:4]),
                )
            unsupported_focus_grades = {
                grade
                for grade in re.findall(r"중[123]", middle_math_record_profile.focus)
                if grade not in middle_math_grades
            }
            if unsupported_focus_grades:
                safe_focus = middle_math_record_profile.focus
                for grade in sorted(unsupported_focus_grades):
                    safe_focus = safe_focus.replace(grade, "중등 과정")
                safe_focus = clean(safe_focus.replace("중등 과정 중등 과정", "중등 과정"))
                middle_math_record_profile = HighMathProfile(
                    focus=safe_focus,
                    source_title=middle_math_record_profile.source_title,
                    intents=middle_math_record_profile.intents,
                    source_markers=middle_math_record_profile.source_markers,
                )
            used_focuses = {
                str(existing["middle_math_profile"].focus)
                for existing in records
                if isinstance(existing.get("middle_math_profile"), HighMathProfile)
            }
            if middle_math_record_profile.focus in used_focuses:
                primary, secondary, support, extra = middle_math_record_profile.intents
                focus_candidates = (
                    f"{primary.label}에서 {secondary.label}으로 이어지는 중등 수학 학습 순서",
                    f"{primary.label}{particle_for(primary.label, '과', '와')} {support.label}을 {secondary.label}로 확인하는 방법",
                    f"{secondary.label} 기록으로 {primary.label}을 다시 점검하는 방법",
                    f"{primary.label}{particle_for(primary.label, '과', '와')} {secondary.label} 뒤 {extra.label}을 확인하는 학습 기준",
                    f"{center['locality']} 수학 자료로 {primary.label}과 {secondary.label}을 구분하는 방법",
                )
                focus_tokens = tuple(intent.label for intent in middle_math_record_profile.intents)
                normalized_candidates = tuple(
                    repair_middle_math_focus(
                        naturalize_middle_math_text(
                            normalize_particle_joins(value, focus_tokens),
                            middle_math_record_profile,
                        )
                    )
                    for value in focus_candidates
                )
                replacement = next(
                    (value for value in normalized_candidates if value not in used_focuses),
                    None,
                )
                if replacement is None:
                    raise ValueError(f"unable to create unique middle Math focus: {title}")
                middle_math_record_profile = HighMathProfile(
                    focus=replacement,
                    source_title=middle_math_record_profile.source_title,
                    intents=middle_math_record_profile.intents,
                    source_markers=middle_math_record_profile.source_markers,
                )
            student = middle_math_student_type(
                config, center, middle_math_record_profile, seed
            )
            profile_tokens = list(middle_math_particle_tokens(middle_math_record_profile))
            if middle_math_grades:
                profile_tokens.append("·".join(middle_math_grades))
            tokens = (*page_particle_tokens(config, center, signals, student), *profile_tokens)
            student = normalize_particle_joins(student, tokens)
            sections = normalize_generated_tree(
                middle_math_sections(config, center, middle_math_record_profile, seed),
                tokens,
            )
            faq = normalize_generated_tree(
                middle_math_faq(config, center, middle_math_record_profile, seed),
                tokens,
            )
            scenarios = normalize_generated_tree(
                middle_math_scenarios(config, center, middle_math_record_profile, seed),
                tokens,
            )
            meta = normalize_particle_joins(
                middle_math_meta_description(config, center, middle_math_record_profile),
                tokens,
            )
            quick = normalize_particle_joins(
                middle_math_quick_answer(config, center, middle_math_record_profile, seed),
                tokens,
            )
        elif config.slug == "고등수학학원":
            math_profile = high_math_profile(raw, str(center["locality"]))
            high_math_grades = relevant_grades(config, center, "수학")
            unsupported_focus_grades = {
                grade
                for grade in re.findall(r"고[123]", math_profile.focus)
                if grade not in high_math_grades
            }
            if unsupported_focus_grades:
                safe_focus = math_profile.focus
                for grade in sorted(unsupported_focus_grades):
                    safe_focus = safe_focus.replace(grade, "고등 과정")
                safe_focus = clean(safe_focus.replace("고등 과정 고등 과정", "고등 과정"))
                math_profile = HighMathProfile(
                    focus=safe_focus,
                    source_title=math_profile.source_title,
                    intents=math_profile.intents,
                    source_markers=math_profile.source_markers,
                )
            used_focuses = {
                str(existing["high_math_profile"].focus)
                for existing in records
                if isinstance(existing.get("high_math_profile"), HighMathProfile)
            }
            if math_profile.focus in used_focuses:
                primary, secondary, support, extra = math_profile.intents
                focus_candidates = (
                    f"{primary.label}에서 {secondary.label}으로 이어지는 수학 학습 순서",
                    f"{primary.label}과 {support.label}을 {secondary.label}로 확인하는 방법",
                    f"{secondary.label} 기록으로 {primary.label}을 다시 점검하는 방법",
                    f"{primary.label}·{secondary.label} 뒤 {extra.label}을 확인하는 학습 기준",
                )
                focus_tokens = tuple(intent.label for intent in math_profile.intents)
                normalized_candidates = tuple(normalize_particle_joins(value, focus_tokens) for value in focus_candidates)
                replacement = next((value for value in normalized_candidates if value not in used_focuses), None)
                if replacement is None:
                    raise ValueError(f"unable to create unique high Math focus: {title}")
                math_profile = HighMathProfile(
                    focus=replacement,
                    source_title=math_profile.source_title,
                    intents=math_profile.intents,
                    source_markers=math_profile.source_markers,
                )
            student = high_math_student_type(config, center, math_profile, seed)
            profile_tokens = list(high_math_particle_tokens(math_profile))
            if high_math_grades:
                profile_tokens.append("·".join(high_math_grades))
            tokens = (*page_particle_tokens(config, center, signals, student), *profile_tokens)
            student = normalize_particle_joins(student, tokens)
            sections = normalize_generated_tree(high_math_sections(config, center, math_profile, seed), tokens)
            faq = normalize_generated_tree(high_math_faq(config, center, math_profile, seed), tokens)
            scenarios = normalize_generated_tree(high_math_scenarios(config, center, math_profile, seed), tokens)
            meta = normalize_particle_joins(high_math_meta_description(config, center, math_profile), tokens)
            quick = normalize_particle_joins(high_math_quick_answer(config, center, math_profile, seed), tokens)
        else:
            student = student_type(config, signals, seed, str(center["locality"]))
            tokens = page_particle_tokens(config, center, signals, student)
            student = normalize_particle_joins(student, tokens)
            sections = normalize_generated_value(build_sections(config, center, signals, seed), tokens)
            faq = normalize_generated_value(build_faq(config, center, signals, seed), tokens)
            scenarios = normalize_generated_value(build_scenarios(config, center, signals, seed), tokens)
            meta = normalize_particle_joins(meta_description(config, center, signals), tokens)
            quick = normalize_particle_joins(quick_answer(config, center, signals, seed), tokens)
        record: dict[str, object] = {
            "sequence": index,
            "center": center,
            "title": title,
            "slug": center["slug"],
            "rep_name": rep_name,
            "signals": signals,
            "student": student,
            "meta": meta,
            "quick": quick,
            "sections": sections,
            "faq": faq,
            "scenarios": scenarios,
        }
        if english_profile is not None:
            record["high_english_profile"] = english_profile
        if middle_profile is not None:
            record["middle_english_profile"] = middle_profile
        if middle_math_record_profile is not None:
            record["middle_math_profile"] = middle_math_record_profile
        if math_profile is not None:
            record["high_math_profile"] = math_profile
        records.append(record)
    return records


def high_english_collision_tail(
    record: dict[str, object],
    slot: int,
    attempt: int = 0,
) -> str:
    """Add a purpose-matched follow-up when a useful block would repeat.

    Selection depends on the page and slot, not just the section position.
    This keeps the added sentence informative while preventing a single fixed
    sentence from spreading across the large reading/sentence intent cluster.
    """

    profile = record["high_english_profile"]
    primary, secondary, support, extra = profile.intents
    seed = f"{record['title']}|{record['sequence']}|{slot}|collision-tail"
    prefixes = (
        "다음 판단을 위해,",
        "첫 주 기록을 남길 때,",
        "학생 설명을 확인한 뒤,",
        "계획을 조정하기 전에,",
        "자료를 다시 펼쳤을 때,",
        "상담 메모를 정리하면서,",
        "학교 일정과 대조할 때,",
        "가정에서 확인할 때,",
    )

    direct_tails = (
        f"첫 판단 뒤에는 ‘{primary.evidence}’의 표시와 ‘{secondary.checkpoint}’에 대한 학생 설명이 일치하는지 대조하세요.",
        f"진단 메모에는 ‘{primary.concern}’의 답과 ‘{secondary.evidence}’가 확인된 자료 위치를 함께 적으세요.",
        f"두 원인을 나눈 다음 ‘{primary.action}’을 먼저 할지 ‘{secondary.action}’을 먼저 할지 한 가지만 선택하세요.",
        f"학생의 말과 자료가 다르면 ‘{primary.checkpoint}’에 다시 답할 날짜를 정한 뒤 진도를 결정하세요.",
        f"첫 행동을 정할 때는 ‘{primary.evidence}’가 남은 자료와 ‘{secondary.concern}’에 대한 학생의 답을 나란히 보세요.",
        f"설명할 영역은 ‘{primary.checkpoint}’, 혼자 연습할 영역은 ‘{secondary.checkpoint}’로 구분해 판단하세요.",
        f"정답 수가 같아도 ‘{primary.concern}’과 ‘{secondary.concern}’의 답이 다르면 과제의 순서를 나눠야 합니다.",
        f"이번 점검에서는 ‘{primary.action}’의 수행 여부와 ‘{secondary.checkpoint}’의 답을 같은 날짜에 기록하세요.",
    )
    evidence_tails = (
        f"재확인표에는 ‘{primary.evidence}’가 달라진 위치와 ‘{secondary.checkpoint}’를 확인할 날짜를 함께 적으세요.",
        f"첫 자료에는 ‘{primary.evidence}’를, 다음 자료에는 ‘{secondary.evidence}’를 표시해 판단 변화가 보이게 하세요.",
        f"기록 옆에 ‘{primary.checkpoint}’의 학생 답과 ‘{secondary.action}’을 수행한 날짜를 나란히 남기세요.",
        f"자료를 다시 볼 때는 ‘{primary.concern}’의 답이 ‘{secondary.evidence}’와 같은 원인을 가리키는지 확인하세요.",
        f"도움을 받기 전에는 ‘{primary.evidence}’, 도움을 받은 뒤에는 ‘{secondary.checkpoint}’를 기준으로 차이를 적으세요.",
        f"표의 시작 칸에는 ‘{primary.evidence}’를, 마지막 칸에는 ‘{secondary.checkpoint}’의 재확인 결과를 넣으세요.",
        f"다음 점검에서는 ‘{primary.action}’을 한 흔적과 ‘{secondary.evidence}’에서 바뀐 근거만 비교하세요.",
        f"학생이 직접 남긴 ‘{primary.evidence}’와 ‘{secondary.evidence}’에는 각각 자료 위치와 날짜를 붙이세요.",
    )
    exam_tails = (
        f"달력에는 ‘{primary.exam_use}’를 적용할 자료와 ‘{secondary.exam_use}’를 확인할 날짜를 따로 적으세요.",
        f"시험 뒤에는 ‘{primary.checkpoint}’와 ‘{secondary.checkpoint}’에 대한 답을 자료별로 다시 비교하세요.",
        f"내신과 모의고사 어느 쪽이든 ‘{primary.action}’과 ‘{secondary.action}’ 중 필요한 최소 행동만 배치하세요.",
        f"학교 범위의 마감일과 누적 학습의 재확인일에는 {primary.label}·{secondary.label} 기록을 모두 연결하세요.",
        f"시험별 분량을 정하기 전 ‘{primary.exam_use}’와 ‘{secondary.exam_use}’가 실제 자료에 맞는지 확인하세요.",
        f"한 시험에 한 영역을 고정하지 말고 {primary.label}과 {secondary.label}의 흔적을 두 자료에서 각각 찾아보세요.",
        f"시험 일정이 겹치면 ‘{primary.checkpoint}’에 필요한 행동과 ‘{secondary.checkpoint}’의 최소 복습을 다른 날에 둡니다.",
        f"다음 주 비중은 ‘{primary.evidence}’와 ‘{secondary.evidence}’에서 실제로 끝낸 양을 보고 조정하세요.",
    )
    fact_tails = (
        "참고 학교명은 실제 개설을 보장하지 않으므로 학생의 범위표와 현재 센터 답변을 함께 대조하세요.",
        "가능 학년 표기와 실제 시간표는 확인 시점이 다를 수 있어 각각 확인한 날짜를 남겨야 합니다.",
        "센터 주소·교습비·시간표는 학습 판단과 분리해 사실 확인 목록에 적으세요.",
        "학교 목록보다 학생이 가져온 시험 범위표가 자료 반영 여부를 묻는 우선 근거입니다.",
        "등록 전에는 제공 주소의 통학 동선과 현재 열려 있는 고등 영어 시간을 함께 확인하세요.",
        "공개된 센터 정보와 상담에서 새로 확인한 조건은 출처와 확인일을 나누어 기록하세요.",
        "학년·학교·시간표는 한 문장으로 묶지 말고 표기된 정보와 현재 답변을 다른 줄에 두세요.",
        "교습비 자료가 연결돼 있어도 실제 과정과 기준 시점이 같은지는 등록 전에 다시 물어보세요.",
    )
    plan_tails = (
        f"첫 주가 끝나면 ‘{support.evidence}’와 ‘{extra.checkpoint}’를 보고 다음 분량을 유지할지 정하세요.",
        f"계획표에는 ‘{support.action}’의 완료일과 ‘{extra.action}’을 시작할 조건을 다른 칸에 적으세요.",
        f"미완료가 생기면 ‘{support.concern}’의 답을 확인한 뒤 시간과 분량 중 한 가지만 바꾸세요.",
        f"다음 점검일에는 ‘{support.checkpoint}’와 ‘{extra.checkpoint}’에 학생이 직접 답하게 하세요.",
        f"일주일 기록에서 ‘{support.evidence}’가 바뀐 지점과 ‘{extra.evidence}’가 남은 지점을 구분하세요.",
        f"첫 행동은 ‘{support.action}’으로 제한하고 완료 뒤 ‘{extra.action}’을 다음 주에 넣을지 판단하세요.",
        f"계획을 수정한 이유는 ‘{support.checkpoint}’의 결과와 ‘{extra.concern}’에 대한 학생 답으로 남기세요.",
        f"완료량보다 ‘{support.checkpoint}’와 ‘{extra.checkpoint}’의 변화가 보이는 자료를 다음 계획에 사용하세요.",
    )
    consult_tails = (
        f"상담 메모에는 ‘{primary.consult_question}’의 답과 이를 다시 확인할 자료를 함께 적으세요.",
        f"답변을 들은 뒤 ‘{secondary.action}’의 시작일과 ‘{primary.checkpoint}’의 재확인일을 정하세요.",
        f"가져갈 자료에는 ‘{primary.evidence}’가 보이는 위치와 ‘{secondary.concern}’을 묻는 질문을 표시하세요.",
        f"상담이 끝나면 ‘{primary.action}’을 할 학생 일정과 ‘{secondary.checkpoint}’를 확인할 날짜를 대조하세요.",
        f"학습 질문에는 ‘{secondary.consult_question}’를, 이용 조건에는 시간표·교습비 확인일을 따로 남기세요.",
        f"추상적인 설명은 ‘{primary.evidence}’에 어떻게 적용할지 물어 실제 행동과 연결하세요.",
        f"첫 주에 가져올 기록은 ‘{primary.checkpoint}’의 답과 ‘{secondary.evidence}’의 변화가 보이는 자료입니다.",
        f"상담 뒤 학생에게 ‘{primary.action}’과 ‘{secondary.action}’의 순서를 자기 말로 설명하게 하세요.",
    )

    paragraph_banks = (
        direct_tails,
        evidence_tails,
        exam_tails,
        fact_tails,
        plan_tails,
        consult_tails,
    )
    if slot < 20:
        bank = paragraph_banks[min(slot // 2, len(paragraph_banks) - 1)]
        base = bank[(stable_int(seed, "paragraph-tail") + attempt) % len(bank)]
    else:
        faq_index = slot - 20
        faq_banks = (
            direct_tails,
            evidence_tails,
            exam_tails,
            (
                f"가져간 자료에는 ‘{support.evidence}’와 ‘{extra.evidence}’의 표시 위치를 각각 남기세요.",
                f"답변 뒤에는 ‘{support.checkpoint}’와 ‘{extra.checkpoint}’를 다시 확인할 날짜도 정하세요.",
                f"상담 메모에서 ‘{support.action}’과 ‘{extra.action}’의 시작 조건을 다른 줄에 적으세요.",
                f"자료를 고를 때는 {support.label}의 현재선과 {extra.label}의 재확인 흔적이 모두 보이는지 확인하세요.",
                f"첫 질문은 ‘{support.consult_question}’, 다음 질문은 ‘{extra.consult_question}’로 나누어 적으세요.",
                f"학생 자료에서 ‘{support.evidence}’와 ‘{extra.evidence}’가 확인되는 부분만 골라 가져가세요.",
            ),
            fact_tails,
        )
        bank = faq_banks[min(faq_index, len(faq_banks) - 1)]
        base = bank[(stable_int(seed, "faq-tail") + attempt) % len(bank)]
    prefix = prefixes[
        (
            stable_int(seed, "tail-prefix")
            + attempt // len(bank)
        )
        % len(prefixes)
    ]
    tail = f"{prefix} {base}"
    return normalize_particle_joins(tail, high_english_particle_tokens(profile))


def contextualize_duplicate_paragraphs(records_by_category: dict[str, list[dict[str, object]]]) -> None:
    """Add page context only where a generated paragraph would otherwise repeat."""
    occurrences: dict[str, list[tuple[str, dict[str, object], dict[str, object], int]]] = defaultdict(list)
    occupied: set[str] = set()
    used_contexts: dict[tuple[str, str], set[str]] = defaultdict(set)
    for slug, records in records_by_category.items():
        for record in records:
            for section in record["sections"]:
                for paragraph_index, paragraph in enumerate(section["paragraphs"]):
                    value = clean(paragraph)
                    occupied.add(value)
                    occurrences[value].append((slug, record, section, paragraph_index))

    for paragraph, references in occurrences.items():
        if len(references) < 2:
            continue
        for slug, record, section, paragraph_index in references:
            config = CONFIG_BY_SLUG[slug]
            center = record["center"]
            locality = str(center["locality"])
            if slug in {"중등영어학원", "중등수학학원", "고등영어학원", "고등수학학원"}:
                # High-English pages use purpose-written section combinations.
                # Rewriting every shared advisory paragraph with a locality
                # prefix made otherwise useful Korean read like a template.
                # Corpus QA therefore evaluates whole-section/document
                # uniqueness and bounded shared blocks instead.
                continue
            subject_text = "·".join(config.subjects)
            contexts = (
                f"이 기준은 {locality}에서 {config.level} {subject_text} 상담을 준비할 때 학생 자료와 함께 확인합니다.",
                f"{locality} {config.level} {subject_text} 계획에서는 이 내용을 실제 학습 기록과 대조합니다.",
                f"{locality} 학부모가 {config.level} {subject_text} 상담을 준비한다면 이 항목도 따로 메모해 두세요.",
                f"{locality} 학생의 {config.level} {subject_text} 흐름에 적용할 때는 현재 교재와 실행 기록을 함께 살핍니다.",
                f"이 내용은 {locality} {config.level} {subject_text} 수업 조건을 비교할 때 확인할 질문으로 사용할 수 있습니다.",
                f"{locality}에서 {config.level} {subject_text} 학습을 알아볼 때도 확인된 자료와 상담 질문을 구분합니다.",
                f"{locality}의 {config.level} {subject_text} 학습에서는 이 기준을 다음 점검일과 연결해 살핍니다.",
                f"상담에는 {locality} 학생의 {config.level} {subject_text} 기록도 함께 가져가세요.",
                f"이 확인 순서는 {locality} {config.level} {subject_text} 학습 자료를 비교할 때 사용할 수 있습니다.",
                f"{locality} {config.level} {subject_text} 상담 메모에서는 이 내용과 실제 이용 조건을 나누어 적습니다.",
                f"현재 자료를 비교하는 {locality} {config.level} {subject_text} 과정에서도 이 항목을 별도로 확인합니다.",
                f"이 질문은 {locality}에서 {config.level} {subject_text} 계획을 정하기 전에 학생과 함께 살펴봅니다.",
            )
            context_seed = f"{slug}|{locality}|{section['key']}|{paragraph_index}|paragraph-context"
            start = stable_int(context_seed, "context") % len(contexts)
            candidate = ""
            page_key = (slug, str(record["slug"]))
            for offset in range(len(contexts)):
                context = contexts[(start + offset) % len(contexts)]
                candidate = f"{paragraph} {context}"
                if context not in used_contexts[page_key] and candidate not in occupied:
                    used_contexts[page_key].add(context)
                    break
            else:
                raise ValueError(f"unable to contextualize duplicate paragraph: {record['title']}")
            section["paragraphs"][paragraph_index] = candidate
            occupied.add(candidate)


def contextualize_duplicate_faq_answers(records_by_category: dict[str, list[dict[str, object]]]) -> None:
    """Keep FAQ answers distinct without repeating the locality on every answer."""
    occurrences: dict[str, list[tuple[str, dict[str, object], int]]] = defaultdict(list)
    occupied: set[str] = set()
    used_contexts: dict[tuple[str, str], set[str]] = defaultdict(set)
    for slug, records in records_by_category.items():
        for record in records:
            for faq_index, (_question, answer) in enumerate(record["faq"]):
                value = clean(answer)
                occupied.add(value)
                occurrences[value].append((slug, record, faq_index))

    for answer, references in occurrences.items():
        if len(references) < 2:
            continue
        for slug, record, faq_index in references:
            config = CONFIG_BY_SLUG[slug]
            center = record["center"]
            locality = str(center["locality"])
            if slug in {"중등영어학원", "중등수학학원", "고등영어학원", "고등수학학원"}:
                # Keep the concise authored answer.  Page-level FAQ sets remain
                # unique and repeated answers are bounded by corpus QA.
                continue
            subject_text = "·".join(config.subjects)
            contexts = (
                f"{locality}의 {config.level} {subject_text} 상담에서는 학생 자료와 센터 조건을 함께 확인합니다.",
                f"이 답변은 {locality}에서 {config.level} {subject_text} 상담을 준비할 때 적용할 수 있습니다.",
                f"{locality} {config.level} {subject_text} 학습에서는 실제 자료를 확인한 뒤 판단해야 합니다.",
                f"{locality} 학부모는 {config.level} {subject_text} 질문과 이용 조건을 나누어 기록해 두세요.",
                f"{center['center_name']}의 {locality} {config.level} {subject_text} 운영 여부는 상담 시점에 다시 확인합니다.",
            )
            seed = f"{slug}|{locality}|{faq_index}|faq-context"
            start = stable_int(seed, "context") % len(contexts)
            candidate = ""
            page_key = (slug, str(record["slug"]))
            for offset in range(len(contexts)):
                context = contexts[(start + offset) % len(contexts)]
                candidate = f"{answer} {context}"
                if context not in used_contexts[page_key] and candidate not in occupied:
                    used_contexts[page_key].add(context)
                    break
            else:
                raise ValueError(f"unable to contextualize duplicate FAQ answer: {record['title']}")
            question, _old_answer = record["faq"][faq_index]
            record["faq"][faq_index] = (question, candidate)
            occupied.add(candidate)


def individualize_frequent_middle_english_paragraphs(
    records_by_category: dict[str, list[dict[str, object]]],
) -> None:
    """Add one decision sentence only to paragraphs shared above the corpus limit."""

    records = records_by_category.get("중등영어학원", [])
    occurrences: dict[str, list[tuple[dict[str, object], dict[str, object], int]]] = defaultdict(list)
    for record in records:
        for section in record["sections"]:
            for paragraph_index, paragraph in enumerate(section["paragraphs"]):
                occurrences[clean(paragraph)].append((record, section, paragraph_index))

    for references in occurrences.values():
        if len(references) <= HIGH_ENGLISH_PARAGRAPH_DF_LIMIT:
            continue
        for record, section, paragraph_index in references:
            profile = record["middle_english_profile"]
            primary, secondary, support, extra = profile.intents
            locality = str(record["center"]["locality"])
            key = str(section["key"])
            notes = {
                "direct-answer": f"{locality}에서는 ‘{profile.focus}’을 {primary.label}의 자료와 {secondary.label}의 재확인 결과로 구체화하세요.",
                "diagnostic-evidence": f"{locality} 기록에는 ‘{primary.evidence}’가 보인 위치와 ‘{secondary.checkpoint}’의 확인 날짜를 함께 남기세요.",
                "exam-strategy": f"{locality} 시험표에는 {primary.label}의 학교 범위 마감일과 {secondary.label}의 누적 학습 확인일을 나누어 적으세요.",
                "center-facts": f"{locality}의 학습 판단과 센터 이용 조건은 같은 문장에 섞지 말고 확인한 자료와 날짜를 각각 남기세요.",
                "four-week-plan": f"{locality} 계획에서는 ‘{support.action}’의 완료일과 ‘{extra.checkpoint}’의 재확인일을 다른 칸에 두세요.",
                "consultation-checklist": f"{locality} 상담 뒤에는 {primary.label}의 첫 행동과 {secondary.label}의 확인 기준을 학생이 직접 설명하게 하세요.",
            }
            note = notes.get(
                key,
                f"{locality}에서는 {primary.label}의 실행 기록과 {secondary.label}의 다음 확인일을 함께 남기세요.",
            )
            old = str(section["paragraphs"][paragraph_index])
            section["paragraphs"][paragraph_index] = clean(f"{old} {note}")


def individualize_frequent_middle_english_faq_answers(
    records_by_category: dict[str, list[dict[str, object]]],
) -> None:
    """Keep FAQ answers under the same exact-paragraph corpus limit."""

    records = records_by_category.get("중등영어학원", [])
    occurrences: dict[str, list[tuple[dict[str, object], int]]] = defaultdict(list)
    for record in records:
        for faq_index, (_question, answer) in enumerate(record["faq"]):
            occurrences[clean(answer)].append((record, faq_index))
    for references in occurrences.values():
        if len(references) <= HIGH_ENGLISH_PARAGRAPH_DF_LIMIT:
            continue
        for record, faq_index in references:
            profile = record["middle_english_profile"]
            primary, secondary = profile.intents[:2]
            locality = str(record["center"]["locality"])
            question, answer = record["faq"][faq_index]
            notes = (
                f"{locality}에서는 {primary.label}의 자료 위치와 {secondary.label}의 재확인 날짜도 함께 적으세요.",
                f"{locality} 학생의 첫 기록에는 {primary.label}의 근거와 {secondary.label}의 다음 행동을 나누어 남기세요.",
                f"{locality} 상담 메모에는 {primary.label}을 확인한 자료와 {secondary.label}을 다시 볼 시점을 표시하세요.",
            )
            note = stable_pick(
                f"{record['title']}|{faq_index}",
                "middle-faq-frequency-note",
                notes,
            )
            question_contexts = (
                "첫 자료 질문:",
                "기록 구분 질문:",
                "일정 배분 질문:",
                "상담 준비 질문:",
                "센터 확인 질문:",
            )
            note = f"{question_contexts[faq_index % len(question_contexts)]} {note}"
            record["faq"][faq_index] = (question, clean(f"{answer} {note}"))


def individualize_duplicate_middle_english_sections(
    records_by_category: dict[str, list[dict[str, object]]],
) -> None:
    """Resolve whole-section collisions with a useful middle-English decision note."""

    records = records_by_category.get("중등영어학원", [])
    occurrences: dict[str, list[tuple[dict[str, object], dict[str, object]]]] = defaultdict(list)
    for record in records:
        for section in record["sections"]:
            section_text = clean(" ".join([str(section["heading"]), *section["paragraphs"]]))
            occurrences[section_text].append((record, section))

    for references in occurrences.values():
        if len(references) < 2:
            continue
        for record, section in references:
            center = record["center"]
            profile = record["middle_english_profile"]
            primary, secondary = profile.intents[:2]
            locality = str(center["locality"])
            key = str(section["key"])
            notes = {
                "diagnostic-evidence": (
                    f"{locality}에서는 {primary.label}에서 혼자 설명한 부분과 {secondary.label}에서 도움이 필요했던 부분을 구분해 다음 분량의 근거로 사용하세요.",
                    f"{locality} 점검표에는 {primary.label}의 수정 이유와 {secondary.label}의 새 문제 확인 결과를 따로 적어 병목을 구분하세요.",
                    f"{locality} 학생의 기록은 {primary.label}의 첫 판단과 {secondary.label}의 재확인 답을 다른 칸에 남겨 다음 학습 순서를 정하세요.",
                ),
                "exam-strategy": (
                    f"{locality} 학생의 학교 일정표에는 시험 범위 마감일과 새 지문 재확인일을 따로 적고 {primary.label}의 최소 행동이 빠지지 않았는지 확인하세요.",
                    f"{locality} 시험 계획에서는 학교 범위 완료일과 누적 독해 확인일을 나눈 뒤 {secondary.label} 기록이 어느 쪽에서 흔들렸는지 표시하세요.",
                    f"{locality} 학생이라면 시험 뒤 {primary.label}의 근거와 {secondary.label}의 소요 시간을 비교해 다음 주 비중을 정하세요.",
                ),
                "four-week-plan": (
                    f"{locality} 계획표에는 ‘{primary.action}’의 시작일과 ‘{secondary.checkpoint}’의 재확인일을 나란히 적어 실행과 판단을 연결하세요.",
                    f"{locality} 학생의 첫 주 기록에서는 ‘{primary.evidence}’의 변화와 ‘{secondary.action}’의 완료 여부를 다른 칸에 남기세요.",
                    f"{locality} 계획은 문제 수보다 ‘{primary.checkpoint}’에 답한 날과 ‘{secondary.checkpoint}’를 다시 볼 날이 이어지는지로 점검하세요.",
                ),
                "consultation-checklist": (
                    f"{locality} 상담 메모에는 ‘{primary.action}’의 실행 날짜와 ‘{secondary.checkpoint}’의 재확인 날짜를 함께 적으세요.",
                    f"{locality}에서는 상담 답변을 최근 시험지의 실제 표시와 연결해 설명할 수 있는지까지 확인하세요.",
                    f"{locality} 상담 뒤에는 {primary.label}의 첫 행동과 {secondary.label}의 확인일을 학생이 직접 설명하게 하세요.",
                ),
            }.get(
                key,
                (
                    f"{locality} 학생의 최근 자료에서는 ‘{primary.evidence}’의 위치와 ‘{secondary.checkpoint}’의 재확인일을 함께 남기세요.",
                ),
            )
            note = stable_pick(f"{record['title']}|{key}", "middle-section-note", notes)
            section["paragraphs"][-1] = clean(f"{section['paragraphs'][-1]} {note}")


def individualize_duplicate_high_english_sections(
    records_by_category: dict[str, list[dict[str, object]]],
) -> None:
    """Resolve only whole-section collisions with one useful local decision note."""

    records = records_by_category.get("고등영어학원", [])
    occurrences: dict[str, list[tuple[dict[str, object], dict[str, object]]]] = defaultdict(list)
    for record in records:
        for section in record["sections"]:
            section_text = clean(" ".join([str(section["heading"]), *section["paragraphs"]]))
            occurrences[section_text].append((record, section))

    for references in occurrences.values():
        if len(references) < 2:
            continue
        for record, section in references:
            center = record["center"]
            profile = record["high_english_profile"]
            primary, secondary = profile.intents[:2]
            locality = str(center["locality"])
            key = str(section["key"])
            notes = {
                "diagnostic-evidence": (
                    f"{locality}에서는 비교 결과를 {primary.label}의 설명 단계와 {secondary.label}의 새 문제 확인 단계로 나누어 다음 계획표에 남기세요.",
                    f"{locality} 학생의 기록은 {primary.label}에서 혼자 수행한 부분과 {secondary.label}에서 도움이 필요했던 부분을 구분해 다음 분량의 근거로 사용하세요.",
                    f"{locality} 점검표에는 {primary.label}의 수정 이유와 {secondary.label}의 다음 문제를 따로 적어 같은 자료를 반복해서 설명하지 않도록 합니다.",
                ),
                "exam-strategy": (
                    f"{locality} 학생의 실제 학교 일정표에는 내신 마감일과 모의고사 재풀이 날짜를 따로 적고, {primary.label}의 최소 행동이 빠지지 않았는지 확인하세요.",
                    f"{locality} 시험 계획을 검토할 때는 학교 범위의 완료일과 새 지문의 재확인일을 나눈 뒤 {secondary.label} 기록이 어느 쪽에서 흔들렸는지 표시하세요.",
                    f"{locality} 학생이라면 시험 뒤 {primary.label}의 근거와 {secondary.label}의 소요 시간을 같은 주간표에서 비교해 다음 비중을 정하세요.",
                ),
                "four-week-plan": (
                    f"{locality} 계획표에는 ‘{primary.action}’을 시작할 날짜와 ‘{secondary.checkpoint}’를 다시 물을 날짜를 나란히 적어 실행과 점검이 끊기지 않게 하세요.",
                    f"{locality} 학생의 첫 주 기록에서는 ‘{primary.evidence}’의 변화와 ‘{secondary.action}’의 완료 여부를 다른 칸에 남겨 다음 분량의 근거로 사용하세요.",
                    f"{locality} 계획은 문제 수보다 ‘{primary.checkpoint}’에 혼자 답한 날과 ‘{secondary.checkpoint}’를 다시 볼 날이 연결됐는지로 점검하세요.",
                ),
                "consultation-checklist": (
                    f"{locality} 상담 메모에는 ‘{primary.action}’의 실행 날짜와 ‘{secondary.checkpoint}’의 재확인 날짜를 나란히 적어 답변이 실제 계획으로 이어지는지 확인하세요.",
                    f"{locality}에서 상담할 때는 ‘{primary.consult_question}’라고 묻고, 답을 학생의 최근 자료에 적용할 수 있는지까지 확인하세요.",
                    f"{locality} 상담 뒤에는 {primary.label}의 첫 행동과 {secondary.label}의 확인일을 학생이 직접 설명하게 해 추상적인 안내와 실행 계획을 구분하세요.",
                ),
            }.get(
                key,
                (
                    f"{locality} 학생의 최근 자료에서는 ‘{primary.evidence}’의 위치와 ‘{secondary.checkpoint}’의 재확인일을 함께 남겨 이 판단을 실제 계획에 연결하세요.",
                ),
            )
            note = stable_pick(f"{record['title']}|{key}", "high-section-note", notes)
            section["paragraphs"][-1] = clean(f"{section['paragraphs'][-1]} {note}")


def individualize_duplicate_high_math_sections(
    records_by_category: dict[str, list[dict[str, object]]],
) -> None:
    """Resolve whole-section collisions with a useful mathematical decision note."""

    records = records_by_category.get("고등수학학원", [])
    occurrences: dict[str, list[tuple[dict[str, object], dict[str, object]]]] = defaultdict(list)
    for record in records:
        for section in record["sections"]:
            section_text = clean(" ".join([str(section["heading"]), *section["paragraphs"]]))
            occurrences[section_text].append((record, section))

    for references in occurrences.values():
        if len(references) < 2:
            continue
        for record, section in references:
            center = record["center"]
            profile = record["high_math_profile"]
            primary, secondary = profile.intents[:2]
            locality = str(center["locality"])
            key = str(section["key"])
            notes = {
                "exam-strategy": (
                    f"{locality} 학생의 시험표에는 학교 범위 완료일과 모의고사 새 문제 확인일을 나누고, {primary.label} 풀이에서 바꿀 행동을 한 줄로 적으세요.",
                    f"{locality} 시험 계획을 검토할 때는 {primary.label}의 문제 번호와 {secondary.label}의 검산 결과를 다른 칸에 남겨 다음 비중을 정하세요.",
                    f"{locality} 학생이라면 내신 뒤 {primary.label} 풀이와 모의고사에서의 {secondary.label} 기록을 대조해 공통 약점만 다음 주에 남기세요.",
                ),
                "four-week-plan": (
                    f"{locality} 계획표에는 ‘{primary.action}’을 시작할 날짜와 ‘{secondary.checkpoint}’를 확인할 문제를 나란히 적어 실행과 판단을 연결하세요.",
                    f"{locality} 학생의 첫 주 기록에서는 ‘{primary.evidence}’의 변화와 ‘{secondary.action}’의 완료 여부를 다른 칸에 남기세요.",
                    f"{locality} 계획은 문제 수보다 ‘{primary.checkpoint}’에 답한 날과 {secondary.label}을 새 문제에서 확인한 날이 이어지는지로 점검하세요.",
                ),
                "consultation-checklist": (
                    f"{locality} 상담 메모에는 ‘{primary.action}’을 적용할 문제와 ‘{secondary.checkpoint}’를 다시 확인할 문제를 구분해 적으세요.",
                    f"{locality}에서는 상담 답변을 최근 시험지의 실제 풀이에 적용해 설명할 수 있는지까지 확인하세요.",
                    f"{locality} 상담 뒤에는 {primary.label}의 첫 연습과 {secondary.label}의 확인 문제를 학생이 직접 설명하게 하세요.",
                ),
            }.get(
                key,
                (
                    f"{locality}에서는 {primary.label}의 첫 행동과 {secondary.label}의 재확인 기준을 서로 다른 칸에 남겨 이 판단을 실제 계획에 연결하세요.",
                ),
            )
            note = stable_pick(f"{record['title']}|{key}", "high-math-section-note", notes)
            section["paragraphs"][-1] = clean(f"{section['paragraphs'][-1]} {note}")


def individualize_duplicate_middle_math_sections(
    records_by_category: dict[str, list[dict[str, object]]],
) -> None:
    """Resolve whole-section collisions without adding generic locality prose."""

    records = records_by_category.get("중등수학학원", [])
    occurrences: dict[str, list[tuple[dict[str, object], dict[str, object]]]] = defaultdict(list)
    for record in records:
        for section in record["sections"]:
            section_text = clean(" ".join([str(section["heading"]), *section["paragraphs"]]))
            occurrences[section_text].append((record, section))

    for references in occurrences.values():
        if len(references) < 2:
            continue
        for record, section in references:
            center = record["center"]
            profile = record["middle_math_profile"]
            primary, secondary = profile.intents[:2]
            locality = str(center["locality"])
            key = str(section["key"])
            notes = {
                "diagnostic-evidence": (
                    f"{locality}에서는 {primary.label}의 첫 풀이와 {secondary.label}의 수정 풀이를 다른 칸에 두어 다음 문제 선택의 근거로 사용하세요.",
                    f"{locality} 점검표에는 {primary.label}에서 바꾼 식과 {secondary.label}을 새 문제에서 확인한 결과를 따로 적으세요.",
                    f"{locality} 학생의 기록은 {primary.label}의 오류 위치와 {secondary.label}의 재확인 답을 나누어 다음 순서를 정하세요.",
                ),
                "exam-strategy": (
                    f"{locality} 시험표에는 학교 범위 완료일과 누적 유형 문제 확인일을 나누고 {primary.label}에서 바꿀 행동을 한 줄로 적으세요.",
                    f"{locality} 시험 계획에서는 {primary.label}의 문제 번호와 {secondary.label}의 검산 결과를 다른 칸에 남겨 다음 비중을 정하세요.",
                    f"{locality} 학생이라면 학교 시험 뒤 {primary.label} 풀이와 누적 유형 문제의 {secondary.label} 기록을 대조해 공통 약점만 남기세요.",
                ),
                "four-week-plan": (
                    f"{locality} 계획표에는 ‘{primary.action}’의 시작일과 ‘{secondary.checkpoint}’의 확인일을 나란히 적으세요.",
                    f"{locality} 학생의 첫 주 기록에서는 ‘{primary.evidence}’의 변화와 ‘{secondary.action}’의 완료 여부를 다른 칸에 남기세요.",
                    f"{locality} 계획은 문제 수보다 ‘{primary.checkpoint}’에 답한 날과 {secondary.label}을 새 문제에서 확인한 날이 이어지는지로 점검하세요.",
                ),
                "consultation-checklist": (
                    f"{locality} 상담 메모에는 ‘{primary.action}’을 적용할 문제와 ‘{secondary.checkpoint}’를 확인할 문제를 구분해 적으세요.",
                    f"{locality}에서는 상담 답변을 최근 시험지의 실제 풀이에 적용해 설명할 수 있는지까지 확인하세요.",
                    f"{locality} 상담 뒤에는 {primary.label}의 첫 연습과 {secondary.label}의 확인 문제를 학생이 직접 설명하게 하세요.",
                ),
            }.get(
                key,
                (
                    f"{locality}에서는 {primary.label}의 첫 행동과 {secondary.label}의 재확인 기준을 서로 다른 칸에 남겨 실제 계획에 연결하세요.",
                ),
            )
            note = stable_pick(f"{record['title']}|{key}", "middle-math-section-note", notes)
            section["paragraphs"][-1] = naturalize_middle_math_text(
                clean(f"{section['paragraphs'][-1]} {note}"),
                profile,
            )


def preflight(records_by_category: dict[str, list[dict[str, object]]]) -> dict[str, object]:
    titles: list[str] = []
    metas: list[str] = []
    paragraph_pages: dict[str, set[str]] = defaultdict(set)
    sections: dict[str, str] = {}
    authored_blocks: dict[str, str] = {}
    faq_sets: set[str] = set()
    scenario_sets: set[str] = set()
    middle_english_focuses: set[str] = set()
    middle_math_focuses: set[str] = set()
    high_math_focuses: set[str] = set()
    specialized_slugs = {
        "중등영어학원", "중등수학학원", "고등영어학원", "고등수학학원",
    }
    for slug, records in records_by_category.items():
        if len(records) != EXPECTED_ROWS:
            raise ValueError(f"{slug}: record count={len(records)}")
        for record in records:
            title = str(record["title"])
            titles.append(title)
            meta = str(record["meta"])
            if not 70 <= len(meta) <= 100 or not meta.endswith((".", "요.")):
                raise ValueError(f"invalid meta description: {title} ({len(meta)}): {meta}")
            metas.append(meta)
            if slug == "중등영어학원":
                profile = record["middle_english_profile"]
                focus = str(profile.focus)
                if focus in middle_english_focuses:
                    raise ValueError(f"duplicate middle English focus: {title}: {focus}")
                middle_english_focuses.add(focus)
                quick = str(record["quick"])
                center = record["center"]
                tokens = (
                    *page_particle_tokens(CONFIG_BY_SLUG[slug], center, record["signals"], str(record["student"])),
                    *middle_english_particle_tokens(profile),
                )
                guidance = normalize_generated_value(
                    naturalize_middle_english_tree(high_english_focus_guidance(profile), profile),
                    tokens,
                )
                if focus not in quick or any(sentence not in quick for sentence in guidance):
                    raise ValueError(f"middle English quick answer does not fulfill its focus: {title}")
                focus_sections = sum(
                    focus in clean(" ".join([str(section["heading"]), *section["paragraphs"]]))
                    for section in record["sections"]
                )
                if focus_sections < 2:
                    raise ValueError(f"middle English focus coverage is too thin: {title}")
                intent_sections = 0
                for section in record["sections"]:
                    section_text = clean(" ".join([str(section["heading"]), *section["paragraphs"]]))
                    intent_sections += any(
                        value in section_text
                        for intent in profile.intents
                        for value in (intent.evidence, intent.action, intent.checkpoint)
                    )
                if intent_sections < 2:
                    raise ValueError(f"middle English evidence/action/checkpoint coverage is too thin: {title}")
                if any(str(section["heading"]).count("·") > 3 for section in record["sections"]):
                    raise ValueError(f"middle English heading has excessive middle-dot joins: {title}")
                faq_question, faq_answer = record["faq"][3]
                support, extra = profile.intents[2:4]
                if not all(value in faq_question and value in faq_answer for value in (support.label, extra.label)):
                    raise ValueError(f"middle English FAQ4 intent mismatch: {title}")
            elif slug == "중등수학학원":
                profile = record["middle_math_profile"]
                focus = str(profile.focus)
                if focus in middle_math_focuses:
                    raise ValueError(f"duplicate middle Math focus: {title}: {focus}")
                middle_math_focuses.add(focus)
                quick = str(record["quick"])
                center = record["center"]
                tokens = (
                    *page_particle_tokens(
                        CONFIG_BY_SLUG[slug], center, record["signals"], str(record["student"])
                    ),
                    *middle_math_particle_tokens(profile),
                )
                guidance = normalize_generated_value(
                    naturalize_middle_math_tree(middle_math_focus_guidance(profile), profile),
                    tokens,
                )
                if focus not in quick or any(sentence not in quick for sentence in guidance):
                    raise ValueError(f"middle Math quick answer does not fulfill its focus: {title}")
                focus_sections = sum(
                    focus in clean(" ".join([str(section["heading"]), *section["paragraphs"]]))
                    for section in record["sections"]
                )
                if focus_sections < 2:
                    raise ValueError(f"middle Math focus coverage is too thin: {title}")
                intent_sections = 0
                for section in record["sections"]:
                    section_text = clean(" ".join([str(section["heading"]), *section["paragraphs"]]))
                    intent_sections += any(
                        value in section_text
                        for intent in profile.intents
                        for value in (intent.evidence, intent.action, intent.checkpoint)
                    )
                if intent_sections < 2:
                    raise ValueError(f"middle Math evidence/action/checkpoint coverage is too thin: {title}")
                if any(str(section["heading"]).count("·") > 3 for section in record["sections"]):
                    raise ValueError(f"middle Math heading has excessive middle-dot joins: {title}")
                faq_question, faq_answer = record["faq"][3]
                support, extra = profile.intents[2:4]
                if not all(
                    value in faq_question and value in faq_answer
                    for value in (support.label, extra.label)
                ):
                    raise ValueError(f"middle Math FAQ4 intent mismatch: {title}")
                authored = clean(" ".join((
                    str(record["student"]),
                    str(record["meta"]),
                    str(record["quick"]),
                    *(str(section["heading"]) for section in record["sections"]),
                    *(
                        str(paragraph)
                        for section in record["sections"]
                        for paragraph in section["paragraphs"]
                    ),
                    *(str(question) for question, _answer in record["faq"]),
                    *(str(answer) for _question, answer in record["faq"]),
                    *(str(value) for value in record["scenarios"]),
                )))
                contamination = sorted(set(re.findall(
                    r"(?:영어|어휘|문법|독해|구문|영작|듣기|말하기|수능|모의고사|고[23]|본문\s*변형)",
                    authored,
                )))
                if contamination:
                    raise ValueError(f"middle Math subject/grade contamination: {title}: {contamination}")
            elif slug == "고등영어학원":
                profile = record["high_english_profile"]
                focus = str(profile.focus)
                quick = str(record["quick"])
                center = record["center"]
                tokens = (
                    *page_particle_tokens(CONFIG_BY_SLUG[slug], center, record["signals"], str(record["student"])),
                    *high_english_particle_tokens(profile),
                )
                guidance = normalize_generated_value(high_english_focus_guidance(profile), tokens)
                if focus not in quick or any(sentence not in quick for sentence in guidance):
                    raise ValueError(f"high English quick answer does not fulfill its focus: {title}")
                focus_sections = sum(
                    focus in clean(" ".join([str(section["heading"]), *section["paragraphs"]]))
                    for section in record["sections"]
                )
                if focus_sections < 1:
                    raise ValueError(f"high English focus is missing from the direct answer: {title}")
                intent_sections = 0
                for section in record["sections"]:
                    section_text = clean(" ".join([str(section["heading"]), *section["paragraphs"]]))
                    intent_sections += any(
                        value in section_text
                        for intent in profile.intents
                        for value in (intent.evidence, intent.action, intent.checkpoint)
                    )
                if intent_sections < 2:
                    raise ValueError(f"high English evidence/action/checkpoint coverage is too thin: {title}")
                if any(str(section["heading"]).count("·") > 3 for section in record["sections"]):
                    raise ValueError(f"high English heading has excessive middle-dot joins: {title}")
                faq_question, faq_answer = record["faq"][3]
                support, extra = profile.intents[2:4]
                if not all(value in faq_question and value in faq_answer for value in (support.label, extra.label)):
                    raise ValueError(f"high English FAQ4 intent mismatch: {title}")
            elif slug == "고등수학학원":
                profile = record["high_math_profile"]
                focus = str(profile.focus)
                if focus in high_math_focuses:
                    raise ValueError(f"duplicate high Math focus: {title}: {focus}")
                high_math_focuses.add(focus)
                quick = str(record["quick"])
                center = record["center"]
                tokens = (
                    *page_particle_tokens(CONFIG_BY_SLUG[slug], center, record["signals"], str(record["student"])),
                    *high_math_particle_tokens(profile),
                )
                guidance = normalize_generated_value(high_math_focus_guidance(profile), tokens)
                if focus not in quick or any(sentence not in quick for sentence in guidance):
                    raise ValueError(f"high Math quick answer does not fulfill its focus: {title}")
                focus_sections = sum(
                    focus in clean(" ".join([str(section["heading"]), *section["paragraphs"]]))
                    for section in record["sections"]
                )
                if focus_sections < 1:
                    raise ValueError(f"high Math focus is missing from the direct answer: {title}")
                intent_sections = 0
                for section in record["sections"]:
                    section_text = clean(" ".join([str(section["heading"]), *section["paragraphs"]]))
                    intent_sections += any(
                        value in section_text
                        for intent in profile.intents
                        for value in (intent.evidence, intent.action, intent.checkpoint)
                    )
                if intent_sections < 2:
                    raise ValueError(f"high Math evidence/action/checkpoint coverage is too thin: {title}")
                if any(str(section["heading"]).count("·") > 3 for section in record["sections"]):
                    raise ValueError(f"high Math heading has excessive middle-dot joins: {title}")
                faq_question, faq_answer = record["faq"][3]
                support, extra = profile.intents[2:4]
                if not all(value in faq_question and value in faq_answer for value in (support.label, extra.label)):
                    raise ValueError(f"high Math FAQ4 intent mismatch: {title}")
            page_blocks: list[str] = [str(record["student"])]
            if slug not in specialized_slugs:
                page_blocks.append(str(record["quick"]))
            page_paragraphs: set[str] = set()
            for section in record["sections"]:
                if slug not in specialized_slugs:
                    page_blocks.append(str(section["heading"]))
                section_text = clean(" ".join([str(section["heading"]), *section["paragraphs"]]))
                if section_text in sections:
                    raise ValueError(f"duplicate authored section: {title} / {sections[section_text]}")
                sections[section_text] = title
                for paragraph in section["paragraphs"]:
                    value = clean(paragraph)
                    if value in page_paragraphs:
                        raise ValueError(f"duplicate authored paragraph within page: {title}: {value}")
                    page_paragraphs.add(value)
                    paragraph_pages[value].add(title)
                    page_blocks.append(value)
                if slug not in specialized_slugs:
                    page_blocks.extend(clean(value) for _label, value in section.get("checklist", []))
            for question, answer in record["faq"]:
                if slug not in specialized_slugs:
                    page_blocks.extend((clean(question), clean(answer)))
                else:
                    page_blocks.append(clean(answer))
            page_blocks.extend(clean(value) for value in record["scenarios"])
            page_authored_blocks: set[str] = set()
            for value in page_blocks:
                if value in page_authored_blocks:
                    raise ValueError(f"duplicate authored block within page: {title}: {value}")
                page_authored_blocks.add(value)
                if slug not in specialized_slugs and value in authored_blocks:
                    raise ValueError(f"duplicate authored block: {title} / {authored_blocks[value]}: {value}")
                authored_blocks.setdefault(value, title)
            faq_key = json.dumps(record["faq"], ensure_ascii=False)
            scenario_key = json.dumps(record["scenarios"], ensure_ascii=False)
            if faq_key in faq_sets or scenario_key in scenario_sets:
                raise ValueError(f"duplicate faq/scenario set: {title}")
            faq_sets.add(faq_key)
            scenario_sets.add(scenario_key)
    if len(titles) != len(set(titles)):
        raise ValueError("duplicate titles")
    if len(metas) != len(set(metas)):
        raise ValueError("duplicate meta descriptions")
    paragraph_max_df = max((len(pages) for pages in paragraph_pages.values()), default=0)
    if paragraph_max_df > HIGH_ENGLISH_PARAGRAPH_DF_LIMIT:
        raise ValueError(
            f"authored paragraph document frequency exceeds {HIGH_ENGLISH_PARAGRAPH_DF_LIMIT}: {paragraph_max_df}"
        )
    return {
        "detail_pages": len(titles),
        "unique_titles": len(set(titles)),
        "unique_meta": len(set(metas)),
        "unique_paragraphs": len(paragraph_pages),
        "paragraph_max_df": paragraph_max_df,
        "unique_sections": len(sections),
        "unique_faq_sets": len(faq_sets),
        "unique_scenario_sets": len(scenario_sets),
        "unique_authored_blocks": len(authored_blocks),
    }


def update_selected_sitemap_lastmods(selected_slugs: set[str], lastmod: str) -> int:
    """Refresh only selected subject-detail entries without rewriting site-wide dates."""

    sitemap = ROOT / "sitemap.xml"
    source = sitemap.read_text(encoding="utf-8")
    entry_re = re.compile(
        r"(?P<prefix><url><loc>(?P<loc>.*?)</loc><lastmod>)(?P<date>\d{4}-\d{2}-\d{2})(?P<suffix></lastmod></url>)"
    )
    changed = 0

    def replace(match: re.Match[str]) -> str:
        nonlocal changed
        loc = html.unescape(match.group("loc"))
        parts = unquote(urlsplit(loc).path).strip("/").split("/")
        if len(parts) != 3 or parts[0] != "과목별학원" or parts[1] not in selected_slugs:
            return match.group(0)
        if match.group("date") != lastmod:
            changed += 1
        return match.group("prefix") + lastmod + match.group("suffix")

    updated = entry_re.sub(replace, source)
    if updated != source:
        sitemap.write_text(updated, encoding="utf-8", newline="")
    return changed


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate seven subject academy categories from XLSX source signals.")
    parser.add_argument("--category", action="append", choices=tuple(CONFIG_BY_SLUG), help="Generate only selected category; repeatable.")
    parser.add_argument("--skip-production-files", action="store_true", help="Do not rebuild canonical/sitemap/rss/social metadata.")
    parser.add_argument("--skip-hubs", action="store_true", help="Preserve existing category and root hub files.")
    args = parser.parse_args()
    selected = tuple(CONFIG_BY_SLUG[slug] for slug in args.category) if args.category else CONFIGS
    centers = load_centers()
    assignments = assign_representatives(selected)
    records_by_category: dict[str, list[dict[str, object]]] = {}
    for config in selected:
        records_by_category[config.slug] = build_records(config, centers, load_source_rows(config), assignments[config.slug])
    contextualize_duplicate_paragraphs(records_by_category)
    contextualize_duplicate_faq_answers(records_by_category)
    individualize_frequent_middle_english_paragraphs(records_by_category)
    individualize_frequent_middle_english_faq_answers(records_by_category)
    individualize_duplicate_middle_english_sections(records_by_category)
    individualize_duplicate_middle_math_sections(records_by_category)
    individualize_duplicate_high_english_sections(records_by_category)
    individualize_duplicate_high_math_sections(records_by_category)
    for record in records_by_category.get("중등영어학원", []):
        profile = record["middle_english_profile"]
        tokens = (
            *page_particle_tokens(
                CONFIG_BY_SLUG["중등영어학원"],
                record["center"],
                record["signals"],
                str(record["student"]),
            ),
            *middle_english_particle_tokens(profile),
        )
        record["student"] = normalize_particle_joins(
            naturalize_middle_english_text(str(record["student"]), profile), tokens
        )
        record["meta"] = normalize_particle_joins(
            naturalize_middle_english_text(str(record["meta"]), profile), tokens
        )
        record["quick"] = naturalize_middle_english_text(str(record["quick"]), profile)
        record["sections"] = normalize_generated_value(
            naturalize_middle_english_tree(record["sections"], profile), tokens
        )
        record["faq"] = normalize_generated_value(
            naturalize_middle_english_tree(record["faq"], profile), tokens
        )
        record["scenarios"] = normalize_generated_value(
            naturalize_middle_english_tree(record["scenarios"], profile), tokens
        )
    for record in records_by_category.get("고등영어학원", []):
        profile = record["high_english_profile"]
        tokens = (
            *page_particle_tokens(
                CONFIG_BY_SLUG["고등영어학원"],
                record["center"],
                record["signals"],
                str(record["student"]),
            ),
            *high_english_particle_tokens(profile),
        )
        record["student"] = naturalize_high_english_text(str(record["student"]), profile)
        record["meta"] = naturalize_high_english_text(str(record["meta"]), profile)
        record["quick"] = naturalize_high_english_text(str(record["quick"]), profile)
        record["sections"] = normalize_generated_value(
            naturalize_high_english_tree(record["sections"], profile), tokens
        )
        record["faq"] = normalize_generated_value(
            naturalize_high_english_tree(record["faq"], profile), tokens
        )
        record["scenarios"] = normalize_generated_value(
            naturalize_high_english_tree(record["scenarios"], profile), tokens
        )
    for record in records_by_category.get("중등수학학원", []):
        profile = record["middle_math_profile"]
        tokens = (
            *page_particle_tokens(
                CONFIG_BY_SLUG["중등수학학원"],
                record["center"],
                record["signals"],
                str(record["student"]),
            ),
            *middle_math_particle_tokens(profile),
        )
        record["student"] = normalize_particle_joins(
            naturalize_middle_math_text(str(record["student"]), profile), tokens
        )
        record["meta"] = normalize_particle_joins(
            naturalize_middle_math_text(str(record["meta"]), profile), tokens
        )
        record["quick"] = naturalize_middle_math_text(str(record["quick"]), profile)
        record["sections"] = normalize_generated_value(
            naturalize_middle_math_tree(record["sections"], profile), tokens
        )
        record["faq"] = normalize_generated_value(
            naturalize_middle_math_tree(record["faq"], profile), tokens
        )
        record["scenarios"] = normalize_generated_value(
            naturalize_middle_math_tree(record["scenarios"], profile), tokens
        )
    for record in records_by_category.get("고등수학학원", []):
        profile = record["high_math_profile"]
        tokens = (
            *page_particle_tokens(
                CONFIG_BY_SLUG["고등수학학원"],
                record["center"],
                record["signals"],
                str(record["student"]),
            ),
            *high_math_particle_tokens(profile),
        )
        record["sections"] = normalize_generated_value(record["sections"], tokens)
        record["faq"] = normalize_generated_value(record["faq"], tokens)
    report = preflight(records_by_category)
    from add_national_anchor_tocs import render_page as add_anchor_toc
    from normalize_internal_links_and_social_meta import transform_html

    targeted_postprocess = {
        "pages": 0,
        "anchor_targets": 0,
        "links_rewritten": 0,
        "social_pages": 0,
        "dimension_pages": 0,
    }
    for config in selected:
        records = records_by_category[config.slug]
        category_root = TARGET_ROOT / config.slug
        category_root.mkdir(parents=True, exist_ok=True)
        for index, record in enumerate(records):
            target = category_root / str(record["slug"]) / "index.html"
            target.parent.mkdir(parents=True, exist_ok=True)
            rendered = render_page(config, record, records[(index - 1) % len(records)], records[(index + 1) % len(records)])
            if config.slug in {"중등영어학원", "중등수학학원", "고등영어학원", "고등수학학원"} and target.is_file():
                rendered = preserve_high_english_center_schema(rendered, target.read_text(encoding="utf-8"))
            if config.slug in {"중등영어학원", "중등수학학원", "고등영어학원", "고등수학학원"}:
                rendered, toc_count = add_anchor_toc(rendered)
                if toc_count != 9:
                    raise ValueError(f"{target}: expected 9 anchor targets, found {toc_count}")
                rendered, page_stats = transform_html(rendered, target, ROOT)
                targeted_postprocess["pages"] += 1
                targeted_postprocess["anchor_targets"] += toc_count
                for key in ("links_rewritten", "social_pages", "dimension_pages"):
                    targeted_postprocess[key] += page_stats[key]
            target.write_text(rendered, encoding="utf-8")
        if not args.skip_hubs:
            (category_root / "index.html").write_text(render_category_hub(config, records), encoding="utf-8")

    # The legacy three-category generator owns the established root-hub
    # markup. Its hub now reads the shared ten-entry catalog, so reuse only
    # that pure renderer without invoking its page-generation main().
    if not args.skip_hubs:
        original_argv = sys.argv[:]
        try:
            sys.argv = [sys.argv[0], "high"]
            import generate_highschool_korean_english_math as legacy_generator
        finally:
            sys.argv = original_argv
        (TARGET_ROOT / "index.html").write_text(legacy_generator.render_root_hub(), encoding="utf-8")

    if not args.skip_production_files:
        from datetime import date

        report["sitemap_lastmods_updated"] = update_selected_sitemap_lastmods(
            {config.slug for config in selected},
            date.today().isoformat(),
        )
    report["targeted_postprocess"] = targeted_postprocess
    report["category_hubs"] = 0 if args.skip_hubs else len(selected)
    report["generated_categories"] = [config.slug for config in selected]
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
